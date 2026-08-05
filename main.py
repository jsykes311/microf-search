from fastapi import FastAPI, HTTPException, Query, Depends, Body, BackgroundTasks, UploadFile, File, Form
from pydantic import BaseModel as _BaseModel
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, RedirectResponse, JSONResponse, HTMLResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.requests import Request as _Request
from starlette.middleware.base import BaseHTTPMiddleware
from typing import Optional, List
import httpx
import os
import csv
import io
import asyncio
import secrets
import time as _time
import base64
import urllib.parse
from datetime import datetime, timedelta, date
from dotenv import load_dotenv
import json
import io as _io
from collections import defaultdict
try:
    import pandas as _pd
except ImportError:
    _pd = None
import aiosmtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders as _enc
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import uuid as _uuid

load_dotenv()

app = FastAPI()

# Return unhandled exceptions as JSON so errors are readable in the browser/curl
from fastapi.responses import JSONResponse
from fastapi.requests import Request
import traceback

@app.exception_handler(Exception)
async def _unhandled_exception_handler(request: Request, exc: Exception):
    tb = traceback.format_exc()
    print(f"[ERROR] {request.url}\n{tb}")
    return JSONResponse(status_code=500, content={"detail": str(exc), "type": type(exc).__name__})

# Allow Azure Static Web Apps + SharePoint + localhost for dev.
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8000",
        "http://127.0.0.1:8000",
    ],
    allow_origin_regex=r"https://.*\.(azurestaticapps\.net|sharepoint\.com|sharepoint\.us)$",
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

# ── Microsoft OAuth (Azure AD) ────────────────────────────────────────────────
# Set these on Render. AZURE_CLIENT_ID is required to enable auth;
# if missing the app runs open (local dev mode).
_AZ_CLIENT_ID  = os.getenv("AZURE_CLIENT_ID",     "")
_AZ_CLIENT_SEC = os.getenv("AZURE_CLIENT_SECRET",  "")
_AZ_TENANT_ID  = os.getenv("AZURE_TENANT_ID",      "common")
_SESSION_SECRET = os.getenv("SESSION_SECRET",      "dev-secret-please-set-me")
_ALLOWED_DOMAIN = os.getenv("ALLOWED_EMAIL_DOMAIN", "microf.com")
_COOKIE_MAX_AGE = 8 * 3600   # 8 hours

_signer = URLSafeTimedSerializer(_SESSION_SECRET)

# Separate token for internal/automated endpoints (no special chars needed).
# Set SYNC_TOKEN on Render and in GitHub Secrets.
_SYNC_TOKEN = os.getenv("SYNC_TOKEN", "")
_WEBHOOK_TOKEN = (os.getenv("AC_WEBHOOK_TOKEN") or
                  os.getenv("DEAL_WEBHOOK_TOKEN") or "")

# ── Scheduled email reports ───────────────────────────────────────────────
# Set these env vars on Render to enable report delivery.
# SMTP_USER + REPORT_RECIPIENTS are required; everything else has defaults.
_ANTHROPIC_KEY = os.getenv("ANTHROPIC_API_KEY", "")

_SMTP_HOST  = os.getenv("SMTP_HOST",      "smtp.gmail.com")
_SMTP_PORT  = int(os.getenv("SMTP_PORT",  "587"))
_SMTP_USER  = os.getenv("SMTP_USER",      "")
_SMTP_PASS  = os.getenv("SMTP_PASS",      "")
_SMTP_FROM  = os.getenv("SMTP_FROM_NAME", "Microf Reports")
_RECIPIENTS = [r.strip() for r in os.getenv("REPORT_RECIPIENTS", "").split(",") if r.strip()]

# ── Admin / Scheduler ─────────────────────────────────────────────────────
_ADMIN_EMAILS       = {e.strip().lower() for e in os.getenv("ADMIN_EMAIL",       "jsykes@microf.com,bsanders@microf.com,parnold@microf.com,lfutrell@microf.com").split(",") if e.strip()}
_SALES_ADMIN_EMAILS = {e.strip().lower() for e in os.getenv("SALES_ADMIN_EMAIL", "parnold@microf.com").split(",") if e.strip()}
_CONTRACTOR_SUPPORT_EMAILS = {e.strip().lower() for e in os.getenv("CONTRACTOR_SUPPORT_EMAIL", "elove@microf.com,cristian.perez@microf.com,rlugo@microf.com").split(",") if e.strip()}
_ONBOARDING_EMAILS  = {e.strip().lower() for e in os.getenv("ONBOARDING_EMAIL",  "tbillings@microf.com,cristian.perez@microf.com,elove@microf.com").split(",") if e.strip()}
_ACCT_MGMT_EMAILS   = {e.strip().lower() for e in os.getenv("ACCT_MGMT_EMAIL",   "jtiplady@microf.com,ajones@microf.com,lfutrell@microf.com,abergen@microf.com,wneely@microf.com,charden@microf.com,zolbrys@microf.com,rolbrys@microf.com,ctwiggs@microf.com").split(",") if e.strip()}
# All groups that can access the Apps tab
_APPS_EMAILS        = _ADMIN_EMAILS | _CONTRACTOR_SUPPORT_EMAILS | _ONBOARDING_EMAILS | _ACCT_MGMT_EMAILS | _SALES_ADMIN_EMAILS | {e.strip().lower() for e in os.getenv("APPS_EMAIL", "").split(",") if e.strip()}
_SCHEDULES_FILE  = os.getenv("SCHEDULES_FILE",  os.path.join(os.path.dirname(__file__), "schedules.json"))
_APEX_FILE       = os.getenv("APEX_DATA_FILE",  os.path.join(os.path.dirname(__file__), "apex_data.json"))
_LOGINS_FILE     = os.getenv("LOGINS_FILE",     os.path.join(os.path.dirname(__file__), "logins.json"))
_scheduler      = AsyncIOScheduler()
_schedules: dict = {}   # job_id → schedule dict

# ── Last-login tracker ────────────────────────────────────────────────────
# Persisted to logins.json so it survives restarts.
# Format: { "email@microf.com": "2026-05-18T14:32:00+00:00", ... }
_last_login: dict = {}

def _load_logins() -> None:
    global _last_login
    try:
        with open(_LOGINS_FILE) as f:
            _last_login = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        _last_login = {}

def _save_logins() -> None:
    try:
        with open(_LOGINS_FILE, "w") as f:
            json.dump(_last_login, f, indent=2)
    except Exception as e:
        print(f"[logins] failed to save: {e}")

def _record_last_seen(email: str) -> None:
    """Update last-seen for an authenticated user. Writes to disk only when the date changes."""
    if not email:
        return
    from datetime import timezone as _tz
    now_str = datetime.now(_tz.utc).strftime("%Y-%m-%dT%H:%M:%S+00:00")
    existing = _last_login.get(email, "")
    # Only save to disk if the date portion has changed (at most one write per user per day)
    if existing[:10] != now_str[:10]:
        _last_login[email] = now_str
        _save_logins()
    elif not existing:
        _last_login[email] = now_str

_load_logins()


def _load_schedules_from_disk():
    """Load schedules from file, falling back to SCHEDULES_JSON env var (for Render deploys)."""
    print(f"[scheduler] schedules file path: {_SCHEDULES_FILE} (exists={os.path.exists(_SCHEDULES_FILE)})")
    saved = None
    if os.path.exists(_SCHEDULES_FILE):
        try:
            with open(_SCHEDULES_FILE) as f:
                saved = json.load(f)
            print(f"[scheduler] Loaded {len(saved)} schedule(s) from disk")
        except Exception as e:
            print(f"[scheduler] Failed to load from disk: {e}")
    if saved is None:
        env_json = os.getenv("SCHEDULES_JSON", "").strip()
        if env_json:
            try:
                saved = json.loads(env_json)
                print(f"[scheduler] Loaded {len(saved)} schedule(s) from SCHEDULES_JSON env var")
                # Write to disk so future in-process saves work normally
                _save_schedules_to_disk_raw(saved)
            except Exception as e:
                print(f"[scheduler] Failed to parse SCHEDULES_JSON env var: {e}")
    if saved:
        for s in saved:
            _register_schedule(s, persist=False)


def _save_schedules_to_disk_raw(data: list):
    """Write a schedule list to disk (used by both save paths)."""
    try:
        os.makedirs(os.path.dirname(_SCHEDULES_FILE), exist_ok=True)
        with open(_SCHEDULES_FILE, "w") as f:
            json.dump(data, f, indent=2)
        print(f"[scheduler] Saved {len(data)} schedule(s) to {_SCHEDULES_FILE}")
    except Exception as e:
        print(f"[scheduler] Failed to write schedules to disk ({_SCHEDULES_FILE}): {e}")

def _save_schedules_to_disk():
    _save_schedules_to_disk_raw(list(_schedules.values()))


def _register_schedule(s: dict, persist: bool = True):
    job_id = s["id"]
    freq   = s["frequency"]       # daily | weekly | monthly
    hour   = int(s.get("hour", 9))
    minute = int(s.get("minute", 0))

    tz = "America/New_York"
    if freq == "daily":
        trigger = CronTrigger(hour=hour, minute=minute, timezone=tz)
    elif freq == "weekly":
        trigger = CronTrigger(day_of_week=s.get("day_of_week", "mon"), hour=hour, minute=minute, timezone=tz)
    else:  # monthly
        trigger = CronTrigger(day=int(s.get("day_of_month", 1)), hour=hour, minute=minute, timezone=tz)

    report_type = s["report_type"]
    recipients  = s["recipients"]
    period      = s.get("period") or None

    async def _run():
        job = _REPORT_JOBS.get(report_type)
        if not job:
            print(f"[scheduler] Unknown report type: {report_type}")
            return
        try:
            kwargs = {"recipients": recipients}
            if period:
                kwargs["preset"] = period
            await job(**kwargs)
            print(f"[scheduler] Sent '{report_type}' → {recipients}")
        except Exception as exc:
            print(f"[scheduler] Job '{report_type}' failed: {exc}")

    _scheduler.add_job(_run, trigger=trigger, id=job_id, replace_existing=True)
    _schedules[job_id] = s
    if persist:
        _save_schedules_to_disk()


def _require_admin(request: _Request):
    email = _get_session_email(request)
    if not _AZ_CLIENT_ID:          # no Azure → local dev, allow all
        return "local"
    if not email or email.lower() not in _ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Admin only")
    return email


def _require_welcome(request: _Request):
    """Admin or Onboarding group can access the Welcome Email tool."""
    email = _get_session_email(request)
    if not _AZ_CLIENT_ID:
        return "local"
    em = (email or "").lower()
    if not em or (em not in _ADMIN_EMAILS and em not in _ONBOARDING_EMAILS):
        raise HTTPException(status_code=403, detail="Access restricted")
    return email


def _redirect_uri() -> str:
    base = os.getenv("RENDER_EXTERNAL_URL", "http://localhost:8000").rstrip("/")
    return f"{base}/auth/callback"


def _get_session_email(request: _Request) -> Optional[str]:
    """Return the authenticated email from the session cookie, or None."""
    token = request.cookies.get("session")
    if not token:
        return None
    try:
        return _signer.loads(token, max_age=_COOKIE_MAX_AGE)
    except (BadSignature, SignatureExpired):
        return None


def require_auth(request: _Request):
    """Dependency: returns the current user email. Middleware enforces blocking."""
    if not _AZ_CLIENT_ID:          # no Azure configured → open (local dev)
        return "dev@microf.com"
    return _get_session_email(request) or "unknown"


class _MSAuthMiddleware(BaseHTTPMiddleware):
    """Block unauthenticated requests. Redirects pages → /login, 401s for APIs."""
    _PUBLIC = {"/login", "/auth/start", "/auth/callback", "/logout", "/health",
               "/api/health", "/api/dealer-index/status", "/dealer-locator", "/dealer-locator-beta",
               "/api/accounts/nearest", "/api/accounts/by-state",
               "/webhook/deal-created", "/webhook/debug-sp", "/webhook/reset-sp-file"}

    async def dispatch(self, request: _Request, call_next):
        path = request.url.path
        # Always allow public paths and static assets
        if path in self._PUBLIC or path.startswith("/static"):
            return await call_next(request)
        # Dev mode — no Azure client ID configured
        if not _AZ_CLIENT_ID:
            return await call_next(request)
        # Automated endpoints: accept SYNC_TOKEN header/query instead of cookie
        bearer = request.headers.get("Authorization", "")
        sync_q = request.query_params.get("token", "")
        if _SYNC_TOKEN and (bearer == f"Bearer {_SYNC_TOKEN}" or sync_q == _SYNC_TOKEN):
            return await call_next(request)
        # Check session cookie
        email = _get_session_email(request)
        if not email:
            if path.startswith("/api/"):
                return JSONResponse(status_code=401, content={"detail": "Not authenticated"})
            return RedirectResponse(url="/login", status_code=302)
        # Record last-seen for authenticated users (write to disk only when date changes)
        _record_last_seen(email)
        return await call_next(request)

# Register auth middleware after the class is defined, after CORS so CORS
# headers are still applied before the 401/redirect response is returned.
app.add_middleware(_MSAuthMiddleware)

app.mount("/static", StaticFiles(directory="static"), name="static")


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.get("/login")
async def login_page():
    return FileResponse("static/login.html")


@app.get("/auth/start")
async def auth_start():
    """Redirect browser to Microsoft login."""
    params = {
        "client_id":     _AZ_CLIENT_ID,
        "response_type": "code",
        "redirect_uri":  _redirect_uri(),
        "scope":         "openid email profile",
        "response_mode": "query",
    }
    ms_url = (
        f"https://login.microsoftonline.com/{_AZ_TENANT_ID}"
        f"/oauth2/v2.0/authorize?{urllib.parse.urlencode(params)}"
    )
    return RedirectResponse(url=ms_url)


@app.get("/auth/callback")
async def auth_callback(
    code:  Optional[str] = Query(None),
    error: Optional[str] = Query(None),
):
    """Exchange code for token, validate @microf.com, set session cookie."""
    if error or not code:
        return RedirectResponse(url="/login?error=cancelled")

    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            f"https://login.microsoftonline.com/{_AZ_TENANT_ID}/oauth2/v2.0/token",
            data={
                "client_id":     _AZ_CLIENT_ID,
                "client_secret": _AZ_CLIENT_SEC,
                "code":          code,
                "redirect_uri":  _redirect_uri(),
                "grant_type":    "authorization_code",
                "scope":         "openid email profile",
            },
        )
    token_data = resp.json()

    # Decode the id_token JWT (middle segment) — no signature verification needed
    # since we got it directly from Microsoft over TLS.
    id_token = token_data.get("id_token", "")
    try:
        seg = id_token.split(".")[1]
        seg += "=" * (-len(seg) % 4)           # re-pad base64
        payload = json.loads(base64.urlsafe_b64decode(seg))
    except Exception:
        return RedirectResponse(url="/login?error=token")

    email = (payload.get("email") or payload.get("preferred_username", "")).lower()

    if not email.endswith(f"@{_ALLOWED_DOMAIN}"):
        return RedirectResponse(url=f"/login?error=domain&email={urllib.parse.quote(email)}")

    _record_last_seen(email)

    session_token = _signer.dumps(email)
    response = RedirectResponse(url="/search", status_code=302)
    response.set_cookie(
        "session", session_token,
        max_age=_COOKIE_MAX_AGE, httponly=True, samesite="lax", secure=True,
    )
    return response


@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("session")
    return response


@app.get("/search")
async def search_page(_: None = Depends(require_auth)):
    return FileResponse("static/search.html")

@app.get("/dealer-locator")
@app.get("/dealer-locator-beta")
async def dealer_locator_page():
    """Public-facing dealer locator — no auth required."""
    return FileResponse("static/dealer-locator.html")

@app.on_event("startup")
async def _startup():
    """Kick off background tasks.
    Dealer index runs immediately (uses CF data, no SLP dependency at this stage).
    SLP-dependent indexes (location, SLP state) are triggered by _slp_cache_loop after
    the first successful SLP load — avoids a race on the _slp_cache_lock at boot.
    """
    asyncio.create_task(_build_dealer_id_index())
    # asyncio.create_task(_keep_alive())  # disabled — not needed on paid tier
    asyncio.create_task(_slp_cache_loop())  # waits 90s, then fetches SLPs and kicks off location/state indexes
    asyncio.create_task(_lc_cache_loop())   # waits 90s then builds last-contacted cache
    asyncio.create_task(_ta_cache_loop())   # waits 90s then caches raw notes+activity for team report
    asyncio.create_task(_acct_cf_cache_loop())  # waits 90s, then keeps account custom-field cache warm
    _load_schedules_from_disk()
    _scheduler.start()
    print(f"[scheduler] Started with {len(_schedules)} job(s)")

async def _keep_alive() -> None:
    """Ping this app's own health endpoint every 10 minutes to prevent Render from
    spinning down the instance due to inactivity."""
    import os as _os
    self_url = _os.getenv("RENDER_EXTERNAL_URL", "").rstrip("/")
    if not self_url:
        print("[keep-alive] RENDER_EXTERNAL_URL not set — skipping keep-alive pings")
        return
    await asyncio.sleep(120)   # wait 2 min after boot before first ping
    while True:
        try:
            async with httpx.AsyncClient(timeout=10) as _hc:
                await _hc.get(f"{self_url}/api/dealer-index/status")
            print("[keep-alive] pinged OK")
        except Exception as _e:
            print(f"[keep-alive] ping failed: {_e}")
        await asyncio.sleep(600)   # 10 minutes

@app.get("/api/dealer-index/refresh")
async def dealer_index_refresh(_: None = Depends(require_auth)):
    """Manually trigger a full rebuild of the dealer_id ↔ account index."""
    asyncio.create_task(_build_dealer_id_index())
    return {"status": "rebuild started", "accounts_indexed": len(_dealer_id_index)}

@app.get("/api/dealer-index/status")
async def dealer_index_status():
    """Return current index size and age."""
    age = int(_time.time() - _dealer_index_ts) if _dealer_index_ts else None
    return {
        "indexed_dealers":   len(_dealer_id_index),
        "indexed_accounts":  len(_account_to_dealer),
        "indexed_platforms": len(_account_to_platform),
        "indexed_bdrs":      len(_account_to_bdr),
        "age_seconds":       age,
        "last_error":        _dealer_index_error or None,
    }

@app.get("/api/dealer-index/diagnose")
async def dealer_index_diagnose(_: None = Depends(_require_admin)):
    """Fetch first page of accountCustomFieldData and first SLP record raw — debug only."""
    cf_page  = await ac_get("accountCustomFieldData", {"limit": 5, "offset": 0})
    slp_page = await ac_get(f"customObjects/records/{SLP_SCHEMA_ID}", {"limit": 1})
    slp_rec  = slp_page.get("records", [{}])[0] if slp_page.get("records") else {}
    return {
        "cf_total":    cf_page.get("meta", {}).get("total"),
        "cf_sample":   cf_page.get("accountCustomFieldData", [])[:3],
        "slp_sample":  slp_rec,
        "index_sizes": {
            "dealers":   len(_account_to_dealer),
            "platforms": len(_account_to_platform),
            "bdrs":      len(_account_to_bdr),
        },
    }

AC_BASE_URL = (os.getenv("AC_BASE_URL") or os.getenv("PROD_URL", "")).rstrip("/")
AC_API_KEY  = os.getenv("AC_API_KEY") or os.getenv("PROD_KEY", "")
HEADERS     = {"Api-Token": AC_API_KEY, "Content-Type": "application/json"}

import re as _re
import time as _time
_subdomain  = _re.match(r"https?://([^.]+)", AC_BASE_URL)
AC_UI_BASE  = f"https://{_subdomain.group(1)}.activehosted.com" if _subdomain else ""

def ac_account_url(account_id) -> str:
    return f"{AC_UI_BASE}/app/accounts/{account_id}" if account_id else ""

def ac_contact_url(contact_id) -> str:
    return f"{AC_UI_BASE}/app/contacts/{contact_id}" if contact_id else ""

# ── Account custom field metadata (cached) ──────────────────────────────────
_cf_meta_cache: dict = {}   # str(field_id) -> label
_cf_meta_ts: float   = 0.0

# ── Dealer ID ↔ Account index (built at startup, refreshed hourly) ───────────
_dealer_id_index:  dict  = {}   # dealer_id (str) → {"id": account_id, "name": account_name}
_account_to_dealer: dict = {}   # account_id (str) → dealer_id (str) from account CF18
_account_to_slp_dealer: dict = {} # account_id (str) → dealer_id (str) from SLP record (authoritative for SLP reports)
_account_to_platform: dict = {} # account_id (str) → platform/Dealer Program (customfield 29)
_account_to_bdr: dict = {}      # account_id (str) → Assigned BDR (customfield 119)
_account_to_name: dict = {}     # account_id (str) → account name
_account_to_owner: dict = {}    # account_id (str) → owner user_id (str)
_account_to_states: dict = {}   # account_id (str) → "TX,FL,GA" (customfield 22)
_account_to_zip: dict = {}      # account_id (str) → postal code (customfield 6)
_account_to_city: dict = {}     # account_id (str) → city (customfield 4)
_account_to_state_prov: dict = {}  # account_id (str) → state/province (customfield 5)
_account_to_phone: dict = {}    # account_id (str) → phone number (customfield 11)
_account_to_website: dict = {}  # account_id (str) → website (customfield 39)
_account_to_address: dict = {}  # account_id (str) → address 1 (customfield 2)
_account_to_last_app: dict = {} # account_id (str) → last app date (SLP last-app-date; CF140 fallback)
_account_to_last_rpa: dict = {} # account_id (str) → last RPA date (SLP last-rpa-date; CF38 fallback)
_account_to_type: dict = {}     # account_id (str) → account type (customfield 76)
_account_to_region: dict = {}   # account_id (str) → sales region (customfield 23)
_account_to_dba: dict = {}      # account_id (str) → DBA Name (customfield 15)
_account_to_status: dict = {}   # account_id (str) → Account Status (customfield 19)
_account_to_tax_id: dict = {}   # account_id (str) → Vendor Tax-ID (customfield 40)
_account_to_group: dict = {}    # account_id (str) → Group Name (CF146)
_account_to_legal_name: dict = {}       # account_id (str) → Account Name (Legal Business Name) CF36
_account_to_revenue: dict = {}          # account_id (str) → Annual Revenue dropdown value CF9
_account_to_strategic_partners: dict = {} # account_id (str) → Strategic Partners multiselect CF132
_account_to_contractor_reactivation: dict = {} # account_id (str) → "Yes" or "" CF32
_account_to_reactivation_date: dict = {}       # account_id (str) → Reactivation Date CF28
_account_to_oracle_id: dict = {}               # account_id (str) → Oracle Producer ID (CF118)
_account_to_activation_date: dict = {}         # account_id (str) → contractor-activated-date from SLP
_account_to_slp_states: dict = {}              # account_id (str) → doing-business-in-states from SLP
_user_id_to_name: dict = {}     # AC user_id (str) → "First Last"
_program_to_accounts: dict = {} # lowercase(dealer_program) → set of account_ids
_dealer_index_ts:  float = 0.0
_dealer_index_error: str = ""   # last build error message, for /api/dealer-index/status

async def _get_account_cf_meta() -> dict:
    """Return {str(field_id): label} cached for 1 hour."""
    global _cf_meta_ts
    if _cf_meta_cache and (_time.time() - _cf_meta_ts) < 3600:
        return _cf_meta_cache
    data = await ac_get("accountCustomFieldMeta", {"limit": 200})
    _cf_meta_cache.clear()
    for f in data.get("accountCustomFieldMeta", []):
        fid = str(f.get("id", ""))
        _cf_meta_cache[fid] = f.get("fieldLabel") or f.get("perstag") or f"field_{fid}"
    _cf_meta_ts = _time.time()
    return _cf_meta_cache

def _normalize_platform(val: str) -> str:
    """Normalize platform/program display names for reporting.
    Microf (LTO Only), LTO → Microf
    Optimus 2.0, OPTIMUS 3.0, etc. → OPTIMUS
    """
    v = val.strip().lower()
    if "optimus" in v:
        return "OPTIMUS"
    if "microf" in v or v == "lto":
        return "Microf"
    return val.strip()


def norm(x) -> str:
    """Normalize a string for case/whitespace-insensitive comparisons."""
    return (x or "").strip().lower()

def norm_id(x) -> str:
    """Normalize a dealer ID: strip whitespace and leading zeros."""
    return str(x or "").strip().lstrip("0")

def get_account_id(slp) -> str | None:
    """Extract account ID from an SLP record, handling all known AC relationship formats."""
    rel = slp.get("relationships") or {}

    # Case 1: simple list  {"account": ["123", ...]}
    if isinstance(rel.get("account"), list) and rel["account"]:
        return str(rel["account"][0])

    # Case 2: plural key   {"accounts": ["123", ...]}
    if isinstance(rel.get("accounts"), list) and rel["accounts"]:
        return str(rel["accounts"][0])

    # Case 3: nested dict  {"account": {"data": [{"id": "123"}, ...]}}
    acct = rel.get("account")
    if isinstance(acct, dict):
        data = acct.get("data")
        if isinstance(data, list) and data:
            return str(data[0].get("id"))

    # Case 4: nested plural {"accounts": {"data": [{"id": "123"}, ...]}}
    accts = rel.get("accounts")
    if isinstance(accts, dict):
        data = accts.get("data")
        if isinstance(data, list) and data:
            return str(data[0].get("id"))

    return None


def _extract_cf_value(cf: dict) -> str:
    """Read the first non-empty value across all custom field value types."""
    for key in ("custom_field_text_value", "custom_field_date_value",
                "custom_field_datetime_value", "custom_field_number_value",
                "custom_field_currency_value"):
        v = cf.get(key)
        if v and v not in ("0000-00-00", "0000-00-00 00:00:00"):
            return str(v).strip()
    return ""

async def _build_dealer_id_index() -> None:
    """Build dealer_id ↔ account index using the bulk accountCustomFieldData endpoint.
    Phase 1: page through all CF data concurrently (20 pages at a time) to extract
             customFieldId-18 (dealer ID) records — ~10-15 s for ~190k records.
    Phase 2: paginate accounts to get names.
    Runs on server startup; re-triggered via /api/dealer-index/refresh."""
    global _dealer_index_ts, _dealer_index_error
    DEALER_CF_ID   = 18    # customFieldId for "Parent Dealer ID"
    # PLATFORM_CF_ID 29 (Dealer Program) deleted from PROD — channel lives on SLP records now
    BDR_CF_ID      = 119   # customFieldId for "Assigned BDR"
    STATES_CF_ID   = 22    # customFieldId for "Doing Business in States"
    ZIP_CF_ID      = 6     # customFieldId for "Postal Code"
    CITY_CF_ID     = 4     # customFieldId for "City"
    STATE_PROV_CF  = 5     # customFieldId for "State/Province"
    PHONE_CF_ID    = 11    # customFieldId for "Phone Number"
    WEBSITE_CF_ID  = 39    # customFieldId for "Website"
    ADDRESS_CF_ID  = 2     # customFieldId for "Address 1"
    LAST_RPA_CF_ID = 38    # customFieldId for "Last RPA Date" (stays in PROD)
    # NOTE: Last App Date lives only on SLP records (last-app-date field), NOT on accounts
    ACCT_TYPE_CF   = 76    # customFieldId for "Account Type"
    REGION_CF_ID   = 23    # customFieldId for "Sales Region"
    DBA_NAME_CF    = 15    # customFieldId for "DBA Name"
    ACCT_STATUS_CF = 19    # customFieldId for "Account Status"
    VENDOR_TAX_CF  = 40    # customFieldId for "Vendor Tax-ID"
    GROUP_NAME_CF  = 146   # customFieldId for "Group Name"
    LEGAL_NAME_CF  = 36    # customFieldId for "Account Name (Legal Business Name)"
    REVENUE_CF     = 9     # customFieldId for "Annual Revenue" (dropdown)
    STRAT_PART_CF  = 132   # customFieldId for "Strategic Partners" (multiselect)
    CONTRACTOR_REACT_CF = 32  # customFieldId for "Contractor Reactivation" (checkbox)
    REACT_DATE_CF  = 28    # customFieldId for "Reactivation Date"
    ORACLE_CF_ID   = 118   # customFieldId for "Oracle Producer ID"
    CF_PAGE        = 1000  # 1000 records/page → ~190 pages instead of ~1900
    CONCURRENCY    = 8     # 8 concurrent requests → index builds in ~10s instead of ~5min

    try:
        print("[dealer-index] Starting build…")

        # ── Phase 1: bulk accountCustomFieldData ──────────────────────────
        first_page = await ac_get("accountCustomFieldData", {"limit": CF_PAGE, "offset": 0})
        total_cf   = int(first_page.get("meta", {}).get("total", 0))
        print(f"[dealer-index] {total_cf} CF records total, scanning…")

        acct_to_dealer:     dict = {}
        acct_to_platform:   dict = {}
        acct_to_bdr:        dict = {}
        acct_to_states:     dict = {}
        acct_to_zip:        dict = {}
        acct_to_city:       dict = {}
        acct_to_state_prov: dict = {}
        acct_to_phone:      dict = {}
        acct_to_website:    dict = {}
        acct_to_address:    dict = {}
        acct_to_last_app:   dict = {}
        acct_to_last_rpa:   dict = {}
        acct_to_type:       dict = {}
        acct_to_region:     dict = {}
        acct_to_dba:        dict = {}
        acct_to_status:     dict = {}
        acct_to_tax_id:     dict = {}
        acct_to_group:      dict = {}
        acct_to_legal_name: dict = {}
        acct_to_revenue:    dict = {}
        acct_to_strat_part: dict = {}
        acct_to_react:      dict = {}
        acct_to_react_date: dict = {}
        acct_to_oracle_id:  dict = {}

        def _ingest(items: list) -> None:
            for item in items:
                cf_id   = int(item.get("customFieldId", 0))
                aid     = str(item.get("accountId", ""))
                raw     = item.get("fieldValue")
                if isinstance(raw, list):
                    raw = ", ".join(str(v) for v in raw if v)
                val = (str(raw) if raw is not None else "").strip()
                if not (aid and val):
                    continue
                if cf_id == DEALER_CF_ID:
                    acct_to_dealer[aid]     = val
                elif cf_id == BDR_CF_ID:
                    acct_to_bdr[aid]        = val
                elif cf_id == STATES_CF_ID:
                    acct_to_states[aid]     = val
                elif cf_id == ZIP_CF_ID:
                    acct_to_zip[aid]        = val
                elif cf_id == CITY_CF_ID:
                    acct_to_city[aid]       = val
                elif cf_id == STATE_PROV_CF:
                    acct_to_state_prov[aid] = val
                elif cf_id == PHONE_CF_ID:
                    acct_to_phone[aid]      = val
                elif cf_id == WEBSITE_CF_ID:
                    acct_to_website[aid]    = val
                elif cf_id == ADDRESS_CF_ID:
                    acct_to_address[aid]    = val
                elif cf_id == ACCT_TYPE_CF:
                    acct_to_type[aid]       = val
                elif cf_id == REGION_CF_ID:
                    acct_to_region[aid]     = val
                elif cf_id == DBA_NAME_CF:
                    acct_to_dba[aid]        = val
                elif cf_id == ACCT_STATUS_CF:
                    acct_to_status[aid]     = val
                elif cf_id == VENDOR_TAX_CF:
                    acct_to_tax_id[aid]     = val
                elif cf_id == GROUP_NAME_CF:
                    acct_to_group[aid]      = val
                elif cf_id == LEGAL_NAME_CF:
                    acct_to_legal_name[aid] = val
                elif cf_id == REVENUE_CF:
                    acct_to_revenue[aid]    = val
                elif cf_id == STRAT_PART_CF:
                    acct_to_strat_part[aid] = val
                elif cf_id == CONTRACTOR_REACT_CF:
                    acct_to_react[aid]      = val   # "Yes" if checked
                elif cf_id == REACT_DATE_CF:
                    acct_to_react_date[aid] = val
                elif cf_id == ORACLE_CF_ID:
                    acct_to_oracle_id[aid]  = val

        _ingest(first_page.get("accountCustomFieldData", []))

        # Remaining pages — fetch in parallel batches
        remaining_offsets = list(range(CF_PAGE, total_cf, CF_PAGE))
        for i in range(0, len(remaining_offsets), CONCURRENCY):
            batch = remaining_offsets[i : i + CONCURRENCY]
            pages = await asyncio.gather(
                *[ac_get("accountCustomFieldData", {"limit": CF_PAGE, "offset": off})
                  for off in batch],
                return_exceptions=True,
            )
            for page in pages:
                if not isinstance(page, Exception):
                    _ingest(page.get("accountCustomFieldData", []))

        print(f"[dealer-index] {len(acct_to_dealer)} dealer IDs, "
              f"{len(acct_to_platform)} platforms, {len(acct_to_bdr)} BDRs indexed; "
              f"fetching account names…")

        # ── Phase 2: paginate accounts for names + fetch AC users ────────
        all_accounts = await ac_get_all("accounts", "accounts", {})
        acct_to_name  = {str(a.get("id", "")): a.get("name", "")              for a in all_accounts}
        acct_to_owner = {str(a.get("id", "")): str(a.get("owner", "") or "")  for a in all_accounts}
        print(f"[dealer-index] {len(all_accounts)} account names loaded")

        # ── Phase 3: fetch AC users for owner name lookup ─────────────────
        try:
            users_resp = await ac_get("users", {"limit": 100})
            new_user_map: dict = {}
            for u in users_resp.get("users", []):
                uid  = str(u.get("id", ""))
                name = f"{u.get('firstName', '')} {u.get('lastName', '')}".strip()
                if uid:
                    new_user_map[uid] = name or u.get("email", uid)
            _user_id_to_name.clear(); _user_id_to_name.update(new_user_map)
            print(f"[dealer-index] {len(new_user_map)} AC users loaded")
        except Exception as _ue:
            print(f"[dealer-index] user fetch failed: {_ue}")

        # ── Phase 4: last-app/rpa dates from SLP records ─────────────────────
        # SLP fields last-app-date and last-rpa-date are the sole source of truth.
        # last-app-date exists only on SLP, not on accounts. last-rpa-date (CF38)
        # is also supplemented here — SLP value wins if more recent.
        try:
            # Non-blocking snapshot — never await get_slp_cache() here because
            # the SLP loop holds _slp_cache_lock during its fetch, which would
            # deadlock Phase 4 and prevent _account_to_name from ever being set.
            # If SLP isn't loaded yet, Phase 4 is skipped (non-fatal); the
            # _update_app_rpa_from_slp_cache() call after each SLP refresh
            # keeps these fields current once SLP data arrives.
            slp_recs_for_dates = list(_slp_cache_records)

            def _slp_fv(slp_rec, fid):
                for _f in slp_rec.get("fields", []):
                    if _f.get("id") == fid:
                        _v = (_f.get("value") or "").strip()
                        return _v[:10] if len(_v) >= 10 else ""
                return ""

            slp_app_n = slp_rpa_n = 0
            for slp_rec in slp_recs_for_dates:
                for acct_id in slp_rec.get("relationships", {}).get("account", []):
                    _aid = str(acct_id)
                    app_v = _slp_fv(slp_rec, "last-app-date")
                    if app_v:
                        if not acct_to_last_app.get(_aid) or app_v > acct_to_last_app[_aid]:
                            acct_to_last_app[_aid] = app_v
                            slp_app_n += 1
                    rpa_v = _slp_fv(slp_rec, "last-rpa-date")
                    if rpa_v:
                        if not acct_to_last_rpa.get(_aid) or rpa_v > acct_to_last_rpa[_aid]:
                            acct_to_last_rpa[_aid] = rpa_v
                            slp_rpa_n += 1
            # Also build channel index from SLP channel field
            slp_ch_n = 0
            for slp_rec in slp_recs_for_dates:
                for acct_id in slp_rec.get("relationships", {}).get("account", []):
                    _aid = str(acct_id)
                    for _f in slp_rec.get("fields", []):
                        if _f.get("id") == "channel":
                            _ch = (_f.get("value") or "").strip()
                            if _ch and not acct_to_platform.get(_aid):
                                acct_to_platform[_aid] = _ch
                                slp_ch_n += 1
                            break
            print(f"[dealer-index] SLP supplement: {slp_app_n} last-app, {slp_rpa_n} last-rpa, {slp_ch_n} channel from {len(slp_recs_for_dates)} SLPs")
        except Exception as _slp_date_exc:
            print(f"[dealer-index] SLP last-app/rpa supplement failed (non-fatal): {_slp_date_exc}")

        # ── Publish index from bulk scan immediately so app is usable ─────
        new_did: dict = {}
        new_atd: dict = {}
        for aid, did in acct_to_dealer.items():
            new_atd[aid] = did
            new_did[did] = {"id": aid, "name": acct_to_name.get(aid, "")}

        _dealer_id_index.clear();    _dealer_id_index.update(new_did)
        _account_to_dealer.clear();  _account_to_dealer.update(new_atd)
        _account_to_platform.clear(); _account_to_platform.update(acct_to_platform)
        _account_to_bdr.clear();     _account_to_bdr.update(acct_to_bdr)
        _account_to_name.clear();    _account_to_name.update(acct_to_name)
        _account_to_owner.clear();   _account_to_owner.update(acct_to_owner)
        _account_to_states.clear();      _account_to_states.update(acct_to_states)
        _account_to_zip.clear();         _account_to_zip.update(acct_to_zip)
        _account_to_city.clear();        _account_to_city.update(acct_to_city)
        _account_to_state_prov.clear();  _account_to_state_prov.update(acct_to_state_prov)
        _account_to_phone.clear();       _account_to_phone.update(acct_to_phone)
        _account_to_website.clear();     _account_to_website.update(acct_to_website)
        _account_to_address.clear();     _account_to_address.update(acct_to_address)
        _account_to_last_app.clear();    _account_to_last_app.update(acct_to_last_app)
        _account_to_last_rpa.clear();    _account_to_last_rpa.update(acct_to_last_rpa)
        _account_to_type.clear();        _account_to_type.update(acct_to_type)
        _account_to_region.clear();      _account_to_region.update(acct_to_region)
        _account_to_dba.clear();         _account_to_dba.update(acct_to_dba)
        _account_to_status.clear();      _account_to_status.update(acct_to_status)
        _account_to_tax_id.clear();      _account_to_tax_id.update(acct_to_tax_id)
        _account_to_group.clear();       _account_to_group.update(acct_to_group)
        _account_to_legal_name.clear();       _account_to_legal_name.update(acct_to_legal_name)
        _account_to_revenue.clear();          _account_to_revenue.update(acct_to_revenue)
        _account_to_strategic_partners.clear(); _account_to_strategic_partners.update(acct_to_strat_part)
        global _apex_partners_cache; _apex_partners_cache = None  # invalidate on index rebuild
        _account_to_contractor_reactivation.clear(); _account_to_contractor_reactivation.update(acct_to_react)
        _account_to_reactivation_date.clear();       _account_to_reactivation_date.update(acct_to_react_date)
        _account_to_oracle_id.clear();               _account_to_oracle_id.update(acct_to_oracle_id)

        # Reverse index: lowercase dealer program → set of account IDs
        new_prog: dict = {}
        for aid, prog in acct_to_platform.items():
            key = prog.lower().strip()
            if key:
                new_prog.setdefault(key, set()).add(aid)
        _program_to_accounts.clear(); _program_to_accounts.update(new_prog)

        _dealer_index_ts = _time.time()
        print(f"[dealer-index] Done. {len(new_did)} dealer IDs, "
              f"{len(new_prog)} dealer programs indexed across {len(new_atd)} accounts.")

    except Exception as _build_exc:
        import traceback
        _dealer_index_error = f"{type(_build_exc).__name__}: {_build_exc}"
        print(f"[dealer-index] BUILD FAILED: {_build_exc}")
        traceback.print_exc()


SLP_SCHEMA_ID           = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
LICENSE_SCHEMA_ID       = "4bc17cb1-31be-4c15-a186-853ea85b1d40"
TRAINING_SCHEMA_ID      = "9368fee4-ccef-407b-a0d3-4b72c346b2af"
ACCT_ACTIVITY_SCHEMA_ID = "3a11374e-4b3d-47b8-b423-17ebcb7b1f4b"
ALT_CONTACT_SCHEMA_ID   = "b8259d61-10ba-4b15-8b2e-d1c8045712e0"

# Known account custom field IDs (from field_id_mapping.csv)
ACCT_FIELD = {
    "dealer_id":             "18",
    "account_status":        "19",
    "dba_name":              "15",
    "doing_business_in":     "22",
    "sales_region":          "23",
    "partner_activation":    "26",
    # "dealer_program" (CF29) and "platforms" (CF34) deleted from PROD
    "original_owner":        "35",
    "assigned_bdr":          "119",
    "oracle_producer_id":    "118",
}

# ═══════════════════════════════════════════════════════════════════════════
# CACHING
# ═══════════════════════════════════════════════════════════════════════════

CACHE: dict = {
    "account_custom_fields": {},
    "contact_custom_fields": {},
    "deal_custom_fields":    {},
    "field_metadata":        {},
    "schemas":               {},
    "slp_by_state":          {},
}
CACHE_TIMESTAMPS: dict = {
    "account_custom_fields": {},
    "contact_custom_fields": {},
    "deal_custom_fields":    {},
    "field_metadata":        {},
    "schemas":               {},
    "slp_by_state":          {},
}
CACHE_TTL = 300  # seconds

# ── SLP record cache (shared across all report endpoints) ────────────────────
_slp_cache_records: list  = []    # list of raw SLP record dicts
_slp_cache_ts:      float = 0.0   # epoch of last successful refresh
_slp_cache_lock             = asyncio.Lock()
_SLP_CACHE_TTL              = 900  # 15 minutes — fetch takes ~1-2 min, no point hammering every 5
_slp_refreshing:    bool   = False  # True while a refresh is in flight

# ── Account custom-field data cache (shared across all report endpoints) ─────
_acct_cf_raw:    list  = []   # all raw accountCustomFieldData records
_acct_cf_raw_ts: float = 0.0
_ACCT_CF_TTL           = 600  # 10 minutes
_acct_cf_lock             = asyncio.Lock()
_acct_cf_refreshing: bool = False  # True while a refresh is in flight

async def _refresh_acct_cf_cache() -> None:
    """Fetch ALL accountCustomFieldData records and atomically swap into _acct_cf_raw.

    A cold rebuild over the full account custom-field dataset (~160K+ records)
    took ~9.5 minutes at 100 records/page, one page at a time — long enough to
    hang any report depending on this cache (e.g. Contractor Activations)
    whenever it went stale. A first attempt at fixing this used
    ac_client.fetch_all_pages's concurrent path, which builds a {offset: page}
    dict for every page and only flattens it into the final list at the end —
    momentarily holding the entire ~160K-record dataset in memory TWICE. That
    was confirmed (via Render's logs/metrics) to be the direct cause of an
    out-of-memory crash in production. Fetching at 1000 records/page with 8
    concurrent workers, extending a single shared list as each page completes
    (same page size/concurrency already used by _build_dealer_id_index),
    avoids the double-buffering and still finishes in well under a minute."""
    global _acct_cf_raw, _acct_cf_raw_ts, _acct_cf_refreshing
    async with _acct_cf_lock:
        if _acct_cf_raw and (_time.time() - _acct_cf_raw_ts) < _ACCT_CF_TTL:
            return
        _acct_cf_refreshing = True
        print("[acct-cf-cache] Refreshing account custom field data…")
        PAGE, CONCURRENCY = 1000, 8
        raw: list = []
        try:
            first = await ac_get("accountCustomFieldData", {"limit": PAGE, "offset": 0})
            raw.extend(first.get("accountCustomFieldData", []))
            total = int(first.get("meta", {}).get("total", 0))

            if total > PAGE:
                sem = asyncio.Semaphore(CONCURRENCY)
                async def fetch_and_extend(offset: int):
                    async with sem:
                        page = await ac_get("accountCustomFieldData", {"limit": PAGE, "offset": offset})
                        raw.extend(page.get("accountCustomFieldData", []))
                await asyncio.gather(*[fetch_and_extend(o) for o in range(PAGE, total, PAGE)])

            print(f"[acct-cf-cache] fetched {len(raw)} records")
            if raw:
                _acct_cf_raw    = raw
                _acct_cf_raw_ts = _time.time()
            else:
                print("[acct-cf-cache] WARNING: 0 records returned — keeping existing cache, will retry")
        except Exception as _e:
            print(f"[acct-cf-cache] fetch failed: {_e}")
        finally:
            _acct_cf_refreshing = False

async def _acct_cf_cache_loop() -> None:
    """Background task: keep the account custom-field cache warm, refreshing every
    _ACCT_CF_TTL seconds, so report requests never have to wait on a cold rebuild."""
    await asyncio.sleep(90)   # stagger away from other startup fetches
    while True:
        try:
            await _refresh_acct_cf_cache()
        except Exception as _e:
            print(f"[acct-cf-cache] loop error: {_e}")
        if _acct_cf_raw:
            await asyncio.sleep(_ACCT_CF_TTL)
        else:
            print("[acct-cf-cache] cache still empty — retrying in 30s")
            await asyncio.sleep(30)

async def _refresh_slp_cache() -> None:
    """Fetch ALL SLP records from AC and atomically swap into _slp_cache_records.
    Uses ac_client.fetch_all_slps (sequential, dedup-by-id) — stops as soon as
    a page adds no new records, avoids the multi-pass overhead that was causing
    the cache to run almost continuously."""
    global _slp_cache_records, _slp_cache_ts, _slp_refreshing
    async with _slp_cache_lock:
        # Double-check inside the lock — another waiter may have just refreshed
        if _slp_cache_records and (_time.time() - _slp_cache_ts) < _SLP_CACHE_TTL:
            return

        _slp_refreshing = True
        print("[slp-cache] Refreshing SLP records…")
        try:
            from ac_client import fetch_all_slps as _fetch_slps
            temp_records = await _fetch_slps(SLP_SCHEMA_ID)
            print(f"[slp-cache] fetched {len(temp_records)} records")
            if temp_records:
                _slp_cache_records = temp_records
                _slp_cache_ts      = _time.time()
                _update_app_rpa_from_slp_cache()
            else:
                print("[slp-cache] WARNING: 0 records returned — keeping existing cache, will retry")
        except Exception as _e:
            print(f"[slp-cache] fetch failed: {_e}")
        finally:
            _slp_refreshing = False

async def get_slp_cache() -> list:
    """Return cached SLP records, refreshing if stale or empty.

    If a refresh is already in flight AND the cache already has data, return
    the existing complete (slightly stale) data immediately rather than blocking.
    If the cache is empty (initial startup), always wait for the refresh to finish.
    """
    if _slp_refreshing and _slp_cache_records:
        # Background refresh running but we have a previous complete snapshot — use it
        return _slp_cache_records
    if not _slp_cache_records or (_time.time() - _slp_cache_ts) > _SLP_CACHE_TTL:
        await _refresh_slp_cache()
    return _slp_cache_records

# ── Last-Contacted cache ──────────────────────────────────────────────────────
_lc_cache: dict  = {}    # account_id → "YYYY-MM-DD"
_lc_cache_ts: float = 0.0
_LC_CACHE_TTL = 1800     # 30 minutes

async def _refresh_lc_cache() -> None:
    global _lc_cache, _lc_cache_ts
    today = date.today()
    latest: dict = {}   # aid → {"date": "YYYY-MM-DD", "type": str}

    def _update(aid: str, raw_date: str, contact_type: str) -> None:
        if not aid or not raw_date:
            return
        try:
            d = date.fromisoformat(str(raw_date)[:10])
            if d.year < 2000 or d > today:
                return
            ds = d.isoformat()
            if aid not in latest or ds > latest[aid]["date"]:
                latest[aid] = {"date": ds, "type": contact_type}
        except Exception:
            pass

    # ── 1. Account Activity custom object ────────────────────────────────────
    try:
        activity_records = await ac_get_all(
            f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", "records", {}
        )
        for r in activity_records:
            fmap = {f["id"]: (f.get("value") or "") for f in r.get("fields", [])}
            aid  = next(iter(r.get("relationships", {}).get("account", [])), "")
            _update(aid, str(fmap.get("activity-date", ""))[:10], "Activity")
        print(f"[lc-cache] account activity: {len(activity_records)} records")
    except Exception as e:
        print(f"[lc-cache] activity fetch error: {e}")

    # ── 2. Account Notes ──────────────────────────────────────────────────────
    try:
        all_notes = await ac_get_all("notes", "notes",
                                     {"reltype": "CustomerAccount", "limit": 100})
        acct_note_n = 0
        for n in all_notes:
            if (n.get("reltype") or "").lower() != "customeraccount":
                continue
            aid = str(n.get("rel_id") or n.get("relid") or "")
            _update(aid, n.get("cdate", ""), "Note")
            acct_note_n += 1
        print(f"[lc-cache] account notes: {acct_note_n} notes")
    except Exception as e:
        print(f"[lc-cache] notes fetch error: {e}")

    # ── 3. Build contact → account map from accountContacts ──────────────────
    contact_to_acct: dict = {}
    try:
        ac_contacts = await ac_get_all("accountContacts", "accountContacts", {"limit": 100})
        for ac in ac_contacts:
            cid = str(ac.get("contact") or "")
            aid = str(ac.get("account") or "")
            if cid and aid:
                contact_to_acct[cid] = aid
        print(f"[lc-cache] contact→account map: {len(contact_to_acct)} contacts")
    except Exception as e:
        print(f"[lc-cache] accountContacts fetch error: {e}")

    # ── 4. Contact Notes ──────────────────────────────────────────────────────
    if contact_to_acct:
        try:
            contact_notes = await ac_get_all("notes", "notes",
                                             {"reltype": "Subscriber", "limit": 100})
            contact_note_n = 0
            for n in contact_notes:
                if (n.get("reltype") or "").lower() != "subscriber":
                    continue
                cid = str(n.get("rel_id") or n.get("relid") or "")
                aid = contact_to_acct.get(cid)
                if aid:
                    _update(aid, n.get("cdate", ""), "Note")
                    contact_note_n += 1
            print(f"[lc-cache] contact notes: {contact_note_n} notes mapped to accounts")
        except Exception as e:
            print(f"[lc-cache] contact notes fetch error: {e}")

    # Source 5 (email activity) removed — the /activities endpoint without a
    # contact filter pages through every AC event ever recorded and can take
    # 10+ minutes.  Account Activity + Notes already cover AM-initiated contact.

    if latest:
        _lc_cache    = latest
        _lc_cache_ts = _time.time()
        print(f"[lc-cache] refreshed — {len(latest)} accounts with last-contacted date")

async def _lc_cache_loop() -> None:
    await asyncio.sleep(90)    # stagger slightly after SLP + dealer index
    while True:
        try:
            await _refresh_lc_cache()
        except Exception as e:
            print(f"[lc-cache] loop error: {e}")
        await asyncio.sleep(_LC_CACHE_TTL)

# ─── Team Activity raw-data cache ────────────────────────────────────────────
_ta_cache: dict  = {}
_ta_cache_ts: float = 0.0
_TA_CACHE_TTL = 900   # 15 minutes

async def _refresh_ta_cache() -> None:
    global _ta_cache, _ta_cache_ts
    print("[ta-cache] refreshing…")
    from ac_client import fetch_all_pages as _fap

    # Fetch notes and activity concurrently; contacts are paged below to avoid
    # holding 100K+ raw contact objects in memory at the same time.
    users_data, all_notes_raw, all_activity = await asyncio.gather(
        ac_get("users"),
        _fap("notes", key="notes"),
        ac_get_all(f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", "records", {}),
    )

    # Page through contacts, building slim maps on the fly — never accumulate
    # all raw contact objects in memory (100K+ × ~2KB each = ~200MB peak avoided).
    contact_to_account: dict = {}
    contact_email_map:  dict = {}
    offset, PAGE = 0, 100
    while True:
        page_data = await ac_get("contacts", {"limit": PAGE, "offset": offset, "orders[id]": "ASC"})
        batch = page_data.get("contacts", [])
        if not batch:
            break
        for c in batch:
            cid = str(c.get("id", ""))
            aid = str(c.get("account", "") or "")
            if aid and aid != "0":
                contact_to_account[cid] = aid
            email = c.get("email", "")
            if email:
                contact_email_map[cid] = email
        offset += PAGE
        if len(batch) < PAGE:
            break

    # Slim notes — keep only fields used by the team report endpoints.
    slim_notes = [
        {
            "id":      n.get("id"),
            "userid":  n.get("userid"),
            "relid":   n.get("relid") or n.get("rel_id"),
            "reltype": n.get("reltype"),
            "cdate":   n.get("cdate"),
            "note":    n.get("note") or "",
        }
        for n in all_notes_raw
        if (n.get("reltype") or "").lower() in ("contact", "customeraccount", "deal")
    ]

    # Slim activity records — drop unused nested fields.
    slim_activity = [
        {
            "id":            r.get("id"),
            "fields":        {f["id"]: f.get("value") for f in r.get("fields", [])},
            "account":       next(iter(r.get("relationships", {}).get("account", [])), ""),
        }
        for r in all_activity
    ]

    _ta_cache = {
        "users_data":         users_data,
        "all_notes_raw":      slim_notes,
        "contact_to_account": contact_to_account,
        "contact_email_map":  contact_email_map,
        "all_activity":       slim_activity,
    }
    _ta_cache_ts = _time.time()
    print(f"[ta-cache] done — {len(slim_notes)} notes, {len(slim_activity)} activities, "
          f"{len(contact_to_account)} contact→account mappings")

async def _ta_cache_loop() -> None:
    await asyncio.sleep(90)    # contacts+notes now concurrent — builds in ~10s
    while True:
        try:
            await _refresh_ta_cache()
        except Exception as e:
            print(f"[ta-cache] loop error: {e}")
        await asyncio.sleep(_TA_CACHE_TTL)

_slp_dependent_indexes_built: bool = False  # True after first post-SLP location/state build

# ── Partner-channel BDR sync ─────────────────────────────────────────────────
# Every SLP whose channel isn't Microf/Microf Direct should have Assigned BDR
# set to "Partner". Originally a one-off manual bulk push; this keeps it true
# going forward for every new SLP as the cache refreshes, without needing a
# native AC automation (custom-object field changes can't trigger those).
PARTNER_BDR_SYNC_DRY_RUN = False   # live — writes Assigned BDR = Partner to AC
PARTNER_BDR_EXCLUDED_CHANNELS = {"Microf", "Microf Direct"}
PARTNER_BDR_VALUE = "Partner"

async def _sync_partner_bdr() -> None:
    checked = updated = errors = 0
    for rec in _slp_cache_records:
        fields = {f.get("id"): f.get("value") for f in rec.get("fields", [])}
        channel = (fields.get("channel") or "").strip()
        current_bdr = (fields.get("assigned-bdr") or "").strip()
        if not channel or channel in PARTNER_BDR_EXCLUDED_CHANNELS:
            continue
        if current_bdr == PARTNER_BDR_VALUE:
            continue
        checked += 1
        dealer_id = fields.get("dealer-id", "")
        if PARTNER_BDR_SYNC_DRY_RUN:
            print(f"[partner-bdr-sync] DRY RUN would update dealer {dealer_id} "
                  f"(channel={channel!r}, current BDR={current_bdr!r} -> Partner)")
            updated += 1
            continue
        try:
            raw_fields = list(rec.get("fields", []))
            found = False
            for f in raw_fields:
                if f.get("id") == "assigned-bdr":
                    f["value"] = PARTNER_BDR_VALUE
                    found = True
                    break
            if not found:
                raw_fields.append({"id": "assigned-bdr", "value": PARTNER_BDR_VALUE})
            payload = {"record": {"id": rec["id"], "fields": raw_fields,
                                   "relationships": rec.get("relationships", {})}}
            await ac_post(f"customObjects/records/{SLP_SCHEMA_ID}", payload)
            updated += 1
        except Exception as e:
            errors += 1
            print(f"[partner-bdr-sync] failed for dealer {dealer_id}: {e}")

    if checked:
        mode = "DRY RUN — " if PARTNER_BDR_SYNC_DRY_RUN else ""
        print(f"[partner-bdr-sync] {mode}checked {checked} candidates, "
              f"{'would update' if PARTNER_BDR_SYNC_DRY_RUN else 'updated'} {updated}, errors {errors}")

# ── Microf Direct blank-BDR sync ──────────────────────────────────────────────
# Every Microf Direct SLP should have a real Assigned BDR. Where none has ever
# been set (blank), default it to "House" so it's not left blank indefinitely —
# mirrors partner-bdr-sync's role for the Microf Direct side of the same rule
# (Microf Direct = named person or House; everything else = Partner).
HOUSE_BDR_SYNC_DRY_RUN = False   # live — writes Assigned BDR = House to AC
HOUSE_BDR_CHANNEL = "Microf Direct"
HOUSE_BDR_VALUE = "House"

async def _sync_house_bdr() -> None:
    checked = updated = errors = 0
    for rec in _slp_cache_records:
        fields = {f.get("id"): f.get("value") for f in rec.get("fields", [])}
        channel = (fields.get("channel") or "").strip()
        current_bdr = (fields.get("assigned-bdr") or "").strip()
        if channel != HOUSE_BDR_CHANNEL or current_bdr:
            continue
        checked += 1
        dealer_id = fields.get("dealer-id", "")
        if HOUSE_BDR_SYNC_DRY_RUN:
            print(f"[house-bdr-sync] DRY RUN would update dealer {dealer_id} "
                  f"(channel={channel!r}, blank BDR -> House)")
            updated += 1
            continue
        try:
            raw_fields = list(rec.get("fields", []))
            found = False
            for f in raw_fields:
                if f.get("id") == "assigned-bdr":
                    f["value"] = HOUSE_BDR_VALUE
                    found = True
                    break
            if not found:
                raw_fields.append({"id": "assigned-bdr", "value": HOUSE_BDR_VALUE})
            payload = {"record": {"id": rec["id"], "fields": raw_fields,
                                   "relationships": rec.get("relationships", {})}}
            await ac_post(f"customObjects/records/{SLP_SCHEMA_ID}", payload)
            updated += 1
        except Exception as e:
            errors += 1
            print(f"[house-bdr-sync] failed for dealer {dealer_id}: {e}")

    if checked:
        mode = "DRY RUN — " if HOUSE_BDR_SYNC_DRY_RUN else ""
        print(f"[house-bdr-sync] {mode}checked {checked} candidates, "
              f"{'would update' if HOUSE_BDR_SYNC_DRY_RUN else 'updated'} {updated}, errors {errors}")

# ── "Not Started" -> "Pre-activation" status sync ────────────────────────────
# dealer.microf.com still auto-creates new SLPs with the legacy "Not Started"
# status. "Pre-activation" is the current canonical value for that same state
# (Step 1 of the flow), so this keeps new records normalized going forward the
# same way the one-off migration cleaned up the existing backlog.
NOT_STARTED_STATUS_SYNC_DRY_RUN = True
NOT_STARTED_STATUS_OLD = "Not Started"
NOT_STARTED_STATUS_NEW = "Pre-activation"

async def _sync_not_started_status() -> None:
    checked = updated = errors = 0
    for rec in _slp_cache_records:
        fields = {f.get("id"): f.get("value") for f in rec.get("fields", [])}
        if (fields.get("slp-status-detail") or "").strip() != NOT_STARTED_STATUS_OLD:
            continue
        checked += 1
        dealer_id = fields.get("dealer-id", "")
        if NOT_STARTED_STATUS_SYNC_DRY_RUN:
            print(f"[not-started-sync] DRY RUN would update dealer {dealer_id} "
                  f"'{NOT_STARTED_STATUS_OLD}' -> '{NOT_STARTED_STATUS_NEW}'")
            updated += 1
            continue
        try:
            raw_fields = list(rec.get("fields", []))
            found = False
            for f in raw_fields:
                if f.get("id") == "slp-status-detail":
                    f["value"] = NOT_STARTED_STATUS_NEW
                    found = True
                    break
            if not found:
                raw_fields.append({"id": "slp-status-detail", "value": NOT_STARTED_STATUS_NEW})
            payload = {"record": {"id": rec["id"], "fields": raw_fields,
                                   "relationships": rec.get("relationships", {})}}
            await ac_post(f"customObjects/records/{SLP_SCHEMA_ID}", payload)
            updated += 1
        except Exception as e:
            errors += 1
            print(f"[not-started-sync] failed for dealer {dealer_id}: {e}")

    if checked:
        mode = "DRY RUN — " if NOT_STARTED_STATUS_SYNC_DRY_RUN else ""
        print(f"[not-started-sync] {mode}checked {checked} candidates, "
              f"{'would update' if NOT_STARTED_STATUS_SYNC_DRY_RUN else 'updated'} {updated}, errors {errors}")

# ── PandaDoc-signed → Awaiting Training sync ─────────────────────────────────
# The "Agreement Completed" AC automation (native PandaDoc trigger) tags the
# contact "PandaDoc-Signed" when a dealer agreement is signed. AC can't update
# the related SLP custom object from within that automation directly, so this
# picks up the tag here instead: find the account behind the tagged contact,
# advance any of its SLPs sitting in "Awaiting Dealer Agreement Signature" to
# "Awaiting Training", then remove the tag so it isn't reprocessed. If no
# matching SLP is found, the tag is left in place to retry next cycle.
PANDADOC_SIGNED_SYNC_DRY_RUN = False   # live — advances SLPs and removes the tag in AC
PANDADOC_SIGNED_TAG_NAME = "PandaDoc-Signed"
_pandadoc_signed_tag_id: Optional[str] = None

async def _sync_pandadoc_signed() -> None:
    global _pandadoc_signed_tag_id
    if _pandadoc_signed_tag_id is None:
        tags_resp = await ac_get("tags", {"search": PANDADOC_SIGNED_TAG_NAME, "limit": 100})
        for t in tags_resp.get("tags", []):
            if (t.get("tag") or "").strip().lower() == PANDADOC_SIGNED_TAG_NAME.lower():
                _pandadoc_signed_tag_id = t.get("id")
                break
        if _pandadoc_signed_tag_id is None:
            print(f"[pandadoc-signed-sync] tag {PANDADOC_SIGNED_TAG_NAME!r} not found in AC yet — skipping")
            return

    tagged = await ac_get_all("contacts", "contacts", {"tagid": _pandadoc_signed_tag_id})
    if not tagged:
        return

    # account_id -> [SLP records currently Awaiting Dealer Agreement Signature]
    awaiting_by_account: dict[str, list] = {}
    for rec in _slp_cache_records:
        fields = {f.get("id"): f.get("value") for f in rec.get("fields", [])}
        if (fields.get("slp-status-detail") or "").strip() != "Awaiting Dealer Agreement Signature":
            continue
        for acct_id in rec.get("relationships", {}).get("account", []):
            awaiting_by_account.setdefault(str(acct_id), []).append(rec)

    checked = updated = errors = untagged_only = 0
    for contact in tagged:
        cid = contact.get("id")
        checked += 1
        try:
            acct_links = await ac_get_all("accountContacts", "accountContacts", {"contact": cid})
            account_ids = {str(a.get("account")) for a in acct_links if a.get("account")}  # set — dedupe repeat links
        except Exception as e:
            errors += 1
            print(f"[pandadoc-signed-sync] failed to look up account for contact {cid}: {e}")
            continue

        seen_rec_ids: set = set()
        matching_slps = []
        for aid in account_ids:
            for rec in awaiting_by_account.get(aid, []):
                if rec["id"] not in seen_rec_ids:
                    seen_rec_ids.add(rec["id"])
                    matching_slps.append(rec)
        if not matching_slps:
            print(f"[pandadoc-signed-sync] contact {cid} tagged but no SLP in "
                  f"'Awaiting Dealer Agreement Signature' found for its account(s) {account_ids} — leaving tag, will retry")
            untagged_only += 1
            continue

        for rec in matching_slps:
            dealer_id = next((f.get("value") for f in rec.get("fields", []) if f.get("id") == "dealer-id"), "")
            if PANDADOC_SIGNED_SYNC_DRY_RUN:
                print(f"[pandadoc-signed-sync] DRY RUN would advance dealer {dealer_id} "
                      f"(contact {cid}) -> Awaiting Training, then remove tag")
                updated += 1
                continue
            try:
                raw_fields = list(rec.get("fields", []))
                for f in raw_fields:
                    if f.get("id") == "slp-status-detail":
                        f["value"] = "Awaiting Training"
                        break
                payload = {"record": {"id": rec["id"], "fields": raw_fields,
                                       "relationships": rec.get("relationships", {})}}
                await ac_post(f"customObjects/records/{SLP_SCHEMA_ID}", payload)
                updated += 1
            except Exception as e:
                errors += 1
                print(f"[pandadoc-signed-sync] failed to advance dealer {dealer_id}: {e}")
                continue

        if not PANDADOC_SIGNED_SYNC_DRY_RUN:
            try:
                tag_assoc = await ac_get("contactTags", {"contact": cid, "tag": _pandadoc_signed_tag_id})
                for assoc in tag_assoc.get("contactTags", []):
                    await ac_delete(f"contactTags/{assoc['id']}")
            except Exception as e:
                print(f"[pandadoc-signed-sync] advanced dealer(s) but failed to remove tag from contact {cid}: {e}")

    if checked:
        mode = "DRY RUN — " if PANDADOC_SIGNED_SYNC_DRY_RUN else ""
        print(f"[pandadoc-signed-sync] {mode}checked {checked} tagged contacts, "
              f"{'would advance' if PANDADOC_SIGNED_SYNC_DRY_RUN else 'advanced'} {updated} SLP(s), "
              f"{untagged_only} left tagged (no match yet), errors {errors}")

async def _slp_cache_loop() -> None:
    """Background task: keep SLP cache warm, refreshing every 5 minutes.
    On failure (0 records), retries every 30s until data is loaded, then
    switches to the normal 5-minute refresh interval.
    After the first successful SLP load, kicks off the location and SLP-state
    index builds (which depend on SLP data) so they don't race at startup.
    """
    global _slp_dependent_indexes_built
    await asyncio.sleep(90)   # give dealer index a head-start before first SLP fetch
    while True:
        try:
            await _refresh_slp_cache()
        except Exception as _e:
            print(f"[slp-cache] loop error: {_e}")
        # After first successful SLP load, kick off SLP-dependent index builders once
        if _slp_cache_records and not _slp_dependent_indexes_built:
            _slp_dependent_indexes_built = True
            print("[slp-cache] SLP data ready — triggering location + state index builds")
            asyncio.create_task(_build_location_index())
            asyncio.create_task(_build_slp_state_index())
        if _slp_cache_records:
            try:
                await _sync_partner_bdr()
            except Exception as _e:
                print(f"[partner-bdr-sync] loop error: {_e}")
            try:
                await _sync_house_bdr()
            except Exception as _e:
                print(f"[house-bdr-sync] loop error: {_e}")
            try:
                await _sync_not_started_status()
            except Exception as _e:
                print(f"[not-started-sync] loop error: {_e}")
            try:
                await _sync_pandadoc_signed()
            except Exception as _e:
                print(f"[pandadoc-signed-sync] loop error: {_e}")
        # If cache is still empty, retry quickly; otherwise use normal TTL
        if _slp_cache_records:
            await asyncio.sleep(_SLP_CACHE_TTL)
        else:
            print("[slp-cache] cache still empty — retrying in 30s")
            await asyncio.sleep(30)

def _update_app_rpa_from_slp_cache() -> None:
    """Update _account_to_last_app, _account_to_last_rpa, and _account_to_platform
    from the current SLP cache.  Called after every SLP cache refresh so the indexes
    stay current without waiting for the 24h dealer index rebuild.

    NOTE: channel is updated here (not just in _build_dealer_id_index) so that
    it survives cases where Phase-4 of the dealer index fails silently at startup."""
    app_n = rpa_n = ch_n = did_n = 0
    for slp_rec in _slp_cache_records:
        for acct_id in slp_rec.get("relationships", {}).get("account", []):
            aid = str(acct_id)
            for _f in slp_rec.get("fields", []):
                fid = _f.get("id", "")
                val = (_f.get("value") or "").strip()
                if fid == "channel":
                    if val:
                        _account_to_platform[aid] = val   # always overwrite — latest SLP wins
                        ch_n += 1
                    continue
                if fid == "dealer-id":
                    if val:
                        _account_to_slp_dealer[aid] = val  # SLP dealer-id (authoritative)
                        did_n += 1
                    continue
                if fid == "contractor-activated-date":
                    if val:
                        _account_to_activation_date[aid] = val[:10]
                    continue
                if fid == "doing-business-in-states":
                    if val:
                        _account_to_slp_states[aid] = val
                    continue
                v10 = val[:10] if len(val) >= 10 else ""
                if not v10:
                    continue
                if fid == "last-app-date":
                    if not _account_to_last_app.get(aid) or v10 > _account_to_last_app[aid]:
                        _account_to_last_app[aid] = v10
                        app_n += 1
                elif fid == "last-rpa-date":
                    if not _account_to_last_rpa.get(aid) or v10 > _account_to_last_rpa[aid]:
                        _account_to_last_rpa[aid] = v10
                        rpa_n += 1
    print(f"[slp-cache] index updated — {app_n} app dates, {rpa_n} rpa dates, {ch_n} channels, {did_n} dealer IDs")

def get_cached(cache_type: str, key: str):
    if key in CACHE[cache_type]:
        ts = CACHE_TIMESTAMPS[cache_type].get(key, 0)
        if datetime.now().timestamp() - ts < CACHE_TTL:
            return CACHE[cache_type][key]
    return None

def set_cached(cache_type: str, key: str, value):
    CACHE[cache_type][key] = value
    CACHE_TIMESTAMPS[cache_type][key] = datetime.now().timestamp()

MAX_CONCURRENT_REQUESTS = 20
semaphore = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)


# ═══════════════════════════════════════════════════════════════════════════
# AC API HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def ac_url(path: str) -> str:
    return f"{AC_BASE_URL}/api/3/{path.lstrip('/')}"

async def ac_get(path: str, params: dict = None):
    async with httpx.AsyncClient(timeout=60) as client:
        r = await client.get(ac_url(path), headers=HEADERS, params=params or {})
        r.raise_for_status()
        return r.json()

async def ac_post(path: str, body: dict):
    async with httpx.AsyncClient(timeout=60) as client:
        r = await client.post(ac_url(path), headers=HEADERS, json=body)
        if not r.is_success:
            raise Exception(f"HTTP {r.status_code} {r.text[:300]}")
        return r.json()

async def ac_put(path: str, body: dict):
    async with httpx.AsyncClient(timeout=60) as client:
        r = await client.put(ac_url(path), headers=HEADERS, json=body)
        if not r.is_success:
            raise Exception(f"HTTP {r.status_code} {r.text[:300]}")
        return r.json()

async def ac_delete(path: str) -> int:
    async with httpx.AsyncClient(timeout=60) as client:
        r = await client.delete(ac_url(path), headers=HEADERS)
        return r.status_code

async def ac_get_all(path: str, key: str, params: dict = None) -> list:
    """Paginate through all records, deduplicating by id.

    For all endpoints: sequential pages, stops when a page adds no new records
    (handles AC's non-deterministic pagination without relying on meta.total).
    For custom-object endpoints: up to 3 full passes to catch records that shift
    between pages, but exits early if a pass yields nothing new.
    """
    is_custom_obj = "customObjects" in path
    max_passes    = 3 if is_custom_obj else 1

    seen  = {}
    p     = params or {}
    limit = 100

    for pass_num in range(max_passes):
        offset        = 0
        new_this_pass = 0
        while True:
            data = await ac_get(path, {**p, "limit": limit, "offset": offset})
            page = data.get(key, [])
            if not page:
                break
            for item in page:
                item_id = item.get("id")
                if item_id is not None:
                    if item_id not in seen:
                        seen[item_id] = item
                        new_this_pass += 1
                else:
                    seen[len(seen)] = item
                    new_this_pass += 1
            offset += limit
        if is_custom_obj:
            print(f"[ac_get_all] {path} pass {pass_num+1} done, unique={len(seen)} new={new_this_pass}")
        # Stop early if this pass found nothing new — no point doing another
        if new_this_pass == 0:
            break

    return list(seen.values())


# ═══════════════════════════════════════════════════════════════════════════
# FIELD DISCOVERY
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/objects")
async def list_objects():
    return {
        "objects": [
            {"id": "slp",             "name": "Strategic Lending Partners",   "icon": "📊"},
            {"id": "accounts",        "name": "Accounts",                     "icon": "🏢"},
            {"id": "contacts",        "name": "Contacts",                     "icon": "👤"},
            {"id": "deals",           "name": "Deals",                        "icon": "💰"},
            {"id": "trainings",       "name": "Trainings",                    "icon": "🎓"},
            {"id": "license_details", "name": "Contractor License Details",   "icon": "📜"},
            {"id": "notes",           "name": "Notes",                        "icon": "📝"},
        ]
    }


async def _schema_fields(schema_id: str) -> tuple:
    """Return (fields_list, field_types_dict) for a custom object schema."""
    data   = await ac_get(f"customObjects/schemas/{schema_id}")
    schema = data.get("schema", {})
    fields, ftypes = [], {}
    for f in schema.get("fields", []):
        fid   = f.get("id", f.get("slug"))
        label = f.get("labels", {}).get("singular", f.get("slug", ""))
        ftype = f.get("type", "text")
        fields.append({"id": fid, "label": label, "type": "primary", "dataType": ftype})
        ftypes[fid] = {"type": ftype, "options": f.get("options", [])}
    return fields, ftypes


async def _account_custom_field_defs() -> list:
    """Return account custom field meta as list of field dicts."""
    try:
        data = await ac_get("accountCustomFieldMeta")
        return data.get("accountCustomFieldMeta", [])
    except Exception:
        return []


async def _contact_custom_field_defs() -> list:
    try:
        data = await ac_get("fields")
        return data.get("fields", [])
    except Exception:
        return []


@app.get("/api/fields/{object_type}")
async def get_fields(object_type: str):
    fields: list = []
    ftypes: dict = {}

    if object_type == "slp":
        f, ft = await _schema_fields(SLP_SCHEMA_ID)
        fields.extend(f); ftypes.update(ft)
        fields.extend(await _related_account_fields("Account"))
        fields.extend(_related_contact_summary_fields("Contacts"))
        fields.extend(_related_training_summary_fields("Trainings"))
        fields.extend(_related_deal_summary_fields("Deals"))

    elif object_type == "license_details":
        f, ft = await _schema_fields(LICENSE_SCHEMA_ID)
        fields.extend(f); ftypes.update(ft)
        fields.extend(await _related_account_fields("Account"))

    elif object_type == "trainings":
        f, ft = await _schema_fields(TRAINING_SCHEMA_ID)
        fields.extend(f); ftypes.update(ft)
        fields.extend(await _related_account_fields("Account"))

    elif object_type == "accounts":
        # Built-in account fields
        sample = await _sample("accounts", "accounts")
        for key, val in sample.items():
            if key not in ("links", "fieldValues") and not isinstance(val, dict):
                fields.append({"id": key, "label": key, "type": "primary", "dataType": "text"})
        # Account custom fields
        for cf in await _account_custom_field_defs():
            fid   = f"customfield_{cf['id']}"
            label = cf.get("fieldLabel", cf.get("fieldName", str(cf["id"])))
            ftype = cf.get("fieldType", "text")
            obj   = {"id": fid, "label": label, "type": "primary", "dataType": ftype}
            if ftype in ("dropdown", "listbox", "radio"):
                opts = cf.get("fieldOptions", "")
                if isinstance(opts, str) and opts:
                    obj["options"] = [o.strip() for o in opts.replace("\n", ",").split(",") if o.strip()]
                elif isinstance(opts, list):
                    obj["options"] = [str(o.get("value", o)) if isinstance(o, dict) else str(o) for o in opts if o]
            fields.append(obj)
            ftypes[fid] = obj
        # Cross-object fields
        fields.extend(await _related_slp_fields("SLP"))
        fields.extend(_related_contact_summary_fields("Contacts"))
        fields.extend(_related_deal_summary_fields("Deals"))
        fields.extend(_related_training_summary_fields("Trainings"))
        fields.extend(_related_notes_summary_fields("Notes"))

    elif object_type == "contacts":
        sample = await _sample("contacts", "contacts")
        for key, val in sample.items():
            if key not in ("links", "fieldValues") and not isinstance(val, dict):
                fields.append({"id": key, "label": key, "type": "primary", "dataType": "text"})
        for cf in await _contact_custom_field_defs():
            fid   = f"customfield_{cf['id']}"
            label = cf.get("title", str(cf["id"]))
            fields.append({"id": fid, "label": label, "type": "primary", "dataType": "text"})
        fields.extend(await _related_account_fields("Account"))
        fields.extend(await _related_slp_fields("SLP"))
        fields.extend(_related_deal_summary_fields("Deals"))
        fields.extend(_related_notes_summary_fields("Notes"))

    elif object_type == "notes":
        fields = [
            {"id": "id",                "label": "Note ID",            "type": "primary", "dataType": "text"},
            {"id": "note",              "label": "Note Content",       "type": "primary", "dataType": "text"},
            {"id": "cdate",             "label": "Created Date",       "type": "primary", "dataType": "date"},
            {"id": "mdate",             "label": "Modified Date",      "type": "primary", "dataType": "date"},
            {"id": "userid",            "label": "Author User ID",     "type": "primary", "dataType": "text"},
            {"id": "reltype",           "label": "Related To Type",    "type": "primary", "dataType": "text"},
            {"id": "rel_id",            "label": "Related Object ID",  "type": "primary", "dataType": "text"},
            {"id": "pinned",            "label": "Pinned",             "type": "primary", "dataType": "text"},
            {"id": "contact.firstName", "label": "Contact: First Name","type": "related", "dataType": "text"},
            {"id": "contact.lastName",  "label": "Contact: Last Name", "type": "related", "dataType": "text"},
            {"id": "contact.email",     "label": "Contact: Email",     "type": "related", "dataType": "text"},
        ]
        fields.extend(await _related_account_fields("Account"))

    elif object_type == "deals":
        sample = await _sample("deals", "deals")
        for key, val in sample.items():
            if key not in ("links",) and not isinstance(val, dict):
                fields.append({"id": key, "label": key, "type": "primary", "dataType": "text"})
        fields.extend(await _related_account_fields("Account"))
        fields.extend([
            {"id": "contact.email",     "label": "Contact: Email",      "type": "related", "dataType": "text"},
            {"id": "contact.firstName", "label": "Contact: First Name", "type": "related", "dataType": "text"},
            {"id": "contact.lastName",  "label": "Contact: Last Name",  "type": "related", "dataType": "text"},
            {"id": "contact.phone",     "label": "Contact: Phone",      "type": "related", "dataType": "text"},
        ])

    return {"fields": fields, "fieldTypes": ftypes}


async def _sample(path: str, key: str) -> dict:
    try:
        data = await ac_get(path, {"limit": 1})
        return data.get(key, [{}])[0]
    except Exception:
        return {}


async def _related_account_fields(prefix: str) -> list:
    fields = []
    sample = await _sample("accounts", "accounts")
    for key, val in sample.items():
        if key not in ("links", "fieldValues") and not isinstance(val, dict):
            fields.append({"id": f"account.{key}", "label": f"{prefix}: {key}", "type": "related", "dataType": "text"})
    for cf in await _account_custom_field_defs():
        fid   = f"account.customfield_{cf['id']}"
        label = cf.get("fieldLabel", cf.get("fieldName", str(cf["id"])))
        fields.append({"id": fid, "label": f"{prefix}: {label}", "type": "related", "dataType": "text"})
    return fields


async def _related_slp_fields(prefix: str) -> list:
    fields = []
    try:
        f, _ = await _schema_fields(SLP_SCHEMA_ID)
        for field in f:
            fields.append({"id": f"slp.{field['id']}", "label": f"{prefix}: {field['label']}", "type": "related", "dataType": field.get("dataType", "text")})
        fields.append({"id": "slp._count", "label": f"{prefix}: Record Count", "type": "related", "dataType": "number"})
    except Exception:
        pass
    return fields


def _related_contact_summary_fields(prefix: str) -> list:
    return [
        {"id": "primary_contact.email",     "label": f"{prefix}: Primary Email",      "type": "related", "dataType": "text"},
        {"id": "primary_contact.firstName", "label": f"{prefix}: Primary First Name", "type": "related", "dataType": "text"},
        {"id": "primary_contact.lastName",  "label": f"{prefix}: Primary Last Name",  "type": "related", "dataType": "text"},
        {"id": "primary_contact.phone",     "label": f"{prefix}: Primary Phone",      "type": "related", "dataType": "text"},
        {"id": "contact_count",             "label": f"{prefix}: Count",              "type": "related", "dataType": "number"},
    ]


def _related_deal_summary_fields(prefix: str) -> list:
    return [
        {"id": "deal.title",  "label": f"{prefix}: Latest Title",  "type": "related", "dataType": "text"},
        {"id": "deal.stage",  "label": f"{prefix}: Latest Stage",  "type": "related", "dataType": "text"},
        {"id": "deal.status", "label": f"{prefix}: Latest Status", "type": "related", "dataType": "text"},
        {"id": "deal.cdate",  "label": f"{prefix}: Latest Date",   "type": "related", "dataType": "date"},
        {"id": "deal_count",  "label": f"{prefix}: Count",         "type": "related", "dataType": "number"},
    ]


def _related_training_summary_fields(prefix: str) -> list:
    return [
        {"id": "training.training-type",   "label": f"{prefix}: Latest Type",    "type": "related", "dataType": "text"},
        {"id": "training.training-agenda", "label": f"{prefix}: Latest Agenda",  "type": "related", "dataType": "text"},
        {"id": "training.date-of-training","label": f"{prefix}: Latest Date",    "type": "related", "dataType": "date"},
        {"id": "training.trained-by",      "label": f"{prefix}: Trained By",     "type": "related", "dataType": "text"},
        {"id": "training_count",           "label": f"{prefix}: Count",          "type": "related", "dataType": "number"},
    ]


def _related_notes_summary_fields(prefix: str) -> list:
    return [
        {"id": "note_count",       "label": f"{prefix}: Count",       "type": "related", "dataType": "number"},
        {"id": "latest_note_date", "label": f"{prefix}: Latest Date", "type": "related", "dataType": "date"},
        {"id": "latest_note",      "label": f"{prefix}: Latest Text", "type": "related", "dataType": "text"},
    ]


@app.get("/api/field-values/{object_type}/{field_id}")
async def get_field_values(object_type: str, field_id: str):
    """Return unique values for a field (for dropdown filters in UI)."""
    values: set = set()
    try:
        if object_type == "slp":
            # Use in-memory SLP cache — fast, already loaded at startup
            records = _slp_cache_records if _slp_cache_records else await get_slp_cache()
            for r in records:
                for fo in r.get("fields", []):
                    if fo.get("id") == field_id and fo.get("value"):
                        values.add(str(fo["value"]))
        elif object_type in {"trainings", "license_details"}:
            schema_map = {"trainings": TRAINING_SCHEMA_ID, "license_details": LICENSE_SCHEMA_ID}
            records = await ac_get_all(f"customObjects/records/{schema_map[object_type]}", "records", {})
            for r in records[:2000]:
                for fo in r.get("fields", []):
                    if fo.get("id") == field_id and fo.get("value"):
                        values.add(str(fo["value"]))
        elif object_type == "accounts":
            records = await ac_get_all("accounts", "accounts", {})
            for r in records[:2000]:
                val = r.get(field_id)
                if val:
                    values.add(str(val))
    except Exception:
        pass
    return {"values": sorted(values)}


# ═══════════════════════════════════════════════════════════════════════════
# FILTER EVALUATION
# ═══════════════════════════════════════════════════════════════════════════

def evaluate_filter(record: dict, f: dict) -> bool:
    from datetime import timezone
    field      = f.get("field")
    ftype      = f.get("type", "text")
    operator   = f.get("operator", "equals")
    value      = f.get("value")
    values     = f.get("values", [])
    date_range = f.get("dateRange")

    if not field:
        return True
    if ftype != "date" and not value and not values:
        return True

    rv = record.get(field)

    if ftype == "text":
        if rv is None:
            return False
        rs = str(rv).lower()
        if values:
            return any(str(v).lower() in rs for v in values)
        vs = str(value).lower()
        if operator == "equals":      return rs == vs
        if operator == "contains":    return vs in rs
        if operator == "starts_with": return rs.startswith(vs)
        if operator == "not_equals":  return rs != vs

    elif ftype == "dropdown":
        if rv is None:
            return False
        if values: return str(rv) in values
        if value:  return str(rv) == str(value)

    elif ftype == "date":
        if not rv or rv == "null":
            return False
        try:
            if isinstance(rv, str):
                if not rv.strip():
                    return False
                rd = datetime.fromisoformat(rv.replace("Z", "+00:00")) if "T" in rv else datetime.strptime(rv[:10], "%Y-%m-%d")
            else:
                rd = rv
            if rd.tzinfo is None:
                rd = rd.replace(tzinfo=timezone.utc)
            now = datetime.now(timezone.utc)

            if date_range == "today":         return rd.date() == now.date()
            if date_range == "yesterday":     return rd.date() == (now - timedelta(days=1)).date()
            if date_range == "this_week":
                start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
                return rd >= start
            if date_range == "last_7_days":   return rd >= now - timedelta(days=7)
            if date_range == "last_30_days":  return rd >= now - timedelta(days=30)
            if date_range == "last_90_days":  return rd >= now - timedelta(days=90)
            if date_range == "next_30_days":  return now <= rd <= now + timedelta(days=30)
            if date_range == "next_90_days":  return now <= rd <= now + timedelta(days=90)
            if date_range == "this_month":
                start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                end   = (start.replace(month=start.month % 12 + 1, day=1) if start.month < 12
                         else start.replace(year=start.year + 1, month=1, day=1))
                return start <= rd < end
            if date_range == "last_month":
                end   = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                start = (end.replace(month=end.month - 1) if end.month > 1
                         else end.replace(year=end.year - 1, month=12))
                return start <= rd < end
            if date_range == "this_quarter":
                q     = (now.month - 1) // 3
                start = now.replace(month=q * 3 + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
                return rd >= start
            if date_range == "last_quarter":
                q = (now.month - 1) // 3
                if q == 0:
                    start = now.replace(year=now.year - 1, month=10, day=1, hour=0, minute=0, second=0, microsecond=0)
                    end   = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
                else:
                    start = now.replace(month=(q - 1) * 3 + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
                    end   = now.replace(month=q * 3 + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
                return start <= rd < end
            if date_range == "ytd":
                start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
                return rd >= start
            if date_range == "this_year":     return rd.year == now.year
            if date_range == "last_year":     return rd.year == now.year - 1
            if date_range == "custom":
                fd = f.get("fromDate")
                td = f.get("toDate")
                if fd and rd < datetime.strptime(fd, "%Y-%m-%d").replace(tzinfo=timezone.utc):
                    return False
                if td and rd > datetime.strptime(td, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc):
                    return False
                return True
        except Exception:
            return False

    return True


# ═══════════════════════════════════════════════════════════════════════════
# PRIMARY RECORD FETCHERS
# ═══════════════════════════════════════════════════════════════════════════

def _flatten_custom_object(r: dict) -> dict:
    flat = {"id": r.get("id"), "_relationships": r.get("relationships", {})}
    for fo in r.get("fields", []):
        flat[fo.get("id")] = fo.get("value")
    return flat

async def fetch_slp_records() -> list:
    raw = await get_slp_cache()
    return [_flatten_custom_object(r) for r in raw]

async def fetch_license_records() -> list:
    raw = await ac_get_all(f"customObjects/records/{LICENSE_SCHEMA_ID}", "records", {})
    return [_flatten_custom_object(r) for r in raw]

async def fetch_training_records() -> list:
    raw = await ac_get_all(f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {})
    return [_flatten_custom_object(r) for r in raw]

async def fetch_account_records() -> list:
    accounts_data = await ac_get_all("accounts", "accounts", {})
    records = []
    for acc in accounts_data:
        flat = {"id": acc.get("id")}
        for key, val in acc.items():
            if key != "links" and not isinstance(val, dict):
                flat[key] = val
        records.append(flat)
    print(f"Fetching custom fields for {len(records)} accounts...")
    for i in range(0, len(records), 500):
        batch = records[i:i+500]
        print(f"  {i}/{len(records)}", end="\r")
        await asyncio.gather(*[_fetch_account_cf(rec) for rec in batch], return_exceptions=True)
    print(f"  ✓ {len(records)} accounts done")
    return records

async def fetch_account_records_basic() -> list:
    accounts_data = await ac_get_all("accounts", "accounts", {})
    records = []
    for acc in accounts_data:
        flat = {"id": acc.get("id")}
        for key, val in acc.items():
            if key != "links" and not isinstance(val, dict):
                flat[key] = val
        records.append(flat)
    print(f"Fetched {len(records)} accounts (basic)")
    return records

async def _fetch_account_cf(record: dict):
    aid    = record["id"]
    cached = get_cached("account_custom_fields", aid)
    if cached:
        record.update(cached)
        return
    async with semaphore:
        try:
            data = await ac_get(f"accounts/{aid}/accountCustomFieldData")
            cfs  = {}
            for cf in data.get("customerAccountCustomFieldData", []):
                fid = cf.get("custom_field_id")
                val = (cf.get("custom_field_text_value") or cf.get("custom_field_date_value") or
                       cf.get("custom_field_number_value") or cf.get("custom_field_currency_value"))
                if fid and val is not None:
                    key = f"customfield_{fid}"
                    record[key] = val
                    cfs[key]    = val
            set_cached("account_custom_fields", aid, cfs)
        except Exception:
            pass

async def fetch_contact_records() -> list:
    contacts_data = await ac_get_all("contacts", "contacts", {})
    records = []
    for c in contacts_data:
        flat = {"id": c.get("id"), "_account_id": c.get("account")}
        for key, val in c.items():
            if key not in ("links", "fieldValues") and not isinstance(val, dict):
                flat[key] = val
        records.append(flat)
    print(f"Fetching custom fields for {len(records)} contacts...")
    for i in range(0, len(records), 500):
        await asyncio.gather(*[_fetch_contact_cf(rec) for rec in records[i:i+500]], return_exceptions=True)
    print(f"  ✓ {len(records)} contacts done")
    return records

async def fetch_contact_records_basic() -> list:
    contacts_data = await ac_get_all("contacts", "contacts", {})
    records = []
    for c in contacts_data:
        flat = {"id": c.get("id"), "_account_id": c.get("account")}
        for key, val in c.items():
            if key not in ("links", "fieldValues") and not isinstance(val, dict):
                flat[key] = val
        records.append(flat)
    print(f"Fetched {len(records)} contacts (basic)")
    return records

async def _fetch_contact_cf(record: dict):
    cid    = record["id"]
    cached = get_cached("contact_custom_fields", cid)
    if cached:
        record.update(cached)
        return
    async with semaphore:
        try:
            data = await ac_get(f"contacts/{cid}/fieldValues")
            cfs  = {}
            for fv in data.get("fieldValues", []):
                fid = fv.get("field")
                val = fv.get("value")
                if fid and val is not None:
                    key = f"customfield_{fid}"
                    record[key] = val
                    cfs[key]    = val
            set_cached("contact_custom_fields", cid, cfs)
        except Exception:
            pass

async def fetch_deal_records() -> list:
    deals_data = await ac_get_all("deals", "deals", {})
    records = []
    for d in deals_data:
        flat = {"id": d.get("id"), "_contact_id": d.get("contact"), "_account_id": d.get("account")}
        for key, val in d.items():
            if key != "links" and not isinstance(val, dict):
                flat[key] = val
        records.append(flat)
    print(f"Fetching custom fields for {len(records)} deals...")
    for i in range(0, len(records), 500):
        await asyncio.gather(*[_fetch_deal_cf(rec) for rec in records[i:i+500]], return_exceptions=True)
    print(f"  ✓ {len(records)} deals done")
    return records

async def fetch_deal_records_basic() -> list:
    deals_data = await ac_get_all("deals", "deals", {})
    records = []
    for d in deals_data:
        flat = {"id": d.get("id"), "_contact_id": d.get("contact"), "_account_id": d.get("account")}
        for key, val in d.items():
            if key != "links" and not isinstance(val, dict):
                flat[key] = val
        records.append(flat)
    print(f"Fetched {len(records)} deals (basic)")
    return records

async def _fetch_deal_cf(record: dict):
    did    = record["id"]
    cached = get_cached("deal_custom_fields", did)
    if cached:
        record.update(cached)
        return
    async with semaphore:
        try:
            data = await ac_get(f"deals/{did}/dealCustomFieldData")
            cfs  = {}
            for cf in data.get("dealCustomFieldData", []):
                fid = cf.get("customFieldId")
                val = cf.get("fieldValue")
                if fid and val is not None:
                    key = f"customfield_{fid}"
                    record[key] = val
                    cfs[key]    = val
            set_cached("deal_custom_fields", did, cfs)
        except Exception:
            pass


async def fetch_note_records() -> list:
    """Fetch all notes from AC, enriched with contact name/email and account ID."""
    print("Fetching notes...")
    raw = await ac_get_all("notes", "notes", {})

    records      = []
    contact_ids  = set()
    deal_ids     = set()

    for n in raw:
        reltype = (n.get("reltype") or "").lower()
        rec = {
            "id":      n.get("id"),
            "note":    n.get("note", ""),
            "cdate":   n.get("cdate", ""),
            "mdate":   n.get("mdate", ""),
            "userid":  n.get("userid", ""),
            "reltype": n.get("reltype", ""),
            "rel_id":  str(n.get("rel_id", "")),
            "pinned":  str(n.get("pinned", "0")),
            "_contact_id": str(n.get("rel_id", "")) if reltype == "contact" else None,
            "_deal_id":    str(n.get("rel_id", "")) if reltype == "deal"    else None,
            "_account_id": None,
        }
        if reltype == "contact":
            contact_ids.add(str(n.get("rel_id", "")))
        elif reltype == "deal":
            deal_ids.add(str(n.get("rel_id", "")))
        records.append(rec)

    print(f"  {len(records)} notes ({len(contact_ids)} contacts, {len(deal_ids)} deals)")

    # Batch-fetch contacts to get name + their account ID
    contact_map: dict = {}
    async def _fc(cid):
        try:
            async with semaphore:
                d = await ac_get(f"contacts/{cid}")
                c = d.get("contact", {})
                contact_map[cid] = {
                    "firstName":   c.get("firstName", ""),
                    "lastName":    c.get("lastName", ""),
                    "email":       c.get("email", ""),
                    "_account_id": str(c.get("account", "")) if c.get("account") else "",
                }
        except Exception:
            contact_map[cid] = {}

    if contact_ids:
        await asyncio.gather(*[_fc(cid) for cid in contact_ids], return_exceptions=True)

    # Batch-fetch deals to get their account ID
    deal_map: dict = {}
    async def _fd(did):
        try:
            async with semaphore:
                d = await ac_get(f"deals/{did}")
                dl = d.get("deal", {})
                deal_map[did] = {"_account_id": str(dl.get("account", "")) if dl.get("account") else ""}
        except Exception:
            deal_map[did] = {}

    if deal_ids:
        await asyncio.gather(*[_fd(did) for did in deal_ids], return_exceptions=True)

    # Enrich records with contact/account info
    for rec in records:
        cid = rec.pop("_contact_id", None)
        did = rec.pop("_deal_id",    None)
        if cid and cid in contact_map:
            cm = contact_map[cid]
            rec["contact.firstName"] = cm.get("firstName", "")
            rec["contact.lastName"]  = cm.get("lastName", "")
            rec["contact.email"]     = cm.get("email", "")
            rec["_account_id"]       = cm.get("_account_id", "")
        if did and did in deal_map:
            rec["_account_id"] = deal_map[did].get("_account_id", "")

    print(f"  ✓ Notes enriched")
    return records


# ═══════════════════════════════════════════════════════════════════════════
# ENRICHMENT HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _account_id_for(rec: dict, source_type: str) -> str | None:
    if source_type in ("slp", "license_details", "trainings"):
        rel = rec.get("_relationships", {}).get("account", [])
        return str(rel[0]) if isinstance(rel, list) and rel else None
    if source_type in ("contacts", "deals", "notes"):
        v = rec.get("_account_id")
        return str(v) if v else None
    if source_type == "accounts":
        return str(rec.get("id", ""))
    return None


async def enrich_with_accounts(records: list, source_type: str, field_list: list = []) -> list:
    """Merge account fields into records."""
    account_ids = {aid for rec in records if (aid := _account_id_for(rec, source_type))}
    if not account_ids:
        return records

    need_cf = any("account.customfield_" in f for f in field_list)
    print(f"Enriching with {len(account_ids)} accounts (custom_fields={need_cf})...")

    accounts_map: dict = {}
    cf_map: dict       = {}

    # Fetch in parallel with semaphore
    async def _fetch_one(aid):
        try:
            d = await ac_get(f"accounts/{aid}")
            accounts_map[aid] = d.get("account", {})
            if need_cf:
                d2 = await ac_get(f"accounts/{aid}/accountCustomFieldData")
                cf_map[aid] = {}
                for cf in d2.get("customerAccountCustomFieldData", []):
                    fid = cf.get("custom_field_id")
                    val = (cf.get("custom_field_text_value") or cf.get("custom_field_date_value") or
                           cf.get("custom_field_number_value") or cf.get("custom_field_currency_value"))
                    if fid and val is not None:
                        cf_map[aid][fid] = val
        except Exception:
            accounts_map[aid] = {}

    await asyncio.gather(*[_fetch_one(aid) for aid in account_ids], return_exceptions=True)

    for rec in records:
        aid = _account_id_for(rec, source_type)
        if aid and aid in accounts_map:
            for k, v in accounts_map[aid].items():
                if k != "links" and not isinstance(v, dict):
                    rec[f"account.{k}"] = v
            if need_cf and aid in cf_map:
                for fid, val in cf_map[aid].items():
                    rec[f"account.customfield_{fid}"] = val

    for rec in records:
        rec.pop("_relationships", None)
        rec.pop("_account_id",   None)

    return records


async def enrich_with_contacts(records: list, source_type: str) -> list:
    """Add 1:1 contact fields to deal records."""
    contact_ids = set()
    for rec in records:
        if source_type == "deals":
            cid = rec.get("_contact_id")
            if cid:
                contact_ids.add(str(cid))
    if not contact_ids:
        return records

    contacts_map: dict = {}
    async def _fetch(cid):
        try:
            d = await ac_get(f"contacts/{cid}")
            contacts_map[cid] = d.get("contact", {})
        except Exception:
            contacts_map[cid] = {}

    await asyncio.gather(*[_fetch(cid) for cid in contact_ids], return_exceptions=True)

    for rec in records:
        cid = rec.get("_contact_id")
        if cid and str(cid) in contacts_map:
            for k, v in contacts_map[str(cid)].items():
                if k not in ("links", "fieldValues") and not isinstance(v, dict):
                    rec[f"contact.{k}"] = v

    for rec in records:
        rec.pop("_contact_id", None)

    return records


async def enrich_with_contacts_list(records: list, source_type: str) -> list:
    """Add primary contact + contact_count to account/SLP/training records."""
    account_ids = {aid for rec in records if (aid := _account_id_for(rec, source_type))}
    if not account_ids:
        return records

    print(f"Fetching contacts for {len(account_ids)} accounts...")
    all_contacts = await ac_get_all("contacts", "contacts", {})

    by_account: dict = defaultdict(list)
    for c in all_contacts:
        aid = str(c.get("account", ""))
        if aid in account_ids:
            by_account[aid].append(c)

    for rec in records:
        aid      = _account_id_for(rec, source_type)
        contacts = by_account.get(aid, [])
        rec["contact_count"] = len(contacts)
        if contacts:
            p = contacts[0]
            rec["primary_contact.email"]     = p.get("email", "")
            rec["primary_contact.firstName"] = p.get("firstName", "")
            rec["primary_contact.lastName"]  = p.get("lastName", "")
            rec["primary_contact.phone"]     = p.get("phone", "")
        else:
            rec["primary_contact.email"]     = ""
            rec["primary_contact.firstName"] = ""
            rec["primary_contact.lastName"]  = ""
            rec["primary_contact.phone"]     = ""

    return records


async def enrich_with_slp(records: list, source_type: str) -> list:
    """Add SLP fields + count to account/contact records."""
    account_ids = {aid for rec in records if (aid := _account_id_for(rec, source_type))}
    if not account_ids:
        return records

    print(f"Fetching SLP records for {len(account_ids)} accounts...")
    all_slp = await get_slp_cache()

    by_account: dict = defaultdict(list)
    for r in all_slp:
        for aid in r.get("relationships", {}).get("account", []):
            aid = str(aid)
            if aid in account_ids:
                by_account[aid].append({fo["id"]: fo.get("value") for fo in r.get("fields", [])})

    for rec in records:
        aid  = _account_id_for(rec, source_type)
        slps = by_account.get(aid, [])
        rec["slp._count"] = len(slps)
        if slps:
            for fid, val in slps[0].items():
                rec[f"slp.{fid}"] = val

    return records


async def enrich_with_deals_summary(records: list, source_type: str) -> list:
    """Add deal count + latest deal info to account/SLP/contact records."""
    account_ids = set()
    contact_ids = set()
    for rec in records:
        if source_type == "contacts":
            cid = rec.get("id")
            if cid:
                contact_ids.add(str(cid))
        else:
            aid = _account_id_for(rec, source_type)
            if aid:
                account_ids.add(aid)

    if not account_ids and not contact_ids:
        return records

    print("Fetching deals for enrichment...")
    all_deals = await ac_get_all("deals", "deals", {})

    by_account: dict = defaultdict(list)
    by_contact: dict = defaultdict(list)
    for d in all_deals:
        aid = str(d.get("account", ""))
        cid = str(d.get("contact", ""))
        if aid in account_ids:
            by_account[aid].append(d)
        if cid in contact_ids:
            by_contact[cid].append(d)

    for rec in records:
        if source_type == "contacts":
            deals = by_contact.get(str(rec.get("id", "")), [])
        else:
            deals = by_account.get(_account_id_for(rec, source_type) or "", [])

        rec["deal_count"] = len(deals)
        if deals:
            latest = max(deals, key=lambda d: d.get("cdate", ""))
            rec["deal.title"]  = latest.get("title", "")
            rec["deal.status"] = latest.get("status", "")
            rec["deal.stage"]  = latest.get("stage", "")
            rec["deal.cdate"]  = latest.get("cdate", "")

    return records


async def enrich_with_trainings_summary(records: list, source_type: str) -> list:
    """Add training count + latest training info to account/SLP records."""
    account_ids = {aid for rec in records if (aid := _account_id_for(rec, source_type))}
    if not account_ids:
        return records

    print("Fetching trainings for enrichment...")
    all_trainings = await ac_get_all(f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {})

    by_account: dict = defaultdict(list)
    for r in all_trainings:
        for aid in r.get("relationships", {}).get("account", []):
            aid = str(aid)
            if aid in account_ids:
                by_account[aid].append({fo["id"]: fo.get("value") for fo in r.get("fields", [])})

    for rec in records:
        aid       = _account_id_for(rec, source_type)
        trainings = by_account.get(aid, [])
        rec["training_count"] = len(trainings)
        if trainings:
            latest = max(trainings, key=lambda t: t.get("date-of-training", ""))
            rec["training.training-type"]    = latest.get("training-type", "")
            rec["training.training-agenda"]  = latest.get("training-agenda", "")
            rec["training.date-of-training"] = latest.get("date-of-training", "")
            rec["training.trained-by"]       = latest.get("trained-by", "")

    return records


async def enrich_with_notes_summary(records: list, source_type: str) -> list:
    """Add note_count, latest_note_date, latest_note to contact/account/SLP/etc. records."""
    print("Fetching notes for enrichment...")
    all_notes = await ac_get_all("notes", "notes", {})

    if source_type == "contacts":
        # Match notes directly to contact IDs
        entity_ids = {str(rec.get("id", "")) for rec in records if rec.get("id")}
        by_entity: dict = defaultdict(list)
        for n in all_notes:
            if (n.get("reltype") or "").lower() == "contact":
                cid = str(n.get("rel_id", ""))
                if cid in entity_ids:
                    by_entity[cid].append(n)

        for rec in records:
            notes = sorted(by_entity.get(str(rec.get("id", "")), []),
                           key=lambda n: n.get("cdate", ""), reverse=True)
            rec["note_count"]       = len(notes)
            rec["latest_note_date"] = notes[0].get("cdate", "") if notes else ""
            rec["latest_note"]      = (notes[0].get("note", "") or "")[:300] if notes else ""

    else:
        # For accounts / SLP / trainings / license_details:
        # notes belong to contacts → need contact→account map
        account_ids = {aid for rec in records if (aid := _account_id_for(rec, source_type))}
        if not account_ids:
            return records

        all_contacts = await ac_get_all("contacts", "contacts", {})
        contact_to_account: dict = {}
        for c in all_contacts:
            aid = str(c.get("account", ""))
            if aid in account_ids:
                contact_to_account[str(c.get("id", ""))] = aid

        by_account: dict = defaultdict(list)
        for n in all_notes:
            if (n.get("reltype") or "").lower() == "contact":
                cid = str(n.get("rel_id", ""))
                if cid in contact_to_account:
                    by_account[contact_to_account[cid]].append(n)

        for rec in records:
            aid   = _account_id_for(rec, source_type) or ""
            notes = sorted(by_account.get(aid, []),
                           key=lambda n: n.get("cdate", ""), reverse=True)
            rec["note_count"]       = len(notes)
            rec["latest_note_date"] = notes[0].get("cdate", "") if notes else ""
            rec["latest_note"]      = (notes[0].get("note", "") or "")[:300] if notes else ""

    return records


def deduplicate_records(records: list, dedup_field: str) -> list:
    seen: dict  = {}
    deduped = []
    for rec in records:
        key = rec.get(dedup_field)
        if not key:
            deduped.append(rec)
            continue
        if key not in seen:
            seen[key] = len(deduped)
            deduped.append(rec)
        else:
            # Keep the record with the most recent activation date
            existing_idx  = seen[key]
            existing_date = deduped[existing_idx].get("contractor-activated-date", "")
            current_date  = rec.get("contractor-activated-date", "")
            if current_date > existing_date:
                deduped[existing_idx] = rec
    return deduped


# ═══════════════════════════════════════════════════════════════════════════
# MAIN REPORT ENDPOINT
# ═══════════════════════════════════════════════════════════════════════════

RELATED_PREFIXES = ("account.", "contact.", "deal.", "slp.", "training.", "primary_contact.")

@app.get("/api/report")
async def generate_report(
    object_type: str          = Query(...),
    fields:      str          = Query(...),
    filters:     Optional[str]= Query(None),
    dedup_field: Optional[str]= Query(None),
):
    field_list  = [f for f in fields.split(",") if f] if fields else []
    filter_list = json.loads(filters) if filters else []

    print(f"\n{'='*60}\nREPORT: {object_type} | fields={len(field_list)} filters={len(filter_list)}\n{'='*60}")

    # ── Fetch primary records ──
    needs_cf    = any(f.startswith("customfield_") for f in field_list)
    needs_cf_f  = any(f.get("field", "").startswith("customfield_") for f in filter_list)

    if object_type == "slp":
        records = await fetch_slp_records()
    elif object_type == "license_details":
        records = await fetch_license_records()
    elif object_type == "trainings":
        records = await fetch_training_records()
    elif object_type == "accounts":
        records = await fetch_account_records() if (needs_cf or needs_cf_f) else await fetch_account_records_basic()
    elif object_type == "contacts":
        records = await fetch_contact_records() if (needs_cf or needs_cf_f) else await fetch_contact_records_basic()
    elif object_type == "deals":
        records = await fetch_deal_records() if (needs_cf or needs_cf_f) else await fetch_deal_records_basic()
    elif object_type == "notes":
        records = await fetch_note_records()
    else:
        raise HTTPException(status_code=400, detail=f"Unknown object type: {object_type}")

    print(f"Fetched {len(records)} primary records")

    # ── Primary filters ──
    primary_filters = [f for f in filter_list if not any(f.get("field","").startswith(p) for p in RELATED_PREFIXES)]
    related_filters = [f for f in filter_list if     any(f.get("field","").startswith(p) for p in RELATED_PREFIXES)]

    if primary_filters:
        records = [r for r in records if all(evaluate_filter(r, f) for f in primary_filters)]
        print(f"After primary filters: {len(records)}")

    # ── Cross-object enrichment (only fetch what's needed) ──
    fl = set(field_list)
    rf = set(f.get("field", "") for f in related_filters)
    all_fields = fl | rf

    if any(f.startswith("account.") for f in all_fields):
        records = await enrich_with_accounts(records, object_type, field_list)

    if any(f.startswith("contact.") for f in all_fields) and object_type == "deals":
        records = await enrich_with_contacts(records, object_type)

    if any(f.startswith("primary_contact.") or f == "contact_count" for f in all_fields):
        records = await enrich_with_contacts_list(records, object_type)

    if any(f.startswith("slp.") or f == "slp._count" for f in all_fields):
        records = await enrich_with_slp(records, object_type)

    if any(f.startswith("deal.") or f == "deal_count" for f in all_fields):
        records = await enrich_with_deals_summary(records, object_type)

    if any(f.startswith("training.") or f == "training_count" for f in all_fields):
        records = await enrich_with_trainings_summary(records, object_type)

    if any(f in ("note_count", "latest_note", "latest_note_date") for f in all_fields):
        records = await enrich_with_notes_summary(records, object_type)

    # ── Related filters ──
    if related_filters:
        records = [r for r in records if all(evaluate_filter(r, f) for f in related_filters)]
        print(f"After related filters: {len(records)}")

    # ── Dedup ──
    if dedup_field:
        records = deduplicate_records(records, dedup_field)
        print(f"After dedup: {len(records)}")

    # ── Project fields ──
    final = [{fid: r.get(fid, "") for fid in field_list} for r in records]
    print(f"Returning {len(final)} records\n")
    return {"count": len(final), "records": final}


# ═══════════════════════════════════════════════════════════════════════════
# CSV EXPORT
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/report/csv")
async def export_csv(
    object_type: str          = Query(...),
    fields:      str          = Query(...),
    filters:     Optional[str]= Query(None),
    dedup_field: Optional[str]= Query(None),
):
    result  = await generate_report(object_type, fields, filters, dedup_field)
    records = result["records"]
    if not records:
        raise HTTPException(status_code=404, detail="No records to export")

    filter_list = json.loads(filters) if filters else []
    fields_data = await get_fields(object_type)
    field_labels = {f["id"]: f["label"] for f in fields_data["fields"]}

    output = io.StringIO()
    output.write(f"AC Reporter Export\n")
    output.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
    output.write(f"Object: {object_type.upper()} | Records: {len(records)}\n")

    if filter_list:
        output.write("\nFilters:\n")
        for f in filter_list:
            lbl     = field_labels.get(f.get("field"), f.get("field"))
            dr      = f.get("dateRange")
            display = dr or f"{f.get('operator', '=')} {f.get('value', '')}"
            output.write(f"  - {lbl}: {display}\n")

    if dedup_field:
        output.write(f"\nDedup by: {field_labels.get(dedup_field, dedup_field)}\n")

    output.write("\n" + "=" * 80 + "\n\n")

    if records:
        fieldnames = list(records[0].keys())
        headers    = [field_labels.get(fid, fid) for fid in fieldnames]
        writer = csv.writer(output)
        writer.writerow(headers)
        for rec in records:
            writer.writerow([rec.get(k, "") for k in fieldnames])

    filename = f"report_{object_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: PARTNER ACTIVATIONS
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/report/activations")
async def activations_report(
    from_date:        Optional[str] = Query(None, description="YYYY-MM-DD"),
    to_date:          Optional[str] = Query(None, description="YYYY-MM-DD"),
    platform:         Optional[str] = Query(None),
    bdr:              Optional[str] = Query(None),
    state:            Optional[str] = Query(None, description="2-letter state abbreviation"),
    exclude_platforms:Optional[str] = Query(None, description="Comma-separated"),
    format:           str           = Query("json"),
):
    """Partner activations: SLP records with Contractor Activated status, joined to accounts."""
    from datetime import timezone
    print("\nActivations report...")
    slp_records = await get_slp_cache()
    exclude_set = {p.strip() for p in exclude_platforms.split(",")} if exclude_platforms else set()

    account_ids: set = set()
    candidates  = []

    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}

        if fields.get("slp-status-detail") != "Contractor Activated":
            continue

        plat      = str(fields.get("channel", "")).strip()
        plat_norm = _normalize_platform(plat)
        if platform and plat_norm != _normalize_platform(platform):
            continue
        if plat_norm in exclude_set or plat in exclude_set:
            continue

        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None

        slp_bdr = str(fields.get("assigned-bdr", "")).strip()
        # Fall back to account-level BDR (CF119) when SLP has none
        eff_bdr = slp_bdr or _account_to_bdr.get(acc_id or "", "")
        if bdr == "__unassigned__":
            if eff_bdr:
                continue   # skip records that DO have a BDR
        elif bdr and eff_bdr != bdr:
            continue

        if state:
            states_val = str(fields.get("doing-business-in-states", "") or "").upper()
            if state.upper() not in [s.strip() for s in states_val.split(",")]:
                continue

        act_str = str(fields.get("contractor-activated-date", "")).strip()
        if not act_str:
            continue
        try:
            act_dt = (datetime.fromisoformat(act_str.replace("Z", "+00:00")) if "T" in act_str
                      else datetime.strptime(act_str[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc))
        except Exception:
            continue

        if from_date and act_dt < datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc):
            continue
        if to_date and act_dt > datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc):
            continue

        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id, "slp_id": r.get("id"), "eff_bdr": eff_bdr})

    print(f"  {len(candidates)} candidates")

    # Bulk-fetch only the CF fields we need in one paginated sweep — far faster
    # than one API call per account, especially for large date ranges.
    _needed_cf_ids = {
        ACCT_FIELD["dba_name"], ACCT_FIELD["doing_business_in"],
        ACCT_FIELD["sales_region"], ACCT_FIELD["oracle_producer_id"],
    }
    cf_map = await _fetch_acct_cf_map(_needed_cf_ids)
    acct_cache = {
        aid: {"name": _account_to_name.get(aid, ""), "cfs": cf_map.get(aid, {})}
        for aid in account_ids
    }

    results = []
    for c in candidates:
        f   = c["fields"]
        acc = acct_cache.get(c["account_id"], {"name": "", "cfs": {}}) if c["account_id"] else {"name": "", "cfs": {}}
        cfs = acc["cfs"]
        results.append({
            "slp_id":                    c["slp_id"],
            "account_id":                c["account_id"],
            "account_name":              acc["name"],
            "dba_name":                  cfs.get(ACCT_FIELD["dba_name"], ""),
            "dealer_id":                 f.get("dealer-id", ""),
            "channel":                   f.get("channel", ""),
            "slp_status":                f.get("slp-status-detail", ""),
            "contractor_activated_date": f.get("contractor-activated-date", ""),
            "original_owner":            f.get("original-owner", ""),
            "assigned_bdr":              c.get("eff_bdr") or f.get("assigned-bdr", ""),
            "sales_region":              cfs.get(ACCT_FIELD["sales_region"], ""),
            "oracle_producer_id":        cfs.get(ACCT_FIELD["oracle_producer_id"], ""),
            "doing_business_in_states":  cfs.get(ACCT_FIELD["doing_business_in"], "") or f.get("doing-business-in-states", ""),
            "ein":                       f.get("ein", ""),
            "contractor_reactivation":   f.get("contractor-reactivation", ""),
        })

    results.sort(key=lambda x: x.get("contractor_activated_date", ""), reverse=True)

    if format == "csv":
        out = io.StringIO()
        if results:
            w = csv.DictWriter(out, fieldnames=results[0].keys())
            w.writeheader(); w.writerows(results)
        fn = f"activations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})
    return {"count": len(results), "records": results}


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: ENROLLMENT (all SLP statuses, same date filters as activations)
# ═══════════════════════════════════════════════════════════════════════════

_SLP_STATUSES = [
    # Current workflow, in flow order
    "Pre-activation",
    "BDR Review",
    "Onboarding Review",
    "Mgmt Review",
    "BDR Follow-up",
    "Approved",
    "Awaiting Dealer Agreement Signature",
    "Awaiting Training",
    "Training Completed",
    "Contractor Activated",
    "Awaiting Contractor Response",
    "Declined",
    # Legacy / other statuses still present on older SLP records in AC
    "Declined by Onboarding",
    "Account on Hold (Suspended)",
    "Being Reviewed By Ops Manager",
    "Deactivated",
    "Deactivated for Dormancy",
    "Deactivated w/ Cause",
    "Declined by Lender",
    "In Progress – Other",
    "In Progress – Signed Contract Needed",
    "Not Active",
    "Not Started",
    "On Hold",
    "On Definite Hold – Signed Training Attestation Not Returned",
    "On Indefinite Hold - Agreement/Documents Not Signed",
    "On Indefinite Hold - Lender ID Not Provided",
    "On Indefinite Hold - Online Reviews Not Available",
    "On Indefinite Hold - Training Not Completed",
    "On Indefinite Hold - Valid BL or PL Not Provided",
    "Pending - Lender ID Verification",
    "Pending - Training Not Completed",
    "Pending - Waiting on Contractor to provide BL or PL License",
    "Pending - Waiting on Online Reviews",
    "Pending Activation - GS UCA Agreement",
    "Pending Lender Training",
    "Waiting_on_BDR_Approval",
    "Withdrawn",
]

@app.get("/api/report/enrollment")
async def enrollment_report(
    from_date:         Optional[str] = Query(None, description="YYYY-MM-DD — filter by contractor-activated-date or enrollment-request-date"),
    to_date:           Optional[str] = Query(None, description="YYYY-MM-DD"),
    slp_status:        Optional[str] = Query(None, description="Filter to a specific SLP status (leave blank for all)"),
    platform:          Optional[str] = Query(None),
    bdr:               Optional[str] = Query(None),
    state:             Optional[str] = Query(None, description="2-letter state abbreviation"),
    exclude_platforms: Optional[str] = Query(None, description="Comma-separated platforms to exclude"),
    format:            str           = Query("json"),
):
    """All SLP enrollments regardless of status, with optional status filter. Date range filters
    on contractor-activated-date when set, falling back to enrollment-request-date."""
    from datetime import timezone
    print("\nEnrollment report...")
    exclude_set = {p.strip() for p in exclude_platforms.split(",")} if exclude_platforms else set()
    from_dt = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) if from_date else None
    to_dt   = datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc) if to_date else None

    slp_records = await get_slp_cache()
    candidates  = []

    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None

        # Status filter
        rec_status = fields.get("slp-status-detail", "")
        if slp_status and rec_status != slp_status:
            continue

        # Platform filter
        plat      = str(fields.get("channel", "")).strip()
        plat_norm = _normalize_platform(plat)
        if platform and plat_norm != _normalize_platform(platform):
            continue
        if plat_norm in exclude_set or plat in exclude_set:
            continue

        # BDR filter — check SLP field first, fall back to account-level BDR cache
        slp_bdr = str(fields.get("assigned-bdr", "")).strip()
        eff_bdr = slp_bdr or _account_to_bdr.get(acc_id or "", "")
        if bdr == "__unassigned__":
            if eff_bdr:
                continue
        elif bdr and eff_bdr != bdr:
            continue

        # State filter
        if state:
            states_val = str(fields.get("doing-business-in-states", "") or "").upper()
            if state.upper() not in [s.strip() for s in states_val.split(",")]:
                continue

        # Date filter — use contractor-activated-date, fall back to enrollment-request-date
        if from_dt or to_dt:
            date_str = (str(fields.get("contractor-activated-date", "")).strip() or
                        str(fields.get("enrollment-request-date", "")).strip())
            if not date_str:
                continue
            try:
                rec_dt = (datetime.fromisoformat(date_str.replace("Z", "+00:00")) if "T" in date_str
                          else datetime.strptime(date_str[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc))
            except Exception:
                continue
            if from_dt and rec_dt < from_dt:
                continue
            if to_dt and rec_dt > to_dt:
                continue

        candidates.append({"fields": fields, "account_id": acc_id, "slp_id": r.get("id"),
                            "created_ts": r.get("createdTimestamp", ""), "updated_ts": r.get("updatedTimestamp", "")})

    print(f"  {len(candidates)} candidates")

    results = []
    for c in candidates:
        f   = c["fields"]
        aid = c["account_id"] or ""
        results.append({
            "slp_id":                    c["slp_id"],
            "account_id":                aid,
            "account_name":              _account_to_name.get(aid, ""),
            "slp_name":                  f.get("name", ""),
            "dba_name":                  _account_to_dba.get(aid, ""),
            "dealer_id":                 f.get("dealer-id", ""),
            "channel":                   f.get("channel", ""),
            "slp_status":                f.get("slp-status-detail", ""),
            "contractor_activated_date": f.get("contractor-activated-date", ""),
            "enrollment_request_date":   f.get("enrollment-request-date", ""),
            "original_owner":            f.get("original-owner", ""),
            "assigned_bdr":              f.get("assigned-bdr", ""),
            "sales_region":              _account_to_region.get(aid, ""),
            "oracle_producer_id":        _account_to_oracle_id.get(aid, ""),
            "doing_business_in_states":  _account_to_states.get(aid, "") or f.get("doing-business-in-states", ""),
            "ein":                       f.get("ein", ""),
            "contractor_reactivation":   f.get("contractor-reactivation", ""),
            "decline_reason":            f.get("decline-reason", ""),
            "status_changed_date":       f.get("status-changed-date", ""),
            "created_timestamp":         c.get("created_ts", ""),
            "updated_timestamp":         c.get("updated_ts", ""),
        })

    results.sort(key=lambda x: (x.get("contractor_activated_date") or x.get("enrollment_request_date") or ""), reverse=True)

    if format == "csv":
        out = io.StringIO()
        if results:
            w = csv.DictWriter(out, fieldnames=results[0].keys())
            w.writeheader(); w.writerows(results)
        fn = f"enrollment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})
    return {"count": len(results), "records": results, "valid_statuses": _SLP_STATUSES}


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: NOT ACTIVATED (SLPs without Contractor Activated status)
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/report/not-activated")
async def not_activated_report(
    from_date:         Optional[str] = Query(None, description="YYYY-MM-DD — filter by enrollment-request-date"),
    to_date:           Optional[str] = Query(None, description="YYYY-MM-DD"),
    platform:          Optional[str] = Query(None),
    bdr:               Optional[str] = Query(None),
    status:            Optional[str] = Query(None, description="Filter to a specific non-activated status"),
    state:             Optional[str] = Query(None, description="2-letter state abbreviation"),
    exclude_platforms: Optional[str] = Query(None, description="Comma-separated platforms to exclude"),
    format:            str           = Query("json"),
):
    """SLP records whose status is NOT 'Contractor Activated', joined to accounts."""
    from datetime import timezone
    print("\nNot-activated report...")
    exclude_set = {p.strip() for p in exclude_platforms.split(",")} if exclude_platforms else set()
    from_dt = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) if from_date else None
    to_dt   = datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc) if to_date else None

    slp_records = await ac_get_all(
        f"customObjects/records/{SLP_SCHEMA_ID}", "records", {}
    )

    account_ids: set = set()
    candidates  = []

    for r in slp_records:
        fields     = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        status_val = str(fields.get("slp-status-detail", "")).strip()
        if status_val in ("Contractor Activated", "Inactive", "Deactivated", "Deactivated for Dormancy", "Declined by Onboarding", "Not Active"):
            continue
        if status and status_val != status:
            continue

        plat      = str(fields.get("channel", "")).strip()
        plat_norm = _normalize_platform(plat)
        if platform and plat_norm != _normalize_platform(platform):
            continue
        if plat_norm in exclude_set or plat in exclude_set:
            continue

        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None

        slp_bdr = str(fields.get("assigned-bdr", "")).strip()
        eff_bdr = slp_bdr or _account_to_bdr.get(acc_id or "", "")
        if bdr and eff_bdr != bdr:
            continue

        if state:
            states_val = str(fields.get("doing-business-in-states", "") or "").upper()
            if state.upper() not in [s.strip() for s in states_val.split(",")]:
                continue

        if from_dt or to_dt:
            enroll_str = str(fields.get("enrollment-request-date") or "").strip()
            if not enroll_str:
                continue
            try:
                enroll_dt = datetime.fromisoformat(enroll_str.replace("Z", "+00:00"))
                if enroll_dt.tzinfo is None:
                    enroll_dt = enroll_dt.replace(tzinfo=timezone.utc)
                if from_dt and enroll_dt < from_dt:
                    continue
                if to_dt and enroll_dt > to_dt:
                    continue
            except Exception:
                continue

        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id, "eff_bdr": eff_bdr})

    print(f"  {len(candidates)} not-activated candidates")

    # Fetch account names
    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    results = []
    for c in candidates:
        aid = c["account_id"]
        f   = c["fields"]
        results.append({
            "account_name":  acct_cache.get(aid, ""),
            "account_id":    aid or "",
            "channel":       f.get("channel", ""),
            "dealer_id":     f.get("dealer-id", ""),
            "slp_status":    f.get("slp-status-detail", "") or "Not Started",
            "assigned_bdr":  c["eff_bdr"],
        })

    results.sort(key=lambda x: (x.get("slp_status", ""), x.get("account_name", "")))

    if format == "csv":
        out = io.StringIO()
        if results:
            w = csv.DictWriter(out, fieldnames=results[0].keys())
            w.writeheader(); w.writerows(results)
        fn = f"not_activated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})
    return {"count": len(results), "records": results}


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: LICENSE EXPIRATION
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/report/license-expiration")
async def license_expiration_report(
    days_ahead:      int            = Query(90),
    include_expired: bool           = Query(True),
    from_date:       Optional[date] = Query(None, description="Filter by expiration date ≥ this date"),
    to_date:         Optional[date] = Query(None, description="Filter by expiration date ≤ this date"),
    preset:          Optional[str]  = Query(None, description="Date preset: this_month | last_month | last_week | etc."),
    format:          str            = Query("json"),
):
    """License records filtered by expiration date. Supports days_ahead (future window) or
    explicit from_date/to_date or a named preset (this_month, last_month, etc.)."""
    from datetime import timezone
    print("\nLicense expiration report...")
    lic_records = await ac_get_all(f"customObjects/records/{LICENSE_SCHEMA_ID}", "records", {})
    now = datetime.now(timezone.utc)

    # Resolve date-range mode vs days-ahead mode
    use_range = bool(preset or from_date or to_date)
    range_start: Optional[date] = None
    range_end:   Optional[date] = None
    if use_range:
        range_start, range_end = _resolve_date_range(from_date, to_date, preset)
    else:
        cutoff = now + timedelta(days=days_ahead)

    account_ids: set = set()
    candidates  = []

    for r in lic_records:
        fields  = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        exp_str = (fields.get("expiration-date") or fields.get("license-expiration-date")
                   or fields.get("expires") or "")
        if not exp_str:
            continue
        try:
            exp_dt = (datetime.fromisoformat(str(exp_str).replace("Z", "+00:00")) if "T" in str(exp_str)
                      else datetime.strptime(str(exp_str)[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc))
        except Exception:
            continue

        is_expired = exp_dt < now
        if use_range:
            # Filter by whether the expiration date falls in the range
            exp_date_only = exp_dt.date()
            if range_start and exp_date_only < range_start:
                continue
            if range_end and exp_date_only > range_end:
                continue
        else:
            if is_expired and not include_expired:
                continue
            if not is_expired and exp_dt > cutoff:
                continue

        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id, "record_id": r.get("id"),
                           "expiration_date": exp_str, "is_expired": is_expired,
                           "days_until": (exp_dt - now).days})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    results = []
    for c in candidates:
        row = {"record_id": c["record_id"], "account_id": c["account_id"],
               "account_name": acct_cache.get(c["account_id"], ""),
               "expiration_date": c["expiration_date"],
               "days_until_expiration": c["days_until"],
               "status": "EXPIRED" if c["is_expired"] else "EXPIRING"}
        for k, v in c["fields"].items():
            row[f"license.{k}"] = v
        results.append(row)

    results.sort(key=lambda x: x.get("days_until_expiration", 9999))

    if format == "csv":
        out = io.StringIO()
        if results:
            # Union of all keys so rows with extra fields don't crash the writer
            all_keys: list = list(dict.fromkeys(k for row in results for k in row.keys()))
            w = csv.DictWriter(out, fieldnames=all_keys, extrasaction="ignore")
            w.writeheader(); w.writerows(results)
        fn = f"license_expiration_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})
    return {"count": len(results), "records": results}


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: BDR SUMMARY
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/report/bdr-summary")
async def bdr_summary_report(
    from_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    to_date:   Optional[str] = Query(None, description="YYYY-MM-DD"),
    platform:  Optional[str] = Query(None),
    format:    str           = Query("json"),
):
    """Activations, account counts, and platform breakdown per BDR."""
    from datetime import timezone
    print("\nBDR summary report...")
    try:
        slp_records = await get_slp_cache()
    except BaseException as _e:
        print(f"[bdr-summary] CAUGHT {type(_e).__name__}: {_e}")
        return JSONResponse(status_code=500, content={"detail": str(_e), "type": type(_e).__name__})

    bdr_data: dict = defaultdict(lambda: {"total_slps": 0, "activated": 0,
                                           "channels": defaultdict(int), "accounts": set()})
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        plat   = str(fields.get("channel", "")).strip()
        if platform and plat != platform:
            continue
        bdr    = str(fields.get("assigned-bdr", "")).strip() or "Unassigned"
        bdr_data[bdr]["total_slps"] += 1

        if fields.get("slp-status-detail") == "Contractor Activated":
            act_str  = str(fields.get("contractor-activated-date", "")).strip()
            in_range = True
            if act_str and (from_date or to_date):
                try:
                    act_dt = (datetime.fromisoformat(act_str.replace("Z", "+00:00")) if "T" in act_str
                              else datetime.strptime(act_str[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc))
                    if from_date:
                        in_range = in_range and act_dt >= datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    if to_date:
                        in_range = in_range and act_dt <= datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
                except Exception:
                    in_range = False
            if in_range:
                bdr_data[bdr]["activated"] += 1

        if plat:
            bdr_data[bdr]["channels"][plat] += 1

        rel = r.get("relationships", {}).get("account", [])
        if rel:
            bdr_data[bdr]["accounts"].add(str(rel[0]))

    results = [
        {"bdr": bdr, "total_slps": d["total_slps"], "activated": d["activated"],
         "account_count": len(d["accounts"]),
         "channels": ", ".join(f"{k}:{v}" for k, v in sorted(d["channels"].items()))}
        for bdr, d in sorted(bdr_data.items())
    ]
    results.sort(key=lambda x: x["activated"], reverse=True)

    if format == "csv":
        out = io.StringIO()
        if results:
            w = csv.DictWriter(out, fieldnames=results[0].keys())
            w.writeheader(); w.writerows(results)
        fn = f"bdr_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})
    return {"count": len(results), "records": results}


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: TRAINING SUMMARY
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/report/training-summary")
async def training_summary_report(
    from_date:     Optional[str] = Query(None),
    to_date:       Optional[str] = Query(None),
    trainer:       Optional[str] = Query(None),
    training_type: Optional[str] = Query(None),
    format:        str           = Query("json"),
):
    """Training records joined with account data."""
    from datetime import timezone
    print("\nTraining summary report...")
    training_records = await ac_get_all(f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {})

    account_ids: set = set()
    candidates  = []

    for r in training_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}

        if trainer       and str(fields.get("trained-by", "")).strip() != trainer:
            continue
        if training_type and str(fields.get("training-type", "")).strip() != training_type:
            continue

        date_str = str(fields.get("date-of-training", "")).strip()
        if date_str and (from_date or to_date):
            try:
                td = (datetime.fromisoformat(date_str.replace("Z", "+00:00")) if "T" in date_str
                      else datetime.strptime(date_str[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc))
                if from_date and td < datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc):
                    continue
                if to_date and td > datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc):
                    continue
            except Exception:
                pass

        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id, "record_id": r.get("id")})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            ad  = await ac_get(f"accounts/{aid}")
            acd = await ac_get(f"accounts/{aid}/accountCustomFieldData")
            cfs = {str(cf["custom_field_id"]): cf.get("custom_field_text_value") or ""
                   for cf in acd.get("customerAccountCustomFieldData", [])}
            acct_cache[aid] = {"name": ad.get("account", {}).get("name", ""), "cfs": cfs}
        except Exception:
            acct_cache[aid] = {"name": "", "cfs": {}}

    results = []
    for c in candidates:
        f   = c["fields"]
        acc = acct_cache.get(c["account_id"], {"name": "", "cfs": {}}) if c["account_id"] else {"name": "", "cfs": {}}
        results.append({
            "record_id":        c["record_id"],
            "account_id":       c["account_id"],
            "account_name":     acc["name"],
            "dealer_id":        acc["cfs"].get(ACCT_FIELD["dealer_id"], ""),
            "channel":          acc["cfs"].get(ACCT_FIELD["dealer_program"], ""),
            "training_type":    f.get("training-type", ""),
            "training_agenda":  f.get("training-agenda", ""),
            "date_of_training": f.get("date-of-training", ""),
            "trained_by":       f.get("trained-by", ""),
            "training_notes":   f.get("training-notes", ""),
        })

    results.sort(key=lambda x: x.get("date_of_training", ""), reverse=True)

    if format == "csv":
        out = io.StringIO()
        if results:
            w = csv.DictWriter(out, fieldnames=results[0].keys())
            w.writeheader(); w.writerows(results)
        fn = f"training_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})
    return {"count": len(results), "records": results}


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: DEALER 360 PROFILE
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/report/dealer-profile")
async def dealer_profile(
    account_id: Optional[str] = Query(None),
    dealer_id:  Optional[str] = Query(None),
):
    """360° view of one dealer: account + SLPs + contacts + deals + trainings."""
    if not account_id and not dealer_id:
        raise HTTPException(status_code=400, detail="Provide account_id or dealer_id")

    # Resolve dealer_id → account_id
    if not account_id and dealer_id:
        all_accounts = await ac_get_all("accounts", "accounts", {})
        for acc in all_accounts:
            try:
                acd = await ac_get(f"accounts/{acc['id']}/accountCustomFieldData")
                for cf in acd.get("customerAccountCustomFieldData", []):
                    if (str(cf.get("custom_field_id")) == ACCT_FIELD["dealer_id"] and
                            str(cf.get("custom_field_text_value", "")).strip() == str(dealer_id).strip()):
                        account_id = str(acc["id"])
                        break
            except Exception:
                pass
            if account_id:
                break

    if not account_id:
        raise HTTPException(status_code=404, detail="Account not found")

    # Fetch account + contacts in parallel
    acc_data, acc_cf_data, acc_contacts = await asyncio.gather(
        ac_get(f"accounts/{account_id}"),
        ac_get(f"accounts/{account_id}/accountCustomFieldData"),
        ac_get(f"accounts/{account_id}/contacts"),
        return_exceptions=True,
    )

    account   = acc_data.get("account", {})  if isinstance(acc_data, dict)     else {}
    acct_cfs  = {str(cf["custom_field_id"]): cf.get("custom_field_text_value") or ""
                 for cf in (acc_cf_data.get("customerAccountCustomFieldData", []) if isinstance(acc_cf_data, dict) else [])}
    contact_ids = [ac.get("contact") for ac in (acc_contacts.get("accountContacts", []) if isinstance(acc_contacts, dict) else [])]

    # Fetch SLPs, trainings, deals, contacts in parallel
    slp_task      = get_slp_cache()
    training_task = ac_get_all(f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {})
    deal_task     = ac_get_all("deals", "deals", {})

    all_slps, all_trainings, all_deals = await asyncio.gather(slp_task, training_task, deal_task)

    slps      = [{fo["id"]: fo.get("value") for fo in r.get("fields", [])} for r in all_slps
                 if str(account_id) in [str(x) for x in r.get("relationships", {}).get("account", [])]]
    trainings = [{fo["id"]: fo.get("value") for fo in r.get("fields", [])} for r in all_trainings
                 if str(account_id) in [str(x) for x in r.get("relationships", {}).get("account", [])]]
    deals     = [{"id": d.get("id"), "title": d.get("title"), "stage": d.get("stage"), "status": d.get("status")}
                 for d in all_deals if str(d.get("account", "")) == str(account_id)]

    contacts = []
    for cid in contact_ids:
        try:
            cd = await ac_get(f"contacts/{cid}")
            c  = cd.get("contact", {})
            contacts.append({"id": c.get("id"), "email": c.get("email"),
                             "firstName": c.get("firstName"), "lastName": c.get("lastName")})
        except Exception:
            pass

    return {
        "account":  {"id": account_id, "name": account.get("name", ""), "custom_fields": acct_cfs},
        "slps":     slps,
        "contacts": contacts,
        "deals":    deals,
        "trainings":trainings,
        "summary":  {"slp_count": len(slps), "contact_count": len(contacts),
                     "deal_count": len(deals), "training_count": len(trainings)},
    }


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: ACCOUNT ACTIVITY
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/report/account-activity")
async def account_activity_report(
    from_date:     Optional[str] = Query(None, description="YYYY-MM-DD"),
    to_date:       Optional[str] = Query(None, description="YYYY-MM-DD"),
    activity_type: Optional[str] = Query(None, description="Filter by activity type"),
    performed_by:  Optional[str] = Query(None, description="Filter by performed-by value"),
    format:        str           = Query("json"),
):
    """Account Activity custom object — counts grouped by type and performed-by."""
    print("\nAccount activity report (custom object)...")

    all_records = await ac_get_all(
        f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", "records", {}
    )

    from_d = datetime.strptime(from_date, "%Y-%m-%d").date() if from_date else None
    to_d   = datetime.strptime(to_date,   "%Y-%m-%d").date() if to_date   else None

    counts_by_type:   dict = defaultdict(int)
    accounts_by_type: dict = defaultdict(set)
    counts_by_person: dict = defaultdict(int)
    type_by_person:   dict = defaultdict(lambda: defaultdict(int))
    total = 0

    for r in all_records:
        fmap         = {f["id"]: f.get("value") for f in r.get("fields", [])}
        act_type     = (fmap.get("activity-type") or "Unknown").strip()
        act_date     = (fmap.get("activity-date") or "")[:10]
        performed    = (fmap.get("performed-by")  or "").strip()
        account_id   = next(iter(r.get("relationships", {}).get("account", [])), "")

        # Date filter
        if act_date and (from_d or to_d):
            try:
                ad = datetime.strptime(act_date, "%Y-%m-%d").date()
                if from_d and ad < from_d: continue
                if to_d   and ad > to_d:   continue
            except Exception:
                pass

        # Activity type filter
        if activity_type and act_type.lower() != activity_type.lower():
            continue

        # Performed-by filter
        if performed_by and performed_by.lower() not in performed.lower():
            continue

        counts_by_type[act_type] += 1
        accounts_by_type[act_type].add(account_id)
        if performed:
            counts_by_person[performed] += 1
            type_by_person[performed][act_type] += 1
        total += 1

    by_type = sorted(
        [{"activity_type": t, "count": c, "unique_accounts": len(accounts_by_type[t])}
         for t, c in counts_by_type.items()],
        key=lambda x: -x["count"]
    )
    by_person = sorted(
        [{"performed_by": p, "count": c,
          "breakdown": dict(sorted(type_by_person[p].items(), key=lambda x: -x[1]))}
         for p, c in counts_by_person.items()],
        key=lambda x: -x["count"]
    )

    if format == "csv":
        rows = []
        for row in by_person:
            for atype, cnt in row["breakdown"].items():
                rows.append({"performed_by": row["performed_by"],
                             "activity_type": atype, "count": cnt})
        if not rows:
            rows = [{"performed_by": "", "activity_type": t, "count": c}
                    for t, c in counts_by_type.items()]
        out = io.StringIO()
        if rows:
            w = csv.DictWriter(out, fieldnames=rows[0].keys())
            w.writeheader(); w.writerows(rows)
        fn = f"account_activity_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})

    return {"total": total, "by_type": by_type, "by_person": by_person}


# ═══════════════════════════════════════════════════════════════════════════
# PRE-BUILT REPORT: TEAM ACTIVITY / PERFORMANCE
# ═══════════════════════════════════════════════════════════════════════════

def _extract_performer(fmap: dict) -> str:
    """Return the performer name from an activity record.

    Tries `performed-by` and `performed-by-1` (actual AC field ID) first;
    if empty, parses the `name` field which often has the pattern
    'Call — Warren Neely' or 'Email — Amanda Jones'.
    """
    performed = (fmap.get("performed-by") or fmap.get("performed-by-1") or "").strip()
    if performed:
        return performed
    name = (fmap.get("name") or "").strip()
    for sep in (" — ", " – ", " - "):   # em-dash, en-dash, hyphen
        if sep in name:
            candidate = name.split(sep, 1)[1].strip()
            if candidate:
                return candidate
    return ""


@app.get("/api/report/team-activity")
async def team_activity_report(
    from_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    to_date:   Optional[str] = Query(None, description="YYYY-MM-DD"),
    format:    str           = Query("json"),
):
    """Per-user activity summary combining Notes (reliable author) + Account Activity (performed-by)."""
    from datetime import timezone
    print("\nTeam activity report...")

    if _ta_cache and (_time.time() - _ta_cache_ts) < _TA_CACHE_TTL:
        users_data         = _ta_cache["users_data"]
        all_notes_raw      = _ta_cache["all_notes_raw"]
        contact_to_account = _ta_cache["contact_to_account"]
        all_activity       = _ta_cache["all_activity"]
        print("[ta-cache] using cached data")
    else:
        print("[ta-cache] cache miss — fetching fresh")
        users_data, all_notes_raw_raw, all_contacts_raw, all_activity_raw = await asyncio.gather(
            ac_get("users"),
            ac_get_all("notes", "notes", {}),
            ac_get_all("contacts", "contacts", {}),
            ac_get_all(f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", "records", {}),
        )
        contact_to_account = {str(c.get("id","")): str(c.get("account","") or "")
                              for c in all_contacts_raw if str(c.get("account","") or "") not in ("","0")}
        all_notes_raw = [{"id": n.get("id"), "userid": n.get("userid"), "relid": n.get("relid") or n.get("rel_id"),
                          "reltype": n.get("reltype"), "cdate": n.get("cdate"), "note": n.get("note") or ""}
                         for n in all_notes_raw_raw
                         if (n.get("reltype") or "").lower() in ("contact","customeraccount","deal")]
        all_activity = [{"id": r.get("id"),
                         "fields": {f["id"]: f.get("value") for f in r.get("fields", [])},
                         "account": str(next(iter(r.get("relationships",{}).get("account",[])), "") or "")}
                        for r in all_activity_raw]

    # Build user map: userid → display name
    users: dict = {}
    for u in (users_data.get("users", []) if isinstance(users_data, dict) else []):
        uid  = str(u.get("id", ""))
        name = f"{u.get('firstName','').strip()} {u.get('lastName','').strip()}".strip()
        users[uid] = name or u.get("email", f"User {uid}")

    # Try to match a free-text performed-by value to a known user name
    def match_user(val: str) -> Optional[str]:
        if not val: return None
        v = val.strip().lower()
        for uid, name in users.items():
            if name.lower() == v: return uid           # exact match
        if len(v) == 2 and v.isalpha():                # initials e.g. "TB"
            for uid, name in users.items():
                parts = name.split()
                if (len(parts) >= 2
                        and parts[0][:1].lower() == v[0]
                        and parts[-1][:1].lower() == v[1]):
                    return uid
        for uid, name in users.items():                # first-name or contains
            parts = name.split()
            if parts and parts[0].lower() == v: return uid
            if v in name.lower(): return uid
        return None

    # account_id → owner user_id — use pre-built startup index (no extra API call)
    account_owner = _account_to_owner

    # Build user email map: uid → email (for matching last-login records)
    user_emails: dict = {}
    for u in (users_data.get("users", []) if isinstance(users_data, dict) else []):
        uid   = str(u.get("id", ""))
        email = (u.get("email") or "").lower()
        if uid and email:
            user_emails[uid] = email

    # ── All-time latest activity (unfiltered) — computed once before date window loop ──
    all_time_latest: dict = {}   # uid → "YYYY-MM-DD"
    for n in all_notes_raw:
        uid = str(n.get("userid", "") or "")
        d   = (n.get("cdate") or "")[:10]
        if uid and d and d > all_time_latest.get(uid, ""):
            all_time_latest[uid] = d
    for r in all_activity:
        fmap      = r["fields"]
        act_date  = (fmap.get("activity-date") or "")[:10]
        performed = _extract_performer(fmap)
        uid = match_user(performed)
        if uid and act_date and act_date > all_time_latest.get(uid, ""):
            all_time_latest[uid] = act_date

    from_dt = (datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
               if from_date else None)
    to_dt   = (datetime.strptime(to_date, "%Y-%m-%d").replace(
                   hour=23, minute=59, second=59, tzinfo=timezone.utc)
               if to_date else None)
    from_d  = from_dt.date() if from_dt else None
    to_d    = to_dt.date()   if to_dt   else None

    user_stats: dict = defaultdict(lambda: {
        "notes": 0, "activities": 0, "accounts": set(), "latest_date": ""
    })

    # ── Notes (reliable author via userid) ───────────────────────────────
    for n in all_notes_raw:
        reltype = (n.get("reltype") or "").lower()
        if reltype not in ("contact", "customeraccount", "deal"):
            continue
        raw_date = n.get("cdate", "")
        if from_dt or to_dt:
            try:
                nd = (datetime.fromisoformat(raw_date.replace("Z", "+00:00")) if "T" in raw_date
                      else datetime.strptime(raw_date[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc))
                if from_dt and nd < from_dt: continue
                if to_dt   and nd > to_dt:   continue
            except Exception:
                continue
        uid = str(n.get("userid", "") or "")
        if not uid: continue
        cid = str(n.get("relid") or n.get("rel_id") or "")
        # Account notes use cid as the account ID directly
        aid = cid if reltype == "customeraccount" else contact_to_account.get(cid, "")
        s   = user_stats[uid]
        s["notes"] += 1
        if aid: s["accounts"].add(aid)
        if raw_date > s["latest_date"]: s["latest_date"] = raw_date

    # ── Account Activity (performed-by, fuzzy-matched to users) ──────────
    unmatched_activity: dict = defaultdict(int)   # raw performed-by → count
    for r in all_activity:
        fmap       = r["fields"]
        act_date   = (fmap.get("activity-date") or "")[:10]
        performed  = _extract_performer(fmap)
        account_id = str(r.get("account", "") or "")

        if from_d or to_d:
            if not act_date:
                continue   # no date on record — exclude when filtering by date
            try:
                ad = datetime.strptime(act_date, "%Y-%m-%d").date()
                if from_d and ad < from_d: continue
                if to_d   and ad > to_d:   continue
            except Exception:
                continue   # unparseable date — exclude

        uid = match_user(performed)
        # Removed account-owner fallback: attributing unmatched activities to the
        # AM inflates their counts and hides activity from non-AM users like admins.
        if uid:
            s = user_stats[uid]
            s["activities"] += 1
            if account_id: s["accounts"].add(account_id)
            if act_date and act_date > s["latest_date"][:10]:
                s["latest_date"] = act_date
        elif performed:
            unmatched_activity[performed] += 1

    # Build result rows — include ALL known users so zero-activity users are visible
    user_rows = []
    all_uids  = set(user_stats.keys()) | set(users.keys())
    # Users excluded from the Team Performance report (admins, IT, inactive accounts, ghosts)
    _TP_EXCLUDED_IDS = {
        "1",   # Jeremy Sykes
        "5",   # Charles Posey
        "14",  # Barb Yeskey
        "15",  # Renee Mitchell
        "17",  # Ansley Bergen
        "18",  # Tristen Smithey
        "19",  # Lauren Futrell
        "21",  # Cher Shell
        "23",  # Microf IT
        "24",  # Dallas Munkus
    }

    for uid in all_uids:
        if uid in _TP_EXCLUDED_IDS:
            continue
        name = users.get(uid, f"User {uid}")
        # Skip ghost/deleted users (no longer in AC, show up as "User N")
        if name.startswith("User ") and not users.get(uid):
            continue
        s     = user_stats.get(uid, {"notes": 0, "activities": 0, "accounts": set(), "latest_date": ""})
        total = s["notes"] + s["activities"]
        email = user_emails.get(uid, "")
        user_rows.append({
            "user_name":            name,
            "notes_written":        s["notes"],
            "activities_logged":    s["activities"],
            "total_actions":        total,
            "accounts_touched":     len(s["accounts"]),
            "latest_activity_date": all_time_latest.get(uid, ""),   # always all-time, regardless of date filter
            "last_login":           (_last_login.get(email, "")[:10] if email else ""),
        })
    user_rows.sort(key=lambda x: (-x["total_actions"], x["user_name"]))

    # Unmatched performed-by values (couldn't tie to a user)
    unmatched = sorted(
        [{"performed_by": k, "activity_count": v} for k, v in unmatched_activity.items()],
        key=lambda x: -x["activity_count"]
    )

    if format == "csv":
        out = io.StringIO()
        if user_rows:
            w = csv.DictWriter(out, fieldnames=user_rows[0].keys())
            w.writeheader(); w.writerows(user_rows)
        fn = f"team_activity_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})

    return {"count": len(user_rows), "records": user_rows, "unmatched_activity": unmatched}


@app.get("/api/report/team-activity/breakdown")
async def team_activity_breakdown(
    user_name:  str           = Query(..., description="Display name of the user"),
    from_date:  Optional[str] = Query(None),
    to_date:    Optional[str] = Query(None),
    _: None = Depends(require_auth),
):
    """Per-account breakdown for a single user in the given date range."""
    from datetime import timezone

    if _ta_cache and (_time.time() - _ta_cache_ts) < _TA_CACHE_TTL:
        users_data         = _ta_cache["users_data"]
        all_notes_raw      = _ta_cache["all_notes_raw"]
        contact_to_account = _ta_cache["contact_to_account"]
        all_activity       = _ta_cache["all_activity"]
    else:
        users_data, all_notes_raw_raw, all_contacts_raw, all_activity_raw = await asyncio.gather(
            ac_get("users"),
            ac_get_all("notes", "notes", {}),
            ac_get_all("contacts", "contacts", {}),
            ac_get_all(f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", "records", {}),
        )
        contact_to_account = {str(c.get("id","")): str(c.get("account","") or "")
                              for c in all_contacts_raw if str(c.get("account","") or "") not in ("","0")}
        all_notes_raw = [{"id": n.get("id"), "userid": n.get("userid"), "relid": n.get("relid") or n.get("rel_id"),
                          "reltype": n.get("reltype"), "cdate": n.get("cdate"), "note": n.get("note") or ""}
                         for n in all_notes_raw_raw
                         if (n.get("reltype") or "").lower() in ("contact","customeraccount","deal")]
        all_activity = [{"id": r.get("id"),
                         "fields": {f["id"]: f.get("value") for f in r.get("fields", [])},
                         "account": str(next(iter(r.get("relationships",{}).get("account",[])), "") or "")}
                        for r in all_activity_raw]

    # Build user map and find target uid(s)
    users: dict = {}
    for u in (users_data.get("users", []) if isinstance(users_data, dict) else []):
        uid  = str(u.get("id", ""))
        name = f"{u.get('firstName','').strip()} {u.get('lastName','').strip()}".strip()
        users[uid] = name or u.get("email", f"User {uid}")

    target_uids = {uid for uid, name in users.items() if name == user_name}

    def match_user(val: str) -> Optional[str]:
        if not val: return None
        v = val.strip().lower()
        for uid, name in users.items():
            if name.lower() == v: return uid
        if len(v) == 2 and v.isalpha():
            for uid, name in users.items():
                parts = name.split()
                if len(parts) >= 2 and parts[0][:1].lower() == v[0] and parts[-1][:1].lower() == v[1]:
                    return uid
        for uid, name in users.items():
            parts = name.split()
            if parts and parts[0].lower() == v: return uid
            if v in name.lower(): return uid
        return None

    from_dt = (datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) if from_date else None)
    to_dt   = (datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc) if to_date else None)
    from_d  = from_dt.date() if from_dt else None
    to_d    = to_dt.date()   if to_dt   else None

    # account_id → {notes, activities, latest_date}
    acct_stats: dict = defaultdict(lambda: {"notes": 0, "activities": 0, "latest_date": ""})

    for n in all_notes_raw:
        reltype = (n.get("reltype") or "").lower()
        if reltype not in ("contact", "customeraccount", "deal"):
            continue
        uid = str(n.get("userid", "") or "")
        if uid not in target_uids:
            continue
        raw_date = n.get("cdate", "")
        if from_dt or to_dt:
            try:
                nd = (datetime.fromisoformat(raw_date.replace("Z", "+00:00")) if "T" in raw_date
                      else datetime.strptime(raw_date[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc))
                if from_dt and nd < from_dt: continue
                if to_dt   and nd > to_dt:   continue
            except Exception:
                continue
        cid = str(n.get("relid") or n.get("rel_id") or "")
        # Account notes (customeraccount) use cid as the account ID directly
        if reltype == "customeraccount":
            aid = cid
        else:
            aid = contact_to_account.get(cid, "")
        if not aid:
            continue
        s = acct_stats[aid]
        s["notes"] += 1
        if raw_date > s["latest_date"]: s["latest_date"] = raw_date

    for r in all_activity:
        fmap       = r["fields"]
        act_date   = (fmap.get("activity-date") or "")[:10]
        performed  = _extract_performer(fmap)
        account_id = str(r.get("account", "") or "")
        if not account_id:
            continue
        if from_d or to_d:
            if not act_date:
                continue
            try:
                ad = datetime.strptime(act_date, "%Y-%m-%d").date()
                if from_d and ad < from_d: continue
                if to_d   and ad > to_d:   continue
            except Exception:
                continue
        matched_uid = match_user(performed)
        if matched_uid not in target_uids:
            continue
        s = acct_stats[account_id]
        s["activities"] += 1
        if act_date and act_date > s["latest_date"][:10]:
            s["latest_date"] = act_date

    accounts = []
    for aid, s in acct_stats.items():
        accounts.append({
            "account_id":    aid,
            "account_name":  _account_to_name.get(aid, f"Account {aid}"),
            "dealer_id":     _account_to_dealer.get(aid, ""),
            "channel":       _account_to_platform.get(aid, ""),
            "region":        _account_to_region.get(aid, ""),
            "notes":         s["notes"],
            "activities":    s["activities"],
            "total":         s["notes"] + s["activities"],
            "latest_date":   s["latest_date"][:10] if s["latest_date"] else "",
        })
    accounts.sort(key=lambda x: (-x["total"], x["account_name"]))

    return {"user_name": user_name, "accounts": accounts, "total_accounts": len(accounts)}


@app.get("/api/report/team-activity/account-detail")
async def team_activity_account_detail(
    user_name:  str           = Query(...),
    account_id: str           = Query(...),
    from_date:  Optional[str] = Query(None),
    to_date:    Optional[str] = Query(None),
    format:     str           = Query("json"),
    _: None = Depends(require_auth),
):
    """Notes and activities by a specific user on a specific account."""
    from datetime import timezone

    if _ta_cache and (_time.time() - _ta_cache_ts) < _TA_CACHE_TTL:
        users_data         = _ta_cache["users_data"]
        all_notes_raw      = _ta_cache["all_notes_raw"]
        contact_to_account = _ta_cache["contact_to_account"]
        contact_email_map  = _ta_cache["contact_email_map"]
        all_activity       = _ta_cache["all_activity"]
    else:
        users_data, all_notes_raw_raw, all_contacts_raw, all_activity_raw = await asyncio.gather(
            ac_get("users"),
            ac_get_all("notes", "notes", {}),
            ac_get_all("contacts", "contacts", {}),
            ac_get_all(f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", "records", {}),
        )
        contact_to_account = {str(c.get("id","")): str(c.get("account","") or "")
                              for c in all_contacts_raw if str(c.get("account","") or "") not in ("","0")}
        contact_email_map  = {str(c.get("id","")): c.get("email","") for c in all_contacts_raw if c.get("email")}
        all_notes_raw = [{"id": n.get("id"), "userid": n.get("userid"), "relid": n.get("relid") or n.get("rel_id"),
                          "reltype": n.get("reltype"), "cdate": n.get("cdate"), "note": n.get("note") or ""}
                         for n in all_notes_raw_raw
                         if (n.get("reltype") or "").lower() in ("contact","customeraccount","deal")]
        all_activity = [{"id": r.get("id"),
                         "fields": {f["id"]: f.get("value") for f in r.get("fields", [])},
                         "account": str(next(iter(r.get("relationships",{}).get("account",[])), "") or "")}
                        for r in all_activity_raw]

    # Build user map and find target UIDs
    users: dict = {}
    for u in (users_data.get("users", []) if isinstance(users_data, dict) else []):
        uid  = str(u.get("id", ""))
        name = f"{u.get('firstName','').strip()} {u.get('lastName','').strip()}".strip()
        users[uid] = name or u.get("email", f"User {uid}")

    target_uids = {uid for uid, name in users.items() if name == user_name}

    def match_user(val: str) -> Optional[str]:
        if not val: return None
        v = val.strip().lower()
        for uid, name in users.items():
            if name.lower() == v: return uid
        if len(v) == 2 and v.isalpha():
            for uid, name in users.items():
                parts = name.split()
                if len(parts) >= 2 and parts[0][:1].lower() == v[0] and parts[-1][:1].lower() == v[1]:
                    return uid
        for uid, name in users.items():
            parts = name.split()
            if parts and parts[0].lower() == v: return uid
            if v in name.lower(): return uid
        return None

    from_dt = (datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) if from_date else None)
    to_dt   = (datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc) if to_date else None)
    from_d  = from_dt.date() if from_dt else None
    to_d    = to_dt.date()   if to_dt   else None

    # Contacts that belong to this account
    account_contact_ids = [cid for cid, aid in contact_to_account.items() if aid == account_id]

    notes_out = []
    for n in all_notes_raw:
        reltype = (n.get("reltype") or "").lower()
        if reltype not in ("contact", "customeraccount", "deal"):
            continue
        uid = str(n.get("userid", "") or "")
        if uid not in target_uids:
            continue
        raw_date = n.get("cdate", "")
        if from_dt or to_dt:
            try:
                nd = (datetime.fromisoformat(raw_date.replace("Z", "+00:00")) if "T" in raw_date
                      else datetime.strptime(raw_date[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc))
                if from_dt and nd < from_dt: continue
                if to_dt   and nd > to_dt:   continue
            except Exception:
                continue
        cid = str(n.get("relid") or n.get("rel_id") or "")
        if reltype == "customeraccount":
            note_aid = cid
        else:
            note_aid = contact_to_account.get(cid, "")
        if note_aid != account_id:
            continue
        notes_out.append({
            "id":   n.get("id"),
            "note": (n.get("note") or ""),
            "date": raw_date[:10] if raw_date else "",
            "contact_id": cid if reltype != "customeraccount" else None,
        })
    notes_out.sort(key=lambda x: x["date"], reverse=True)

    activities_out = []
    for r in all_activity:
        fmap       = r["fields"]
        act_date   = (fmap.get("activity-date") or "")[:10]
        performed  = _extract_performer(fmap)
        rec_aid    = str(r.get("account", "") or "")
        if rec_aid != account_id:
            continue
        if from_d or to_d:
            if not act_date:
                continue
            try:
                ad = datetime.strptime(act_date, "%Y-%m-%d").date()
                if from_d and ad < from_d: continue
                if to_d   and ad > to_d:   continue
            except Exception:
                continue
        matched_uid = match_user(performed)
        if matched_uid not in target_uids:
            continue
        activities_out.append({
            "id":          r.get("id"),
            "type":        fmap.get("activity-type", "") or fmap.get("subject", ""),
            "description": (fmap.get("body") or fmap.get("activity-description") or ""),
            "name":        fmap.get("name", ""),
            "date":        act_date,
            "performed_by": performed,
        })
    activities_out.sort(key=lambda x: x["date"], reverse=True)

    # ── Contact email activity logs (activityLogs per contact) ───────────────
    EMAIL_TYPES = {"send", "sms", "call", "reply", "forward"}

    async def _fetch_contact_logs(cid: str) -> list:
        try:
            data = await ac_get(f"contacts/{cid}/activityLogs", {"limit": 100})
            return data.get("contactActivities", [])
        except Exception:
            return []

    # Fetch concurrently for all contacts on this account (cap at 20 contacts)
    sem = asyncio.Semaphore(10)
    async def _fetch_with_sem(cid):
        async with sem:
            return await _fetch_contact_logs(cid)

    logs_nested = await asyncio.gather(*[_fetch_with_sem(c) for c in account_contact_ids[:20]])
    all_logs = [log for logs in logs_nested for log in logs]

    emails_out = []
    for log in all_logs:
        a_type  = (log.get("type") or "").lower()
        if a_type not in EMAIL_TYPES:
            continue
        ts = log.get("tstamp") or log.get("cdate") or ""
        log_date = ts[:10] if ts else ""
        if log_date and (from_d or to_d):
            try:
                ld = datetime.strptime(log_date, "%Y-%m-%d").date()
                if from_d and ld < from_d: continue
                if to_d   and ld > to_d:   continue
            except Exception:
                pass
        cid = str(log.get("contact", "") or "")
        subject = (log.get("subject") or
                   (log.get("campaign", {}).get("name", "") if isinstance(log.get("campaign"), dict) else "") or
                   "")
        emails_out.append({
            "type":       a_type,
            "label":      ACTIVITY_LABELS.get(a_type, a_type.title()),
            "subject":    subject,
            "date":       log_date,
            "contact_email": contact_email_map.get(cid, ""),
        })
    emails_out.sort(key=lambda x: x["date"], reverse=True)

    account_name = _account_to_name.get(account_id, f"Account {account_id}")
    dealer_id    = _account_to_dealer.get(account_id, "")

    if format == "csv":
        out = io.StringIO()
        rows = []
        for a in activities_out:
            rows.append({"type": "Activity", "date": a["date"], "kind": a["type"],
                         "title": a["name"], "body": a["description"],
                         "contact": "", "account": account_name, "dealer_id": dealer_id})
        for n in notes_out:
            rows.append({"type": "Note", "date": n["date"], "kind": "",
                         "title": "", "body": n["note"],
                         "contact": "", "account": account_name, "dealer_id": dealer_id})
        for e in emails_out:
            rows.append({"type": "Email", "date": e["date"], "kind": e["label"],
                         "title": e["subject"], "body": "",
                         "contact": e["contact_email"], "account": account_name, "dealer_id": dealer_id})
        rows.sort(key=lambda x: x["date"], reverse=True)
        if rows:
            w = csv.DictWriter(out, fieldnames=rows[0].keys())
            w.writeheader(); w.writerows(rows)
        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in account_name)[:40]
        fn = f"{safe_name}_{from_date or 'all'}_{to_date or 'all'}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})

    return {
        "user_name":    user_name,
        "account_id":   account_id,
        "account_name": account_name,
        "dealer_id":    dealer_id,
        "notes":        notes_out,
        "activities":   activities_out,
        "emails":       emails_out,
    }


# ═══════════════════════════════════════════════════════════════════════════
# ACCOUNT BROWSER
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/accounts/search")
async def accounts_search(q: str = Query(""), limit: int = Query(20)):
    """List or search accounts. Empty q returns alphabetical accounts.
    Numeric queries also check the dealer-id index for direct dealer ID lookups."""
    q = q.strip()
    is_numeric = q.isdigit()

    # ── Numeric query: check dealer-id index first ─────────────────────────
    if is_numeric and q in _dealer_id_index:
        entry = _dealer_id_index[q]
        return {
            "accounts": [{"id": entry["id"], "name": entry["name"], "dealer_id": q}],
            "total": 1,
        }

    # ── Normal AC name search ──────────────────────────────────────────────
    params = {"limit": limit, "orders[name]": "ASC"}
    if q:
        params["search"] = q
    data = await ac_get("accounts", params)
    raw  = data.get("accounts", [])

    # Resolve dealer IDs: check in-memory index before making API calls
    dealer_id_field = ACCT_FIELD["dealer_id"]   # "18"

    async def get_dealer_id(account_id: str) -> str:
        aid = str(account_id)
        if aid in _account_to_dealer:
            return _account_to_dealer[aid]
        try:
            cf_data = await ac_get(f"accounts/{aid}/accountCustomFieldData")
            for cf in cf_data.get("customerAccountCustomFieldData", []):
                if str(cf.get("custom_field_id")) == dealer_id_field:
                    return cf.get("custom_field_text_value") or ""
        except Exception:
            pass
        return ""

    dealer_ids = await asyncio.gather(*[get_dealer_id(a.get("id", "")) for a in raw])

    accounts = [
        {"id": a.get("id"), "name": a.get("name", ""), "dealer_id": dealer_ids[i]}
        for i, a in enumerate(raw)
    ]
    return {"accounts": accounts, "total": len(accounts)}


_slp_state_index: dict = {}       # state (str) → {account_id: {name, dealer_id}}
_slp_state_index_ts: float = 0.0
_SLP_STATE_TTL = 86400            # rebuild at most once per 24 hours
def _is_microf_channel(ch: str) -> bool:
    """True for any channel that is a Microf/LTO program.
    Matches 'Microf', 'Microf Direct', 'Microf (LTO Only)', 'LTO', etc.
    Uses 'contains microf' so new variants are caught automatically."""
    c = (ch or "").strip().lower()
    return "microf" in c or c == "lto"

# ── Shared qualifying-accounts cache ─────────────────────────────────────────
# Both the state index and the location index need "accounts with a Contractor
# Activated Microf/LTO SLP".  Rather than each fetching all SLP records
# independently, they share a single cached set rebuilt at most once per day.
_qualifying_accounts_cache: set   = set()
_qualifying_accounts_ts:    float = 0.0

async def _get_qualifying_microf_accounts() -> set:
    """Return the set of AC account IDs that have ≥1 Contractor Activated
    Microf/LTO SLP record.  Result is cached for 24 h."""
    global _qualifying_accounts_cache, _qualifying_accounts_ts
    now = _time.time()
    if _qualifying_accounts_cache and (now - _qualifying_accounts_ts) < _SLP_STATE_TTL:
        return _qualifying_accounts_cache

    raw = await get_slp_cache()
    result: set = set()
    for r in raw:
        fields = {f.get("field") or f.get("id"): f.get("value") for f in r.get("fields", [])}
        if str(fields.get("slp-status-detail", "")).strip() != "Contractor Activated":
            continue
        if not _is_microf_channel(str(fields.get("channel", ""))):
            continue
        rel    = r.get("relationships", {}).get("account", [])
        a0     = rel[0] if rel else None
        acc_id = str(a0) if isinstance(a0, (int, str)) else str(a0.get("id", "")) if a0 else None
        if acc_id:
            result.add(acc_id)

    _qualifying_accounts_cache = result
    _qualifying_accounts_ts    = now
    print(f"[slp-cache] {len(result)} qualifying Microf/LTO accounts cached")
    return result

# ── City-centroid lookup (built once from pgeocode's bundled dataset) ─────────
# Keyed "CITY|ST" → (lat, lon).  Building this inside _build_location_index
# on every 24-h rebuild wastes time — pgeocode's data never changes.
def _build_city_coords_once() -> dict:
    import math as _m
    coords: dict = {}
    try:
        import pgeocode
        df = pgeocode.Nominatim('us')._data.dropna(subset=['latitude', 'longitude']).copy()
        df['_k'] = df['place_name'].str.strip().str.upper() + '|' + df['state_code'].str.strip().str.upper()
        for key, row in df.groupby('_k')[['latitude', 'longitude']].mean().iterrows():
            try:
                lat, lon = float(row['latitude']), float(row['longitude'])
                if not (_m.isnan(lat) or _m.isnan(lon)):
                    coords[key] = (lat, lon)
            except Exception:
                pass
    except Exception:
        pass
    print(f"[city-coords] {len(coords)} city centroids loaded")
    return coords

_CITY_COORDS: dict = _build_city_coords_once()


async def _build_slp_state_index() -> dict:
    """Build state → active Microf/LTO accounts index from SLP records.
    Uses account-level CF22 (Doing Business in States) for geography.
    Filters to SLPs with platform in Microf/LTO/Microf(LTO Only) + Contractor Activated."""
    global _slp_state_index, _slp_state_index_ts
    now = _time.time()
    if _slp_state_index and (now - _slp_state_index_ts) < _SLP_STATE_TTL:
        return _slp_state_index

    qualifying_accounts = await _get_qualifying_microf_accounts()

    # Build state index from account-level CF22 (already in memory from dealer index)
    index: dict = {}
    for acc_id in qualifying_accounts:
        states_val = str(_account_to_states.get(acc_id, "") or "").upper()
        if not states_val:
            continue
        name      = _account_to_name.get(acc_id, "")
        dealer_id = _account_to_dealer.get(acc_id, "")
        for s in [x.strip() for x in states_val.split(",") if x.strip()]:
            index.setdefault(s, {})[acc_id] = {"name": name, "dealer_id": dealer_id}

    _slp_state_index    = index
    _slp_state_index_ts = now
    return index


@app.get("/api/accounts/by-state")
async def accounts_by_state(state: str = "", limit: int = 10):
    if not state:
        return {"accounts": [], "state": state}
    state_upper = state.upper().strip()
    index = await _build_slp_state_index()
    bucket = index.get(state_upper, {})
    results = sorted(
        [{"id": aid, **info} for aid, info in bucket.items()],
        key=lambda x: x["name"]
    )
    return {"accounts": results[:limit], "state": state_upper, "total": len(bucket)}


# ── Nearest contractor lookup ─────────────────────────────────────────────────

# Approximate geographic center of each US state (lat, lon)
_STATE_CENTROIDS: dict = {
    "AL": (32.806671, -86.791130), "AK": (61.370716, -152.404419),
    "AZ": (33.729759, -111.431221), "AR": (34.969704, -92.373123),
    "CA": (36.116203, -119.681564), "CO": (39.059811, -105.311104),
    "CT": (41.597782, -72.755371), "DE": (39.318523, -75.507141),
    "FL": (27.766279, -81.686783), "GA": (33.040619, -83.643074),
    "HI": (21.094318, -157.498337), "ID": (44.240459, -114.478828),
    "IL": (40.349457, -88.986137), "IN": (39.849426, -86.258278),
    "IA": (42.011539, -93.210526), "KS": (38.526600, -96.726486),
    "KY": (37.668140, -84.670067), "LA": (31.169960, -91.867805),
    "ME": (44.693947, -69.381927), "MD": (39.063946, -76.802101),
    "MA": (42.230171, -71.530106), "MI": (43.326618, -84.536095),
    "MN": (45.694454, -93.900192), "MS": (32.741646, -89.678696),
    "MO": (38.456085, -92.288368), "MT": (46.921925, -110.454353),
    "NE": (41.125370, -98.268082), "NV": (38.313515, -117.055374),
    "NH": (43.452492, -71.563896), "NJ": (40.298904, -74.521011),
    "NM": (34.840515, -106.248482), "NY": (42.165726, -74.948051),
    "NC": (35.630066, -79.806419), "ND": (47.528912, -99.784012),
    "OH": (40.388783, -82.764915), "OK": (35.565342, -96.928917),
    "OR": (44.572021, -122.070938), "PA": (40.590752, -77.209755),
    "RI": (41.680893, -71.511780), "SC": (33.856892, -80.945007),
    "SD": (44.299782, -99.438828), "TN": (35.747845, -86.692345),
    "TX": (31.054487, -97.563461), "UT": (40.150032, -111.862434),
    "VT": (44.045876, -72.710686), "VA": (37.769337, -78.169968),
    "WA": (47.400902, -121.490494), "WV": (38.491226, -80.954453),
    "WI": (44.268543, -89.616508), "WY": (42.755966, -107.302490),
}

def _haversine(lat1, lon1, lat2, lon2) -> float:
    """Return distance in miles between two lat/lon points."""
    import math
    R = 3958.8
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

_location_index: dict = {}        # account_id → {lat, lon, name, dealer_id, city, state, zip}
_location_index_ts: float = 0.0
_location_index_building: bool = False
_LOCATION_TTL = 86400

async def _build_location_index() -> dict:
    global _location_index, _location_index_ts, _location_index_building
    now = _time.time()
    if _location_index and (now - _location_index_ts) < _LOCATION_TTL:
        return _location_index

    _location_index_building = True
    try:
        # Use whatever SLP records are in memory right now — do NOT await get_slp_cache()
        # because that blocks on _slp_cache_lock while the background SLP fetch runs,
        # causing warming_up to persist for minutes. If cache is empty, we build a fast
        # empty index (so warming_up clears immediately); _slp_cache_loop will trigger a
        # proper rebuild once SLP data is loaded.
        slp_snapshot = list(_slp_cache_records)

        # Derive qualifying accounts directly from the snapshot (Contractor Activated Microf/LTO)
        qualifying: set = set()
        for _r in slp_snapshot:
            _flds = {_f.get("field") or _f.get("id"): _f.get("value") for _f in _r.get("fields", [])}
            if str(_flds.get("slp-status-detail", "")).strip() != "Contractor Activated":
                continue
            if not _is_microf_channel(str(_flds.get("channel", ""))):
                continue
            _rel = _r.get("relationships", {}).get("account", [])
            _a0  = _rel[0] if _rel else None
            _acc = str(_a0) if isinstance(_a0, (int, str)) else (str(_a0.get("id", "")) if _a0 else None)
            if _acc:
                qualifying.add(_acc)

        # Build channel map directly from SLP records
        slp_raw = slp_snapshot
        acct_to_channel: dict = {}
        for _r in slp_raw:
            _flds = {_f.get("field") or _f.get("id"): _f.get("value") for _f in _r.get("fields", [])}
            _ch = str(_flds.get("channel", "") or "").strip()
            if not _ch:
                continue
            _rel = _r.get("relationships", {}).get("account", [])
            _a0  = _rel[0] if _rel else None
            _aid = str(_a0) if isinstance(_a0, (int, str)) else (str(_a0.get("id", "")) if _a0 else None)
            if _aid and _aid not in acct_to_channel:
                acct_to_channel[_aid] = _ch

        import math as _math, random

        # zip-level lookup for the small subset of accounts that actually have CF6
        zip_coords: dict = {}
        try:
            import pgeocode
            geo  = pgeocode.Nominatim('us')
            zips = {(_account_to_zip.get(aid, "") or "")[:5] for aid in qualifying
                    if (_account_to_zip.get(aid, "") or "")[:5].isdigit()}
            zips.discard("")
            if zips:
                result = geo.query_postal_code(list(zips))
                rows   = result if hasattr(result, 'iterrows') else result.to_frame().T
                for idx, row in rows.iterrows():
                    try:
                        lat, lon = float(row['latitude']), float(row['longitude'])
                        if not (_math.isnan(lat) or _math.isnan(lon)):
                            zip_coords[str(idx)] = (lat, lon)
                    except Exception:
                        pass
        except Exception:
            pass

        index: dict = {}
        for aid in qualifying:
            z    = (_account_to_zip.get(aid, "") or "")[:5]
            city = (_account_to_city.get(aid, "") or "").strip()
            st   = (_account_to_state_prov.get(aid, "") or "").strip().upper()[:2]

            city_key = f"{city.upper()}|{st}"

            if z in zip_coords:
                lat, lon  = zip_coords[z]
                precision = "zip"
            elif city_key in _CITY_COORDS:
                lat, lon  = _CITY_COORDS[city_key]
                precision = "city"
            elif st in _STATE_CENTROIDS:
                # Last resort — spread pins randomly across the state
                clat, clon = _STATE_CENTROIDS[st]
                lat = clat + random.uniform(-1.0, 1.0)
                lon = clon + random.uniform(-1.5, 1.5)
                precision  = "state"
            else:
                continue  # no location data at all

            index[aid] = {
                "lat":       round(lat, 5),
                "lon":       round(lon, 5),
                "name":      _account_to_name.get(aid, ""),
                "dealer_id": _account_to_dealer.get(aid, ""),
                "city":      city,
                "state":     st,
                "zip":       z,
                "platform":  acct_to_channel.get(aid, ""),
                "approx":    precision != "zip",
            }

        _location_index    = index
        _location_index_ts = now
        by_prec = {"zip": 0, "city": 0, "state": 0}
        for v in index.values():
            if not v["approx"]:
                by_prec["zip"] += 1
            elif v.get("city"):
                by_prec["city"] += 1
            else:
                by_prec["state"] += 1
        print(f"[location-index] {len(index)} accounts — "
              f"{by_prec['zip']} exact-zip, {by_prec['city']} city-centroid, {by_prec['state']} state-fallback")
        return index
    except Exception as _loc_exc:
        import traceback
        print(f"[location-index] BUILD FAILED: {_loc_exc}")
        traceback.print_exc()
        return _location_index  # return whatever we had before
    finally:
        _location_index_building = False  # always unset so dealer-locator can retry


@app.get("/api/accounts/nearest")
async def accounts_nearest(address: str = "", limit: int = 10):
    if not address:
        return {"accounts": [], "lat": None, "lon": None}
    try:
        import pgeocode
        geo = pgeocode.Nominatim('us')
    except ImportError:
        return {"error": "pgeocode not installed"}

    # Geocode the search input — try as zip first, fall back to Nominatim
    search_lat, search_lon, search_label = None, None, address
    zip_clean = address.strip()[:5]
    if zip_clean.isdigit():
        r = geo.query_postal_code(zip_clean)
        try:
            search_lat = float(r['latitude'])
            search_lon = float(r['longitude'])
            search_label = f"{r.get('place_name', zip_clean)}, {r.get('state_code', '')}"
        except Exception:
            pass

    if search_lat is None:
        # Fall back to Nominatim for city/address strings
        import httpx
        async with httpx.AsyncClient(timeout=10) as client:
            nr = await client.get("https://nominatim.openstreetmap.org/search",
                params={"q": address, "countrycodes": "us", "limit": 1, "format": "json"},
                headers={"Accept-Language": "en"})
            results = nr.json()
            if results:
                search_lat = float(results[0]["lat"])
                search_lon = float(results[0]["lon"])
                search_label = results[0].get("display_name", address).split(",")[0]

    if search_lat is None:
        return {"error": f"Could not geocode: {address}"}

    # If the index hasn't been built yet, tell the client to retry.
    # Use _location_index_ts to distinguish "never built" from "built but empty".
    if _location_index_ts == 0:
        if not _location_index_building:
            asyncio.create_task(_build_location_index())
        return {"warming_up": True, "accounts": [], "total": 0,
                "lat": search_lat, "lon": search_lon, "label": search_label}
    loc_index = _location_index
    distances = []
    for aid, info in loc_index.items():
        d = _haversine(search_lat, search_lon, info["lat"], info["lon"])
        distances.append({
            **info,
            "id":           aid,
            "distance_miles": round(d, 1),
            "phone":        _account_to_phone.get(aid, ""),
            "website":      _account_to_website.get(aid, ""),
            "address":      _account_to_address.get(aid, ""),
            "last_app_date": _account_to_last_app.get(aid, ""),
        })

    distances.sort(key=lambda x: x["distance_miles"])
    within_15 = [d for d in distances if d["distance_miles"] <= 15]
    return {
        "accounts": distances[:limit],
        "lat": search_lat,
        "lon": search_lon,
        "label": search_label,
        "total": len(distances),
        "within_15": len(within_15),
    }


@app.get("/api/accounts/{account_id}/detail")
async def account_detail(account_id: str):
    """Fast 360° account view — fetches data scoped to this account only."""

    # Stage 1: account core data + custom fields + contacts + notes (parallel)
    acc_data, acc_cf_data, acc_contacts, acc_notes_data, cf_meta = await asyncio.gather(
        ac_get(f"accounts/{account_id}"),
        ac_get(f"accounts/{account_id}/accountCustomFieldData"),
        ac_get(f"accounts/{account_id}/contacts"),
        ac_get("notes", {"reltype": "account", "rel_id": account_id, "limit": 25}),
        _get_account_cf_meta(),
        return_exceptions=True,
    )

    account = acc_data.get("account", {}) if isinstance(acc_data, dict) else {}

    # Build named custom field map — use AC labels, read all value types
    named_cfs = {}
    if isinstance(acc_cf_data, dict) and isinstance(cf_meta, dict):
        for cf in acc_cf_data.get("customerAccountCustomFieldData", []):
            fid = str(cf.get("custom_field_id", ""))
            val = _extract_cf_value(cf)
            if val:
                label = cf_meta.get(fid, f"field_{fid}")
                named_cfs[label] = val

    # Dealer ID drives SLP lookup (field 18 = "Parent Dealer ID")
    dealer_id = named_cfs.get("Parent Dealer ID", "")

    contact_ids = []
    if isinstance(acc_contacts, dict):
        contact_ids = [ac.get("contact") for ac in acc_contacts.get("accountContacts", [])]

    notes = []
    if isinstance(acc_notes_data, dict):
        for n in acc_notes_data.get("notes", []):
            notes.append({
                "id":    n.get("id"),
                "note":  n.get("note", "")[:300],
                "cdate": n.get("cdate", ""),
            })

    # Stage 2: SLPs (filtered by account relationship), contacts (by ID), alternate contacts
    slp_task       = ac_get(f"customObjects/records/{SLP_SCHEMA_ID}",
                            {"filters[relationships.account]": account_id, "limit": 100})
    deal_task      = ac_get("deals", {"filters[account]": account_id, "limit": 50})
    alt_con_task   = ac_get(f"customObjects/records/{ALT_CONTACT_SCHEMA_ID}",
                            {"filters[relationships.account]": account_id, "limit": 50})

    slp_r, deal_r, alt_con_r = await asyncio.gather(slp_task, deal_task, alt_con_task, return_exceptions=True)

    def flatten_co(records):
        seen_ids, result = set(), []
        for r in records:
            rid = r.get("id")
            if rid in seen_ids:
                continue
            seen_ids.add(rid)
            row = {"record_id": rid}
            for f in r.get("fields", []):
                row[f.get("id", "")] = f.get("value", "")
            result.append(row)
        return result

    slps  = flatten_co((slp_r.get("records", []) if isinstance(slp_r, dict) else []))

    alt_contacts = []
    if isinstance(alt_con_r, dict):
        for r in alt_con_r.get("records", []):
            fmap = {f.get("id"): f.get("value") for f in r.get("fields", [])}
            alt_contacts.append({
                "id":            r.get("id"),
                "firstName":     fmap.get("name", ""),
                "lastName":      fmap.get("last-name", ""),
                "phone":         fmap.get("phone-number", ""),
                "contact_status": fmap.get("contact-status", ""),
            })
    deals = []
    if isinstance(deal_r, dict):
        for d in deal_r.get("deals", []):
            deals.append({
                "id":     d.get("id"),
                "title":  d.get("title", ""),
                "value":  d.get("value", "0"),
                "status": d.get("status", ""),
                "stage":  d.get("stage", ""),
                "cdate":  d.get("cdate", ""),
            })

    # Fetch contacts
    contacts = []
    if contact_ids:
        contact_tasks = [ac_get(f"contacts/{cid}") for cid in contact_ids[:15]]
        contact_results = await asyncio.gather(*contact_tasks, return_exceptions=True)
        for cr in contact_results:
            if isinstance(cr, dict):
                c = cr.get("contact", {})
                contacts.append({
                    "id":        c.get("id"),
                    "firstName": c.get("firstName", ""),
                    "lastName":  c.get("lastName", ""),
                    "email":     c.get("email", ""),
                    "phone":     c.get("phone", ""),
                })

    return {
        "account": {
            "id":      account_id,
            "name":    account.get("name", ""),
            "url":     account.get("accountUrl", ""),
            "owner":   account.get("owner", ""),
            "created": account.get("created_utc_timestamp", ""),
            "updated": account.get("updated_utc_timestamp", ""),
            "fields":  named_cfs,
            "ac_url":  ac_account_url(account_id),
        },
        "slps":          slps,
        "contacts":      contacts,
        "alt_contacts":  alt_contacts,
        "deals":         deals,
        "summary": {
            "slp_count":     len(slps),
            "contact_count": len(contacts),
            "deal_count":    len(deals),
        },
    }


# ═══════════════════════════════════════════════════════════════════════════
# ACCOUNT NOTES  (Account Activity custom object)
# ═══════════════════════════════════════════════════════════════════════════

class _NoteIn(_BaseModel):
    subject:       str
    note_body:     str
    activity_type: str = "Internal Note"   # Internal Note | Call | Email | Text

@app.post("/api/accounts/{account_id}/notes")
async def create_account_note(account_id: str, note: _NoteIn, request: _Request,
                               user=Depends(require_auth)):
    """Create an Account Activity (note) record linked to an account."""
    from datetime import timezone
    performed_by = _get_session_email(request) or user or "Microf Reports"
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    payload = {
        "record": {
            "fields": [
                {"id": "activity-type",  "value": note.activity_type},
                {"id": "subject",        "value": note.subject},
                {"id": "body",           "value": note.note_body},
                {"id": "activity-date",  "value": now_iso},
                {"id": "performed-by",   "value": performed_by},
                {"id": "source",         "value": "Microf Reports"},
            ],
            "relationships": {"account": [int(account_id)]},
        }
    }
    data = await ac_post(
        f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", payload
    )
    return {"ok": True, "record": data.get("record", {})}


@app.get("/api/accounts/{account_id}/notes")
async def get_account_notes(account_id: str, user=Depends(require_auth)):
    """Fetch Account Activity records linked to an account."""
    data = await ac_get(
        f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}",
        {"filters[relationships.account]": account_id, "limit": 50},
    )
    all_records = data.get("records", []) if isinstance(data, dict) else []
    results = []
    for r in all_records:
        fields = {f["id"]: f.get("value", "") for f in r.get("fields", [])}
        results.append({
            "id":            r.get("id"),
            "activity_type": fields.get("activity-type", ""),
            "subject":       fields.get("subject", ""),
            "body":          fields.get("body", ""),
            "activity_date": fields.get("activity-date", ""),
            "performed_by":  fields.get("performed-by", ""),
        })
    results.sort(key=lambda x: x.get("activity_date", ""), reverse=True)
    return {"notes": results}


# ── Training Records ────────────────────────────────────────────────────────

TRAINING_SCHEMA_ID = "9368fee4-ccef-407b-a0d3-4b72c346b2af"

class _TrainingIn(_BaseModel):
    training_type:   str = ""
    trained_by:      str = ""
    date_of_training: str = ""
    training_agenda: str = ""
    dealer_id:       str = ""
    training_notes:  str = ""
    name:            str = ""

@app.post("/api/accounts/{account_id}/training")
async def create_training_record(account_id: str, rec: _TrainingIn,
                                  user=Depends(require_auth)):
    payload = {
        "record": {
            "fields": [
                {"id": "name",              "value": rec.name},
                {"id": "training-type",     "value": rec.training_type},
                {"id": "trained-by",        "value": rec.trained_by},
                {"id": "date-of-training",  "value": rec.date_of_training},
                {"id": "training-agenda",   "value": rec.training_agenda},
                {"id": "dealer-id",         "value": rec.dealer_id},
                {"id": "training-notes",    "value": rec.training_notes},
            ],
            "relationships": {"account": [int(account_id)]},
        }
    }
    data = await ac_post(f"customObjects/records/{TRAINING_SCHEMA_ID}", payload)
    return {"ok": True, "record": data.get("record", {})}


@app.get("/api/accounts/{account_id}/training")
async def get_training_records(account_id: str, user=Depends(require_auth)):
    data = await ac_get(
        f"customObjects/records/{TRAINING_SCHEMA_ID}",
        {"filters[relationships.account]": account_id, "limit": 50},
    )
    results = []
    for r in (data.get("records", []) if isinstance(data, dict) else []):
        fields = {f["id"]: f.get("value", "") for f in r.get("fields", [])}
        results.append({
            "id":               r.get("id"),
            "name":             fields.get("name", ""),
            "training_type":    fields.get("training-type", ""),
            "trained_by":       fields.get("trained-by", ""),
            "date_of_training": fields.get("date-of-training", ""),
            "training_agenda":  fields.get("training-agenda", ""),
            "dealer_id":        fields.get("dealer-id", ""),
            "training_notes":   fields.get("training-notes", ""),
        })
    results.sort(key=lambda x: x.get("date_of_training", ""), reverse=True)
    return {"training": results}


# ═══════════════════════════════════════════════════════════════════════════
# GLOBAL SEARCH
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/global-search")
async def global_search(q: str = Query(..., min_length=1),
                        program: Optional[str] = Query(None),
                        bdr: Optional[str] = Query(None),
                        owner_id: Optional[str] = Query(None),
                        group: Optional[str] = Query(None)):
    """Search accounts (by name), contacts (by email only, text queries), and SLPs (by dealer ID or name)."""
    q = q.strip()

    # ── In-memory-only path: filters set but no text query ────────────────────
    if not q and (bdr or owner_id or group):
        candidate_ids: set = set(_account_to_name.keys())
        if bdr:
            candidate_ids &= {aid for aid, b in _account_to_bdr.items() if b == bdr}
        if owner_id:
            candidate_ids &= {aid for aid, o in _account_to_owner.items() if o == owner_id}
        if group:
            candidate_ids &= {aid for aid, g in _account_to_group.items() if g == group}
        if program and _program_to_accounts:
            prog_key = program.lower().strip()
            if prog_key in _program_to_accounts:
                candidate_ids &= set(_program_to_accounts[prog_key])
        accs = [{"id": aid, "name": _account_to_name.get(aid, ""),
                  "dealer_id": _account_to_dealer.get(aid, ""),
                  "account_url": ac_account_url(aid)} for aid in candidate_ids]
        accs.sort(key=lambda x: x["name"].lower())
        return {"accounts": accs, "slps": [], "contacts": [], "query": "", "total": len(accs)}

    # Normalize phone-like queries: strip dashes, spaces, dots, parens so
    # "225-681-1638" → "2256811638" matches how AC stores phone numbers.
    import re as _re
    q_digits = _re.sub(r"[\s\-().+]", "", q)
    is_phone_like = (not q.isdigit()) and q_digits.isdigit() and len(q_digits) >= 7
    if is_phone_like:
        q = q_digits   # search AC with the stripped version

    # ── In-memory intersection search (program filter or multi-term query) ────
    # When a program param is supplied, or when the query embeds a known program
    # name (e.g. "ARS 360 Finance"), use _program_to_accounts + _account_to_name
    # for name matching instead of the AC API so we don't hit the limit-50 cap
    # or miss accounts whose names don't come back from the AC search.
    words = q.split()

    # Determine which program key to use
    _use_program_key: str | None = None
    if program and _program_to_accounts:
        _use_program_key = program.lower().strip()
        if _use_program_key not in _program_to_accounts:
            _use_program_key = None   # unknown program — fall through to AC search
    elif len(words) >= 2 and _program_to_accounts:
        q_lower = q.lower()
        _use_program_key = max(
            (pk for pk in _program_to_accounts if pk in q_lower),
            key=len, default=None
        )
        if _use_program_key is None:
            for w in words:
                if w.lower() in _program_to_accounts:
                    _use_program_key = w.lower()
                    break

    if _use_program_key and _account_to_name:
        # Name terms = query minus the program token (if it was embedded in query)
        q_lower = q.lower()
        if _use_program_key in q_lower:
            name_part  = q_lower.replace(_use_program_key, "").strip()
        else:
            name_part  = q_lower   # program came from ?program= param, full query is name
        name_terms = name_part.split() if name_part else []

        prog_ids: set = set(_program_to_accounts[_use_program_key])
        if name_terms:
            final_ids = {
                aid for aid in prog_ids
                if all(nt in _account_to_name.get(aid, "").lower() for nt in name_terms)
            }
        else:
            final_ids = prog_ids

        accounts_out = []
        for aid in sorted(final_ids):
            accounts_out.append({
                "id":             aid,
                "name":           _account_to_name.get(aid, ""),
                "dealer_id":      _account_to_dealer.get(aid, ""),
                "dealer_program": _account_to_platform.get(aid, ""),
                "account_url":    ac_account_url(aid),
                "matched_on":     "dealer program + name",
            })
        accounts_out.sort(key=lambda x: x["name"].lower())
        return {"accounts": accounts_out, "slps": [], "contacts": [], "query": q, "total": len(accounts_out)}

    is_numeric = q.isdigit()

    # For numeric queries:
    #   - Short  (< 7 digits): likely a dealer ID — skip contacts to avoid phone-digit noise
    #   - Longer (≥ 7 digits): likely a phone number — include contacts so the search works
    search_contacts = (not is_numeric) or (is_numeric and len(q) >= 7)

    # SLP filter: exact dealer-id for numbers, name contains for text
    if is_numeric:
        slp_params = {"filters[fields.dealer-id]": q, "limit": 50}
    else:
        slp_params = {"filters[fields.name][contains]": q, "limit": 50}

    tasks = [
        ac_get("accounts", {"search": q, "limit": 50}),
        ac_get(f"customObjects/records/{SLP_SCHEMA_ID}", slp_params),
    ]
    if search_contacts:
        tasks.append(ac_get("contacts", {"search": q, "limit": 50}))

    results = await asyncio.gather(*tasks, return_exceptions=True)
    acc_data = results[0]
    slp_data = results[1]
    con_data = results[2] if search_contacts else {}

    # ── SLP records ───────────────────────────────────────────────────────
    matched_slps = []
    for r in ([] if isinstance(slp_data, Exception) else slp_data.get("records", [])):
        fmap       = {f.get("id"): f.get("value") for f in r.get("fields", [])}
        account_id = next(iter(r.get("relationships", {}).get("account", [])), "")
        matched_slps.append({
            "record_id":      r.get("id"),
            "dealer_id":      fmap.get("dealer-id", ""),
            "channel":        fmap.get("channel", ""),
            "account_id":     account_id,
            "account_url":    ac_account_url(account_id),
            "slp_status":     fmap.get("slp-status-detail", ""),
            "activated_date": str(fmap.get("contractor-activated-date", ""))[:10],
            "oracle_ids":     fmap.get("oracle-producer-ids", ""),
            "assigned_bdr":   fmap.get("assigned-bdr", ""),
        })

    # ── Contacts (text queries only) ───────────────────────────────────────
    matched_contacts = []
    if not isinstance(con_data, Exception):
        for c in con_data.get("contacts", []):
            aid = c.get("account", "")
            cid = c.get("id", "")
            matched_contacts.append({
                "id":           cid,
                "name":         f"{c.get('firstName','')} {c.get('lastName','')}".strip(),
                "email":        c.get("email", ""),
                "account_id":   aid,
                "account_url":  ac_account_url(aid),
                "contact_url":  ac_contact_url(cid),
            })

    # ── Accounts: name search results + index lookup + accounts linked to matched SLPs ───
    seen_account_ids: set = set()
    matched_accounts      = []

    # ① Dealer ID index lookup for numeric queries (fastest path — hits cache, no API call)
    if is_numeric and q in _dealer_id_index:
        entry = _dealer_id_index[q]
        aid   = str(entry["id"])
        seen_account_ids.add(aid)
        matched_accounts.append({
            "id":          aid,
            "name":        entry["name"],
            "dealer_id":   q,
            "matched_on":  "dealer id",
            "account_url": ac_account_url(aid),
        })

    # Name-search results — fetch dealer ID for each in parallel
    raw_accounts = [] if isinstance(acc_data, Exception) else acc_data.get("accounts", [])
    dealer_id_field = ACCT_FIELD["dealer_id"]

    async def fetch_dealer_id(account_id: str) -> str:
        # Check in-memory index first (fast, no API call)
        if account_id in _account_to_dealer:
            return _account_to_dealer[account_id]
        try:
            cf = await ac_get(f"accounts/{account_id}/accountCustomFieldData")
            for f in cf.get("customerAccountCustomFieldData", []):
                if str(f.get("custom_field_id")) == dealer_id_field:
                    return f.get("custom_field_text_value") or ""
        except Exception:
            pass
        return ""

    dealer_ids = await asyncio.gather(*[fetch_dealer_id(str(a.get("id",""))) for a in raw_accounts])

    for a, did in zip(raw_accounts, dealer_ids):
        aid = str(a.get("id", ""))
        seen_account_ids.add(aid)
        matched_accounts.append({
            "id":          aid,
            "name":        a.get("name", ""),
            "dealer_id":   did,
            "matched_on":  "name",
            "account_url": ac_account_url(aid),
        })

    # Accounts linked to matched SLPs (we already know their dealer_id from the SLP)
    for slp in matched_slps:
        aid = str(slp["account_id"])
        if not aid or aid in seen_account_ids:
            continue
        seen_account_ids.add(aid)
        matched_accounts.append({
            "id":          aid,
            "name":        "",           # filled in below
            "dealer_id":   slp["dealer_id"],
            "matched_on":  "dealer id",
            "account_url": slp["account_url"],
            "_needs_name": True,
        })

    # Accounts linked to matched contacts
    for con in matched_contacts:
        aid = str(con["account_id"])
        if not aid or aid in seen_account_ids:
            continue
        seen_account_ids.add(aid)
        matched_accounts.append({
            "id":          aid,
            "name":        "",
            "dealer_id":   "",
            "matched_on":  "contact",
            "account_url": con["account_url"],
            "_needs_name": True,
        })

    # Fetch names for SLP/contact-linked accounts we don't have yet
    needs_name = [a for a in matched_accounts if a.get("_needs_name")]
    if needs_name:
        name_results = await asyncio.gather(
            *[ac_get(f"accounts/{a['id']}") for a in needs_name],
            return_exceptions=True,
        )
        for acct, res in zip(needs_name, name_results):
            acct.pop("_needs_name", None)
            if not isinstance(res, Exception):
                acct["name"] = res.get("account", {}).get("name", "")
    for a in matched_accounts:
        a.pop("_needs_name", None)

    # If a numeric query returned nothing and the index hasn't finished building yet,
    # flag it so the UI can show a helpful "still loading" message instead of "no results".
    index_loading = False
    if is_numeric and not matched_accounts and _dealer_index_ts == 0:
        index_loading = True

    # ── Program filter: keep only accounts whose SLP platform matches ─────────
    if program:
        prog_lower = program.lower()
        # Build set of account IDs that have a matching SLP platform
        prog_account_ids: set = set()
        for slp in matched_slps:
            if (slp.get("channel") or "").lower() == prog_lower:
                prog_account_ids.add(str(slp["account_id"]))
        # Also check in-memory index for accounts not yet in matched_slps
        for aid, plat in _account_to_platform.items():
            if plat.lower() == prog_lower:
                prog_account_ids.add(str(aid))
        matched_accounts = [a for a in matched_accounts if str(a["id"]) in prog_account_ids]
        matched_slps     = [s for s in matched_slps     if (s.get("channel") or "").lower() == prog_lower]
        matched_contacts = [c for c in matched_contacts if str(c.get("account_id","")) in prog_account_ids]

    # ── BDR / owner / group post-filters ──────────────────────────────────────
    if bdr or owner_id or group:
        _allow: set | None = None
        def _intersect(s: set) -> None:
            nonlocal _allow
            _allow = s if _allow is None else _allow & s
        if bdr:
            _intersect({aid for aid, b in _account_to_bdr.items() if b == bdr})
        if owner_id:
            _intersect({aid for aid, o in _account_to_owner.items() if o == owner_id})
        if group:
            _intersect({aid for aid, g in _account_to_group.items() if g == group})
        if _allow is not None:
            matched_accounts = [a for a in matched_accounts if str(a.get("id","")) in _allow]
            matched_slps     = [s for s in matched_slps     if str(s.get("account_id","")) in _allow]
            matched_contacts = [c for c in matched_contacts if str(c.get("account_id","")) in _allow]

    total = len(matched_accounts) + len(matched_contacts) + len(matched_slps)
    print(f"[GLOBAL SEARCH] query={q} results={total}")
    return {
        "query":         q,
        "total":         total,
        "accounts":      matched_accounts,
        "contacts":      matched_contacts,
        "slps":          matched_slps,
        "index_loading": index_loading,
    }


@app.get("/api/accounts/filter-options")
async def accounts_filter_options():
    """Return distinct BDR names, owner name+id pairs, and group names for sidebar filters."""
    bdrs   = sorted({v for v in _account_to_bdr.values()   if v}, key=str.lower)
    owners = sorted(
        [{"id": uid, "name": _user_id_to_name.get(uid, uid)}
         for uid in {v for v in _account_to_owner.values() if v}],
        key=lambda x: x["name"].lower()
    )
    groups = sorted({v for v in _account_to_group.values() if v}, key=str.lower)
    return {"bdrs": bdrs, "owners": owners, "groups": groups}


@app.get("/api/global-search/export")
async def global_search_export(q: str = Query(default=" "),
                               program: Optional[str] = Query(None),
                               bdr: Optional[str] = Query(None),
                               owner_id: Optional[str] = Query(None),
                               group: Optional[str] = Query(None)):
    """Export global search results as CSV using the same in-memory data as the sidebar."""
    effective_q = q.strip() or " "
    search_data = await global_search(q=effective_q, program=program, bdr=bdr, owner_id=owner_id, group=group)

    matched_accounts = {str(a["id"]): a for a in search_data.get("accounts", [])}
    if not matched_accounts:
        fname = f"search_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return _csv_response([], fname)

    # Group already-filtered SLPs by account_id (same data the sidebar shows)
    slps_by_account: dict = defaultdict(list)
    for s in search_data.get("slps", []):
        slps_by_account[str(s.get("account_id", ""))].append(s)

    rows = []
    for aid, acct in matched_accounts.items():
        acct_name = acct.get("name") or _account_to_name.get(aid, "")
        slp_list  = slps_by_account.get(aid, [])

        if slp_list:
            for s in slp_list:
                rows.append({
                    "account_name":   acct_name,
                    "account_id":     aid,
                    "dealer_id":      s.get("dealer_id", ""),
                    "channel":        s.get("channel", ""),
                    "slp_status":     s.get("slp_status", ""),
                    "activated_date": str(s.get("activated_date", ""))[:10],
                    "oracle_ids":     s.get("oracle_ids", ""),
                    "assigned_bdr":   s.get("assigned_bdr", ""),
                })
        else:
            # Account matched (by name or SLP) but no SLP data in index — still include it
            rows.append({
                "account_name":   acct_name,
                "account_id":     aid,
                "dealer_id":      acct.get("dealer_id", ""),
                "dealer_program": _account_to_platform.get(aid, ""),
                "slp_status":     "",
                "activated_date": "",
                "oracle_ids":     "",
                "assigned_bdr":   "",
            })

    rows.sort(key=lambda x: x["account_name"].lower())
    fname = f"search_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return _csv_response(rows, fname)


@app.get("/api/global-search/export-contacts")
async def global_search_export_contacts(q: str = Query(default=" "),
                                        program: Optional[str] = Query(None),
                                        bdr: Optional[str] = Query(None),
                                        owner_id: Optional[str] = Query(None),
                                        group: Optional[str] = Query(None)):
    """Export all contacts for the matched accounts as CSV."""
    effective_q = q.strip() or " "
    search_data = await global_search(q=effective_q, program=program, bdr=bdr, owner_id=owner_id, group=group)

    matched_accounts = {str(a["id"]): a for a in search_data.get("accounts", [])}
    if not matched_accounts:
        fname = f"contacts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return _csv_response([], fname)

    # Fetch contacts for each matched account in parallel
    con_tasks = [
        ac_get("contacts", {"filters[account]": aid, "limit": 100})
        for aid in matched_accounts
    ]
    con_results = await asyncio.gather(*con_tasks, return_exceptions=True)

    rows = []
    for aid, con_resp in zip(matched_accounts.keys(), con_results):
        acct = matched_accounts[aid]
        acct_name = acct.get("name") or _account_to_name.get(aid, "")
        if isinstance(con_resp, dict):
            for c in con_resp.get("contacts", []):
                rows.append({
                    "account_name": acct_name,
                    "account_id":   aid,
                    "first_name":   c.get("firstName", ""),
                    "last_name":    c.get("lastName", ""),
                    "email":        c.get("email", ""),
                    "phone":        c.get("phone", ""),
                })

    rows.sort(key=lambda x: (x["account_name"].lower(), x["last_name"].lower()))
    fname = f"contacts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return _csv_response(rows, fname)


@app.post("/api/global-search/email")
async def global_search_email(
    recipients:  str          = Query(..., description="Comma-separated email addresses"),
    q:           str          = Query(default=" "),
    program:     Optional[str]= Query(None),
    report_type: str          = Query(default="accounts", description="accounts or contacts"),
):
    """Generate a search export CSV and email it via Gmail SMTP."""
    if not _SMTP_USER or not _SMTP_PASS:
        raise HTTPException(status_code=503, detail="Email not configured (SMTP_USER / SMTP_PASS missing)")

    to_list = [r.strip() for r in recipients.split(",") if r.strip()]
    if not to_list:
        raise HTTPException(status_code=400, detail="No valid recipients provided")

    effective_q = q.strip() or " "

    # Generate CSV using existing export logic
    if report_type == "contacts":
        resp        = await global_search_export_contacts(q=effective_q, program=program)
        fname       = f"contacts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        subject_tag = "Contacts"
    else:
        resp        = await global_search_export(q=effective_q, program=program)
        fname       = f"accounts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        subject_tag = "Accounts"

    csv_bytes = b"".join([chunk async for chunk in resp.body_iterator])

    label   = f'"{effective_q.strip()}"' + (f" · {program}" if program else "")
    subject = f"Microf Reports Export — {subject_tag} {label}"

    msg = MIMEMultipart()
    msg["From"]    = f"Microf Reports <{_SMTP_USER}>"
    msg["To"]      = ", ".join(to_list)
    msg["Subject"] = subject
    msg.attach(MIMEText(
        f"<p>Please find the attached {subject_tag.lower()} export for search: "
        f"<strong>{label}</strong>.</p>"
        f"<p style='color:#6b7280;font-size:0.85em;'>Sent from microf-search</p>",
        "html"
    ))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(csv_bytes)
    _enc.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
    msg.attach(part)

    await aiosmtplib.send(
        msg,
        hostname="smtp.gmail.com",
        port=587,
        start_tls=True,
        username=_SMTP_USER,
        password=_SMTP_PASS,
    )

    return {"ok": True, "to": to_list, "filename": fname}


# ═══════════════════════════════════════════════════════════════════════════
# ADMIN / SCHEDULER
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/me")
async def get_me(request: _Request):
    email = _get_session_email(request)
    em = email.lower() if email else ""
    if (not _AZ_CLIENT_ID) or em in _ADMIN_EMAILS:
        group = "admin"
    elif em in _CONTRACTOR_SUPPORT_EMAILS:
        group = "contractor_support"
    elif em in _ONBOARDING_EMAILS:
        group = "onboarding"
    elif em in _SALES_ADMIN_EMAILS:
        group = "sales_admin"
    elif em in _ACCT_MGMT_EMAILS:
        group = "account_management"
    else:
        group = "account_management"   # default — everyone sees basic reports
    is_admin     = group == "admin"
    is_apps_user = True
    return {"email": email or "anonymous", "is_admin": bool(is_admin), "is_apps_user": bool(is_apps_user), "group": group}


@app.get("/api/admin/schedules")
async def list_schedules(admin=Depends(_require_admin)):
    return {
        "schedules":      list(_schedules.values()),
        "_debug_file":    _SCHEDULES_FILE,
        "_debug_exists":  os.path.exists(_SCHEDULES_FILE),
        "_debug_writable": os.access(os.path.dirname(_SCHEDULES_FILE) or ".", os.W_OK),
    }


@app.post("/api/admin/schedules")
async def create_schedule(
    report_type:  str          = Query(...),
    frequency:    str          = Query(..., description="daily | weekly | monthly"),
    hour:         int          = Query(9),
    minute:       int          = Query(0),
    day_of_week:  Optional[str]= Query(None, description="mon-sun for weekly"),
    day_of_month: Optional[int]= Query(None, description="1-28 for monthly"),
    recipients:   str          = Query(..., description="Comma-separated emails"),
    label:        Optional[str]= Query(None),
    period:       Optional[str]= Query(None, description="Date preset applied at run time: yesterday | last_week | this_month | last_month | last_quarter | ytd"),
    admin=Depends(_require_admin),
):
    if report_type not in _REPORT_JOBS:
        raise HTTPException(400, f"Unknown report type. Valid: {list(_REPORT_JOBS)}")
    if frequency not in ("daily", "weekly", "monthly"):
        raise HTTPException(400, "frequency must be daily, weekly, or monthly")

    job_id = str(_uuid.uuid4())[:8]
    s = {
        "id":           job_id,
        "report_type":  report_type,
        "frequency":    frequency,
        "hour":         hour,
        "minute":       minute,
        "day_of_week":  day_of_week or "mon",
        "day_of_month": day_of_month or 1,
        "recipients":   [r.strip() for r in recipients.split(",") if r.strip()],
        "label":        label or report_type,
        "period":       period or "",
        "created_at":   datetime.now().isoformat(),
    }
    _register_schedule(s)
    return {"ok": True, "schedule": s}


@app.delete("/api/admin/schedules/{job_id}")
async def delete_schedule(job_id: str, admin=Depends(_require_admin)):
    if job_id not in _schedules:
        raise HTTPException(404, "Schedule not found")
    try:
        _scheduler.remove_job(job_id)
    except Exception:
        pass
    del _schedules[job_id]
    _save_schedules_to_disk()
    return {"ok": True}


@app.get("/api/admin/schedules/export-json")
async def export_schedules_json(admin=Depends(_require_admin)):
    """Return current schedules as a JSON string suitable for pasting into SCHEDULES_JSON env var."""
    data = list(_schedules.values())
    return {"json": json.dumps(data), "count": len(data)}


# ═══════════════════════════════════════════════════════════════════════════
# GROUP-BY / SUMMARY ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/summary")
async def group_by_summary(
    object_type:  str           = Query(...),
    group_by:     str           = Query(...),
    filters:      Optional[str] = Query(None),
    count_field:  Optional[str] = Query(None, description="Count distinct values of this field"),
):
    """Count records grouped by any field. Supports filters."""
    filter_list = json.loads(filters) if filters else []

    fetchers = {
        "slp":             fetch_slp_records,
        "trainings":       fetch_training_records,
        "license_details": fetch_license_records,
        "accounts":        fetch_account_records_basic,
        "contacts":        fetch_contact_records_basic,
        "deals":           fetch_deal_records_basic,
        "notes":           fetch_note_records,
    }
    if object_type not in fetchers:
        raise HTTPException(status_code=400, detail=f"Unknown object type: {object_type}")

    records = await fetchers[object_type]()
    if filter_list:
        records = [r for r in records if all(evaluate_filter(r, f) for f in filter_list)]

    groups: dict = defaultdict(lambda: {"count": 0, "distinct": set()})
    for r in records:
        key = str(r.get(group_by) or "(empty)")
        groups[key]["count"] += 1
        if count_field and (v := r.get(count_field)):
            groups[key]["distinct"].add(str(v))

    results = []
    for grp, data in sorted(groups.items(), key=lambda x: x[1]["count"], reverse=True):
        row = {"group": grp, "count": data["count"]}
        if count_field:
            row["distinct_count"] = len(data["distinct"])
        results.append(row)

    return {"group_by": group_by, "total_groups": len(results), "total_records": len(records), "records": results}


# ═══════════════════════════════════════════════════════════════════════════
# CONTACT LOOKUP
# ═══════════════════════════════════════════════════════════════════════════

ALLOWED_ACTIVITY_TYPES = {"send", "open", "click", "bounce", "forward", "unsubscribe", "note", "task"}
ACTIVITY_LABELS = {"send": "Email Sent", "open": "Email Opened", "click": "Email Clicked",
                   "bounce": "Email Bounced", "forward": "Email Forwarded",
                   "unsubscribe": "Unsubscribed", "note": "Note", "task": "Task / Call"}
ACTIVITY_ICONS  = {"send": "📧", "open": "📬", "click": "🔗", "bounce": "⚠️",
                   "forward": "↩️", "unsubscribe": "🚫", "note": "📝", "task": "📞"}


@app.get("/api/contact-search")
async def contact_search(q: str = Query(..., min_length=2)):
    results, seen_ids = [], set()
    async with httpx.AsyncClient(timeout=30) as client:
        try:
            r = await client.get(ac_url("contacts"), headers=HEADERS, params={"search": q, "limit": 20})
            r.raise_for_status()
            for c in r.json().get("contacts", []):
                if c["id"] not in seen_ids:
                    seen_ids.add(c["id"])
                    results.append({"id": c["id"], "firstName": c.get("firstName", ""),
                                    "lastName": c.get("lastName", ""), "email": c.get("email", ""),
                                    "phone": c.get("phone", ""), "orgName": c.get("orgname", "")})
        except Exception:
            pass
    return {"contacts": results[:20]}


@app.get("/api/contact-profile/{contact_id}")
async def contact_profile(contact_id: str):
    contact_data, activity_data, notes_data = await asyncio.gather(
        ac_get(f"contacts/{contact_id}"),
        ac_get(f"contacts/{contact_id}/activityLogs", {"limit": 100}),
        ac_get(f"contacts/{contact_id}/notes", {"limit": 50}),
        return_exceptions=True,
    )

    contact = {}
    if isinstance(contact_data, dict):
        c = contact_data.get("contact", {})
        contact = {"id": c.get("id"), "firstName": c.get("firstName", ""),
                   "lastName": c.get("lastName", ""), "email": c.get("email", ""),
                   "phone": c.get("phone", ""), "orgName": c.get("orgname", ""),
                   "created": c.get("cdate", ""), "updated": c.get("udate", "")}

    activity = []
    if isinstance(activity_data, dict):
        for log in activity_data.get("contactActivities", []):
            a_type = log.get("type", "").lower()
            if a_type in ALLOWED_ACTIVITY_TYPES:
                desc = log.get("subject") or (log.get("campaign", {}).get("name", "")
                       if isinstance(log.get("campaign"), dict) else "")
                activity.append({"type": a_type, "label": ACTIVITY_LABELS.get(a_type, a_type),
                                  "icon": ACTIVITY_ICONS.get(a_type, "•"), "description": desc,
                                  "timestamp": log.get("tstamp", log.get("cdate", ""))})

    if isinstance(notes_data, dict):
        for note in notes_data.get("notes", []):
            activity.append({"type": "note", "label": "Note", "icon": "📝",
                              "description": note.get("note", ""), "timestamp": note.get("cdate", "")})

    def _ts(item):
        try:
            return datetime.fromisoformat(item.get("timestamp", "").replace("Z", "+00:00"))
        except Exception:
            return datetime.min.replace(tzinfo=None)

    activity.sort(key=_ts, reverse=True)
    return {"contact": contact, "activity": activity}


# ═══════════════════════════════════════════════════════════════════════════
# FRONTEND
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/mover")
async def serve_mover():
    return FileResponse("static/mover.html")

@app.get("/contacts")
async def serve_contact_lookup():
    return FileResponse("static/contact_lookup.html")

@app.get("/")
async def serve_ui(_: None = Depends(require_auth)):
    return FileResponse("static/index.html")


# ═══════════════════════════════════════════════════════════════════════════
# BROWSER-VIEW REPORTS  (JSON + CSV download, no email)
# ═══════════════════════════════════════════════════════════════════════════

def _csv_response(records: list, filename: str):
    out = io.StringIO()
    if records:
        w = csv.DictWriter(out, fieldnames=list(records[0].keys()))
        w.writeheader(); w.writerows(records)
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.get("/api/report/training-activity")
async def report_training_activity(
    from_date: Optional[str] = Query(None),
    to_date:   Optional[str] = Query(None),
    format:    str           = Query("json"),
):
    from datetime import timezone
    tz = timezone.utc
    from_dt = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=tz) if from_date else None
    to_dt   = datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=tz) if to_date else None

    tr_records  = await ac_get_all(f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {})
    account_ids: set = set()
    candidates = []
    for r in tr_records:
        fields   = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        date_str = str(fields.get("date-of-training", "")).strip()
        if date_str and (from_dt or to_dt):
            try:
                td = (datetime.fromisoformat(date_str.replace("Z", "+00:00")) if "T" in date_str
                      else datetime.strptime(date_str[:10], "%Y-%m-%d").replace(tzinfo=tz))
                if from_dt and td < from_dt: continue
                if to_dt   and td > to_dt:   continue
            except Exception:
                continue
        elif not date_str and (from_dt or to_dt):
            continue
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if acc_id: account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    results = []
    for c in candidates:
        f   = c["fields"]
        aid = c["account_id"] or ""
        results.append({
            "account":       acct_cache.get(aid, ""),
            "dealer_id":     _account_to_dealer.get(aid, ""),
            "trained_by":    f.get("trained-by", ""),
            "training_type": f.get("training-type", ""),
            "agenda":        f.get("training-agenda", ""),
            "date":          str(f.get("date-of-training", ""))[:10],
            "notes":         (f.get("training-notes", "") or "")[:200],
        })
    results.sort(key=lambda x: x["date"], reverse=True)
    if format == "csv":
        return _csv_response(results, f"training_activity_{datetime.now().strftime('%Y%m%d')}.csv")
    return {"count": len(results), "records": results}


@app.get("/api/report/stale-untrained")
async def report_stale_untrained(
    from_date:  Optional[str] = Query(None, description="Filter by activation date from"),
    to_date:    Optional[str] = Query(None, description="Filter by activation date to"),
    stale_days: int           = Query(90),
    platform:   Optional[str] = Query(None),
    bdr:        Optional[str] = Query(None),
    format:     str           = Query("json"),
):
    today        = date.today()
    stale_cutoff = str(today - timedelta(days=stale_days))

    slp_records = await get_slp_cache()
    tr_records  = await ac_get_all(f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {})

    training_by_acct: dict = defaultdict(list)
    for r in tr_records:
        fields   = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        date_str = str(fields.get("date-of-training", "")).strip()
        if not date_str: continue
        for aid in r.get("relationships", {}).get("account", []):
            training_by_acct[str(aid)].append(date_str[:10])

    account_ids: set = set()
    candidates = []
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        if fields.get("slp-status-detail") != "Contractor Activated": continue
        slp_plat = str(fields.get("channel", "")).strip()
        slp_bdr  = str(fields.get("assigned-bdr", "")).strip()
        rel      = r.get("relationships", {}).get("account", [])
        acc_id   = str(rel[0]) if rel else None
        eff_plat = slp_plat
        eff_bdr  = slp_bdr  or _account_to_bdr.get(acc_id or "", "")
        if platform and eff_plat != platform: continue
        if bdr      and eff_bdr  != bdr:      continue
        act_str  = str(fields.get("contractor-activated-date", "")).strip()
        act_date = act_str[:10] if act_str else ""
        if from_date and act_date and act_date < from_date: continue
        if to_date   and act_date and act_date > to_date:   continue
        if acc_id: account_ids.add(acc_id)
        trainings  = training_by_acct.get(acc_id or "", [])
        last_train = max(trainings) if trainings else None
        if last_train and last_train >= stale_cutoff: continue
        days_stale = (today - date.fromisoformat(last_train)).days if last_train else None
        candidates.append({"fields": fields, "account_id": acc_id, "act_date": act_date,
                           "training_count": len(trainings), "last_training": last_train or "",
                           "days_stale": days_stale})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    results = []
    for c in sorted(candidates, key=lambda x: x["days_stale"] or 99999, reverse=True):
        f   = c["fields"]
        aid = c["account_id"] or ""
        results.append({
            "account":         acct_cache.get(aid, ""),
            "dealer_id":       f.get("dealer-id")    or _account_to_dealer.get(aid, ""),
            "channel":         f.get("channel", ""),
            "bdr":             f.get("assigned-bdr") or _account_to_bdr.get(aid, ""),
            "activation_date": c["act_date"],
            "training_count":  c["training_count"],
            "last_training":   c["last_training"] or "Never",
            "days_stale":      c["days_stale"] if c["days_stale"] is not None else "Never trained",
        })
    if format == "csv":
        return _csv_response(results, f"stale_untrained_{datetime.now().strftime('%Y%m%d')}.csv")
    return {"count": len(results), "records": results}


@app.get("/api/report/account-status")
async def report_account_status(format: str = Query("json")):
    all_accounts = await ac_get_all("accounts", "accounts", {})
    cf_map       = await _fetch_acct_cf_map({"19", "23"})

    results = []
    for a in all_accounts:
        aid  = str(a.get("id", ""))
        cfs  = cf_map.get(aid, {})
        results.append({
            "account":      a.get("name", ""),
            "dealer_id":    _account_to_dealer.get(aid, ""),
            "platform":     _account_to_platform.get(aid, ""),
            "bdr":          _account_to_bdr.get(aid, ""),
            "status":       cfs.get("19", ""),
            "sales_region": cfs.get("23", ""),
        })
    results.sort(key=lambda x: (x["status"], x["sales_region"], x["account"]))
    if format == "csv":
        return _csv_response(results, f"account_status_{datetime.now().strftime('%Y%m%d')}.csv")
    return {"count": len(results), "records": results}


@app.get("/api/report/platform-breakdown")
async def report_platform_breakdown(
    from_date: Optional[str] = Query(None),
    to_date:   Optional[str] = Query(None),
    format:    str           = Query("json"),
):
    from datetime import timezone
    tz = timezone.utc
    from_dt = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=tz) if from_date else None
    to_dt   = datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=tz) if to_date else None

    slp_records = await get_slp_cache()
    plat_data: dict = defaultdict(lambda: {"new_activations": 0, "active_slps": 0,
                                           "total_slps": 0, "bdrs": defaultdict(int)})
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        plat   = str(fields.get("channel", "")).strip() or "Unknown"
        bdr    = (str(fields.get("assigned-bdr", "")).strip()
                  or _account_to_bdr.get(acc_id or "", "") or "Unassigned")
        plat_data[plat]["total_slps"] += 1
        if fields.get("slp-status-detail") == "Contractor Activated":
            plat_data[plat]["active_slps"] += 1
            act_str = str(fields.get("contractor-activated-date", "")).strip()
            if act_str and (from_dt or to_dt):
                try:
                    act_dt = (datetime.fromisoformat(act_str.replace("Z", "+00:00")) if "T" in act_str
                              else datetime.strptime(act_str[:10], "%Y-%m-%d").replace(tzinfo=tz))
                    if from_dt and act_dt < from_dt: pass
                    elif to_dt and act_dt > to_dt:   pass
                    else:
                        plat_data[plat]["new_activations"] += 1
                        plat_data[plat]["bdrs"][bdr] += 1
                except Exception:
                    pass
            elif not from_dt and not to_dt:
                plat_data[plat]["new_activations"] += 1
                plat_data[plat]["bdrs"][bdr] += 1

    results = []
    for plat, d in sorted(plat_data.items()):
        top_bdr = max(d["bdrs"], key=d["bdrs"].get) if d["bdrs"] else ""
        results.append({
            "channel":         plat,
            "new_activations": d["new_activations"],
            "active_slps":     d["active_slps"],
            "total_slps":      d["total_slps"],
            "top_bdr":         top_bdr,
        })
    results.sort(key=lambda x: x["new_activations"], reverse=True)
    if format == "csv":
        return _csv_response(results, f"channel_breakdown_{datetime.now().strftime('%Y%m%d')}.csv")
    return {"count": len(results), "records": results}


@app.get("/api/report/partner-activation")
async def report_partner_activation(
    from_date: Optional[str] = Query(None),
    to_date:   Optional[str] = Query(None),
    format:    str           = Query("json"),
):
    cf_map       = await _fetch_acct_cf_map({"26"})
    all_accounts = await ac_get_all("accounts", "accounts", {})
    acct_by_id   = {str(a.get("id", "")): a for a in all_accounts}

    results = []
    for aid, cfs in cf_map.items():
        pa_val = cfs.get("26", "")
        if not pa_val: continue
        pa_str = str(pa_val)[:10]
        try:
            if from_date and pa_str < from_date: continue
            if to_date   and pa_str > to_date:   continue
        except Exception:
            pass
        a = acct_by_id.get(aid, {})
        results.append({
            "account":            a.get("name", ""),
            "dealer_id":          _account_to_dealer.get(aid, ""),
            "platform":           _account_to_platform.get(aid, ""),
            "bdr":                _account_to_bdr.get(aid, ""),
            "partner_activation": pa_str,
        })
    results.sort(key=lambda x: x["partner_activation"], reverse=True)
    if format == "csv":
        return _csv_response(results, f"partner_activation_{datetime.now().strftime('%Y%m%d')}.csv")
    return {"count": len(results), "records": results}


@app.get("/api/report/oracle-missing")
async def report_oracle_missing(
    from_date: Optional[str] = Query(None, description="Filter by activation date from"),
    to_date:   Optional[str] = Query(None, description="Filter by activation date to"),
    platform:  Optional[str] = Query(None),
    bdr:       Optional[str] = Query(None),
    format:    str           = Query("json"),
):
    slp_records = await get_slp_cache()
    cf_map      = await _fetch_acct_cf_map({"118"})

    account_ids: set = set()
    candidates = []
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        if fields.get("slp-status-detail") != "Contractor Activated": continue
        rel      = r.get("relationships", {}).get("account", [])
        acc_id   = str(rel[0]) if rel else None
        if cf_map.get(acc_id or "", {}).get("118"): continue
        slp_plat = str(fields.get("channel", "")).strip()
        slp_bdr  = str(fields.get("assigned-bdr", "")).strip()
        eff_plat = slp_plat
        eff_bdr  = slp_bdr  or _account_to_bdr.get(acc_id or "", "")
        if platform and eff_plat != platform: continue
        if bdr      and eff_bdr  != bdr:      continue
        act_str  = str(fields.get("contractor-activated-date", "")).strip()
        act_date = act_str[:10] if act_str else ""
        if from_date and act_date and act_date < from_date: continue
        if to_date   and act_date and act_date > to_date:   continue
        if acc_id: account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id, "act_date": act_date})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    results = []
    for c in candidates:
        f   = c["fields"]
        aid = c["account_id"] or ""
        results.append({
            "account":         acct_cache.get(aid, ""),
            "dealer_id":       f.get("dealer-id")    or _account_to_dealer.get(aid, ""),
            "channel":         f.get("channel", ""),
            "bdr":             f.get("assigned-bdr") or _account_to_bdr.get(aid, ""),
            "activation_date": c["act_date"],
        })
    results.sort(key=lambda x: (x["channel"], x["bdr"], x["account"]))
    if format == "csv":
        return _csv_response(results, f"oracle_missing_{datetime.now().strftime('%Y%m%d')}.csv")
    return {"count": len(results), "records": results}


@app.get("/api/report/account-program-search")
async def report_account_program_search(
    account_name: str           = Query("", description="Fuzzy account name filter"),
    program:      Optional[str] = Query(None, description="SLP platform/dealer program"),
    format:       str           = Query("json"),
):
    """Search accounts by name (fuzzy) and dealer program (from SLP platform)."""
    import re as _re

    name_q   = account_name.strip().lower()
    # Split into words for multi-term matching (e.g. "ARS optimus" → ["ars", "optimus"])
    name_terms = [t for t in _re.split(r'\s+', name_q) if t] if name_q else []

    slp_records = await get_slp_cache()

    # Build account_id → list of SLP summaries
    slp_by_account: dict = {}
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        slp_plat = str(fields.get("channel", "")).strip()
        if program and slp_plat.lower() != program.lower():
            continue
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if not acc_id:
            continue
        if acc_id not in slp_by_account:
            slp_by_account[acc_id] = []
        slp_by_account[acc_id].append({
            "dealer_id":        fields.get("dealer-id", ""),
            "channel":          slp_plat,
            "slp_status":       fields.get("slp-status-detail", ""),
            "activated_date":   str(fields.get("contractor-activated-date", ""))[:10],
            "program_name":     fields.get("program-name-1", ""),
            "oracle_ids":       fields.get("oracle-producer-ids", ""),
            "assigned_bdr":     fields.get("assigned-bdr", ""),
        })

    if not slp_by_account:
        if format == "csv":
            return _csv_response([], "account_program_search.csv")
        return {"count": 0, "records": []}

    # Fetch account names — use in-memory index where possible
    results = []
    for acc_id, slps in slp_by_account.items():
        acct_name = _account_to_name.get(acc_id, "")
        if not acct_name:
            try:
                d = await ac_get(f"accounts/{acc_id}")
                acct_name = d.get("account", {}).get("name", "")
            except Exception:
                pass

        # Fuzzy name filter — all terms must appear in the name
        if name_terms:
            name_lower = acct_name.lower()
            if not all(t in name_lower for t in name_terms):
                continue

        for slp in slps:
            results.append({
                "account_name":    acct_name,
                "account_id":      acc_id,
                "dealer_id":       slp["dealer_id"],
                "channel":         slp["channel"],
                "slp_status":      slp["slp_status"],
                "activated_date":  slp["activated_date"],
                "program_name":    slp["program_name"],
                "oracle_ids":      slp["oracle_ids"],
                "assigned_bdr":    slp["assigned_bdr"],
            })

    results.sort(key=lambda x: (x["account_name"].lower(), x["channel"]))
    if format == "csv":
        return _csv_response(results, f"account_program_search_{datetime.now().strftime('%Y%m%d')}.csv")
    return {"count": len(results), "records": results}


# ═══════════════════════════════════════════════════════════════════════════
# SCHEDULED EMAIL REPORTS
# Triggered by GitHub Actions cron → /api/send-report/{type}
# Can also be triggered manually via the same endpoint (Basic Auth required).
# ═══════════════════════════════════════════════════════════════════════════

_HTML_WRAPPER = """\
<!DOCTYPE html><html>
<head><style>
  body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
        color:#1a1a2e;background:#f5f7fc;margin:0;padding:20px}}
  .card{{background:white;border-radius:12px;padding:24px 28px;max-width:960px;
         margin:0 auto;box-shadow:0 2px 12px rgba(0,0,0,.07)}}
  h2{{color:#1e44b8;margin:0 0 4px;font-size:1.3rem}}
  .sub{{color:#666;font-size:.85rem;margin-bottom:20px}}
  table{{width:100%;border-collapse:collapse;font-size:.82rem}}
  th{{background:#f0f3fb;color:#1e44b8;font-weight:600;text-align:left;
      padding:8px 10px;border-bottom:2px solid #d0d7f0}}
  td{{padding:7px 10px;border-bottom:1px solid #eef0f8;white-space:nowrap}}
  tr:nth-child(even) td{{background:#fafbff}}
  .footer{{color:#aaa;font-size:.75rem;margin-top:16px;text-align:center}}
</style></head><body><div class="card">
<h2>{title}</h2><div class="sub">{subtitle}</div>
{table}
<div class="footer">Generated by Moogle &middot; {timestamp}</div>
</div></body></html>"""


def _html_table(records: list, cols: list) -> str:
    """Build an HTML table. cols = [(header_label, dict_key), ...]"""
    if not records:
        return "<p style='color:#888;padding:12px 0'>No records found.</p>"
    headers = "".join(f"<th>{h}</th>" for h, _ in cols)
    rows = "".join(
        "<tr>" + "".join(f"<td>{str(r.get(k, '') or '')}</td>" for _, k in cols) + "</tr>"
        for r in records
    )
    return f"<table><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table>"


def _csv_bytes(records: list) -> bytes:
    if not records:
        return b""
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=list(records[0].keys()))
    w.writeheader()
    w.writerows(records)
    return buf.getvalue().encode()


def _enrich_record(record: dict, account_id: str) -> dict:
    """Append standard account fields from in-memory globals to a record dict.
    Uses setdefault so existing values are never overwritten.
    Called by email job functions to ensure CSV attachments match download endpoints."""
    if not account_id:
        return record
    record.setdefault("DBA Name",          _account_to_dba.get(account_id, ""))
    record.setdefault("Account Status",    _account_to_status.get(account_id, ""))
    record.setdefault("Account Type",      _account_to_type.get(account_id, ""))
    record.setdefault("Sales Region",      _account_to_region.get(account_id, ""))
    record.setdefault("Doing Business In", _account_to_states.get(account_id, ""))
    record.setdefault("Last App Date",     _account_to_last_app.get(account_id, ""))
    record.setdefault("Last RPA Date",     _account_to_last_rpa.get(account_id, ""))
    record.setdefault("Vendor Tax-ID",     _account_to_tax_id.get(account_id, ""))
    record.setdefault("Website",           _account_to_website.get(account_id, ""))
    record.setdefault("Phone",             _account_to_phone.get(account_id, ""))
    record.setdefault("Address",           _account_to_address.get(account_id, ""))
    record.setdefault("City",              _account_to_city.get(account_id, ""))
    record.setdefault("State/Prov",        _account_to_state_prov.get(account_id, ""))
    record.setdefault("Zip",               _account_to_zip.get(account_id, ""))
    return record


async def _send_email(subject: str, html: str,
                      csv_data: bytes = None, csv_name: str = None,
                      recipients: list = None):
    """Send an HTML email with an optional CSV attachment via SMTP STARTTLS.
    Pass recipients to override the default REPORT_RECIPIENTS env list."""
    to = recipients or _RECIPIENTS
    if not _SMTP_USER or not to:
        print(f"[reports] Email not configured — skipping: {subject}")
        return
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = f"{_SMTP_FROM} <{_SMTP_USER}>"
    msg["To"]      = ", ".join(to)
    msg.attach(MIMEText(html, "html"))
    if csv_data:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(csv_data)
        _enc.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{csv_name}"')
        msg.attach(part)
    try:
        await aiosmtplib.send(
            msg,
            hostname=_SMTP_HOST,
            port=_SMTP_PORT,
            username=_SMTP_USER,
            password=_SMTP_PASS,
            start_tls=True,
        )
        print(f"[reports] Sent '{subject}' → {to}")
    except Exception as exc:
        print(f"[reports] Email failed: {exc}")


# ── Date-range helpers ────────────────────────────────────────────────────

def _resolve_date_range(
    start: Optional[date],
    end:   Optional[date],
    preset: Optional[str],
    *,
    default_start: Optional[date] = None,
    default_end:   Optional[date] = None,
) -> tuple:
    """Resolve start/end from explicit params, a named preset, or defaults.
    Presets: yesterday | last_week | last_7_days | last_30_days | last_90_days |
             this_week | this_month | last_month
    """
    today = date.today()
    if preset:
        p = preset.lower().replace("-", "_")
        if p == "yesterday":
            d = today - timedelta(days=1); return d, d
        if p == "last_week":
            dow = today.weekday()                      # Mon=0 … Sun=6
            last_sun = today - timedelta(days=dow + 1)
            return last_sun - timedelta(days=6), last_sun
        if p in ("last_7_days", "last_7"):
            return today - timedelta(days=7), today - timedelta(days=1)
        if p in ("last_30_days", "last_30"):
            return today - timedelta(days=30), today - timedelta(days=1)
        if p in ("last_90_days", "last_90"):
            return today - timedelta(days=90), today - timedelta(days=1)
        if p == "this_week":
            return today - timedelta(days=today.weekday()), today
        if p in ("this_month", "current_month"):
            return today.replace(day=1), today
        if p == "last_month":
            first_this = today.replace(day=1)
            last_prev  = first_this - timedelta(days=1)
            return last_prev.replace(day=1), last_prev
        if p == "this_quarter":
            q = (today.month - 1) // 3
            return today.replace(month=q*3+1, day=1), today
        if p == "last_quarter":
            q = (today.month - 1) // 3
            if q == 0:
                qs, qy = 3, today.year - 1
            else:
                qs, qy = q - 1, today.year
            from calendar import monthrange as _mr
            qe_month = qs * 3 + 3
            qe_day   = _mr(qy, qe_month)[1]
            return date(qy, qs*3+1, 1), date(qy, qe_month, qe_day)
        if p in ("this_year", "ytd"):
            return today.replace(month=1, day=1), today
        if p == "last_year":
            y = today.year - 1
            return date(y, 1, 1), date(y, 12, 31)
        if p in ("last_18_months", "last_18mo", "last_18"):
            return today - timedelta(days=548), today
        if p in ("all", "all_time"):
            return date(2000, 1, 1), today
    return (start or default_start), (end or default_end)


async def _fetch_acct_cf_map(field_ids: set) -> dict:
    """Bulk-fetch account custom fields. Returns {account_id: {field_id_str: value}}.
    Backed by the shared _acct_cf_raw cache (kept warm by _acct_cf_cache_loop);
    only blocks on a live refresh if the cache is empty or has gone stale."""
    if not _acct_cf_raw or (_time.time() - _acct_cf_raw_ts) > _ACCT_CF_TTL:
        await _refresh_acct_cf_cache()

    field_ids_int = {int(f) for f in field_ids}
    result: dict  = defaultdict(dict)
    for item in _acct_cf_raw:
        fid = int(item.get("customFieldId", 0))
        if fid not in field_ids_int:
            continue
        aid = str(item.get("accountId", ""))
        fv  = item.get("fieldValue") or ""
        val = (fv if isinstance(fv, str) else (str(fv[0]) if fv else "")).strip()
        if aid and val:
            result[aid][str(fid)] = val
    return dict(result)


# ── Activations (daily Mon–Fri) ──────────────────────────────────────────

async def _job_activations(start_date: Optional[date] = None, end_date: Optional[date] = None,
                           preset: Optional[str] = None, recipients: list = None,
                           platform: Optional[str] = None, bdr: Optional[str] = None,
                           state: Optional[str] = None, exclude_platforms: Optional[str] = None):
    """Email 'Contractor Activated' SLP records for a date range (defaults to yesterday)."""
    from datetime import timezone
    tz_utc = timezone.utc
    today  = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_start=today - timedelta(days=1))
    if _start is None: _start = today - timedelta(days=1)
    if _end   is None: _end   = _start
    from_dt    = datetime(_start.year, _start.month, _start.day, tzinfo=tz_utc)
    to_dt      = datetime(_end.year,   _end.month,   _end.day,   23, 59, 59, tzinfo=tz_utc)
    yesterday  = str(_start)
    date_label = str(_start) if _start == _end else f"{_start} to {_end}"
    print(f"[reports] Activations for {date_label}")
    exclude_set = {p.strip() for p in exclude_platforms.split(",")} if exclude_platforms else set()

    slp_records = await get_slp_cache()
    account_ids: set = set()
    candidates = []
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        if fields.get("slp-status-detail") != "Contractor Activated":
            continue
        plat = str(fields.get("channel", "")).strip()
        plat_norm = _normalize_platform(plat)
        if platform and plat_norm != _normalize_platform(platform):
            continue
        if plat_norm in exclude_set or plat in exclude_set:
            continue
        if state:
            states_val = str(fields.get("doing-business-in-states", "") or "").upper()
            if state.upper() not in [s.strip() for s in states_val.split(",")]:
                continue
        act_str = str(fields.get("contractor-activated-date", "")).strip()
        if not act_str:
            continue
        try:
            act_dt = (datetime.fromisoformat(act_str.replace("Z", "+00:00")) if "T" in act_str
                      else datetime.strptime(act_str[:10], "%Y-%m-%d").replace(tzinfo=tz_utc))
        except Exception:
            continue
        if not (from_dt <= act_dt <= to_dt):
            continue
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        slp_bdr = str(fields.get("assigned-bdr", "")).strip()
        eff_bdr = slp_bdr or _account_to_bdr.get(acc_id or "", "")
        if bdr == "__unassigned__":
            if eff_bdr:
                continue
        elif bdr and eff_bdr != bdr:
            continue
        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id, "eff_bdr": eff_bdr})

    async def _fetch_acct_act(aid: str) -> tuple:
        try:
            name_r, cf_r = await asyncio.gather(
                ac_get(f"accounts/{aid}"),
                ac_get(f"accounts/{aid}/accountCustomFieldData"),
                return_exceptions=True,
            )
            name = name_r.get("account", {}).get("name", "") if isinstance(name_r, dict) else ""
            cfs: dict = {}
            if isinstance(cf_r, dict):
                for item in cf_r.get("accountCustomFieldData", []):
                    cfs[str(item.get("customFieldId", ""))] = (item.get("fieldValue") or "").strip()
            return aid, {"name": name, "channel": "", "bdr": cfs.get("119", ""),
                         "dealer_id": _account_to_dealer.get(aid, "")}
        except Exception:
            return aid, {"name": "", "channel": "", "bdr": "", "dealer_id": _account_to_dealer.get(aid, "")}

    acct_cache: dict = dict(await asyncio.gather(*[_fetch_acct_act(aid) for aid in account_ids]))

    records = []
    for c in candidates:
        f    = c["fields"]
        acct = acct_cache.get(c["account_id"]) or {}
        aid  = c["account_id"] or ""
        rec  = {
            "Account":                   acct.get("name") or f.get("name", ""),
            "Dealer ID":                 f.get("dealer-id") or acct.get("dealer_id", ""),
            "Channel":                   f.get("channel") or acct.get("channel", ""),
            "BDR":                       c.get("eff_bdr") or f.get("assigned-bdr") or acct.get("bdr", ""),
            "Activated":                 str(f.get("contractor-activated-date", "") or "")[:10],
            "SLP Status":                f.get("slp-status-detail", ""),
            "Oracle Producer IDs":       f.get("oracle-producer-ids", ""),
            "Doing Business In States":  f.get("doing-business-in-states", ""),
            "EIN":                       f.get("ein", ""),
            "Contractor Reactivation":   f.get("contractor-reactivation", ""),
            "Original Owner":            f.get("original-owner", ""),
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: x["Activated"], reverse=True)

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"),
            ("Channel","Channel"), ("BDR","BDR"), ("Activated","Activated")]
    html = _HTML_WRAPPER.format(
        title=f"Activations — {date_label}",
        subtitle=f"{len(records)} new activation{'s' if len(records) != 1 else ''}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    csv_label = yesterday if _start == _end else f"{_start}_{_end}"
    await _send_email(
        subject=f"Activations Report — {date_label} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"activations_{csv_label}.csv",
        recipients=recipients,
    )


# ── Enrollment Report (all statuses, optional status filter) ─────────────

async def _job_enrollment(start_date: Optional[date] = None, end_date: Optional[date] = None,
                          preset: Optional[str] = None, recipients: list = None,
                          slp_status: Optional[str] = None):
    """Email all SLP enrollment records for a date range, optionally filtered by status."""
    from datetime import timezone
    tz_utc = timezone.utc
    today  = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_start=today - timedelta(days=1))
    if _start is None: _start = today - timedelta(days=1)
    if _end   is None: _end   = _start
    from_dt    = datetime(_start.year, _start.month, _start.day, tzinfo=tz_utc)
    to_dt      = datetime(_end.year,   _end.month,   _end.day,   23, 59, 59, tzinfo=tz_utc)
    date_label = str(_start) if _start == _end else f"{_start} to {_end}"
    status_label = slp_status or "All Statuses"
    print(f"[reports] Enrollment report for {date_label} | status={status_label}")

    slp_records = await get_slp_cache()
    account_ids: set = set()
    candidates = []
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}

        if slp_status and fields.get("slp-status-detail") != slp_status:
            continue

        # Date filter on contractor-activated-date, fall back to enrollment-request-date
        date_str = (str(fields.get("contractor-activated-date", "")).strip() or
                    str(fields.get("enrollment-request-date", "")).strip())
        if not date_str:
            continue
        try:
            rec_dt = (datetime.fromisoformat(date_str.replace("Z", "+00:00")) if "T" in date_str
                      else datetime.strptime(date_str[:10], "%Y-%m-%d").replace(tzinfo=tz_utc))
        except Exception:
            continue
        if not (from_dt <= rec_dt <= to_dt):
            continue

        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id})

    async def _fetch_acct(aid: str) -> tuple:
        try:
            name_r, cf_r = await asyncio.gather(
                ac_get(f"accounts/{aid}"),
                ac_get(f"accounts/{aid}/accountCustomFieldData"),
                return_exceptions=True,
            )
            name = name_r.get("account", {}).get("name", "") if isinstance(name_r, dict) else ""
            bdr  = ""
            if isinstance(cf_r, dict):
                for item in cf_r.get("customerAccountCustomFieldData", []):
                    if str(item.get("custom_field_id", "")) == "119":
                        bdr = (item.get("custom_field_text_value") or "").strip()
            return aid, {"name": name, "bdr": bdr}
        except Exception:
            return aid, {"name": "", "bdr": ""}

    acct_cache: dict = dict(await asyncio.gather(*[_fetch_acct(aid) for aid in account_ids]))

    records = []
    for c in candidates:
        f    = c["fields"]
        acct = acct_cache.get(c["account_id"]) or {}
        aid  = c["account_id"] or ""
        rec  = {
            "Account":                   acct.get("name") or f.get("name", ""),
            "Dealer ID":                 f.get("dealer-id") or _account_to_dealer.get(aid, ""),
            "Channel":                   f.get("channel", ""),
            "BDR":                       f.get("assigned-bdr") or acct.get("bdr", ""),
            "SLP Status":                f.get("slp-status-detail", ""),
            "Activated":                 str(f.get("contractor-activated-date", "") or "")[:10],
            "Enrollment Request Date":   str(f.get("enrollment-request-date", "") or "")[:10],
            "Oracle Producer IDs":       f.get("oracle-producer-ids", ""),
            "Doing Business In States":  f.get("doing-business-in-states", ""),
            "EIN":                       f.get("ein", ""),
            "Original Owner":            f.get("original-owner", ""),
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: (x["Activated"] or x["Enrollment Request Date"]), reverse=True)

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"), ("Channel","Channel"),
            ("BDR","BDR"), ("SLP Status","SLP Status"), ("Activated","Activated")]
    subtitle = f"{len(records)} enrollment{'s' if len(records) != 1 else ''} — {status_label}"
    html = _HTML_WRAPPER.format(
        title=f"Enrollment Report — {date_label}",
        subtitle=subtitle,
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    csv_label = str(_start) if _start == _end else f"{_start}_{_end}"
    await _send_email(
        subject=f"Enrollment Report — {date_label} ({len(records)} records, {status_label})",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"enrollment_{csv_label}.csv",
        recipients=recipients,
    )


# ── License Expiration (weekly Monday) ───────────────────────────────────

async def _job_license_expiration(start_date: Optional[date] = None, end_date: Optional[date] = None,
                                   preset: Optional[str] = None, recipients: list = None,
                                   days_ahead: int = 90, include_expired: bool = True):
    """Email licenses expiring in a date window (defaults to already-expired through 90 days out)."""
    from datetime import timezone
    tz_utc = timezone.utc
    now    = datetime.now(tz_utc)
    today  = date.today()
    # For license expiration, start/end bound the expiration date itself
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_end=today + timedelta(days=days_ahead))
    if not include_expired and _start is None and not preset:
        _start = today
    cutoff_dt = datetime(_end.year, _end.month, _end.day, 23, 59, 59, tzinfo=tz_utc) if _end else now + timedelta(days=days_ahead)
    floor_dt  = datetime(_start.year, _start.month, _start.day, tzinfo=tz_utc) if _start else None
    today_str = str(today)
    print("[reports] License expiration report")

    lic_records = await ac_get_all(f"customObjects/records/{LICENSE_SCHEMA_ID}", "records", {})
    account_ids: set = set()
    candidates = []
    for r in lic_records:
        fields  = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        exp_str = (fields.get("expiration-date") or fields.get("license-expiration-date")
                   or fields.get("expires") or "")
        if not exp_str:
            continue
        try:
            exp_dt = (datetime.fromisoformat(str(exp_str).replace("Z", "+00:00")) if "T" in str(exp_str)
                      else datetime.strptime(str(exp_str)[:10], "%Y-%m-%d").replace(tzinfo=tz_utc))
        except Exception:
            continue
        if exp_dt > cutoff_dt:
            continue
        if floor_dt and exp_dt < floor_dt:
            continue
        is_expired = exp_dt < now
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id,
                           "exp_str": str(exp_str)[:10],
                           "days_until": (exp_dt - now).days,
                           "is_expired": is_expired})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    records = []
    for c in sorted(candidates, key=lambda x: x["days_until"]):
        status = "EXPIRED" if c["is_expired"] else f"In {c['days_until']}d"
        f   = c["fields"]
        aid = c["account_id"] or ""
        rec = {
            "Account":      acct_cache.get(aid, ""),
            "Expiration":   c["exp_str"],
            "Status":       status,
            "Days Until":   c["days_until"],
            "License #":    f.get("license-number", f.get("license_number", "")),
            "License Type": f.get("license-type", f.get("license_type", "")),
            "State":        f.get("state", f.get("license-state", "")),
            "Dealer ID":    _account_to_dealer.get(aid, ""),
            "Channel":      _account_to_platform.get(aid, ""),
            "BDR":          _account_to_bdr.get(aid, ""),
        }
        _enrich_record(rec, aid)
        records.append(rec)

    cols = [("Account","Account"), ("Expiration","Expiration"),
            ("Status","Status"), ("License #","License #"), ("State","State")]
    end_label   = str(_end)   if _end   else str(today + timedelta(days=90))
    start_label = str(_start) if _start else "past"
    range_label = f"{start_label} – {end_label}"
    html = _HTML_WRAPPER.format(
        title="License Expiration Report",
        subtitle=f"{len(records)} license{'s' if len(records) != 1 else ''} — expiration {range_label}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"License Expiration Report — {today_str} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"license_expiration_{today_str}.csv",
        recipients=recipients,
    )


# ── BDR Summary (weekly Monday) ──────────────────────────────────────────

async def _job_bdr_summary(start_date: Optional[date] = None, end_date: Optional[date] = None,
                           preset: Optional[str] = None, recipients: list = None,
                           platform: Optional[str] = None):
    """Email activations grouped by BDR for a date range (defaults to past 7 days)."""
    from datetime import timezone
    tz_utc = timezone.utc
    _today = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_start=_today - timedelta(days=7),
                                       default_end=_today - timedelta(days=1))
    if _start is None: _start = _today - timedelta(days=7)
    if _end   is None: _end   = _today - timedelta(days=1)
    week_start = _start.strftime("%Y-%m-%d")
    week_end   = _end.strftime("%Y-%m-%d")
    from_dt    = datetime(_start.year, _start.month, _start.day, tzinfo=tz_utc)
    to_dt      = datetime(_end.year,   _end.month,   _end.day,   23, 59, 59, tzinfo=tz_utc)
    print(f"[reports] BDR summary {week_start} → {week_end}")

    slp_records = await get_slp_cache()

    # Pass 1 – collect raw data; find accounts where BDR or platform is missing
    raw_slps = []
    acct_ids_needed: set = set()
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        bdr    = str(fields.get("assigned-bdr", "")).strip()
        plat   = str(fields.get("channel", "")).strip()
        if platform and plat != platform:
            continue
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if (not bdr or not plat) and acc_id:
            acct_ids_needed.add(acc_id)
        raw_slps.append({"fields": fields, "bdr": bdr, "plat": plat, "acc_id": acc_id})

    # Fetch account CFs for accounts where BDR/platform is missing
    acct_cf_cache: dict = {}
    if acct_ids_needed:
        async def _fetch_cf_bdr(aid: str) -> tuple:
            try:
                cf_r = await ac_get(f"accounts/{aid}/accountCustomFieldData")
                cfs  = {str(i.get("customFieldId", "")): (i.get("fieldValue") or "").strip()
                        for i in cf_r.get("accountCustomFieldData", [])}
                return aid, {"bdr": cfs.get("119", "")}
            except Exception:
                return aid, {"bdr": ""}
        acct_cf_cache = dict(await asyncio.gather(*[_fetch_cf_bdr(aid) for aid in acct_ids_needed]))

    # Pass 2 – process with fallbacks
    bdr_data: dict = defaultdict(lambda: {"activated_week": 0, "total_slps": 0,
                                           "channels": defaultdict(int), "accounts": set()})
    for rd in raw_slps:
        fields   = rd["fields"]
        acc_id   = rd["acc_id"]
        fallback = acct_cf_cache.get(acc_id, {}) if acc_id else {}
        bdr      = rd["bdr"] or fallback.get("bdr", "") or "Unassigned"
        bdr_data[bdr]["total_slps"] += 1
        if fields.get("slp-status-detail") == "Contractor Activated":
            act_str = str(fields.get("contractor-activated-date", "")).strip()
            if act_str:
                try:
                    act_dt = (datetime.fromisoformat(act_str.replace("Z", "+00:00")) if "T" in act_str
                              else datetime.strptime(act_str[:10], "%Y-%m-%d").replace(tzinfo=tz_utc))
                    if from_dt <= act_dt <= to_dt:
                        bdr_data[bdr]["activated_week"] += 1
                except Exception:
                    pass
        plat = rd["plat"]
        if plat:
            bdr_data[bdr]["channels"][plat] += 1
        if acc_id:
            bdr_data[bdr]["accounts"].add(acc_id)

    records = [
        {"BDR": bdr,
         "Activations (week)": d["activated_week"],
         "Total SLPs": d["total_slps"],
         "Accounts": len(d["accounts"]),
         "Channels": ", ".join(f"{k}:{v}" for k, v in sorted(d["channels"].items()))}
        for bdr, d in sorted(bdr_data.items())
    ]
    records.sort(key=lambda x: x["Activations (week)"], reverse=True)

    cols = [("BDR","BDR"), ("Activations (week)","Activations (week)"),
            ("Total SLPs","Total SLPs"), ("Accounts","Accounts"), ("Channels","Channels")]
    html = _HTML_WRAPPER.format(
        title=f"BDR Summary — Week of {week_start}",
        subtitle=f"{week_start} through {week_end}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"BDR Summary — Week of {week_start}",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"bdr_summary_{week_start}.csv",
        recipients=recipients,
    )


# ── Training Activity (weekly Monday) ────────────────────────────────────

async def _job_training_activity(start_date: Optional[date] = None, end_date: Optional[date] = None,
                                  preset: Optional[str] = None, recipients: list = None):
    """Email training sessions conducted in the date window, grouped by trainer."""
    from datetime import timezone
    tz_utc = timezone.utc
    today  = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_start=today - timedelta(days=7),
                                       default_end=today - timedelta(days=1))
    if _start is None: _start = today - timedelta(days=7)
    if _end   is None: _end   = today - timedelta(days=1)
    from_dt    = datetime(_start.year, _start.month, _start.day, tzinfo=tz_utc)
    to_dt      = datetime(_end.year,   _end.month,   _end.day,   23, 59, 59, tzinfo=tz_utc)
    date_label = str(_start) if _start == _end else f"{_start} to {_end}"
    print(f"[reports] Training activity {date_label}")

    training_records = await ac_get_all(f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {})
    account_ids: set = set()
    candidates = []
    for r in training_records:
        fields   = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        date_str = str(fields.get("date-of-training", "")).strip()
        if not date_str:
            continue
        try:
            td = (datetime.fromisoformat(date_str.replace("Z", "+00:00")) if "T" in date_str
                  else datetime.strptime(date_str[:10], "%Y-%m-%d").replace(tzinfo=tz_utc))
        except Exception:
            continue
        if not (from_dt <= td <= to_dt):
            continue
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    records = []
    for c in candidates:
        f   = c["fields"]
        aid = c["account_id"] or ""
        rec = {
            "Account":       acct_cache.get(aid, ""),
            "Dealer ID":     _account_to_dealer.get(aid, ""),
            "Channel":       _account_to_platform.get(aid, ""),
            "BDR":           _account_to_bdr.get(aid, ""),
            "Trained By":    f.get("trained-by", ""),
            "Training Type": f.get("training-type", ""),
            "Agenda":        f.get("training-agenda", ""),
            "Date":          str(f.get("date-of-training", ""))[:10],
            "Notes":         (f.get("training-notes", "") or "")[:120],
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: (x["Date"], x["Trained By"]), reverse=True)

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"), ("Trained By","Trained By"),
            ("Training Type","Training Type"), ("Agenda","Agenda"), ("Date","Date")]
    html = _HTML_WRAPPER.format(
        title=f"Training Activity — {date_label}",
        subtitle=f"{len(records)} session{'s' if len(records) != 1 else ''}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    csv_label = str(_start) if _start == _end else f"{_start}_{_end}"
    await _send_email(
        subject=f"Training Activity — {date_label} ({len(records)} sessions)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"training_activity_{csv_label}.csv",
        recipients=recipients,
    )


async def _job_training_summary(start_date: Optional[date] = None, end_date: Optional[date] = None,
                                preset: Optional[str] = None, recipients: list = None,
                                trainer: Optional[str] = None, training_type: Optional[str] = None):
    """Email the same data shown by the Training Summary report."""
    _start, _end = _resolve_date_range(start_date, end_date, preset)
    from_date = str(_start) if _start else None
    to_date = str(_end) if _end else None
    date_label = f"{from_date or 'All'} to {to_date or 'All'}"

    data = await training_summary_report(
        from_date=from_date,
        to_date=to_date,
        trainer=trainer,
        training_type=training_type,
        format="json",
    )
    records = data.get("records", []) if isinstance(data, dict) else []
    cols = [("Account", "account_name"), ("Dealer ID", "dealer_id"),
            ("Training Type", "training_type"), ("Agenda", "training_agenda"),
            ("Trained By", "trained_by"), ("Date", "date_of_training")]
    html = _HTML_WRAPPER.format(
        title=f"Training Summary — {date_label}",
        subtitle=f"{len(records)} training record{'s' if len(records) != 1 else ''}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    csv_label = f"{from_date or 'all'}_{to_date or 'all'}"
    await _send_email(
        subject=f"Training Summary — {date_label} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"training_summary_{csv_label}.csv",
        recipients=recipients,
    )


# ── Stale / Untrained Dealers (monthly) ──────────────────────────────────

async def _job_stale_untrained(start_date: Optional[date] = None, end_date: Optional[date] = None,
                                preset: Optional[str] = None, recipients: list = None,
                                stale_days: int = 90, platform: Optional[str] = None,
                                bdr: Optional[str] = None):
    """Email activated dealers with no training or last training >90 days ago.
    start_date/end_date optionally filter by contractor-activated-date."""
    today = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset)
    stale_cutoff = today - timedelta(days=stale_days)
    print("[reports] Stale/untrained dealers")

    slp_records = await get_slp_cache()
    tr_records  = await ac_get_all(f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {})

    training_by_acct: dict = defaultdict(list)
    for r in tr_records:
        fields   = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        date_str = str(fields.get("date-of-training", "")).strip()
        if not date_str:
            continue
        for aid in r.get("relationships", {}).get("account", []):
            training_by_acct[str(aid)].append(date_str[:10])

    account_ids: set = set()
    candidates = []
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        if fields.get("slp-status-detail") != "Contractor Activated":
            continue
        slp_plat = str(fields.get("channel", "")).strip()
        if platform and slp_plat != platform:
            continue
        act_str  = str(fields.get("contractor-activated-date", "")).strip()
        act_date = act_str[:10] if act_str else ""
        if _start and act_date and act_date < str(_start):
            continue
        if _end   and act_date and act_date > str(_end):
            continue
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        eff_bdr = str(fields.get("assigned-bdr", "")).strip() or _account_to_bdr.get(acc_id or "", "")
        if bdr and eff_bdr != bdr:
            continue
        if acc_id:
            account_ids.add(acc_id)
        trainings  = training_by_acct.get(acc_id or "", [])
        last_train = max(trainings) if trainings else None
        is_stale   = (not last_train) or (last_train < str(stale_cutoff))
        if not is_stale:
            continue
        days_stale = (today - date.fromisoformat(last_train)).days if last_train else None
        candidates.append({"fields": fields, "account_id": acc_id, "act_date": act_date,
                           "training_count": len(trainings), "last_training": last_train or "",
                           "days_stale": days_stale})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    records = []
    for c in sorted(candidates, key=lambda x: x["days_stale"] or 99999, reverse=True):
        f   = c["fields"]
        aid = c["account_id"] or ""
        rec = {
            "Account":         acct_cache.get(aid, ""),
            "Dealer ID":       f.get("dealer-id")    or _account_to_dealer.get(aid, ""),
            "Channel":         f.get("channel", ""),
            "BDR":             f.get("assigned-bdr") or _account_to_bdr.get(aid, ""),
            "Activation Date": c["act_date"],
            "# Trainings":     c["training_count"],
            "Last Training":   c["last_training"] or "Never",
            "Days Stale":      c["days_stale"] if c["days_stale"] is not None else "Never trained",
            "SLP Status":      f.get("slp-status-detail", ""),
            "Oracle Producer IDs": f.get("oracle-producer-ids", ""),
        }
        _enrich_record(rec, aid)
        records.append(rec)

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"), ("Channel","Channel"),
            ("BDR","BDR"), ("Activation Date","Activation Date"),
            ("# Trainings","# Trainings"), ("Last Training","Last Training"),
            ("Days Stale","Days Stale")]
    html = _HTML_WRAPPER.format(
        title="Stale / Untrained Dealers",
        subtitle=f"{len(records)} activated dealer{'s' if len(records) != 1 else ''} with no training or last training >{stale_days} days ago",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Stale/Untrained Dealers — {today} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"stale_untrained_{today}.csv",
        recipients=recipients,
    )


# ── Account Status Summary (weekly Monday) ───────────────────────────────

async def _job_account_status(start_date: Optional[date] = None, end_date: Optional[date] = None,
                               preset: Optional[str] = None, recipients: list = None):
    """Email all accounts with their status and sales region (snapshot, date params unused)."""
    today = date.today()
    print("[reports] Account status summary")

    all_accounts = await ac_get_all("accounts", "accounts", {})
    cf_map       = await _fetch_acct_cf_map({"19", "23"})  # account_status, sales_region

    records = []
    for a in all_accounts:
        aid  = str(a.get("id", ""))
        cfs  = cf_map.get(aid, {})
        rec  = {
            "Account":      a.get("name", ""),
            "Dealer ID":    _account_to_dealer.get(aid, ""),
            "Channel":      _account_to_platform.get(aid, ""),
            "BDR":          _account_to_bdr.get(aid, ""),
            "Status":       cfs.get("19", ""),
            "Sales Region": cfs.get("23", ""),
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: (x["Status"], x["Sales Region"], x["Account"]))

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"), ("Channel","Channel"),
            ("BDR","BDR"), ("Status","Status"), ("Sales Region","Sales Region")]
    html = _HTML_WRAPPER.format(
        title="Account Status Summary",
        subtitle=f"{len(records)} account{'s' if len(records) != 1 else ''} as of {today}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Account Status Summary — {today} ({len(records)} accounts)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"account_status_{today}.csv",
        recipients=recipients,
    )


# ── Channel Breakdown (weekly Monday) ────────────────────────────────────

async def _job_platform_breakdown(start_date: Optional[date] = None, end_date: Optional[date] = None,
                                   preset: Optional[str] = None, recipients: list = None):
    """Email new activations and total SLP counts grouped by channel."""
    from datetime import timezone
    tz_utc = timezone.utc
    today  = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_start=today - timedelta(days=7),
                                       default_end=today - timedelta(days=1))
    if _start is None: _start = today - timedelta(days=7)
    if _end   is None: _end   = today - timedelta(days=1)
    from_dt    = datetime(_start.year, _start.month, _start.day, tzinfo=tz_utc)
    to_dt      = datetime(_end.year,   _end.month,   _end.day,   23, 59, 59, tzinfo=tz_utc)
    date_label = str(_start) if _start == _end else f"{_start} to {_end}"
    print(f"[reports] Platform breakdown {date_label}")

    slp_records = await get_slp_cache()
    plat_data: dict = defaultdict(lambda: {"new_activations": 0, "active_slps": 0,
                                           "total_slps": 0, "bdrs": defaultdict(int)})
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        plat   = str(fields.get("channel", "")).strip() or "Unknown"
        bdr    = (str(fields.get("assigned-bdr", "")).strip()
                  or _account_to_bdr.get(acc_id or "", "") or "Unassigned")
        plat_data[plat]["total_slps"] += 1
        if fields.get("slp-status-detail") == "Contractor Activated":
            plat_data[plat]["active_slps"] += 1
            act_str = str(fields.get("contractor-activated-date", "")).strip()
            if act_str:
                try:
                    act_dt = (datetime.fromisoformat(act_str.replace("Z", "+00:00")) if "T" in act_str
                              else datetime.strptime(act_str[:10], "%Y-%m-%d").replace(tzinfo=tz_utc))
                    if from_dt <= act_dt <= to_dt:
                        plat_data[plat]["new_activations"] += 1
                        plat_data[plat]["bdrs"][bdr] += 1
                except Exception:
                    pass

    records = []
    for plat, d in sorted(plat_data.items()):
        top_bdr = max(d["bdrs"], key=d["bdrs"].get) if d["bdrs"] else ""
        records.append({
            "Channel":         plat,
            "New Activations": d["new_activations"],
            "Active SLPs":     d["active_slps"],
            "Total SLPs":      d["total_slps"],
            "Top BDR":         top_bdr,
        })
    records.sort(key=lambda x: x["New Activations"], reverse=True)

    total_new = sum(r["New Activations"] for r in records)
    cols = [("Channel","Channel"), ("New Activations","New Activations"),
            ("Active SLPs","Active SLPs"), ("Total SLPs","Total SLPs"), ("Top BDR","Top BDR")]
    html = _HTML_WRAPPER.format(
        title=f"Channel Breakdown — {date_label}",
        subtitle=f"{total_new} new activation{'s' if total_new != 1 else ''} across {len(records)} channel{'s' if len(records) != 1 else ''}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    csv_label = str(_start) if _start == _end else f"{_start}_{_end}"
    await _send_email(
        subject=f"Channel Breakdown — {date_label} ({total_new} new activations)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"channel_breakdown_{csv_label}.csv",
        recipients=recipients,
    )


# ── Partner Activation (monthly) ─────────────────────────────────────────

async def _job_partner_activation(start_date: Optional[date] = None, end_date: Optional[date] = None,
                                   preset: Optional[str] = None, recipients: list = None):
    """Email accounts where partner_activation (CF 26) date falls in the window."""
    today = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_start=today.replace(day=1),
                                       default_end=today)
    if _start is None: _start = today.replace(day=1)
    if _end   is None: _end   = today
    date_label = str(_start) if _start == _end else f"{_start} to {_end}"
    print(f"[reports] Partner activation {date_label}")

    cf_map       = await _fetch_acct_cf_map({"26"})
    all_accounts = await ac_get_all("accounts", "accounts", {})
    acct_by_id   = {str(a.get("id", "")): a for a in all_accounts}

    records = []
    for aid, cfs in cf_map.items():
        pa_val = cfs.get("26", "")
        if not pa_val:
            continue
        pa_str = str(pa_val)[:10]
        try:
            pa_date = date.fromisoformat(pa_str)
            if pa_date < _start or pa_date > _end:
                continue
        except Exception:
            if start_date or end_date or preset:
                continue           # skip unparseable dates when a filter is active
            pa_str = str(pa_val)  # show raw value when no filter
        a   = acct_by_id.get(aid, {})
        rec = {
            "Account":            a.get("name", ""),
            "Dealer ID":          _account_to_dealer.get(aid, ""),
            "Channel":            _account_to_platform.get(aid, ""),
            "BDR":                _account_to_bdr.get(aid, ""),
            "Partner Activation": pa_str,
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: x["Partner Activation"], reverse=True)

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"), ("Channel","Channel"),
            ("BDR","BDR"), ("Partner Activation","Partner Activation")]
    html = _HTML_WRAPPER.format(
        title=f"Activations — {date_label}",
        subtitle=f"{len(records)} activation{'s' if len(records) != 1 else ''}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    csv_label = str(_start) if _start == _end else f"{_start}_{_end}"
    await _send_email(
        subject=f"Activations — {date_label} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"partner_activation_{csv_label}.csv",
        recipients=recipients,
    )


# ── Oracle Producer ID Missing (weekly Monday) ────────────────────────────

async def _job_oracle_missing(start_date: Optional[date] = None, end_date: Optional[date] = None,
                               preset: Optional[str] = None, recipients: list = None,
                               platform: Optional[str] = None, bdr: Optional[str] = None):
    """Email activated SLPs whose account has no Oracle Producer ID (CF 118).
    start_date/end_date optionally filter by contractor-activated-date."""
    today = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset)
    print("[reports] Oracle Producer ID missing")

    slp_records = await get_slp_cache()
    cf_map      = await _fetch_acct_cf_map({"118"})  # oracle_producer_id

    account_ids: set = set()
    candidates = []
    for r in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        if fields.get("slp-status-detail") != "Contractor Activated":
            continue
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        if cf_map.get(acc_id or "", {}).get("118"):
            continue   # oracle_producer_id already set
        slp_plat = str(fields.get("channel", "")).strip()
        if platform and slp_plat != platform:
            continue
        eff_bdr = str(fields.get("assigned-bdr", "")).strip() or _account_to_bdr.get(acc_id or "", "")
        if bdr and eff_bdr != bdr:
            continue
        act_str  = str(fields.get("contractor-activated-date", "")).strip()
        act_date = act_str[:10] if act_str else ""
        if _start and act_date and act_date < str(_start):
            continue
        if _end   and act_date and act_date > str(_end):
            continue
        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id, "act_date": act_date})

    acct_cache: dict = {}
    for aid in account_ids:
        try:
            d = await ac_get(f"accounts/{aid}")
            acct_cache[aid] = d.get("account", {}).get("name", "")
        except Exception:
            acct_cache[aid] = ""

    records = []
    for c in candidates:
        f   = c["fields"]
        aid = c["account_id"] or ""
        rec = {
            "Account":                  acct_cache.get(aid, ""),
            "Dealer ID":                f.get("dealer-id")    or _account_to_dealer.get(aid, ""),
            "Channel":                  f.get("channel", ""),
            "BDR":                      f.get("assigned-bdr") or _account_to_bdr.get(aid, ""),
            "Activation Date":          c["act_date"],
            "SLP Status":               f.get("slp-status-detail", ""),
            "Doing Business In States": f.get("doing-business-in-states", ""),
            "EIN":                      f.get("ein", ""),
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: (x["Channel"], x["BDR"], x["Account"]))

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"), ("Channel","Channel"),
            ("BDR","BDR"), ("Activation Date","Activation Date")]
    html = _HTML_WRAPPER.format(
        title="Oracle Producer ID Missing",
        subtitle=f"{len(records)} activated dealer{'s' if len(records) != 1 else ''} missing Oracle Producer ID",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Oracle Producer ID Missing — {today} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"oracle_missing_{today}.csv",
        recipients=recipients,
    )


# ── Account Activity (ad hoc / on-demand) ────────────────────────────────

async def _job_account_activity(start_date=None, end_date=None, preset=None, recipients=None,
                                activity_type: Optional[str] = None, performed_by: Optional[str] = None):
    """Email the same Account Activity summary shown by the prebuilt report."""
    today = str(date.today())
    _start, _end = _resolve_date_range(start_date, end_date, preset)
    date_label = f"{_start} – {_end}" if (_start or _end) else "All Time"
    data = await account_activity_report(
        from_date=str(_start) if _start else None,
        to_date=str(_end) if _end else None,
        activity_type=activity_type,
        performed_by=performed_by,
        format="json",
    )

    rows = []
    for person in data.get("by_person", []):
        for act_type, count in (person.get("breakdown") or {}).items():
            rows.append({
                "performed_by": person.get("performed_by", ""),
                "activity_type": act_type,
                "count": count,
                "total_for_person": person.get("count", 0),
            })
    if not rows:
        rows = data.get("by_type", [])

    cols = [("Performed By", "performed_by"), ("Activity Type", "activity_type"),
            ("Count", "count"), ("Total For Person", "total_for_person")]
    html = _HTML_WRAPPER.format(
        title=f"Account Activity — {date_label}",
        subtitle=f"{data.get('total', 0)} activit{'y' if data.get('total', 0) == 1 else 'ies'}",
        table=_html_table(rows, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Account Activity — {today} ({data.get('total', 0)} activities)",
        html=html,
        csv_data=_csv_bytes(rows),
        csv_name=f"account_activity_{today}.csv",
        recipients=recipients,
    )


# ── Team Activity (ad hoc / on-demand) ───────────────────────────────────

async def _job_team_activity(start_date=None, end_date=None, preset=None, recipients=None):
    """Email the same Team Performance summary shown by the prebuilt report."""
    today = str(date.today())
    _start, _end = _resolve_date_range(start_date, end_date, preset)
    date_label = f"{_start} – {_end}" if (_start or _end) else "All Time"
    data = await team_activity_report(
        from_date=str(_start) if _start else None,
        to_date=str(_end) if _end else None,
        format="json",
    )
    records = data.get("records", []) if isinstance(data, dict) else []
    cols = [("Name", "user_name"), ("Notes", "notes_written"),
            ("Activities", "activities_logged"), ("Total", "total_actions"),
            ("Accounts", "accounts_touched"), ("Last AC Note/Activity", "latest_activity_date"),
            ("Last Microf-Search Login", "last_login")]
    html = _HTML_WRAPPER.format(
        title=f"Team Performance — {date_label}",
        subtitle=f"{len(records)} team member{'s' if len(records) != 1 else ''}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Team Performance — {today} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"team_activity_{today}.csv",
        recipients=recipients,
    )


# ── Last App Date Report ────────────────────────────────────────────────────

@app.get("/api/report/last-app-date")
async def report_last_app_date(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    preset:    Optional[str]  = Query(None),
    format:    str            = Query("json"),
    _: None    = Depends(require_auth),
):
    """Accounts where last-app-date (from SLP records, via dealer index) falls within the given window."""
    today    = date.today()
    _start, _end = _resolve_date_range(from_date, to_date, preset,
                                       default_start=today - timedelta(days=548),
                                       default_end=today)
    if _start is None: _start = today - timedelta(days=548)
    if _end   is None: _end   = today
    date_label = f"{_start} to {_end}"

    records = []
    for aid, date_str in _account_to_last_app.items():
        if not date_str:
            continue
        try:
            d = date.fromisoformat(date_str[:10])
            if d < _start or d > _end:
                continue
        except Exception:
            continue
        records.append({
            "Account":       _account_to_name.get(aid, ""),
            "Dealer ID":     _account_to_dealer.get(aid, ""),
            "Region":        _account_to_region.get(aid, ""),
            "Account Type":  _account_to_type.get(aid, ""),
            "Last App Date": date_str[:10],
        })
    records.sort(key=lambda x: x["Last App Date"], reverse=True)

    if format == "csv":
        out = io.StringIO()
        if records:
            w = csv.DictWriter(out, fieldnames=records[0].keys())
            w.writeheader(); w.writerows(records)
        fn = f"last_app_date_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})
    return {"count": len(records), "date_range": date_label, "records": records}


async def _job_last_app_date(start_date: Optional[date] = None, end_date: Optional[date] = None,
                              preset: Optional[str] = None, recipients: list = None):
    today    = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_start=today - timedelta(days=548),
                                       default_end=today)
    if _start is None: _start = today - timedelta(days=548)
    if _end   is None: _end   = today
    date_label = f"{_start} to {_end}"
    print(f"[reports] Last App Date {date_label}")

    records = []
    for aid, date_str in _account_to_last_app.items():
        if not date_str:
            continue
        try:
            d = date.fromisoformat(date_str[:10])
            if d < _start or d > _end:
                continue
        except Exception:
            continue
        rec = {
            "Account":       _account_to_name.get(aid, ""),
            "Dealer ID":     _account_to_dealer.get(aid, ""),
            "Channel":       _account_to_platform.get(aid, ""),
            "BDR":           _account_to_bdr.get(aid, ""),
            "Region":        _account_to_region.get(aid, ""),
            "Account Type":  _account_to_type.get(aid, ""),
            "Last App Date": date_str[:10],
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: x["Last App Date"], reverse=True)

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"), ("Region","Region"),
            ("Account Type","Account Type"), ("Last App Date","Last App Date")]
    html = _HTML_WRAPPER.format(
        title=f"Last App Date — {date_label}",
        subtitle=f"{len(records)} account{'s' if len(records) != 1 else ''}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Last App Date — {date_label} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"last_app_date_{_start}_{_end}.csv",
        recipients=recipients,
    )


# ── Last RPA Date Report ─────────────────────────────────────────────────────

@app.get("/api/report/last-rpa-date")
async def report_last_rpa_date(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    preset:    Optional[str]  = Query(None),
    format:    str            = Query("json"),
    _: None    = Depends(require_auth),
):
    """Accounts where last-rpa-date (from SLP records + CF38, via dealer index) falls within the given window."""
    today    = date.today()
    _start, _end = _resolve_date_range(from_date, to_date, preset,
                                       default_start=today - timedelta(days=548),
                                       default_end=today)
    if _start is None: _start = today - timedelta(days=548)
    if _end   is None: _end   = today
    date_label = f"{_start} to {_end}"

    records = []
    for aid, date_str in _account_to_last_rpa.items():
        if not date_str:
            continue
        try:
            d = date.fromisoformat(date_str[:10])
            if d < _start or d > _end:
                continue
        except Exception:
            continue
        records.append({
            "Account":       _account_to_name.get(aid, ""),
            "Dealer ID":     _account_to_dealer.get(aid, ""),
            "Region":        _account_to_region.get(aid, ""),
            "Account Type":  _account_to_type.get(aid, ""),
            "Last RPA Date": date_str[:10],
        })
    records.sort(key=lambda x: x["Last RPA Date"], reverse=True)

    if format == "csv":
        out = io.StringIO()
        if records:
            w = csv.DictWriter(out, fieldnames=records[0].keys())
            w.writeheader(); w.writerows(records)
        fn = f"last_rpa_date_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fn}"})
    return {"count": len(records), "date_range": date_label, "records": records}


async def _job_last_rpa_date(start_date: Optional[date] = None, end_date: Optional[date] = None,
                              preset: Optional[str] = None, recipients: list = None):
    today    = date.today()
    _start, _end = _resolve_date_range(start_date, end_date, preset,
                                       default_start=today - timedelta(days=548),
                                       default_end=today)
    if _start is None: _start = today - timedelta(days=548)
    if _end   is None: _end   = today
    date_label = f"{_start} to {_end}"
    print(f"[reports] Last RPA Date {date_label}")

    records = []
    for aid, date_str in _account_to_last_rpa.items():
        if not date_str:
            continue
        try:
            d = date.fromisoformat(date_str[:10])
            if d < _start or d > _end:
                continue
        except Exception:
            continue
        rec = {
            "Account":       _account_to_name.get(aid, ""),
            "Dealer ID":     _account_to_dealer.get(aid, ""),
            "Channel":       _account_to_platform.get(aid, ""),
            "BDR":           _account_to_bdr.get(aid, ""),
            "Region":        _account_to_region.get(aid, ""),
            "Account Type":  _account_to_type.get(aid, ""),
            "Last RPA Date": date_str[:10],
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: x["Last RPA Date"], reverse=True)

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"), ("Region","Region"),
            ("Account Type","Account Type"), ("Last RPA Date","Last RPA Date")]
    html = _HTML_WRAPPER.format(
        title=f"Last RPA Date — {date_label}",
        subtitle=f"{len(records)} account{'s' if len(records) != 1 else ''}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Last RPA Date — {date_label} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"last_rpa_date_{_start}_{_end}.csv",
        recipients=recipients,
    )


# ── Not Activated ────────────────────────────────────────────────────────

async def _job_not_activated(start_date=None, end_date=None,
                              preset: Optional[str] = None, recipients: list = None,
                              platform: Optional[str] = None, bdr: Optional[str] = None,
                              status: Optional[str] = None, state: Optional[str] = None,
                              exclude_platforms: Optional[str] = None):
    """Email all SLP records that are NOT 'Contractor Activated'."""
    print("[reports] Not-activated report")
    _start, _end = _resolve_date_range(start_date, end_date, preset)
    exclude_set = {p.strip() for p in exclude_platforms.split(",")} if exclude_platforms else set()
    slp_records = await get_slp_cache()

    account_ids: set = set()
    candidates = []
    for r in slp_records:
        fields     = {fo["id"]: fo.get("value", "") for fo in r.get("fields", [])}
        status_val = str(fields.get("slp-status-detail", "")).strip()
        if status_val in ("Contractor Activated", "Inactive", "Deactivated", "Deactivated for Dormancy", "Declined by Onboarding", "Not Active"):
            continue
        if status and status_val != status:
            continue
        rel    = r.get("relationships", {}).get("account", [])
        acc_id = str(rel[0]) if rel else None
        plat = str(fields.get("channel", "")).strip()
        plat_norm = _normalize_platform(plat)
        if platform and plat_norm != _normalize_platform(platform):
            continue
        if plat_norm in exclude_set or plat in exclude_set:
            continue
        if state:
            states_val = str(fields.get("doing-business-in-states", "") or "").upper()
            if state.upper() not in [s.strip() for s in states_val.split(",")]:
                continue
        eff_bdr = str(fields.get("assigned-bdr", "")).strip() or _account_to_bdr.get(acc_id or "", "")
        if bdr == "__unassigned__":
            if eff_bdr:
                continue
        elif bdr and eff_bdr != bdr:
            continue
        if _start or _end:
            enroll_date = str(fields.get("enrollment-request-date") or "")[:10]
            if not enroll_date:
                continue
            if _start and enroll_date < str(_start):
                continue
            if _end and enroll_date > str(_end):
                continue
        if acc_id:
            account_ids.add(acc_id)
        candidates.append({"fields": fields, "account_id": acc_id, "eff_bdr": eff_bdr})

    async def _fetch(aid):
        try:
            d = await ac_get(f"accounts/{aid}")
            return aid, d.get("account", {}).get("name", "")
        except Exception:
            return aid, ""

    name_map = dict(await asyncio.gather(*[_fetch(a) for a in account_ids]))

    records = []
    for c in candidates:
        f   = c["fields"]
        aid = c["account_id"] or ""
        rec = {
            "Account":                  name_map.get(aid, ""),
            "Dealer ID":                f.get("dealer-id", ""),
            "Channel":                  f.get("channel", ""),
            "Status":                   f.get("slp-status-detail", "") or "Not Started",
            "BDR":                      c["eff_bdr"],
            "Doing Business In States": f.get("doing-business-in-states", ""),
            "Enrollment Request Date":  str(f.get("enrollment-request-date", "") or "")[:10],
            "Oracle Producer IDs":      f.get("oracle-producer-ids", ""),
        }
        _enrich_record(rec, aid)
        records.append(rec)
    records.sort(key=lambda x: (x.get("Status", ""), x.get("Account", "")))

    cols = [("Account","Account"), ("Dealer ID","Dealer ID"),
            ("Channel","Channel"), ("Status","Status"), ("BDR","BDR")]
    html = _HTML_WRAPPER.format(
        title="Not Activated SLPs",
        subtitle=f"{len(records)} record{'s' if len(records) != 1 else ''} not yet Contractor Activated",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Not Activated SLPs ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"not_activated_{datetime.now().strftime('%Y%m%d')}.csv",
        recipients=recipients,
    )


# ── Verdata email jobs ────────────────────────────────────────────────────

async def _job_verdata_active(start_date=None, end_date=None,
                              preset=None, recipients: list = None):
    """Email the Verdata Active report (Account Status = Active)."""
    today = str(date.today())
    print("[reports] Verdata Active")
    records = []
    slp_records = await get_slp_cache()
    act_dates: dict = {}
    for slp in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in slp.get("fields", [])}
        did = fields.get("dealer-id", "")
        ad  = str(fields.get("contractor-activated-date", "") or "")[:10]
        if did and len(ad) == 10 and (did not in act_dates or ad < act_dates[did]):
            act_dates[did] = ad
    for account_id, status in _account_to_status.items():
        if status.strip().lower() != "active":
            continue
        dealer_id = _account_to_dealer.get(account_id, "")
        records.append({
            "Dealer ID":           dealer_id,
            "Account Name":        _account_to_name.get(account_id, ""),
            "DBA Name":            _account_to_dba.get(account_id, ""),
            "RTO Activation Date": act_dates.get(dealer_id, ""),
            "Account Status":      status,
            "Vendor Tax-ID":       _account_to_tax_id.get(account_id, ""),
            "Website":             _account_to_website.get(account_id, ""),
            "Physical Address":    _account_to_address.get(account_id, ""),
            "Physical City":       _account_to_city.get(account_id, ""),
            "Physical State":      _account_to_state_prov.get(account_id, ""),
            "Physical Zip":        _account_to_zip.get(account_id, ""),
        })
    records.sort(key=lambda r: (r["Account Name"] or "").lower())
    cols = [("Dealer ID","Dealer ID"), ("Account Name","Account Name"),
            ("DBA Name","DBA Name"), ("RTO Activation Date","RTO Activation Date"),
            ("Account Status","Account Status")]
    html = _HTML_WRAPPER.format(
        title="Verdata Active Report",
        subtitle=f"{len(records)} active account{'s' if len(records) != 1 else ''} as of {today}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Verdata Active Report — {today} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"verdata_active_{today}.csv",
        recipients=recipients,
    )


async def _job_verdata_inactive(start_date=None, end_date=None,
                                preset=None, recipients: list = None):
    """Email the Verdata Inactive report (Account Status = Deactivated)."""
    today = str(date.today())
    print("[reports] Verdata Inactive")
    records = []
    slp_records = await get_slp_cache()
    act_dates: dict = {}
    for slp in slp_records:
        fields = {fo["id"]: fo.get("value", "") for fo in slp.get("fields", [])}
        did = fields.get("dealer-id", "")
        ad  = str(fields.get("contractor-activated-date", "") or "")[:10]
        if did and len(ad) == 10 and (did not in act_dates or ad < act_dates[did]):
            act_dates[did] = ad
    for account_id, status in _account_to_status.items():
        if status.strip().lower() != "deactivated":
            continue
        dealer_id = _account_to_dealer.get(account_id, "")
        records.append({
            "Dealer ID":           dealer_id,
            "Account Name":        _account_to_name.get(account_id, ""),
            "DBA Name":            _account_to_dba.get(account_id, ""),
            "RTO Activation Date": act_dates.get(dealer_id, ""),
            "Account Status":      status,
            "Vendor Tax-ID":       _account_to_tax_id.get(account_id, ""),
            "Website":             _account_to_website.get(account_id, ""),
            "Physical Address":    _account_to_address.get(account_id, ""),
            "Physical City":       _account_to_city.get(account_id, ""),
            "Physical State":      _account_to_state_prov.get(account_id, ""),
            "Physical Zip":        _account_to_zip.get(account_id, ""),
        })
    records.sort(key=lambda r: (r["Account Name"] or "").lower())
    cols = [("Dealer ID","Dealer ID"), ("Account Name","Account Name"),
            ("DBA Name","DBA Name"), ("RTO Activation Date","RTO Activation Date"),
            ("Account Status","Account Status")]
    html = _HTML_WRAPPER.format(
        title="Verdata Inactive Report",
        subtitle=f"{len(records)} deactivated account{'s' if len(records) != 1 else ''} as of {today}",
        table=_html_table(records, cols),
        timestamp=datetime.now().strftime("%b %d %Y %H:%M"),
    )
    await _send_email(
        subject=f"Verdata Inactive Report — {today} ({len(records)} records)",
        html=html,
        csv_data=_csv_bytes(records),
        csv_name=f"verdata_inactive_{today}.csv",
        recipients=recipients,
    )


# ── Manual / GitHub Actions trigger ──────────────────────────────────────

_REPORT_JOBS = {
    "activations":          _job_activations,
    "enrollment":           _job_enrollment,
    "license-expiration":   _job_license_expiration,
    "bdr-summary":          _job_bdr_summary,
    "training-summary":     _job_training_summary,
    "training-activity":    _job_training_activity,
    "stale-untrained":      _job_stale_untrained,
    "account-status":       _job_account_status,
    "platform-breakdown":   _job_platform_breakdown,
    "partner-activation":   _job_partner_activation,
    "oracle-missing":       _job_oracle_missing,
    "account-activity":     _job_account_activity,
    "team-activity":        _job_team_activity,
    "last-app-date":        _job_last_app_date,
    "last-rpa-date":        _job_last_rpa_date,
    "not-activated":        _job_not_activated,
    "verdata-active":       _job_verdata_active,
    "verdata-inactive":     _job_verdata_inactive,
}

@app.get("/api/send-report/{report_type}")
async def trigger_report(
    report_type: str,
    start_date:  Optional[date] = Query(None, description="Start of date range (YYYY-MM-DD)"),
    end_date:    Optional[date] = Query(None, description="End of date range (YYYY-MM-DD)"),
    preset:      Optional[str]  = Query(None,
        description="Date preset: yesterday | last_week | last_7_days | last_30_days | "
                    "last_90_days | this_week | this_month | last_month"),
    to:          Optional[str]  = Query(None,
        description="Override recipients — comma-separated email addresses"),
    slp_status:  Optional[str]  = Query(None,
        description="(Enrollment report only) Filter to a specific SLP status. "
                    f"Valid values: {', '.join(_SLP_STATUSES)}"),
    platform:    Optional[str]  = Query(None),
    bdr:         Optional[str]  = Query(None),
    state:       Optional[str]  = Query(None),
    exclude_platforms: Optional[str] = Query(None),
    status:      Optional[str]  = Query(None),
    trainer:     Optional[str]  = Query(None),
    training_type: Optional[str] = Query(None),
    activity_type: Optional[str] = Query(None),
    performed_by: Optional[str] = Query(None),
    stale_days:  Optional[int]  = Query(None),
    days_ahead:  Optional[int]  = Query(None),
    include_expired: Optional[bool] = Query(None),
    _: None = Depends(require_auth),
):
    """Manually trigger a report email. Also called by GitHub Actions on schedule.
    Use preset OR explicit start_date/end_date to override the default date window.
    Pass to= to override the configured REPORT_RECIPIENTS list."""
    job = _REPORT_JOBS.get(report_type)
    if not job:
        raise HTTPException(
            status_code=404,
            detail=f"Unknown report '{report_type}'. Valid: {list(_REPORT_JOBS)}"
        )
    override_recipients = [r.strip() for r in to.split(",") if r.strip()] if to else None
    final_recipients    = override_recipients or _RECIPIENTS
    if not _SMTP_USER or not _SMTP_PASS:
        raise HTTPException(status_code=503, detail="Email not configured — set SMTP_USER and SMTP_PASS in Render environment variables")
    if not final_recipients:
        raise HTTPException(status_code=400, detail="No recipients — enter an email address in the To field")
    try:
        kwargs: dict = dict(start_date=start_date, end_date=end_date, preset=preset,
                            recipients=override_recipients)
        if report_type == "enrollment" and slp_status:
            kwargs["slp_status"] = slp_status
        if report_type in {"activations", "not-activated", "bdr-summary", "stale-untrained", "oracle-missing"} and platform:
            kwargs["platform"] = platform
        if report_type in {"activations", "not-activated", "stale-untrained", "oracle-missing"} and bdr:
            kwargs["bdr"] = bdr
        if report_type in {"activations", "not-activated"} and state:
            kwargs["state"] = state
        if report_type in {"activations", "not-activated"} and exclude_platforms:
            kwargs["exclude_platforms"] = exclude_platforms
        if report_type == "not-activated" and status:
            kwargs["status"] = status
        if report_type == "training-summary":
            if trainer:
                kwargs["trainer"] = trainer
            if training_type:
                kwargs["training_type"] = training_type
        if report_type == "account-activity":
            if activity_type:
                kwargs["activity_type"] = activity_type
            if performed_by:
                kwargs["performed_by"] = performed_by
        if report_type == "stale-untrained" and stale_days is not None:
            kwargs["stale_days"] = stale_days
        if report_type == "license-expiration":
            if days_ahead is not None:
                kwargs["days_ahead"] = days_ahead
            if include_expired is not None:
                kwargs["include_expired"] = include_expired
        await job(**kwargs)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Email failed: {exc}")
    return {"status": "sent", "report": report_type,
            "start_date": str(start_date) if start_date else None,
            "end_date":   str(end_date)   if end_date   else None,
            "preset":     preset,
            "recipients": final_recipients}


# ── SLP field sync ────────────────────────────────────────────────────────────
# Copies missing field values from the linked account's custom fields into
# the SLP custom object record.  Fields synced:
#   dealer-id    ← _account_to_dealer index (account customfield 18)
#   assigned-bdr ← account customfield 119 (Assigned BDR)
# NOTE: platform (CF29 Dealer Program) deleted from PROD — channel lives on SLP directly

_SLP_SYNC_FIELDS = [
    # (slp_field_id, account_cf_id_str)  — None means use the dealer-id index
    ("dealer-id",    None),
    ("assigned-bdr", "119"),
]

# Holds the last sync job result/status in memory
_slp_sync_status: dict = {"status": "idle"}

def _rel_ids(values: list) -> list:
    """Normalize AC relationship arrays that may contain IDs or {id: ...} dicts."""
    out = []
    for v in values or []:
        raw = v.get("id", v) if isinstance(v, dict) else v
        try:
            out.append(int(raw))
        except (TypeError, ValueError):
            if raw:
                out.append(raw)
    return out

async def _merge_update_custom_object_record(schema_id: str, record_id: str,
                                             updates: list[dict]) -> dict:
    """Fetch a custom object record, merge fields, and post the full record back.

    Endpoint quirk: AC's per-record URL (`/customObjects/records/{schema}/{id}`)
    only allows GET and DELETE. Updates must POST to the *collection* endpoint
    (`/customObjects/records/{schema}`) with the existing record id in the body.
    """
    current = await ac_get(f"customObjects/records/{schema_id}/{record_id}")
    record = current.get("record", {})
    fields = {f["id"]: f.get("value", "") for f in record.get("fields", [])}
    for item in updates:
        fields[item["id"]] = item.get("value", "")
    relationships = {
        key: _rel_ids(vals)
        for key, vals in (record.get("relationships") or {}).items()
    }
    payload = {
        "record": {
            "id":     record_id,
            "fields": [{"id": k, "value": v} for k, v in fields.items()],
            "relationships": relationships,
        }
    }
    return await ac_post(f"customObjects/records/{schema_id}", payload)

async def _run_slp_sync(dry_run: bool) -> None:
    """Background worker — pages through SLP records and fills blank fields from account CFs."""
    global _slp_sync_status
    _slp_sync_status = {"status": "running", "dry_run": dry_run,
                        "scanned": 0, "updated": 0, "skipped": 0, "errors": 0,
                        "started": datetime.utcnow().isoformat()}

    PAGE_SIZE = 100

    scanned = updated = skipped = errors = 0
    preview = []
    offset  = 0

    # Ensure the in-memory account indexes are populated before scanning
    if not _account_to_dealer:
        _slp_sync_status["status_detail"] = "building account index…"
        print("[sync-slp] Account index empty — rebuilding before sync")
        await _build_dealer_id_index()
        print(f"[sync-slp] Index ready: {len(_account_to_dealer)} accounts")
    _slp_sync_status["indexed_accounts"] = len(_account_to_dealer)

    try:
        while True:
            page    = await ac_get(f"customObjects/records/{SLP_SCHEMA_ID}",
                                   {"limit": PAGE_SIZE, "offset": offset})
            records = page.get("records", [])
            if not records:
                break

            for r in records:
                scanned += 1
                rec_id = r.get("id")
                fields = {fo["id"]: fo.get("value") for fo in r.get("fields", [])}
                rel    = r.get("relationships", {}).get("account", [])
                # rel may be list of ints OR list of dicts — normalise to string id
                if rel:
                    first = rel[0]
                    acc_id = str(first.get("id", first) if isinstance(first, dict) else first)
                else:
                    acc_id = None

                # Debug first 3 records so we can verify structure
                if scanned <= 3:
                    print(f"[sync-slp][debug] rec={rec_id} acc_id={acc_id} "
                          f"fields={list(fields.keys())} "
                          f"dealer_lookup={_account_to_dealer.get(acc_id,'MISS')} "
                          f"platform_lookup={_account_to_platform.get(acc_id,'MISS')}")

                to_update = []
                for slp_fid, cf_id in _SLP_SYNC_FIELDS:
                    if fields.get(slp_fid):
                        continue
                    if cf_id is None:          # dealer-id → dealer index
                        val = _account_to_dealer.get(acc_id, "") if acc_id else ""
                    elif cf_id == "119":       # BDR → BDR index
                        val = _account_to_bdr.get(acc_id, "") if acc_id else ""
                    else:
                        val = ""
                    if val:
                        to_update.append({"id": slp_fid, "value": val})

                if not to_update:
                    skipped += 1
                    # Track why first few records were skipped for diagnostics
                    if skipped <= 5:
                        reason = "no_account" if not acc_id else "no_index_match"
                        _slp_sync_status.setdefault("skip_samples", []).append(
                            {"rec": rec_id, "acc_id": acc_id, "reason": reason,
                             "field_keys": list(fields.keys())})
                    continue

                if dry_run:
                    if len(preview) < 50:
                        preview.append({"record_id": rec_id, "account_id": acc_id,
                                        "fields": to_update})
                    updated += 1
                    continue

                try:
                    await _merge_update_custom_object_record(SLP_SCHEMA_ID, rec_id, to_update)
                    updated += 1
                except Exception as e:
                    errors += 1
                    err_str = str(e)
                    print(f"[sync-slp] Error updating record {rec_id}: {err_str}")
                    _slp_sync_status["last_error"] = f"record {rec_id}: {err_str}"

            # Update live progress
            _slp_sync_status.update({"scanned": scanned, "updated": updated,
                                     "skipped": skipped, "errors": errors})
            del records
            offset += PAGE_SIZE
            if len(page.get("records", [])) < PAGE_SIZE:
                break

        _slp_sync_status.update({"status": "done", "scanned": scanned, "updated": updated,
                                  "skipped": skipped, "errors": errors,
                                  "finished": datetime.utcnow().isoformat()})
        if dry_run:
            _slp_sync_status["preview"] = preview
        print(f"[sync-slp] Done — scanned={scanned} updated={updated} skipped={skipped} errors={errors}")

    except Exception as e:
        _slp_sync_status.update({"status": "error", "detail": str(e),
                                  "scanned": scanned, "updated": updated,
                                  "skipped": skipped, "errors": errors})
        print(f"[sync-slp] Fatal error: {e}")


def _check_sync_token(token: str = Query(..., description="SYNC_TOKEN value from Render env")):
    if not _SYNC_TOKEN:
        return   # not configured → open (local dev)
    if not secrets.compare_digest(token, _SYNC_TOKEN):
        raise HTTPException(status_code=401, detail="Invalid sync token")


@app.post("/api/sync-slp-fields")
async def sync_slp_fields(
    dry_run: bool = Query(True, description="Preview changes without writing to AC"),
    _: None = Depends(_check_sync_token),
):
    """Start a background sync of missing SLP fields from account data.

    Returns immediately. Poll GET /api/sync-slp-fields/status to track progress.
    Authenticate with ?token=<SYNC_TOKEN>.
    """
    if _slp_sync_status.get("status") == "running":
        return {"status": "already_running", "progress": _slp_sync_status}
    asyncio.create_task(_run_slp_sync(dry_run))
    return {"status": "started", "dry_run": dry_run,
            "poll": "/api/sync-slp-fields/status"}


@app.get("/api/sync-slp-fields/status")
async def sync_slp_fields_status(
    _: None = Depends(_check_sync_token),
):
    """Check the status/results of the last sync-slp-fields run."""
    return _slp_sync_status


@app.get("/api/report/ars-360")
async def report_ars_360(
    format: str = Query("json"),
    _: None = Depends(require_auth),
):
    """ARS dealers on 360 Finance program with all linked contacts.
    Uses SLP channel field directly so accounts with multiple SLPs are not missed."""
    all_accounts = await ac_get_all("accounts", "accounts", {})

    # Build map: account_id → set of SLP channels (from live SLP cache)
    acct_channels: dict = defaultdict(set)
    for slp_rec in _slp_cache_records:
        for acct_id in slp_rec.get("relationships", {}).get("account", []):
            aid = str(acct_id)
            for f in slp_rec.get("fields", []):
                if f.get("id") == "channel" and f.get("value"):
                    acct_channels[aid].add(f["value"].strip())

    # Filter to ARS accounts that have at least one SLP with channel = 360 Finance
    ars_accounts = {
        str(a["id"]): a
        for a in all_accounts
        if "ARS" in a.get("name", "").upper()
        and "360 Finance" in acct_channels.get(str(a["id"]), set())
    }

    # Fetch contacts per account
    by_account: dict = defaultdict(list)
    async def _fetch_contacts_for_account(aid: str):
        try:
            ac_resp = await ac_get(f"accounts/{aid}/accountContacts")
            contact_ids = [
                str(ac.get("contact"))
                for ac in ac_resp.get("accountContacts", [])
                if ac.get("contact")
            ]
            contacts = []
            for cid in contact_ids:
                try:
                    cr = await ac_get(f"contacts/{cid}")
                    contacts.append(cr.get("contact", {}))
                except Exception:
                    pass
            return aid, contacts
        except Exception:
            return aid, []

    contact_results = await asyncio.gather(
        *[_fetch_contacts_for_account(aid) for aid in ars_accounts]
    )
    for aid, contacts in contact_results:
        by_account[aid] = contacts

    results = []
    for aid, acct in sorted(ars_accounts.items(), key=lambda x: x[1].get("name", "")):
        contacts = by_account.get(aid, [])
        # Show all channels this account has SLPs for
        channels = ", ".join(sorted(acct_channels.get(aid, set())))
        base = {
            "dealer_id":   _account_to_dealer.get(aid, ""),
            "dealer_name": acct.get("name", ""),
            "channel":     channels,
            "bdr":         _account_to_bdr.get(aid, ""),
        }
        if contacts:
            for c in contacts:
                results.append({**base,
                    "contact_first": c.get("firstName", ""),
                    "contact_last":  c.get("lastName", ""),
                    "contact_email": c.get("email", ""),
                    "contact_phone": c.get("phone", ""),
                })
        else:
            results.append({**base,
                "contact_first": "", "contact_last": "",
                "contact_email": "", "contact_phone": "",
            })

    if format == "csv":
        return _csv_response(results, f"ars_360_{datetime.now().strftime('%Y%m%d')}.csv")
    return {"count": len(results), "records": results}


# ── Smart Query — Claude-powered NL → report intent ──────────────────────────

_SMART_SYSTEM = """You are a report-routing assistant for Microf, a financing platform.
Your job: parse a natural-language query and return a JSON object describing which report to run.

Available report types (use exactly these values for "report_type"):
  activations       — new contractor/dealer activations
  training-summary  — training records and completions
  license-expiration — licenses expiring soon or already expired
  bdr-summary       — BDR (Business Development Rep) performance summary
  team-activity     — internal team activity / notes
  account-activity  — account engagement / cold accounts with no activity
  dealer-profile    — look up a specific dealer by ID

Available platforms (use exactly as shown, or null):
  "360 Finance", "OPTIMUS", "LTO", "Microf", "SpectrumAC",
  "ACIMA", "FlexShopper", "Snap", "Kornerstone", "GreenSky", "UOWn", "Wells"

Today's date: {today}

Return ONLY valid JSON, no prose. Schema:
{{
  "report_type": "<one of the above, or null if truly ambiguous>",
  "from_date": "<YYYY-MM-DD or null>",
  "to_date": "<YYYY-MM-DD or null>",
  "platform": "<platform name or null>",
  "bdr": "<BDR name/username or null>",
  "dealer_id": "<numeric dealer ID or null>",
  "cold_accounts": <true if user wants accounts with no activity, else false>,
  "days_ahead": <integer if asking about upcoming license expiration, else null>,
  "include_expired": <true if asking about already-expired licenses, else false>,
  "explanation": "<one short sentence describing what you understood, shown to the user>",
  "error": "<only if truly cannot map to a report; leave null otherwise>"
}}

Examples:
  "activations for Optimus last month"
  → {{"report_type":"activations","from_date":"<first of last month>","to_date":"<last of last month>","platform":"OPTIMUS","bdr":null,"dealer_id":null,"cold_accounts":false,"days_ahead":null,"include_expired":false,"explanation":"Activations for OPTIMUS platform last month","error":null}}

  "which BDR signed up the most 360 Finance partners this quarter"
  → {{"report_type":"bdr-summary","from_date":"<Q start>","to_date":"<today>","platform":"360 Finance","bdr":null,"dealer_id":null,"cold_accounts":false,"days_ahead":null,"include_expired":false,"explanation":"BDR summary for 360 Finance this quarter","error":null}}

  "show me licenses expiring in the next 60 days"
  → {{"report_type":"license-expiration","from_date":null,"to_date":null,"platform":null,"bdr":null,"dealer_id":null,"cold_accounts":false,"days_ahead":60,"include_expired":false,"explanation":"Licenses expiring in the next 60 days","error":null}}

  "dealers who haven't had any activity in 6 months"
  → {{"report_type":"account-activity","from_date":"<6 months ago>","to_date":"<today>","platform":null,"bdr":null,"dealer_id":null,"cold_accounts":true,"days_ahead":null,"include_expired":false,"explanation":"Cold accounts with no activity in the last 6 months","error":null}}
"""


@app.get("/api/smart-query")
async def smart_query_endpoint(q: str, user=Depends(require_auth)):
    """Parse a natural-language query with Claude and return structured report intent."""
    if not q.strip():
        raise HTTPException(400, "Query required")

    if not _ANTHROPIC_KEY:
        return {"error": "ANTHROPIC_API_KEY not configured", "fallback": True}

    today = datetime.now().strftime("%Y-%m-%d")
    system = _SMART_SYSTEM.format(today=today)

    try:
        import anthropic as _anthropic
        client = _anthropic.AsyncAnthropic(api_key=_ANTHROPIC_KEY)
        msg = await client.messages.create(
            model="claude-3-5-haiku-20241022",
            max_tokens=512,
            system=system,
            messages=[{"role": "user", "content": q}],
        )
        raw = msg.content[0].text.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        import json as _json
        parsed = _json.loads(raw)
        return parsed
    except Exception as e:
        print(f"[smart-query] Claude parse error: {e}")
        return {"error": str(e), "fallback": True}


# ---------------------------------------------------------------------------
# OPTIMUS Bulk Deactivation
# ---------------------------------------------------------------------------
import re as _re

class _DeactivateConfirmIn(_BaseModel):
    record_ids: list   # SLP record IDs to set Deactivated
    email_text: str = ""  # Original email body to log as Account Activity note

@app.post("/api/admin/optimus-deactivate/preview")
async def optimus_deactivate_preview(body: dict = Body(...), admin=Depends(_require_admin)):
    """
    Parse a GreenSky deactivation email body, find each dealer's OPTIMUS SLP,
    and return a preview list without making any changes.
    """
    text = body.get("text", "")
    # Extract 4-6 digit numbers as candidate dealer IDs
    raw_ids = _re.findall(r'\b(\d{4,6})\b', text)
    dealer_ids = list(dict.fromkeys(raw_ids))  # dedupe, preserve order

    SLP_SCHEMA = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
    rows = []
    not_found = []

    # Determine which dealer IDs are missing from the CF18 index
    missing_dids = {did for did in dealer_ids if did not in _dealer_id_index}

    # For missing IDs: scan all SLP records once; store account_id AND the matching SLP record
    slp_did_to_acct: dict = {}   # did → account_id
    slp_did_to_rec:  dict = {}   # did → SLP record (so we don't need to re-filter by platform)
    if missing_dids:
        offset = 0
        while True:
            page = await ac_get(f"customObjects/records/{SLP_SCHEMA}",
                                {"limit": 100, "offset": offset})
            records = page.get("records", [])
            if not records:
                break
            for r in records:
                fmap = {f["id"]: f.get("value", "") for f in r.get("fields", [])}
                slp_did = str(fmap.get("dealer-id") or "").strip()
                if slp_did in missing_dids and slp_did not in slp_did_to_acct:
                    rels  = r.get("relationships", {})
                    accts = rels.get("account", [])
                    if accts:
                        a0  = accts[0]
                        aid = str(a0) if isinstance(a0, (int, str)) else str(a0.get("id", ""))
                        slp_did_to_acct[slp_did] = aid
                        slp_did_to_rec[slp_did]  = (r, fmap)
            total = int(page.get("meta", {}).get("total", 0))
            offset += len(records)
            if offset >= total or len(slp_did_to_acct) >= len(missing_dids):
                break

        # Fetch names for newly found accounts
        new_acct_ids = list(set(slp_did_to_acct.values()))
        acct_names: dict = {}
        for aid in new_acct_ids:
            try:
                ad = await ac_get(f"accounts/{aid}")
                acct_names[aid] = ad.get("account", {}).get("name", "")
            except Exception:
                acct_names[aid] = ""

    for did in dealer_ids:
        # Phase 1: try CF18 index
        entry     = _dealer_id_index.get(did)
        acct_id   = str(entry["id"])    if entry else None
        acct_name = entry.get("name", "") if entry else ""

        # Phase 2: fall back to SLP dealer-id scan
        via_slp_scan = False
        if not acct_id and did in slp_did_to_acct:
            acct_id      = slp_did_to_acct[did]
            acct_name    = acct_names.get(acct_id, "")
            via_slp_scan = True

        if not acct_id:
            not_found.append(did)
            continue

        if via_slp_scan:
            # Use the SLP record we already found — skip platform filter since
            # platform may be unset on these records
            r, fmap = slp_did_to_rec[did]
            rows.append({
                "record_id":      r["id"],
                "account_id":     acct_id,
                "dealer_id":      did,
                "account_name":   acct_name,
                "current_status": fmap.get("slp-status-detail", ""),
                "channel":        fmap.get("channel", "OPTIMUS"),
            })
        else:
            # Fetch OPTIMUS SLPs for this account (CF18-indexed path)
            slp_data = await ac_get(f"customObjects/records/{SLP_SCHEMA}",
                                    {"filters[relationships.account]": acct_id, "limit": 50})
            found_any = False
            for r in slp_data.get("records", []):
                fmap = {f["id"]: f.get("value", "") for f in r.get("fields", [])}
                if "optimus" not in (fmap.get("channel") or "").lower():
                    continue
                found_any = True
                rows.append({
                    "record_id":      r["id"],
                    "account_id":     acct_id,
                    "dealer_id":      did,
                    "account_name":   acct_name,
                    "current_status": fmap.get("slp-status-detail", ""),
                    "channel":        fmap.get("channel", ""),
                })
            if not found_any:
                not_found.append(did)

    return {"preview": rows, "not_found": not_found, "dealer_ids_parsed": dealer_ids}


@app.post("/api/admin/optimus-deactivate/confirm")
async def optimus_deactivate_confirm(
    body: _DeactivateConfirmIn,
    request: _Request,
    admin=Depends(_require_admin),
):
    """
    Set slp-status-detail = Deactivated on the given SLP record IDs.
    Fetches each record first to avoid wiping other fields.
    Also logs an Account Activity note with the original email body.
    """
    SLP_SCHEMA = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
    results = {"updated": [], "failed": [], "notes": []}
    performed_by = _get_session_email(request) or admin or "Microf Reports"
    today_str = datetime.now().strftime("%Y-%m-%d")
    noted_accounts: set = set()  # track which accounts already got a note

    for rec_id in body.record_ids:
        try:
            # Fetch current record
            rd = await ac_get(f"customObjects/records/{SLP_SCHEMA}/{rec_id}")
            rec = rd.get("record", {})
            existing = {f["id"]: f.get("value", "") for f in rec.get("fields", [])}
            rels = rec.get("relationships", {})
            acct_ids = [int(x) for x in (rels.get("account") or [])]

            # Merge with updated status
            existing["slp-status-detail"] = "Deactivated"

            payload = {
                "record": {
                    "fields": [{"id": k, "value": v} for k, v in existing.items() if v != ""],
                    "relationships": {"account": acct_ids},
                }
            }
            data = await ac_post(f"customObjects/records/{SLP_SCHEMA}/{rec_id}", payload)
            results["updated"].append(rec_id)

            # If no remaining Contractor Activated SLPs, set Account Status = Deactivated
            for acct_id in acct_ids:
                try:
                    all_slps = await ac_get(
                        f"customObjects/records/{SLP_SCHEMA}",
                        {"filters[relationships.account]": acct_id},
                    )
                    still_active = any(
                        f.get("value", "") == "Contractor Activated"
                        for r in all_slps.get("records", [])
                        for f in r.get("fields", [])
                        if f.get("id") == "slp-status-detail"
                    )
                    if not still_active:
                        await ac_post(
                            f"accounts/{acct_id}/accountCustomFieldData",
                            {"accountCustomFieldData": [
                                {"customerAccountFieldId": 19, "fieldValue": "Deactivated"}
                            ]},
                        )
                        results.setdefault("account_status_updated", []).append(acct_id)
                except Exception as ae:
                    print(f"[optimus-deactivate] account status update failed for {acct_id}: {ae}")

            # Post Account Activity note (once per account)
            if body.email_text:
                for acct_id in acct_ids:
                    if acct_id in noted_accounts:
                        continue
                    noted_accounts.add(acct_id)
                    note_payload = {
                        "record": {
                            "fields": [
                                {"id": "activity-type",  "value": "Email"},
                                {"id": "subject",        "value": "GreenSky OPTIMUS Deactivation Notice"},
                                {"id": "body",           "value": body.email_text},
                                {"id": "activity-date",  "value": today_str},
                                {"id": "performed-by",   "value": performed_by},
                                {"id": "source",         "value": "Microf Reports"},
                            ],
                            "relationships": {"account": [acct_id]},
                        }
                    }
                    try:
                        await ac_post(f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", note_payload)
                        results["notes"].append(acct_id)
                    except Exception as ne:
                        print(f"[optimus-deactivate] note failed for acct {acct_id}: {ne}")
        except Exception as e:
            results["failed"].append({"id": rec_id, "error": str(e)})

    return results


# ── OPTIMUS Reactivation ──────────────────────────────────────────────────────

@app.post("/api/admin/optimus-reactivate/preview")
async def optimus_reactivate_preview(body: dict = Body(...), admin=Depends(_require_admin)):
    """Same lookup as deactivation preview but returns all non-activated OPTIMUS SLPs."""
    text = body.get("text", "")
    raw_ids = _re.findall(r'\b(\d{4,6})\b', text)
    dealer_ids = list(dict.fromkeys(raw_ids))
    SLP_SCHEMA = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
    rows, not_found = [], []

    for did in dealer_ids:
        entry   = _dealer_id_index.get(did)
        acct_id = str(entry["id"]) if entry else None
        acct_name = entry.get("name", "") if entry else ""

        if not acct_id:
            slp_fallback = await ac_get(f"customObjects/records/{SLP_SCHEMA}", {"filters[dealer-id]": did, "limit": 10})
            for r in slp_fallback.get("records", []):
                accts = r.get("relationships", {}).get("account", [])
                if accts:
                    a0 = accts[0]
                    acct_id = str(a0) if isinstance(a0, (int, str)) else str(a0.get("id", ""))
                    try:
                        ad = await ac_get(f"accounts/{acct_id}")
                        acct_name = ad.get("account", {}).get("name", "")
                    except Exception:
                        acct_name = ""
                    break

        if not acct_id:
            # Scan all SLP records for this dealer-id
            all_slps = await ac_get_all(f"customObjects/records/{SLP_SCHEMA}", "records", {"limit": 100})
            for r in all_slps:
                fmap = {f["id"]: f.get("value", "") for f in r.get("fields", [])}
                if str(fmap.get("dealer-id", "")).strip() == did:
                    accts = r.get("relationships", {}).get("account", [])
                    if accts:
                        a0 = accts[0]
                        acct_id = str(a0) if isinstance(a0, (int, str)) else str(a0.get("id", ""))
                        try:
                            ad = await ac_get(f"accounts/{acct_id}")
                            acct_name = ad.get("account", {}).get("name", "")
                        except Exception:
                            acct_name = ""
                    break

        if not acct_id:
            not_found.append(did)
            continue

        slp_data = await ac_get(f"customObjects/records/{SLP_SCHEMA}", {"filters[relationships.account]": acct_id, "limit": 50})
        found_any = False
        for r in slp_data.get("records", []):
            fmap = {f["id"]: f.get("value", "") for f in r.get("fields", [])}
            if "optimus" not in (fmap.get("channel") or "").lower():
                continue
            found_any = True
            rows.append({
                "record_id":      r["id"],
                "account_id":     acct_id,
                "dealer_id":      did,
                "account_name":   acct_name,
                "current_status": fmap.get("slp-status-detail", ""),
                "channel":        fmap.get("channel", ""),
                "fields":         fmap,
            })
        if not found_any:
            not_found.append(did)

    return {"preview": rows, "not_found": not_found, "dealer_ids_parsed": dealer_ids}


@app.post("/api/admin/optimus-reactivate/confirm")
async def optimus_reactivate_confirm(
    body: _DeactivateConfirmIn,
    request: _Request,
    admin=Depends(_require_admin),
):
    """Set each SLP to Contractor Activated, set contractor-activated-date to today,
    update account status if this was the only SLP, and log an Account Activity note."""
    SLP_SCHEMA = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S-05:00")
    today_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    performed_by = _get_session_email(request) or admin or "Microf Reports"
    results = {"updated": [], "failed": [], "notes": [], "account_status_updated": []}

    for rec_id in body.record_ids:
        try:
            rec_data = await ac_get(f"customObjects/records/{SLP_SCHEMA}/{rec_id}")
            record = rec_data.get("record", {})
            if not record:
                results["failed"].append({"id": rec_id, "error": "Not found"})
                continue

            acct_id = None
            accts = record.get("relationships", {}).get("account", [])
            if accts:
                a0 = accts[0]
                acct_id = str(a0) if isinstance(a0, (int, str)) else str(a0.get("id", ""))

            # Build updated fields
            existing = {f["id"]: f.get("value", "") for f in record.get("fields", [])}
            existing["slp-status-detail"] = "Contractor Activated"
            existing["contractor-activated-date"] = today_iso
            # Clear deactivation date
            existing["deactivation-date"] = ""

            new_fields = [{"id": k, "value": v} for k, v in existing.items()]
            payload = {
                "record": {
                    "fields": new_fields,
                    "relationships": {"account": [int(acct_id)]} if acct_id else {}
                }
            }
            data = await ac_post(f"customObjects/records/{SLP_SCHEMA}/{rec_id}", payload)
            results["updated"].append(rec_id)

            # Update account status if this is the only active SLP
            if acct_id:
                try:
                    all_acct_slps = await ac_get(f"customObjects/records/{SLP_SCHEMA}",
                                                 {"filters[relationships.account]": acct_id, "limit": 50})
                    other_active = [
                        r for r in all_acct_slps.get("records", [])
                        if r["id"] != rec_id and
                        any(f["id"] == "slp-status-detail" and f.get("value") == "Contractor Activated"
                            for f in r.get("fields", []))
                    ]
                    if not other_active:
                        await ac_post(
                            f"accounts/{acct_id}/accountCustomFieldData",
                            {"accountCustomFieldData": [
                                {"customerAccountFieldId": 19, "fieldValue": "Contractor"}
                            ]},
                        )
                        results["account_status_updated"].append(acct_id)
                except Exception as ae:
                    print(f"[optimus-reactivate] account status update failed for {acct_id}: {ae}")

            # Post Account Activity note
            if body.email_text and acct_id:
                try:
                    note_payload = {
                        "record": {
                            "fields": [
                                {"id": "activity-type",  "value": "Email"},
                                {"id": "subject",        "value": "EGIA OPTIMUS Reactivation Notice"},
                                {"id": "body",           "value": body.email_text},
                                {"id": "activity-date",  "value": today_date},
                                {"id": "performed-by",   "value": performed_by},
                                {"id": "source",         "value": "Microf Reports"},
                            ],
                            "relationships": {"account": [int(acct_id)]}
                        }
                    }
                    await ac_post(f"customObjects/records/{ACCT_ACTIVITY_SCHEMA_ID}", note_payload)
                    results["notes"].append(acct_id)
                except Exception as ne:
                    print(f"[optimus-reactivate] note failed for acct {acct_id}: {ne}")
        except Exception as e:
            results["failed"].append({"id": rec_id, "error": str(e)})

    return results


# ── Move Records (Deal / Contact / SLP) ───────────────────────────────────────

class _MoveIn(_BaseModel):
    record_id: str
    new_account_id: str

@app.post("/api/admin/move-deal")
async def move_deal(body: _MoveIn, admin=Depends(_require_admin)):
    """Reassign a deal to a different account."""
    try:
        data = await ac_put(f"deals/{body.record_id}", {"deal": {"account": body.new_account_id}})
        return {"ok": True, "deal_id": body.record_id, "new_account_id": body.new_account_id, "deal": data.get("deal", {})}
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})

@app.post("/api/admin/move-contact")
async def move_contact(body: _MoveIn, admin=Depends(_require_admin)):
    """Move a contact's account association to a different account."""
    try:
        ac_data = await ac_get("accountContacts", {"contact": body.record_id, "limit": 50})
        associations = ac_data.get("accountContacts", [])

        deleted = []
        for assoc in associations:
            if str(assoc.get("account", "")) != str(body.new_account_id):
                ds = await ac_delete(f"accountContacts/{assoc['id']}")
                if ds in (200, 204):
                    deleted.append(assoc["id"])

        data = await ac_post("accountContacts", {
            "accountContact": {"contact": body.record_id, "account": body.new_account_id}
        })
        return {"ok": True, "contact_id": body.record_id, "new_account_id": body.new_account_id, "removed_associations": deleted}
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})

@app.post("/api/admin/move-slp")
async def move_slp(body: _MoveIn, admin=Depends(_require_admin)):
    """Move an SLP record to a different account.
    AC does not allow changing relationships on existing records, so we
    create a new record on the target account then delete the old one."""
    SLP_SCHEMA = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
    try:
        rec_data = await ac_get(f"customObjects/records/{SLP_SCHEMA}/{body.record_id}")
        record = rec_data.get("record", {})
        if not record:
            return JSONResponse(status_code=404, content={"ok": False, "error": "SLP record not found"})
        fields = record.get("fields", [])
        # Create new record on target account
        create_payload = {
            "record": {
                "fields": fields,
                "relationships": {"account": [int(body.new_account_id)]}
            }
        }
        created = await ac_post(f"customObjects/records/{SLP_SCHEMA}", create_payload)
        new_id = created.get("record", {}).get("id")
        if not new_id:
            return JSONResponse(status_code=500, content={"ok": False, "error": "Failed to create new SLP record"})
        # Delete the old record
        await ac_delete(f"customObjects/records/{SLP_SCHEMA}/{body.record_id}")
        return {"ok": True, "old_slp_id": body.record_id, "new_slp_id": new_id, "new_account_id": body.new_account_id}
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})


# ─────────────────────────────────────────────────────────────────────────────
# Account Consolidation Tool
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/move-account")
async def move_account_page(user=Depends(_require_admin)):
    return FileResponse("static/move-account.html")

@app.get("/api/admin/account-contents/{account_id}")
async def account_contents(account_id: str, admin=Depends(_require_admin)):
    """Return SLPs, contacts, and note count for a given account."""
    SLP_SCHEMA = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
    try:
        acct_data, slp_data, contact_data, note_data = await asyncio.gather(
            ac_get(f"accounts/{account_id}"),
            ac_get(f"customObjects/records/{SLP_SCHEMA}", {"filters[relationships.account][eq]": account_id, "limit": 50}),
            ac_get(f"accounts/{account_id}/contacts", {"limit": 50}),
            ac_get("notes", {"reltype": "CustomerAccount", "rel_id": account_id, "limit": 1}),
        )
        account = acct_data.get("account", {})
        slps = []
        for rec in slp_data.get("records", []):
            fmap = {f["id"]: (f.get("value") or "").strip() for f in rec.get("fields", [])}
            slps.append({
                "id": rec["id"],
                "dealer_id": fmap.get("dealer-id", ""),
                "name": fmap.get("name", ""),
                "channel": fmap.get("channel", ""),
                "status": fmap.get("slp-status-detail", ""),
            })
        contacts = []
        for ac in contact_data.get("accountContacts", []):
            cid = str(ac["contact"])
            cdata = await ac_get(f"contacts/{cid}")
            c = cdata.get("contact", {})
            contacts.append({
                "id": cid,
                "account_contact_id": ac["id"],
                "name": f"{c.get('firstName','')} {c.get('lastName','')}".strip(),
                "email": c.get("email", ""),
            })
        note_count = int(note_data.get("meta", {}).get("total", 0))
        return {
            "account": {"id": account_id, "name": account.get("name", "")},
            "slps": slps,
            "contacts": contacts,
            "note_count": note_count,
        }
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})


class _ConsolidateIn(_BaseModel):
    source_id: str
    target_id: str
    slp_ids: list = []
    contact_ids: list = []
    move_notes: bool = False
    delete_source: bool = False

@app.post("/api/admin/consolidate-account")
async def consolidate_account(body: _ConsolidateIn, admin=Depends(_require_admin)):
    """Move selected SLPs and contacts from source account to target account."""
    SLP_SCHEMA = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
    results = {"slps": [], "contacts": [], "notes": None, "deleted_source": False, "errors": []}

    # Move SLPs — update relationship in-place
    for slp_id in body.slp_ids:
        try:
            rec_data = await ac_get(f"customObjects/records/{SLP_SCHEMA}/{slp_id}")
            record = rec_data.get("record", {})
            fields = record.get("fields", [])
            payload = {"record": {"id": slp_id, "fields": fields, "relationships": {"account": [int(body.target_id)]}}}
            await ac_post(f"customObjects/records/{SLP_SCHEMA}", payload)
            results["slps"].append({"id": slp_id, "ok": True})
        except Exception as e:
            results["errors"].append(f"SLP {slp_id}: {e}")
            results["slps"].append({"id": slp_id, "ok": False, "error": str(e)})

    # Move contacts
    for contact_id in body.contact_ids:
        try:
            # Link to target
            await ac_post("accountContacts", {"accountContact": {"contact": contact_id, "account": body.target_id}})
            # Remove from source
            ac_data = await ac_get("accountContacts", {"contact": contact_id, "limit": 50})
            for assoc in ac_data.get("accountContacts", []):
                if str(assoc.get("account")) == str(body.source_id):
                    await ac_delete(f"accountContacts/{assoc['id']}")
            results["contacts"].append({"id": contact_id, "ok": True})
        except Exception as e:
            results["errors"].append(f"Contact {contact_id}: {e}")
            results["contacts"].append({"id": contact_id, "ok": False, "error": str(e)})

    # Move notes (optional — can be slow for large accounts)
    if body.move_notes:
        moved = errors = 0
        offset = 0
        while True:
            try:
                note_data = await ac_get("notes", {"reltype": "CustomerAccount", "rel_id": body.source_id, "limit": 25, "offset": offset})
                notes = note_data.get("notes", [])
                if not notes:
                    break
                for note in notes:
                    try:
                        await ac_post("notes", {"note": {"note": note.get("note", ""), "reltype": "CustomerAccount", "relid": body.target_id}})
                        await ac_delete(f"notes/{note['id']}")
                        moved += 1
                    except Exception:
                        errors += 1
                offset += len(notes)
            except Exception:
                break
        results["notes"] = {"moved": moved, "errors": errors}

    # Delete source account
    if body.delete_source and not results["errors"]:
        try:
            await ac_delete(f"accounts/{body.source_id}")
            results["deleted_source"] = True
        except Exception as e:
            results["errors"].append(f"Delete source: {e}")

    return {"ok": len(results["errors"]) == 0, **results}


# ─────────────────────────────────────────────────────────────────────────────
# Dealer Fix Tool — split a misfiled dealer (SLP linked to the wrong account,
# usually because a similarly-named account already existed) into its own
# dedicated account, carrying over its contact(s) and deal(s).
# ─────────────────────────────────────────────────────────────────────────────

_DEALER_FIX_SLP_CFS = {
    "dealer-id":                 18,
    "doing-business-in-states":  22,
    "oracle-producer-ids":       118,
    "ein":                       40,
}

def _find_slp_by_dealer_id(slp_records: list, dealer_id: str) -> Optional[dict]:
    for r in slp_records:
        fields = {f.get("id"): f.get("value") for f in r.get("fields", [])}
        if str(fields.get("dealer-id", "")).strip() == dealer_id:
            return {"id": r["id"], "fields": fields, "raw_fields": r.get("fields", []),
                    "relationships": r.get("relationships", {})}
    return None

def _normalize_dealer_name_parts(name: str) -> list:
    """Return normalized candidate names for a dealer/account name string.
    Accounts are commonly named "{Legal Entity} DBA {Common Name}" — either
    side could be the one that matches, so both are returned as candidates."""
    n = (name or "").upper()
    n = n.split(":")[0]                     # drop trailing ": <channel>" suffix
    segments = _re.split(r"\bDBA\b", n) or [n]

    def clean(s: str) -> str:
        s = _re.sub(r"-\d+\s*$", "", s)      # drop trailing "-<digits>"
        s = _re.sub(r"[^A-Z0-9 ]", " ", s)   # strip punctuation
        return _re.sub(r"\s+", " ", s).strip()

    return [c for c in (clean(s) for s in segments) if c]

def _names_look_related(a: str, b: str) -> bool:
    """Best-effort check for whether an SLP name and an account name refer to
    the same company. Not authoritative — a Rojos-enrollment marker deal gets
    created for ANY new dealer-id enrollment, including legitimate re-enrollments
    of an existing dealer under a new channel, so this is surfaced to the admin
    as a caution note rather than used to suppress the mismatch flag."""
    parts_a, parts_b = _normalize_dealer_name_parts(a), _normalize_dealer_name_parts(b)
    for na in parts_a:
        for nb in parts_b:
            if na == nb or na in nb or nb in na:
                return True
    return False

@app.get("/dealer-fix")
async def dealer_fix_page(user=Depends(_require_admin)):
    return FileResponse("static/dealer-fix.html")

@app.get("/api/admin/dealer-fix/check")
async def dealer_fix_check(dealer_id: str = Query(...), admin=Depends(_require_admin)):
    """Look up a dealer's SLP record and check whether it's sharing an account
    with a different (usually similarly-named) dealer. If so, surface the
    contact(s)/deal(s) on that account so an admin can confirm which ones
    actually belong to this dealer before splitting it out."""
    dealer_id = dealer_id.strip()
    if not dealer_id:
        raise HTTPException(status_code=400, detail="Dealer ID required")

    slp_records = await get_slp_cache()
    slp_match = _find_slp_by_dealer_id(slp_records, dealer_id)
    if not slp_match:
        return {"ok": False, "error": f"No SLP record found for Dealer ID {dealer_id}"}

    fields = slp_match["fields"]
    slp_info = {
        "id":                 slp_match["id"],
        "name":               fields.get("name", ""),
        "channel":            fields.get("channel", ""),
        "status":             fields.get("slp-status-detail", ""),
        "ein":                fields.get("ein", ""),
        "states":             fields.get("doing-business-in-states", ""),
        "oracle_producer_id": fields.get("oracle-producer-ids", ""),
        "dealerkey":          fields.get("dealerkey", ""),
    }

    accts = slp_match["relationships"].get("account", [])
    account_id = str(accts[0]) if accts else None

    if not account_id:
        return {"ok": True, "slp": slp_info, "linked": False,
                "message": "This SLP has no linked account at all — nothing to split."}

    acct_data, cfd, deals_resp = await asyncio.gather(
        ac_get(f"accounts/{account_id}"),
        ac_get(f"accounts/{account_id}/accountCustomFieldData"),
        ac_get("deals", {"filters[account]": account_id, "limit": 100}),
    )
    account_name = acct_data.get("account", {}).get("name", "")
    acct_dealer_id = ""
    for cf in cfd.get("customerAccountCustomFieldData", []):
        if str(cf.get("custom_field_id")) == "18":
            acct_dealer_id = (cf.get("custom_field_text_value") or "").strip()
            break

    # A differing Parent Dealer ID field alone does NOT mean this dealer is
    # misfiled — dealers routinely have several SLPs (one per financing
    # channel) legitimately sharing one account, each with its own per-channel
    # dealer-id. The one reliable signal for an actual same-name-collision
    # misfile is the enrollment automation's own marker on the deal
    # description ("Rojos Enrollment for dealer {id}"), combined with the
    # account already belonging to a *different* dealer.
    marker = f"dealer {dealer_id}".lower()
    marker_deals = [d for d in deals_resp.get("deals", []) if marker in (d.get("description") or "").lower()]

    is_misfiled = bool(marker_deals) and bool(acct_dealer_id) and acct_dealer_id != dealer_id

    # Always pull the account's contacts/deals — even when it looks correctly
    # filed — so the UI can offer a "Move Anyway" override in case this
    # detection heuristic is wrong (it's advisory, not authoritative).
    contacts_resp = await ac_get(f"accounts/{account_id}/accountContacts")

    contact_ids = [str(c.get("contact")) for c in contacts_resp.get("accountContacts", [])]
    all_contacts = []
    for cid in contact_ids:
        try:
            cdata = await ac_get(f"contacts/{cid}")
            c = cdata.get("contact", {})
            all_contacts.append({"id": cid, "name": f"{c.get('firstName','')} {c.get('lastName','')}".strip(),
                                  "email": c.get("email", "")})
        except Exception:
            all_contacts.append({"id": cid, "name": "(unknown)", "email": ""})

    slp_name_upper = (fields.get("name") or "").strip().upper()
    all_deals, matched_deal_ids, matched_contact_ids = [], [], []
    for d in deals_resp.get("deals", []):
        desc  = (d.get("description") or "")
        title = (d.get("title") or "").strip()
        is_match = marker in desc.lower() or (slp_name_upper and title.upper() == slp_name_upper)
        deal_contact_id = str(d.get("contact")) if d.get("contact") else None
        all_deals.append({"id": d.get("id"), "title": title, "description": desc,
                           "contact_id": deal_contact_id, "auto_matched": is_match})
        if is_match:
            matched_deal_ids.append(d.get("id"))
            if deal_contact_id:
                matched_contact_ids.append(deal_contact_id)

    message = (
        "This dealer looks correctly filed — no sign of a same-name-collision misfile "
        "(a differing Parent Dealer ID here is normal if this dealer is enrolled under "
        "multiple channels on the same account). You can still move it manually below if "
        "you believe this is wrong."
        if not is_misfiled else None
    )

    # Advisory only: a Rojos-enrollment marker gets created for ANY new dealer-id
    # enrollment, including a legitimate re-enrollment of an EXISTING dealer under a
    # new channel — which looks identical, technically, to a genuine same-name
    # collision. Surface a name-similarity note so the admin can tell the two apart
    # rather than trying to auto-resolve it.
    names_related = _names_look_related(fields.get("name", ""), account_name) if is_misfiled else False

    return {
        "ok": True, "slp": slp_info, "linked": True, "mismatch": is_misfiled,
        "account": {"id": account_id, "name": account_name, "dealer_id": acct_dealer_id, "url": ac_account_url(account_id)},
        "message": message,
        "names_look_related": names_related,
        "all_contacts": all_contacts,
        "all_deals": all_deals,
        "matched_deal_ids": matched_deal_ids,
        "matched_contact_ids": matched_contact_ids,
        "auto_matched": len(matched_deal_ids) > 0,
    }


class _DealerFixIn(_BaseModel):
    dealer_id: str
    source_account_id: str
    contact_ids: list = []
    deal_ids: list = []
    new_account_name: Optional[str] = None

@app.post("/api/admin/dealer-fix/execute")
async def dealer_fix_execute(body: _DealerFixIn, admin=Depends(_require_admin)):
    """Create a dedicated account for a misfiled dealer, move the selected
    contact(s)/deal(s) onto it, and repoint the SLP relationship."""
    SLP_SCHEMA = "d5ccf74f-981f-40ff-8a03-23cd0309808f"
    dealer_id = body.dealer_id.strip()

    slp_records = await get_slp_cache()
    slp_match   = _find_slp_by_dealer_id(slp_records, dealer_id)
    if not slp_match:
        raise HTTPException(status_code=404, detail=f"No SLP record found for Dealer ID {dealer_id}")
    fields     = slp_match["fields"]
    raw_fields = slp_match["raw_fields"]

    base_name = (body.new_account_name or fields.get("name") or f"Dealer {dealer_id}").strip()
    final_name = base_name
    try:
        acc = await ac_post("accounts", {"account": {"name": final_name}})
    except Exception as e:
        if "duplicate" in str(e).lower() or "already exists" in str(e).lower():
            final_name = f"{base_name}-{dealer_id}"
            acc = await ac_post("accounts", {"account": {"name": final_name}})
        else:
            raise HTTPException(status_code=400, detail=f"Failed to create account: {e}")
    new_account_id = acc.get("account", {}).get("id")
    if not new_account_id:
        raise HTTPException(status_code=400, detail=f"Failed to create account: {acc}")

    cf_errors = []
    for field_key, cf_id in _DEALER_FIX_SLP_CFS.items():
        val = (fields.get(field_key) or "").strip()
        if not val:
            continue
        try:
            await ac_post("accountCustomFieldData", {
                "accountCustomFieldDatum": {"customFieldId": cf_id, "customerAccountId": new_account_id, "fieldValue": val}
            })
        except Exception as e:
            cf_errors.append(f"cf {cf_id}: {e}")
    try:
        await ac_post("accountCustomFieldData", {
            "accountCustomFieldDatum": {"customFieldId": 36, "customerAccountId": new_account_id,
                                        "fieldValue": (fields.get("name") or base_name).strip()}
        })
    except Exception as e:
        cf_errors.append(f"cf 36: {e}")

    try:
        old_acct_data = await ac_get(f"accounts/{body.source_account_id}")
        old_name = old_acct_data.get("account", {}).get("name", "")
    except Exception:
        old_name = ""

    note_new = (
        f"Split out from Account {body.source_account_id} ({old_name}), which incorrectly shared this "
        f"account with Dealer {dealer_id} due to a similar company name. Source: SLP record — "
        f"dealerkey {fields.get('dealerkey','')}, channel {fields.get('channel','')}, "
        f"EIN {fields.get('ein','')}, Oracle Producer ID {fields.get('oracle-producer-ids','')}, "
        f"status {fields.get('slp-status-detail','')}."
    )
    note_old = (
        f"Dealer {dealer_id} ({fields.get('name','')}) was incorrectly linked to this account and has "
        f"been split out to its own Account (ID {new_account_id}, {final_name})."
    )
    try:
        await ac_post("notes", {"note": {"note": note_new, "relid": new_account_id, "reltype": "CustomerAccount", "userid": "1"}})
    except Exception as e:
        cf_errors.append(f"note (new account): {e}")
    try:
        await ac_post("notes", {"note": {"note": note_old, "relid": body.source_account_id, "reltype": "CustomerAccount", "userid": "1"}})
    except Exception as e:
        cf_errors.append(f"note (old account): {e}")

    moved_contacts = []
    for cid in body.contact_ids:
        try:
            assoc_data = await ac_get("accountContacts", {"contact": cid, "limit": 50})
            for assoc in assoc_data.get("accountContacts", []):
                if str(assoc.get("account")) == str(body.source_account_id):
                    await ac_delete(f"accountContacts/{assoc['id']}")
            await ac_post("accountContacts", {"accountContact": {"contact": cid, "account": new_account_id}})
            moved_contacts.append({"id": cid, "ok": True})
        except Exception as e:
            moved_contacts.append({"id": cid, "ok": False, "error": str(e)})

    moved_deals = []
    for did in body.deal_ids:
        try:
            await ac_put(f"deals/{did}", {"deal": {"organization": new_account_id}})
            moved_deals.append({"id": did, "ok": True})
        except Exception as e:
            moved_deals.append({"id": did, "ok": False, "error": str(e)})

    slp_repoint_error = None
    try:
        await ac_post(f"customObjects/records/{SLP_SCHEMA}", {
            "record": {"id": slp_match["id"], "fields": raw_fields, "relationships": {"account": [new_account_id]}}
        })
    except Exception as e:
        slp_repoint_error = str(e)

    return {
        "ok": True,
        "new_account_id": new_account_id,
        "new_account_name": final_name,
        "new_account_url": ac_account_url(new_account_id),
        "moved_contacts": moved_contacts,
        "moved_deals": moved_deals,
        "cf_errors": cf_errors,
        "slp_repoint_error": slp_repoint_error,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Webhook: deal-created → append row to SharePoint Excel
# ─────────────────────────────────────────────────────────────────────────────

_GRAPH_TOKEN_CACHE: dict = {}

async def _get_graph_token() -> str:
    """Get a Microsoft Graph access token via client credentials flow, with caching."""
    now = _time.time()
    if _GRAPH_TOKEN_CACHE.get("token") and now < _GRAPH_TOKEN_CACHE.get("expires", 0) - 60:
        return _GRAPH_TOKEN_CACHE["token"]

    tenant = _AZ_TENANT_ID if _AZ_TENANT_ID and _AZ_TENANT_ID != "common" else os.getenv("AZURE_TENANT_ID", "")
    url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
    async with httpx.AsyncClient() as client:
        resp = await client.post(url, data={
            "grant_type":    "client_credentials",
            "client_id":     _AZ_CLIENT_ID,
            "client_secret": _AZ_CLIENT_SEC,
            "scope":         "https://graph.microsoft.com/.default",
        })
        resp.raise_for_status()
        data = resp.json()
    token = data["access_token"]
    _GRAPH_TOKEN_CACHE["token"]   = token
    _GRAPH_TOKEN_CACHE["expires"] = now + int(data.get("expires_in", 3600))
    return token


async def _graph_get(path: str) -> dict:
    token = await _get_graph_token()
    async with httpx.AsyncClient() as client:
        r = await client.get(f"https://graph.microsoft.com/v1.0{path}",
                             headers={"Authorization": f"Bearer {token}"}, timeout=30)
        r.raise_for_status()
        return r.json()


async def _graph_put(path: str, data: bytes, content_type: str = "application/octet-stream") -> dict:
    token = await _get_graph_token()
    async with httpx.AsyncClient() as client:
        r = await client.put(f"https://graph.microsoft.com/v1.0{path}",
                             headers={"Authorization": f"Bearer {token}", "Content-Type": content_type},
                             content=data, timeout=60)
        r.raise_for_status()
        return r.json()


async def _graph_post(path: str, body: dict) -> dict:
    token = await _get_graph_token()
    async with httpx.AsyncClient() as client:
        r = await client.post(f"https://graph.microsoft.com/v1.0{path}",
                              headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                              json=body, timeout=30)
        r.raise_for_status()
        return r.json()


async def _graph_patch(path: str, body: dict) -> dict:
    token = await _get_graph_token()
    async with httpx.AsyncClient() as client:
        r = await client.patch(f"https://graph.microsoft.com/v1.0{path}",
                               headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                               json=body, timeout=30)
        r.raise_for_status()
        return r.json()


_SP_SITE_ID: str = ""
_SP_DRIVE_ID: str = ""
_SP_FILE_ID: str = ""

_DEAL_TRACKER_HEADERS = [
    "Date", "Deal ID", "Account Name", "Pipeline", "Stage", "Status",
    "Deal Description", "Deal Owner", "Lead Source", "Docs Packet Status",
    "Dealer ID", "Dealer Program", "Dealer Activation Timestamp",
    "Contact Name", "Contact Phone", "Contact Email",
]
_SP_HOSTNAME  = "microfllc.sharepoint.com"
_SP_FOLDER    = "2024 Marketing Management"
_SP_FILENAME  = "Deal Tracker.xlsx"


async def _ensure_sp_ids():
    """Resolve and cache SharePoint site ID, drive ID, and file ID."""
    global _SP_SITE_ID, _SP_DRIVE_ID, _SP_FILE_ID

    if not _SP_SITE_ID:
        # DRR is a document library on the root site, not a subsite
        site = await _graph_get(f"/sites/{_SP_HOSTNAME}")
        _SP_SITE_ID = site["id"]

    if not _SP_DRIVE_ID:
        drives = await _graph_get(f"/sites/{_SP_SITE_ID}/drives")
        # Match by webUrl path containing /DRR (URL path may differ from display name)
        for d in drives.get("value", []):
            web_url = d.get("webUrl", "")
            name = d.get("name", "")
            if name.upper() == "DRR" or web_url.rstrip("/").upper().endswith("/DRR"):
                _SP_DRIVE_ID = d["id"]
                break
        if not _SP_DRIVE_ID:
            detail = [(d.get("name"), d.get("webUrl")) for d in drives.get("value", [])]
            raise RuntimeError(f"DRR library not found. Available (name, webUrl): {detail}")

    if not _SP_FILE_ID:
        try:
            item = await _graph_get(
                f"/drives/{_SP_DRIVE_ID}/root:/{_SP_FOLDER}/{_SP_FILENAME}"
            )
            _SP_FILE_ID = item["id"]
        except httpx.HTTPStatusError as e:
            if e.response.status_code == 404:
                _SP_FILE_ID = ""   # will create below
            else:
                raise


async def _ensure_workbook() -> str:
    """Ensure Deal Tracker.xlsx exists. Returns file ID."""
    global _SP_FILE_ID
    await _ensure_sp_ids()

    if not _SP_FILE_ID:
        # Create a minimal XLSX with headers using openpyxl
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Deals"
            ws.append(_DEAL_TRACKER_HEADERS)
            buf = io.BytesIO()
            wb.save(buf)
            content = buf.getvalue()
        except ImportError:
            # Fallback: use a pre-built minimal xlsx bytes if openpyxl not available
            raise RuntimeError("openpyxl is required to create the Excel file. Install it on the server.")

        item = await _graph_put(
            f"/drives/{_SP_DRIVE_ID}/root:/{_SP_FOLDER}/{_SP_FILENAME}:/content",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        _SP_FILE_ID = item["id"]

    return _SP_FILE_ID


async def _append_deal_row(row: list):
    """Upsert one row: if the deal ID already exists, overwrite it if the new data is more complete; otherwise append."""
    import openpyxl
    file_id = await _ensure_workbook()
    token = await _get_graph_token()

    # Download current file
    async with httpx.AsyncClient() as client:
        r = await client.get(
            f"https://graph.microsoft.com/v1.0/drives/{_SP_DRIVE_ID}/items/{file_id}/content",
            headers={"Authorization": f"Bearer {token}"},
            follow_redirects=True,
            timeout=30,
        )
        r.raise_for_status()
        file_bytes = r.content

    # Load, ensure headers, upsert row
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb["Deals"] if "Deals" in wb.sheetnames else wb.active
    # Update header row if it's missing or outdated
    current_headers = [ws.cell(1, c).value for c in range(1, len(_DEAL_TRACKER_HEADERS) + 1)]
    if current_headers != _DEAL_TRACKER_HEADERS:
        for c, h in enumerate(_DEAL_TRACKER_HEADERS, 1):
            ws.cell(1, c).value = h

    # Deal ID is column B (index 1 in row list, column 2 in sheet)
    new_deal_id = str(row[1]) if len(row) > 1 else None
    new_populated = sum(1 for v in row if v)

    existing_row_num = None
    if new_deal_id:
        for r_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(r_idx, 2).value
            if cell_val is not None and str(cell_val) == new_deal_id:
                # Count populated cells in existing row
                existing_populated = sum(
                    1 for c in range(1, len(_DEAL_TRACKER_HEADERS) + 1)
                    if ws.cell(r_idx, c).value not in (None, "")
                )
                if new_populated > existing_populated:
                    # Overwrite with more complete data
                    existing_row_num = r_idx
                else:
                    # Existing row is equally or more complete — skip
                    print(f"[deal-tracker] ⏭ deal={new_deal_id} already exists with {existing_populated} fields, skipping duplicate")
                    return
                break

    if existing_row_num:
        for c_idx, val in enumerate(row, 1):
            ws.cell(existing_row_num, c_idx).value = val
        print(f"[deal-tracker] ↺ deal={new_deal_id} overwriting row {existing_row_num} with more complete data")
    else:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    new_bytes = buf.getvalue()

    # Try checkout to clear any co-authoring lock, then upload, then checkin
    token = await _get_graph_token()
    base = f"https://graph.microsoft.com/v1.0/drives/{_SP_DRIVE_ID}/items/{file_id}"
    async with httpx.AsyncClient() as client:
        # Attempt checkout (ignore errors — file may not support it)
        await client.post(f"{base}/checkout",
                          headers={"Authorization": f"Bearer {token}"}, timeout=15)

        # Upload with retry for residual 423 locks
        for attempt in range(6):
            r = await client.put(
                f"{base}/content",
                headers={
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                content=new_bytes,
                timeout=60,
            )
            if r.status_code == 423 and attempt < 5:
                wait = 15 * (attempt + 1)
                print(f"[webhook] file locked, retrying in {wait}s (attempt {attempt+1}/5)")
                await asyncio.sleep(wait)
                continue
            r.raise_for_status()
            break

        # Checkin so the file is visible/unlocked
        await client.post(f"{base}/checkin",
                          headers={"Authorization": f"Bearer {token}",
                                   "Content-Type": "application/json"},
                          json={"checkInComment": "Deal Tracker row added", "checkInType": "1"},
                          timeout=15)


def _parse_bracket_form(raw_body: bytes) -> dict:
    """Parse AC webhook form-encoded payload like deal[id]=123&deal[field][35]=abc into nested dict."""
    from urllib.parse import parse_qsl
    flat = dict(parse_qsl(raw_body.decode("utf-8", errors="replace")))
    result: dict = {}
    for key, val in flat.items():
        # e.g. "deal[field][35]" → ["deal","field","35"]
        parts = key.replace("]", "").split("[")
        node = result
        for p in parts[:-1]:
            node = node.setdefault(p, {})
        node[parts[-1]] = val
    return result


@app.post("/webhook/reset-sp-file")
async def webhook_reset_sp_file():
    """Delete the cached Deal Tracker file and clear cache so it gets recreated fresh."""
    global _SP_FILE_ID
    try:
        await _ensure_sp_ids()
        if _SP_FILE_ID:
            token = await _get_graph_token()
            async with httpx.AsyncClient() as client:
                r = await client.delete(
                    f"https://graph.microsoft.com/v1.0/drives/{_SP_DRIVE_ID}/items/{_SP_FILE_ID}",
                    headers={"Authorization": f"Bearer {token}"},
                    timeout=30,
                )
            deleted_id = _SP_FILE_ID
            _SP_FILE_ID = ""
            return {"ok": True, "deleted_file_id": deleted_id, "status_code": r.status_code}
        return {"ok": True, "message": "No file cached, nothing to delete"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/webhook/debug-sp")
async def webhook_debug_sp():
    """Debug: show resolved SharePoint IDs and file location."""
    try:
        await _ensure_sp_ids()
        drives = await _graph_get(f"/sites/{_SP_SITE_ID}/drives")
        drive_info = next((d for d in drives.get("value", []) if d["id"] == _SP_DRIVE_ID), {})
        result = {
            "site_id": _SP_SITE_ID,
            "drive_id": _SP_DRIVE_ID,
            "drive_name": drive_info.get("name"),
            "drive_webUrl": drive_info.get("webUrl"),
            "file_id": _SP_FILE_ID,
        }
        if _SP_FILE_ID:
            item = await _graph_get(f"/drives/{_SP_DRIVE_ID}/items/{_SP_FILE_ID}")
            result["file_name"] = item.get("name")
            result["file_webUrl"] = item.get("webUrl")
            result["file_path"] = item.get("parentReference", {}).get("path")
            result["file_sharepoint_url"] = item.get("webUrl")
        # List children of the target folder to see what's actually there
        try:
            folder = await _graph_get(f"/drives/{_SP_DRIVE_ID}/root:/{_SP_FOLDER}")
            result["folder_id"] = folder.get("id")
            result["folder_webUrl"] = folder.get("webUrl")
            children = await _graph_get(f"/drives/{_SP_DRIVE_ID}/items/{folder['id']}/children")
            result["folder_contents"] = [{"name": c.get("name"), "webUrl": c.get("webUrl")} for c in children.get("value", [])]
        except Exception as fe:
            result["folder_error"] = str(fe)
        return result
    except Exception as e:
        return {"error": str(e)}


async def _fetch_full_deal(deal_id: str) -> dict:
    """Fetch all deal data from AC API. Returns a flat dict of all fields."""

    async def safe(coro):
        try:
            return await coro
        except Exception:
            return {}

    # Parallel: deal details + custom field data
    deal_resp, cf_resp = await asyncio.gather(
        safe(ac_get(f"deals/{deal_id}")),
        safe(ac_get(f"deals/{deal_id}/dealCustomFieldData")),
    )

    deal = deal_resp.get("deal", {})

    # Build custom field map: metaId → value
    cf_by_id: dict = {}
    for cf in cf_resp.get("dealCustomFieldData", []):
        mid = str(cf.get("customFieldId") or cf.get("dealCustomFieldMetumId") or "")
        cf_by_id[mid] = cf.get("fieldValue", "") or ""

    # Parallel: pipeline, stage, owner, contact
    pipeline_id = deal.get("group", "")
    stage_id    = deal.get("stage", "")
    owner_id    = deal.get("owner", "")
    contact_id  = deal.get("contact", "")

    async def _empty(): return {}

    pipeline_resp, stage_resp, owner_resp, contact_resp = await asyncio.gather(
        safe(ac_get(f"dealGroups/{pipeline_id}")) if pipeline_id else _empty(),
        safe(ac_get(f"dealStages/{stage_id}"))    if stage_id    else _empty(),
        safe(ac_get(f"users/{owner_id}"))          if owner_id    else _empty(),
        safe(ac_get(f"contacts/{contact_id}"))     if contact_id  else _empty(),
    )

    pipeline_name = pipeline_resp.get("dealGroup", {}).get("title", "")
    stage_name    = stage_resp.get("dealStage", {}).get("title", "")

    user = owner_resp.get("user", {})
    owner_name = f"{user.get('firstName','')} {user.get('lastName','')}".strip()

    contact = contact_resp.get("contact", {})
    contact_name  = f"{contact.get('firstName','')} {contact.get('lastName','')}".strip()
    contact_email = contact.get("email", "")
    contact_phone = contact.get("phone", "")

    status_map = {"0": "Open", "1": "Won", "2": "Lost"}
    status = status_map.get(str(deal.get("status", "0")), str(deal.get("status", "")))

    return {
        "deal_id":       deal.get("id", deal_id),
        "account_name":  deal.get("title", ""),
        "pipeline":      pipeline_name,
        "stage":         stage_name,
        "status":        status,
        "description":   deal.get("description", ""),
        "owner":         owner_name,
        "lead_source":   cf_by_id.get("39", ""),   # CF39
        "docs_packet":   cf_by_id.get("38", ""),   # CF38
        "dealer_id":     cf_by_id.get("35", ""),   # CF35
        "dealer_program":cf_by_id.get("45", ""),   # CF45
        "activation_ts": cf_by_id.get("34", ""),   # CF34
        "contact_name":  contact_name,
        "contact_phone": contact_phone,
        "contact_email": contact_email,
    }


async def _process_deal_to_sharepoint(deal_id: str):
    """Background task: fetch full deal from AC and write row to SharePoint."""
    try:
        # Wait for AC automations to finish populating custom fields
        await asyncio.sleep(60)
        d = await _fetch_full_deal(deal_id)
        # If key fields still empty, wait another 60s and retry once
        if not d.get("dealer_id") and not d.get("lead_source"):
            print(f"[deal-tracker] deal={deal_id} key fields empty, retrying in 60s")
            await asyncio.sleep(60)
            d = await _fetch_full_deal(deal_id)
        row_date = datetime.utcnow().strftime("%Y-%m-%d")
        row = [
            row_date,
            d["deal_id"],
            d["account_name"],
            d["pipeline"],
            d["stage"],
            d["status"],
            d["description"],
            d["owner"],
            d["lead_source"],
            d["docs_packet"],
            d["dealer_id"],
            d["dealer_program"],
            d["activation_ts"],
            d["contact_name"],
            d["contact_phone"],
            d["contact_email"],
        ]
        await _append_deal_row(row)
        print(f"[deal-tracker] ✓ deal={deal_id} acct={d['account_name']} dealer={d['dealer_id']} program={d['dealer_program']}")
    except Exception as e:
        import traceback as _tb
        print(f"[deal-tracker] ✗ deal={deal_id} {e}\n{_tb.format_exc()}")


def _check_webhook_token(request: _Request) -> None:
    if not _WEBHOOK_TOKEN:
        return   # local/dev fallback only
    bearer = request.headers.get("Authorization", "")
    header_token = request.headers.get("X-Webhook-Token", "")
    query_token = request.query_params.get("token", "")
    candidates = []
    if bearer.startswith("Bearer "):
        candidates.append(bearer.removeprefix("Bearer ").strip())
    candidates.extend([header_token, query_token])
    if not any(t and secrets.compare_digest(t, _WEBHOOK_TOKEN) for t in candidates):
        raise HTTPException(status_code=401, detail="Invalid webhook token")


@app.post("/webhook/deal-created")
async def webhook_deal_created(request: _Request, background_tasks: BackgroundTasks):
    """
    Receives ActiveCampaign deal-created webhook. Returns immediately,
    processes the SharePoint write in the background.
    """
    _check_webhook_token(request)
    try:
        body = await request.body()
        data = _parse_bracket_form(body)
        deal_id = data.get("deal", {}).get("id", "")

        if not deal_id:
            print(f"[webhook/deal-created] no deal id in payload: {body[:300]}")
            return {"ok": False, "error": "no deal id"}

        background_tasks.add_task(_process_deal_to_sharepoint, deal_id)
        return {"ok": True, "deal_id": deal_id}

    except Exception as e:
        import traceback as _tb
        print(f"[webhook/deal-created] ✗ {e}\n{_tb.format_exc()}")
        return JSONResponse(status_code=200, content={"ok": False, "error": str(e)})


@app.get("/reports/slp-health")
async def slp_health_page(user=Depends(require_auth)):
    return FileResponse("static/reports/slp-health.html")


# ── Account Summary Report ────────────────────────────────────────────────────

@app.get("/reports/account-summary")
async def account_summary_page(user=Depends(require_auth)):
    return FileResponse("static/reports/account-summary.html")

@app.get("/api/reports/account-summary")
async def account_summary_report(
    owner:      Optional[str] = Query(None),
    acct_type:  Optional[str] = Query(None),
    status:     Optional[str] = Query(None),
    region:     Optional[str] = Query(None),
    bdr:        Optional[str] = Query(None),
    channel:    Optional[str] = Query(None),
    user=Depends(require_auth),
):
    """SLP-based account summary report."""
    rows = []
    for aid, name in _account_to_name.items():
        typ = _account_to_type.get(aid, "")
        if acct_type and typ.lower() != acct_type.lower():
            continue
        acct_status = _account_to_status.get(aid, "")
        if status and acct_status.lower() != status.lower():
            continue
        owner_id = _account_to_owner.get(aid, "")
        if owner and owner_id != owner:
            continue
        assigned_bdr = _account_to_bdr.get(aid, "")
        if bdr and assigned_bdr.lower() != bdr.lower():
            continue
        acct_channel = _account_to_platform.get(aid, "")
        if channel and acct_channel.lower() != channel.lower():
            continue

        dealer_id       = _account_to_slp_dealer.get(aid, "") or _account_to_dealer.get(aid, "")
        activation_date = _account_to_activation_date.get(aid, "")
        slp_states      = _account_to_slp_states.get(aid, "")
        react           = _account_to_contractor_reactivation.get(aid, "")
        react_date      = _account_to_reactivation_date.get(aid, "")
        dba             = _account_to_dba.get(aid, "")
        legal_name      = _account_to_legal_name.get(aid, "")
        revenue         = _account_to_revenue.get(aid, "")
        strat_partners  = _account_to_strategic_partners.get(aid, "")
        region_val      = _account_to_region.get(aid, "")
        if region and region_val.lower() != region.lower():
            continue

        rows.append({
            "id":                   aid,
            "name":                 name,
            "owner_name":           _user_id_to_name.get(owner_id, owner_id or ""),
            "status":               acct_status,
            "bdr":                  assigned_bdr,
            "dba_name":             dba,
            "doing_business_states":slp_states,
            "dealer_id":            dealer_id,
            "activation_date":      activation_date,
            "contractor_reactivation": "Yes" if react else "",
            "reactivation_date":    react_date[:10] if react_date else "",
            "region":               region_val,
            "legal_name":           legal_name,
            "revenue":              revenue,
            "strategic_partners":   strat_partners,
            "channel":              acct_channel,
        })

    rows.sort(key=lambda r: r["name"].lower())

    # Build filter option lists
    owners   = sorted({r["owner_name"] for r in rows if r["owner_name"]})
    statuses = sorted({r["status"]     for r in rows if r["status"]})
    regions  = sorted({r["region"]     for r in rows if r["region"]})
    bdrs     = sorted({r["bdr"]        for r in rows if r["bdr"]})
    channels = sorted({r["channel"]    for r in rows if r["channel"]})

    return {"rows": rows, "owners": owners, "statuses": statuses,
            "regions": regions, "bdrs": bdrs, "channels": channels,
            "total": len(rows)}


@app.get("/api/reports/account-summary/csv")
async def account_summary_csv(
    owner:      Optional[str] = Query(None),
    acct_type:  Optional[str] = Query(None),
    status:     Optional[str] = Query(None),
    region:     Optional[str] = Query(None),
    bdr:        Optional[str] = Query(None),
    channel:    Optional[str] = Query(None),
    user=Depends(require_auth),
):
    data = await account_summary_report(owner=owner, acct_type=acct_type,
                                        status=status, region=region,
                                        bdr=bdr, channel=channel, user=user)
    import io, csv as _csv
    buf = io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["Account Name","Account Owner","Account Status","Assigned BDR",
                "DBA Name","Doing Business In States","Parent Dealer ID",
                "Partner Activation Date","Contractor Reactivation","Sales Region",
                "Reactivation Date","Account Name (Legal Business Name)",
                "Annual Revenue","Strategic Partners","Channel"])
    for r in data["rows"]:
        w.writerow([r["name"], r["owner_name"], r["status"], r["bdr"],
                    r["dba_name"], r["doing_business_states"], r["dealer_id"],
                    r["activation_date"], r["contractor_reactivation"], r["region"],
                    r["reactivation_date"], r["legal_name"], r["revenue"],
                    r["strategic_partners"], r["channel"]])
    from fastapi.responses import StreamingResponse
    buf.seek(0)
    fn = f"account_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(iter([buf.getvalue()]),
                             media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={fn}"})


# ── AM Activity Dashboard ─────────────────────────────────────────────────────

@app.get("/reports/am-activity")
async def am_activity_page(user=Depends(require_auth)):
    return FileResponse("static/reports/am-activity.html")


@app.get("/api/reports/am-activity")
async def am_activity_report(
    owner: Optional[str] = Query(None, description="Filter by AC owner user ID"),
    acct_type: Optional[str] = Query("Contractor", description="Filter by Account Type; empty = all"),
    bdr: Optional[str] = Query(None, description="Filter by Assigned BDR name"),
    channel: Optional[str] = Query(None, description="Filter by channel/platform"),
    user=Depends(require_auth),
):
    """
    Returns all accounts (filtered by owner/type) with last app date, last RPA date,
    and days-since for each. Sorted most-stale first (nulls at top).
    Also returns a managers list for the dropdown.
    """
    today = date.today()

    def clean_date(date_str: str) -> str:
        """
        Normalize AC date values to YYYY-MM-DD.
        AC sometimes stores MM-DD-YY input (e.g. '01-01-22') by treating
        the 2-digit year as the 4-digit year, producing '0001-01-22'.
        When year < 13 we reverse that: stored-year=month, stored-month=day,
        stored-day=2-digit-year → reconstruct the real date.
        """
        if not date_str:
            return ""
        try:
            d = date.fromisoformat(str(date_str)[:10])
            if d.year >= 2000:
                return "" if d.year > today.year + 1 else str(date_str)[:10]
            if 1 <= d.year <= 12:
                # Mangled MM-DD-YY: year slot = real month, month slot = real day,
                # day slot = real 2-digit year
                reconstructed = date(2000 + d.day, d.year, d.month)
                return reconstructed.isoformat() if reconstructed <= today else ""
            return ""  # garbage year (13–1999)
        except Exception:
            return ""

    def days_since(date_str: str) -> Optional[int]:
        if not date_str:
            return None
        try:
            return (today - date.fromisoformat(str(date_str)[:10])).days
        except Exception:
            return None

    accounts = []
    for aid, name in _account_to_name.items():
        # Account type filter
        typ = _account_to_type.get(aid, "")
        if acct_type and typ.lower() != acct_type.lower():
            continue

        owner_id   = _account_to_owner.get(aid, "")
        if owner and owner_id != owner:
            continue

        assigned_bdr = _account_to_bdr.get(aid, "")
        if bdr and assigned_bdr.lower() != bdr.lower():
            continue

        acct_channel = _account_to_platform.get(aid, "")
        if channel and acct_channel.lower() != channel.lower():
            continue

        last_app   = clean_date(_account_to_last_app.get(aid, ""))
        last_rpa   = clean_date(_account_to_last_rpa.get(aid, ""))
        # Use SLP dealer-id (authoritative) — falls back to account CF18 if SLP has none
        dealer_id  = _account_to_slp_dealer.get(aid, "") or _account_to_dealer.get(aid, "")
        region     = _account_to_region.get(aid, "")
        owner_name = _user_id_to_name.get(owner_id, owner_id or "—")

        days_app   = days_since(last_app)
        days_rpa   = days_since(last_rpa)

        accounts.append({
            "id":            aid,
            "name":          name,
            "owner_id":      owner_id,
            "owner_name":    owner_name,
            "bdr":           assigned_bdr,
            "dealer_id":     dealer_id,
            "region":        region,
            "channel":       acct_channel,
            "last_app_date": last_app,
            "days_since_app": days_app,
            "last_rpa_date": last_rpa,
            "days_since_rpa": days_rpa,
        })

    # Sort: null dates first (never touched), then oldest to newest by max staleness
    def _sort_key(a):
        da = a["days_since_app"] if a["days_since_app"] is not None else 99999
        dr = a["days_since_rpa"] if a["days_since_rpa"] is not None else 99999
        return -max(da, dr)

    accounts.sort(key=_sort_key)

    # Build managers + BDR lists from ALL accounts so dropdowns are always complete
    mgrs: dict = {}
    bdrs: set  = set()
    for aid in _account_to_name:
        uid = _account_to_owner.get(aid, "")
        if uid and uid not in mgrs:
            mgrs[uid] = _user_id_to_name.get(uid, uid)
        b = _account_to_bdr.get(aid, "")
        if b:
            bdrs.add(b)

    managers = sorted(
        [{"id": k, "name": v} for k, v in mgrs.items()],
        key=lambda x: x["name"]
    )

    channels = sorted({a["channel"] for a in accounts if a["channel"]})

    return {
        "accounts":  accounts,
        "total":     len(accounts),
        "managers":  managers,
        "bdrs":      sorted(bdrs),
        "channels":  channels,
    }


@app.get("/api/reports/am-activity/last-contacted")
async def am_last_contacted(user=Depends(require_auth)):
    """
    Returns {account_id: "YYYY-MM-DD"} for the most recent contact per account.
    Served from in-memory cache (refreshes every 30 min). If cache is empty
    (e.g. first minute after deploy) triggers a background build and returns empty.
    """
    if not _lc_cache:
        asyncio.create_task(_refresh_lc_cache())
    age = int(_time.time() - _lc_cache_ts) if _lc_cache_ts else None
    # Flatten to {aid: date} and {aid: type} for frontend
    dates = {aid: v["date"] for aid, v in _lc_cache.items()}
    types = {aid: v["type"] for aid, v in _lc_cache.items()}
    return {"last_contacted": dates, "last_contacted_type": types, "cache_age_seconds": age}


# ── Verdata Active Report ─────────────────────────────────────────────────────

@app.get("/api/report/verdata-active")
async def verdata_active_report(
    format: str = Query("json"),
    user=Depends(require_auth),
):
    return await _verdata_report(status_filter="active", format=format)


@app.get("/api/report/verdata-inactive")
async def verdata_inactive_report(
    format: str = Query("json"),
    user=Depends(require_auth),
):
    return await _verdata_report(status_filter="deactivated", format=format)


async def _verdata_report(status_filter: str, format: str):
    """Verdata report filtered by account status — fully in-memory."""
    slp_records = await get_slp_cache()

    def _slp_field(slp, fid):
        for f in slp.get("fields", []):
            if f.get("id") == fid:
                return (f.get("value") or "").strip()
        return ""

    # Build activation-date index keyed by dealer_id (the join key).
    # Take the earliest contractor-activated-date across all SLPs for that dealer,
    # regardless of platform or status.
    act_dates: dict = {}   # dealer_id -> earliest activation date str

    for slp in slp_records:
        dealer_id = _slp_field(slp, "dealer-id")
        if not dealer_id:
            continue
        act_date = _slp_field(slp, "contractor-activated-date")
        if not act_date:
            continue
        act_str = str(act_date)[:10]
        if len(act_str) < 10:
            continue
        if dealer_id not in act_dates or act_str < act_dates[dealer_id]:
            act_dates[dealer_id] = act_str

    records = []
    for account_id, status in _account_to_status.items():
        if status.strip().lower() != status_filter.lower():
            continue

        dealer_id = _account_to_dealer.get(account_id, "")
        acct_name = _account_to_name.get(account_id, "")
        dba_name  = _account_to_dba.get(account_id, "")
        tax_id    = _account_to_tax_id.get(account_id, "")
        website   = _account_to_website.get(account_id, "")
        address   = _account_to_address.get(account_id, "")
        city      = _account_to_city.get(account_id, "")
        state     = _account_to_state_prov.get(account_id, "")
        zip_code  = _account_to_zip.get(account_id, "")

        records.append({
            "Dealer ID":          dealer_id,
            "Account Name":       acct_name,
            "DBA Name":           dba_name,
            "RTO Activation Date": act_dates.get(dealer_id, ""),
            "Account Status":     status,
            "Vendor Tax-ID":      tax_id,
            "Website":            website,
            "Physical Address":   address,
            "Physical City":      city,
            "Physical State":     state,
            "Physical Zip":       zip_code,
        })

    records.sort(key=lambda r: (r["Account Name"] or "").lower())

    if format == "csv":
        output = io.StringIO()
        if records:
            writer = csv.DictWriter(output, fieldnames=list(records[0].keys()))
            writer.writeheader()
            writer.writerows(records)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="verdata_{status_filter}.csv"'},
        )

    return {"count": len(records), "records": records}


@app.get("/api/reports/slp-health")
async def slp_health_report(
    issue: str,  # required: no_dealer_id | no_status | no_platform | no_date | id_mismatch
    region: str = None,
    program: str = None,
    user=Depends(require_auth),
):
    VALID_ISSUES = {"no_dealer_id", "no_status", "no_platform", "no_date", "id_mismatch"}
    if issue not in VALID_ISSUES:
        raise HTTPException(status_code=400, detail=f"issue must be one of {VALID_ISSUES}")

    SLUG_MAP = {
        "no_dealer_id": "dealer-id",
        "no_status": "slp-status-detail",
        "no_platform": "channel",
        "no_date": "contractor-activated-date",
        "id_mismatch": "dealer-id",
    }

    def get_field(slp, slug):
        for f in slp.get("fields", []):
            if f.get("id") == slug:
                return f.get("value") or ""
        return ""

    # Use shared SLP cache — no AC calls needed
    all_slps = await get_slp_cache()

    # Filter in-memory first — only keep SLPs matching the issue
    def has_issue(slp):
        slug = SLUG_MAP[issue]
        val = get_field(slp, slug)
        if issue == "id_mismatch":
            return bool(val)  # has a dealer-id — mismatch checked below
        return not val

    candidates = [s for s in all_slps if has_issue(s)]

    # Optional program filter
    if program:
        candidates = [s for s in candidates if get_field(s, "channel") == program]

    # Build results — all account data from in-memory index (no AC calls)
    records = []
    for slp in candidates:
        account_id = (slp.get("relationships", {}).get("account") or [None])[0]
        acct_region   = _account_to_region.get(str(account_id), "") if account_id else ""
        acct_name     = _account_to_name.get(str(account_id), "Unknown") if account_id else "Unknown"
        cf18          = _account_to_dealer.get(str(account_id), "") if account_id else ""

        if region and norm(acct_region) != norm(region):
            continue

        slp_dealer_id = get_field(slp, "dealer-id")

        # For id_mismatch, skip if they actually match
        if issue == "id_mismatch" and slp_dealer_id == cf18:
            continue

        records.append({
            "slp_id": slp["id"],
            "account_id": account_id,
            "account_name": acct_name,
            "channel": get_field(slp, "channel"),
            "region": acct_region,
            "dealer_id": slp_dealer_id,
            "cf18": cf18,
            "status_detail": get_field(slp, "slp-status-detail"),
            "activation_date": get_field(slp, "contractor-activated-date"),
            "issue": issue,
        })

    return {
        "issue": issue,
        "total_slps_scanned": len(all_slps),
        "total_matching": len(records),
        "records": records,
    }


@app.get("/reports/contractor-states")
async def contractor_states_page(user=Depends(require_auth)):
    return FileResponse("static/reports/contractor-states.html")

@app.get("/api/reports/contractor-states")
async def contractor_states_report(
    acct_state: str = None,
    biz_state: str = None,
    program: str = None,
    user=Depends(require_auth),
):
    def get_field(slp, fid):
        for f in slp.get("fields", []):
            if f.get("id") == fid:
                return f.get("value") or ""
        return ""

    # Use shared SLP cache — no AC calls needed
    all_slps = list(await get_slp_cache())

    if program:
        all_slps = [s for s in all_slps if get_field(s, "channel") == program]

    # Build records — all account data from in-memory index
    records = []
    for slp in all_slps:
        aid = (slp.get("relationships", {}).get("account") or [None])[0]
        if not aid:
            continue
        aid_str   = str(aid)
        acct_type = _account_to_type.get(aid_str, "")
        if norm(acct_type) != "contractor":
            continue
        acct_state_val = _account_to_state_prov.get(aid_str, "")
        if acct_state and norm(acct_state_val) != norm(acct_state):
            continue
        dbi = get_field(slp, "doing-business-in-states")
        if biz_state:
            states_in = [s.strip().upper() for s in (dbi or "").replace(";", ",").split(",") if s.strip()]
            if biz_state.upper() not in states_in:
                continue
        records.append({
            "account_id":               aid,
            "account_name":             _account_to_name.get(aid_str, ""),
            "account_state":            acct_state_val,
            "dealer_id":                get_field(slp, "dealer-id"),
            "channel":                  get_field(slp, "channel"),
            "doing_business_in_states": dbi,
        })

    records.sort(key=lambda x: x["account_name"])
    unique_accounts = len({r["account_id"] for r in records})

    return {
        "total_slps": len(all_slps),
        "total_matching": len(records),
        "unique_accounts": unique_accounts,
        "records": records,
    }


# ── Account + SLP Parent-Child Report ────────────────────────────────────────

@app.get("/reports/account-slp")
async def account_slp_page(user=Depends(require_auth)):
    return FileResponse("static/reports/account-slp.html")

@app.get("/api/reports/account-slp")
async def account_slp_report(user=Depends(require_auth)):
    def get_field(slp, fid):
        for f in slp.get("fields", []):
            if f.get("id") == fid:
                return f.get("value") or ""
        return ""

    # ── 1. Use shared SLP cache ─────────────────────────────────────────────
    all_slps = list(await get_slp_cache())
    cache_age = round(_time.time() - _slp_cache_ts) if _slp_cache_ts else None

    # ── 2. Group SLPs by account_id ─────────────────────────────────────────
    slps_by_account: dict = defaultdict(list)
    for slp in all_slps:
        aid = (slp.get("relationships", {}).get("account") or [None])[0]
        if aid:
            slps_by_account[str(aid)].append(slp)

    # ── 3. Build parent-child structure — all account data from index ────────
    index_ready = _dealer_index_ts > 0
    accounts_out = []
    for aid_str, slps in slps_by_account.items():
        owner_uid  = _account_to_owner.get(aid_str, "")
        owner_name = _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else ""
        slp_list = []
        for slp in slps:
            slp_list.append({
                "slp_id":            slp["id"],
                "channel":           get_field(slp, "channel"),
                "status":            get_field(slp, "slp-status-detail"),
                "activation_date":   get_field(slp, "contractor-activated-date"),
                "oracle_ids":        get_field(slp, "oracle-producer-ids"),
                "bdr":               get_field(slp, "assigned-bdr"),
                "doing_business_in": get_field(slp, "doing-business-in-states"),
                "dealer_id":         get_field(slp, "dealer-id"),
            })
        slp_list.sort(key=lambda x: x["channel"])
        accounts_out.append({
            "account_id": aid_str,
            "name":       _account_to_name.get(aid_str, ""),
            "type":       _account_to_type.get(aid_str, ""),
            "state":      _account_to_state_prov.get(aid_str, ""),
            "region":     _account_to_region.get(aid_str, ""),
            "dealer_id":  _account_to_dealer.get(aid_str, ""),
            "owner":      owner_name,
            "slps":       slp_list,
        })

    accounts_out.sort(key=lambda x: x["name"].upper())

    # Sample IDs for debugging key-format issues
    sample_slp_ids  = list(slps_by_account.keys())[:3]
    sample_idx_ids  = list(_account_to_name.keys())[:3]

    return {
        "total_accounts":    len(accounts_out),
        "total_slps":        len(all_slps),
        "cache_age_seconds": cache_age,
        "index_ready":       index_ready,
        "index_size":        len(_account_to_name),
        "_debug_slp_ids":    sample_slp_ids,
        "_debug_idx_ids":    sample_idx_ids,
        "accounts":          accounts_out,
    }


# ── Parent-Child Report (Account-first, all account types) ───────────────────

@app.get("/reports/parent-child")
async def parent_child_page(user=Depends(require_auth)):
    return FileResponse("static/reports/parent-child.html")

@app.get("/api/reports/parent-child")
async def parent_child_report(
    acct_type:  str  = None,
    region:     str  = None,
    acct_state: str  = None,
    program:    str  = None,
    slp_status: str  = None,
    has_slps:   bool = None,   # None=all, True=only with SLPs, False=only without
    user=Depends(require_auth),
):
    def get_field(slp, fid):
        for f in slp.get("fields", []):
            if f.get("id") == fid:
                return f.get("value") or ""
        return ""

    # Build account→SLPs map from shared cache
    all_slps   = list(await get_slp_cache())
    cache_age  = round(_time.time() - _slp_cache_ts) if _slp_cache_ts else None

    slps_by_account: dict  = defaultdict(list)
    slps_by_account["UNASSIGNED"] = []   # guaranteed bucket for truly orphaned SLPs
    fallback_used          = 0
    unlinked_slps          = 0
    failed_matches         = 0
    missing_relationships  = 0

    for slp in all_slps:
        assigned = False

        # PRIMARY: robust relationship extraction
        aid = get_account_id(slp)
        if not aid:
            missing_relationships += 1
        if aid:
            slps_by_account[aid].append(slp)
            assigned = True

        # FALLBACK: dealer-id → _dealer_id_index (normalized matching)
        # NOTE: _dealer_id_index values are {"id": account_id, "name": ...} dicts
        def _resolve_entry(entry) -> str | None:
            """Extract account ID string from a _dealer_id_index entry (dict or plain str)."""
            if isinstance(entry, dict):
                return str(entry["id"]) if entry.get("id") else None
            return str(entry) if entry else None

        if not assigned:
            dealer_id      = get_field(slp, "dealer-id")
            dealer_id_norm = norm_id(dealer_id)
            acct_id        = None

            # 1) Exact match
            if dealer_id:
                acct_id = _resolve_entry(_dealer_id_index.get(dealer_id))

            # 2) Normalized match (strip leading zeros from SLP dealer-id)
            if not acct_id and dealer_id_norm:
                acct_id = _resolve_entry(_dealer_id_index.get(dealer_id_norm))

            # 3) Reverse-normalize index keys (strip leading zeros from stored keys)
            if not acct_id and dealer_id_norm:
                for k, v in _dealer_id_index.items():
                    if norm_id(k) == dealer_id_norm:
                        acct_id = _resolve_entry(v)
                        break

            if acct_id:
                slps_by_account[acct_id].append(slp)
                fallback_used += 1
                assigned = True
            else:
                if dealer_id:
                    failed_matches += 1
                unlinked_slps += 1

        # FINAL FALLBACK: bucket all truly unresolvable SLPs
        if not assigned:
            slps_by_account["UNASSIGNED"].append(slp)

    print(f"[PARENT-CHILD] total={len(all_slps)} missing_relationships={missing_relationships} "
          f"fallback_used={fallback_used} unlinked={unlinked_slps} failed_dealer_matches={failed_matches} "
          f"unassigned_slps={len(slps_by_account['UNASSIGNED'])}")

    def _build_slp_list(raw_slps):
        slp_list = []
        for slp in raw_slps:
            p = get_field(slp, "channel")
            s = get_field(slp, "slp-status-detail")
            if program    and norm(p) != norm(program):    continue
            if slp_status and norm(s) != norm(slp_status): continue
            slp_list.append({
                "slp_id":            slp["id"],
                "channel":           p,
                "status":            s,
                "activation_date":   get_field(slp, "contractor-activated-date"),
                "oracle_ids":        get_field(slp, "oracle-producer-ids"),
                "bdr":               get_field(slp, "assigned-bdr"),
                "doing_business_in": get_field(slp, "doing-business-in-states"),
                "dealer_id":         get_field(slp, "dealer-id"),
            })
        slp_list.sort(key=lambda x: x["channel"])
        return slp_list

    ALL_ACCOUNT_IDS = set(_account_to_name.keys()) | set(slps_by_account.keys())

    print(f"[PARENT-CHILD] rendered_accounts={len(ALL_ACCOUNT_IDS)}")
    print(f"[PARENT-CHILD] grouped_accounts={len(slps_by_account)}")

    accounts_out = []

    for aid in ALL_ACCOUNT_IDS:
        # UNASSIGNED bucket — skip account filters, always include if it has SLPs
        if aid == "UNASSIGNED":
            raw_slps = slps_by_account.get("UNASSIGNED", [])
            if not raw_slps:
                continue
            slp_list = _build_slp_list(raw_slps)
            if has_slps is False:
                continue
            accounts_out.append({
                "account_id": "UNASSIGNED",
                "name":       "⚠️ Unassigned SLPs",
                "type":       "",
                "state":      "",
                "region":     "",
                "dealer_id":  "",
                "owner":      "",
                "slp_count":  len(raw_slps),
                "slps":       slp_list,
            })
            continue

        row_name   = _account_to_name.get(aid) or f"Unknown Account ({aid})"
        row_type   = _account_to_type.get(aid, "")
        row_state  = _account_to_state_prov.get(aid, "")
        row_region = _account_to_region.get(aid, "")
        owner_uid  = _account_to_owner.get(aid, "")
        row_owner  = _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else ""

        # Account-level filters
        if acct_type  and norm(row_type)   != norm(acct_type):   continue
        if region     and norm(row_region) != norm(region):       continue
        if acct_state and norm(row_state)  != norm(acct_state):   continue

        raw_slps = slps_by_account.get(aid, [])
        slp_list = _build_slp_list(raw_slps)

        if has_slps is True  and len(slp_list) == 0: continue
        if has_slps is False and len(raw_slps) > 0:  continue

        accounts_out.append({
            "account_id": aid,
            "name":       row_name,
            "type":       row_type,
            "state":      row_state,
            "region":     row_region,
            "dealer_id":  _account_to_dealer.get(aid, ""),
            "owner":      row_owner,
            "slp_count":  len(raw_slps),
            "slps":       slp_list,
        })

    accounts_out.sort(key=lambda x: x["name"].upper())

    with_slps      = sum(1 for a in accounts_out if a["slp_count"] > 0)
    total_slps_r   = sum(len(a["slps"]) for a in accounts_out)
    orphaned_count = sum(1 for a in accounts_out if a["name"].startswith("Unknown Account ("))

    return {
        "total_accounts":        len(accounts_out),
        "accounts_with_slps":    with_slps,
        "accounts_without_slps": len(accounts_out) - with_slps,
        "orphaned_slp_accounts": orphaned_count,
        "fallback_linked_slps":  fallback_used,
        "unlinked_slps":         unlinked_slps,
        "total_slps":            total_slps_r,
        "cache_age_seconds":     cache_age,
        "accounts":              accounts_out,
    }


# ── Health Check ─────────────────────────────────────────────────────────────

@app.get("/api/health")
async def health_check():
    from fastapi.responses import JSONResponse
    accounts_ok = len(_account_to_name) > 0
    slp_ok      = len(_slp_cache_records) > 0
    lc_ok       = bool(_lc_cache)
    ta_ok       = bool(_ta_cache)
    # Core ready = accounts + SLP loaded (everything most features need).
    # LC and TA are secondary caches (last-contacted report, team-activity report)
    # and should not block the service from being considered online.
    ready = accounts_ok and slp_ok

    payload = {
        "status":                "online" if ready else "warming",
        "accounts_indexed":      len(_account_to_name),
        "slp_cache_count":       len(_slp_cache_records),
        "slp_cache_age_seconds": round(_time.time() - _slp_cache_ts) if _slp_cache_ts else None,
        "lc_cache_loaded":       lc_ok,
        "ta_cache_loaded":       ta_ok,
    }
    # Always return 200 so Render's health check doesn't force-restart the
    # service while it's still warming up (SLP cache takes ~2 min to load).
    # Warmup state is visible in the "status" field of the response body.
    return JSONResponse(content=payload, status_code=200)


# ── SLP Cache Refresh ─────────────────────────────────────────────────────────

@app.get("/api/slp-cache/refresh")
async def slp_cache_refresh(user=Depends(require_auth)):
    await _refresh_slp_cache()
    return {
        "slp_cache_count":      len(_slp_cache_records),
        "slp_cache_age_seconds": round(_time.time() - _slp_cache_ts) if _slp_cache_ts else None,
    }


# ── Data Integrity Report ─────────────────────────────────────────────────────

@app.get("/api/reports/data-integrity")
async def data_integrity_report(user=Depends(require_auth)):
    def get_field(slp, fid):
        for f in slp.get("fields", []):
            if f.get("id") == fid:
                return f.get("value") or ""
        return ""

    all_slps  = list(await get_slp_cache())
    cache_age = round(_time.time() - _slp_cache_ts) if _slp_cache_ts else None

    # ── SLP checks ────────────────────────────────────────────────────────────
    seen_ids:           set  = set()
    duplicate_ids:      list = []
    missing_dealer_id:  int  = 0
    missing_platform:   int  = 0
    missing_status:     int  = 0
    missing_account_rel: int = 0

    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            duplicate_ids.append(sid)
        else:
            seen_ids.add(sid)

        if not get_field(slp, "dealer-id"):
            missing_dealer_id += 1
        if not get_field(slp, "platform"):
            missing_platform += 1
        if not get_field(slp, "slp-status-detail"):
            missing_status += 1

        rel = slp.get("relationships", {}).get("account") or []
        if not rel:
            missing_account_rel += 1

    # ── Account checks ────────────────────────────────────────────────────────
    total_accounts      = len(_account_to_name)
    missing_acct_type   = sum(1 for v in _account_to_type.values()       if not v.strip())
    missing_acct_state  = sum(1 for v in _account_to_state_prov.values() if not v.strip())

    # Accounts in the index with no type set at all (not in dict)
    missing_acct_type  += total_accounts - len(_account_to_type)
    missing_acct_state += total_accounts - len(_account_to_state_prov)

    return {
        "slp_cache_age_seconds": cache_age,
        "slps": {
            "total":                len(all_slps),
            "duplicate_ids":        duplicate_ids,
            "missing_dealer_id":    missing_dealer_id,
            "missing_platform":     missing_platform,
            "missing_status":       missing_status,
            "missing_account_rel":  missing_account_rel,
        },
        "accounts": {
            "total":               total_accounts,
            "missing_type_cf76":   missing_acct_type,
            "missing_state_cf5":   missing_acct_state,
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# Welcome Email — manual trigger (admin-only)
#
# Phase 1: applies a `welcome-{channel-slug}` tag to every eligible contact on
# an account. AC tag-triggered automations send the actual email per channel.
#
# Setup required in AC before this works:
#   1. Create one tag per channel:           welcome-optimus, welcome-360finance, ...
#   2. Create one tag per channel marker:    welcomed-optimus, welcomed-360finance, ...
#   3. Build one tag-triggered automation per channel that:
#        - Trigger:  Tag added = welcome-{slug}
#        - Action:   Send email  (the welcome template for that channel)
#        - Action:   Add tag    welcomed-{slug}   (so we don't re-send)
#        - Action:   Remove tag welcome-{slug}    (resets the trigger)
# ─────────────────────────────────────────────────────────────────────────────

# Channels we'll send welcome emails for.
# Limited to Microf + OPTIMUS for the initial rollout — expand later as more
# AC tag-triggered automations are built.
WELCOME_CHANNELS = [
    "Microf",
    "OPTIMUS",
]

# Map SLP `channel` field values → the welcome Channel (AC automation).
# SLP channel dropdown has finer values (e.g. "Optimus - RTO") that map to broader welcome channels.
CHANNEL_TO_WELCOME_CHANNEL = {
    "Microf Direct":      "Microf",
    "OPTIMUS":            "OPTIMUS",
    "Optimus 2.0":        "OPTIMUS",
    "Optimus - RTO Only": "OPTIMUS",
    "Optimus - FTL":      "OPTIMUS",
    "Optimus - RTO":      "OPTIMUS",
    "Optimus - Greensky": "OPTIMUS",
}


def _channel_slug(channel: str) -> str:
    """Convert a Channel value to its tag-slug form.
    e.g. 'SpectrumAC (Wells Fargo)' → 'spectrumac-wf'.
    Adjust here to match the tag names you create in AC."""
    s = channel.lower()
    s = s.replace("(wells fargo)", "wf")
    s = s.replace("(lto only)", "lto-only")
    s = s.replace("(no lease integration)", "no-lease")
    s = _re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def _welcome_tag_name(channel: str) -> str:
    return f"welcome-{_channel_slug(channel)}"


def _welcomed_tag_name(channel: str) -> str:
    return f"welcomed-{_channel_slug(channel)}"


def _resend_tag_name(channel: str) -> str:
    return f"resend-{_channel_slug(channel)}"


# Tag-name → tag-id cache (in-memory; refreshed on first lookup)
_WELCOME_TAG_ID_CACHE: dict = {}


async def _get_tag_id(tag_name: str) -> Optional[str]:
    """Resolve an AC tag name to its numeric id. Caches in memory."""
    if tag_name in _WELCOME_TAG_ID_CACHE:
        return _WELCOME_TAG_ID_CACHE[tag_name]
    resp = await ac_get("tags", {"search": tag_name, "limit": 100})
    for t in resp.get("tags", []):
        if (t.get("tag") or "").lower() == tag_name.lower():
            tid = t.get("id")
            _WELCOME_TAG_ID_CACHE[tag_name] = tid
            return tid
    return None


@app.get("/reports/prebuilt")
async def reports_prebuilt_page():
    return FileResponse("static/reports/prebuilt.html")


@app.get("/am-guide")
async def am_guide_page(user=Depends(require_auth)):
    return FileResponse("static/am-guide.html")


@app.get("/onboarding-live")
async def onboarding_live_page(user=Depends(require_auth)):
    return FileResponse("static/reports/onboarding-live.html")


@app.get("/welcome")
async def welcome_page(_admin=Depends(_require_welcome)):
    return FileResponse("static/welcome.html")


@app.get("/api/welcome/channels")
async def welcome_channels(_admin=Depends(_require_welcome)):
    """Return the list of valid Channel values for the UI dropdown."""
    return {"channels": WELCOME_CHANNELS}


async def _eligible_welcome_contacts(account_id: str, channel: str) -> dict:
    """Return the contacts on this account that should/shouldn't receive a welcome email."""
    tag_welcomed = _welcomed_tag_name(channel)

    acc_data, contacts_resp = await asyncio.gather(
        ac_get(f"accounts/{account_id}"),
        ac_get(f"accounts/{account_id}/contacts"),
        return_exceptions=True,
    )
    if not isinstance(acc_data, dict) or not acc_data.get("account"):
        raise HTTPException(404, f"Account {account_id} not found")
    account_name = acc_data["account"].get("name", "")

    contact_ids = []
    if isinstance(contacts_resp, dict):
        contact_ids = [
            ac.get("contact")
            for ac in contacts_resp.get("accountContacts", [])
            if ac.get("contact")
        ]

    # Fetch each contact + their tags in parallel
    detail_tasks = [
        ac_get(f"contacts/{cid}", {"include": "contactTags.tag"})
        for cid in contact_ids
    ]
    details = await asyncio.gather(*detail_tasks, return_exceptions=True)

    eligible, skipped = [], []
    for d in details:
        if not isinstance(d, dict):
            continue
        c = d.get("contact", {})
        cid = c.get("id")
        email = (c.get("email") or "").strip()
        first = c.get("firstName") or ""
        last = c.get("lastName") or ""

        # Tag names this contact already has
        tag_names = set()
        for t in d.get("tags", []):
            if isinstance(t, dict) and t.get("tag"):
                tag_names.add(t["tag"].lower())

        reason = None
        if not email:
            reason = "no email"
        elif str(c.get("status") or "") == "2":  # AC: 2 = unsubscribed
            reason = "unsubscribed"
        elif tag_welcomed.lower() in tag_names:
            reason = "already welcomed"

        row = {
            "contact_id": cid,
            "email":      email,
            "name":       f"{first} {last}".strip(),
        }
        if reason:
            row["skip_reason"] = reason
            skipped.append(row)
        else:
            eligible.append(row)

    return {
        "account_id":            account_id,
        "account_name":          account_name,
        "channel":               channel,
        "tag_to_apply":          _welcome_tag_name(channel),
        "tag_marker_when_done":  tag_welcomed,
        "total_contacts":        len(contact_ids),
        "eligible":              eligible,
        "skipped":               skipped,
    }


@app.get("/api/welcome/account-slps/{account_id}")
async def welcome_account_slps(account_id: str, _admin=Depends(_require_welcome)):
    """Return the SLP rows on this account so the admin can see which Channel(s)
    apply, with one flagged as the recommended pick (most recently activated
    SLP that maps to a supported Channel)."""
    resp = await ac_get(
        f"customObjects/records/{SLP_SCHEMA_ID}",
        {"filters[relationships.account]": account_id, "limit": 100},
    )
    rows = []
    for r in resp.get("records", []):
        f = {x.get("id") or x.get("field"): x.get("value", "") for x in r.get("fields", [])}
        channel_val = (f.get("channel") or "").strip()
        if not channel_val:
            continue
        suggested = CHANNEL_TO_WELCOME_CHANNEL.get(channel_val)
        rows.append({
            "record_id":    r.get("id"),
            "platform":     channel_val,   # raw SLP channel value (shown in panel)
            "status":       (f.get("slp-status-detail") or "").strip(),
            "activated":    (f.get("contractor-activated-date") or "").strip(),
            "dealer_id":    (f.get("dealer-id") or "").strip(),
            "supported":    bool(suggested),
            "channel":      suggested,     # welcome channel for "Use this" button
        })

    # Sort by activated desc (ISO strings sort correctly), then by platform
    rows.sort(key=lambda r: (r["activated"] or "", r["platform"]), reverse=True)

    # Recommended: most recently activated supported row that's Contractor Activated;
    # fall back to most recently activated supported row of any status.
    recommended_id = None
    for row in rows:
        if row["supported"] and row["activated"] and row["status"] == "Contractor Activated":
            recommended_id = row["record_id"]
            break
    if not recommended_id:
        for row in rows:
            if row["supported"] and row["activated"]:
                recommended_id = row["record_id"]
                break

    for row in rows:
        row["recommended"] = (row["record_id"] == recommended_id)

    return {
        "account_id": account_id,
        "slps":       rows,
        "any_supported": any(r["supported"] for r in rows),
    }


@app.get("/api/welcome/preview/{account_id}")
async def welcome_preview(
    account_id: str,
    channel: str = Query(...),
    _admin=Depends(_require_welcome),
):
    """Show who would receive the welcome email for this Channel — does NOT send."""
    if channel not in WELCOME_CHANNELS:
        raise HTTPException(400, f"Unknown channel: {channel}")
    return await _eligible_welcome_contacts(account_id, channel)


class _WelcomeSendRequest(_BaseModel):
    account_id:   str
    channel:      str
    contact_ids:  Optional[List[str]] = None  # if set, only tag these contact IDs
    force_resend: bool = False                # strip welcomed+welcome tags first, then re-tag everyone


@app.post("/api/welcome/send")
async def welcome_send(
    payload: _WelcomeSendRequest,
    user=Depends(_require_welcome),
):
    """Tag eligible contacts on the account so the AC welcome automation fires.
    If contact_ids is provided, only tag those specific contacts.
    If force_resend is True, strip welcomed+welcome tags first then re-tag all contacts.
    """
    if payload.channel not in WELCOME_CHANNELS:
        raise HTTPException(400, f"Unknown channel: {payload.channel}")

    welcome_tag  = _welcome_tag_name(payload.channel)
    welcomed_tag = _welcomed_tag_name(payload.channel)

    tag_id = await _get_tag_id(welcome_tag)
    if not tag_id:
        raise HTTPException(
            500,
            f"Tag '{welcome_tag}' does not exist in ActiveCampaign. "
            "Create it first (Contacts → Manage Tags) and build the matching automation.",
        )

    # ── Force-resend path: apply resend-{channel} tag to trigger AC automation ─
    if payload.force_resend:
        resend_tag  = _resend_tag_name(payload.channel)
        resend_tag_id = await _get_tag_id(resend_tag)
        if not resend_tag_id:
            raise HTTPException(
                500,
                f"Tag '{resend_tag}' does not exist in ActiveCampaign. "
                "Create it first and attach it to the resend automation trigger.",
            )

        acc_data, contacts_resp = await asyncio.gather(
            ac_get(f"accounts/{payload.account_id}"),
            ac_get(f"accounts/{payload.account_id}/contacts"),
            return_exceptions=True,
        )
        if not isinstance(acc_data, dict) or not acc_data.get("account"):
            raise HTTPException(404, f"Account {payload.account_id} not found")
        account_name = acc_data["account"].get("name", "")

        all_cids = []
        if isinstance(contacts_resp, dict):
            all_cids = [ac.get("contact") for ac in contacts_resp.get("accountContacts", []) if ac.get("contact")]

        if payload.contact_ids is not None:
            allowed = set(str(cid) for cid in payload.contact_ids)
            all_cids = [cid for cid in all_cids if str(cid) in allowed]

        # Fetch contact details to filter out unsubscribed / no-email contacts
        details = await asyncio.gather(
            *[ac_get(f"contacts/{cid}") for cid in all_cids],
            return_exceptions=True,
        )

        tagged = errors = 0
        results = []
        for d in details:
            if not isinstance(d, dict):
                continue
            c     = d.get("contact", {})
            cid   = c.get("id")
            email = (c.get("email") or "").strip()
            name  = f"{c.get('firstName','')} {c.get('lastName','')}".strip()

            if not email or str(c.get("status") or "") == "2":
                results.append({"contact_id": cid, "email": email, "name": name,
                                 "status": "skipped",
                                 "skip_reason": "no email" if not email else "unsubscribed"})
                continue

            try:
                await ac_post("contactTags", {"contactTag": {"contact": cid, "tag": resend_tag_id}})
                tagged += 1
                results.append({"contact_id": cid, "email": email, "name": name, "status": "tagged"})
            except Exception as e:
                errors += 1
                results.append({"contact_id": cid, "email": email, "name": name,
                                 "status": "error", "detail": str(e)[:200]})

        print(f"[welcome-resend] {user} → account={payload.account_id} channel={payload.channel} "
              f"tag={resend_tag} tagged={tagged} errors={errors}")
        skipped_list = [r for r in results if r["status"] == "skipped"]
        return {
            "account_id":     payload.account_id,
            "account_name":   account_name,
            "channel":        payload.channel,
            "tag_applied":    resend_tag,
            "total_contacts": len(all_cids),
            "tagged":         tagged,
            "errors":         errors,
            "skipped":        len(skipped_list),
            "skipped_detail": skipped_list,
            "results":        [r for r in results if r["status"] != "skipped"],
            "by":             user,
            "resend":         True,
        }

    # ── Normal send path ───────────────────────────────────────────────────
    preview = await _eligible_welcome_contacts(payload.account_id, payload.channel)

    # Filter to selected contacts if caller passed a list
    to_tag = preview["eligible"]
    if payload.contact_ids is not None:
        allowed = set(str(cid) for cid in payload.contact_ids)
        to_tag = [c for c in to_tag if str(c["contact_id"]) in allowed]

    results = []
    for c in to_tag:
        try:
            await ac_post("contactTags", {"contactTag": {"contact": c["contact_id"], "tag": tag_id}})
            results.append({"contact_id": c["contact_id"], "email": c["email"], "name": c.get("name",""), "status": "tagged"})
        except Exception as e:
            results.append({
                "contact_id": c["contact_id"],
                "email":      c["email"],
                "name":       c.get("name", ""),
                "status":     "error",
                "detail":     str(e)[:200],
            })

    tagged  = sum(1 for r in results if r["status"] == "tagged")
    errors  = sum(1 for r in results if r["status"] == "error")
    print(f"[welcome] {user} → account={payload.account_id} channel={payload.channel} "
          f"tagged={tagged} errors={errors} skipped={len(preview['skipped'])}")

    return {
        "account_id":     payload.account_id,
        "account_name":   preview["account_name"],
        "channel":        payload.channel,
        "tag_applied":    welcome_tag,
        "total_contacts": preview["total_contacts"],
        "tagged":         tagged,
        "errors":         errors,
        "skipped":        len(preview["skipped"]),
        "skipped_detail": preview["skipped"],
        "results":        results,
        "by":             user,
    }


# ── APEX Business Review ──────────────────────────────────────────────────────

# Kept for backward-compat reference; logic now accepts any partner value
_APEX_PARTNERS = {"apex service partners", "southern air"}

# Enrolled dates provided by the partner (keyed by normalised dealer name).
# Normalisation: lowercase, strip punctuation, collapse whitespace.
import re as _re
def _norm_dealer(name: str) -> str:
    name = name.lower()
    name = _re.sub(r"[^a-z0-9 ]", " ", name)
    return _re.sub(r"\s+", " ", name).strip()

_APEX_ENROLLED_DATES: dict = {
    _norm_dealer(k): v for k, v in {
        "SOUTHERN AIR OF SHREVEPORT":                                    "5/7/2024",
        "DILLING HEATING COOLING PLUMBING & ELECTRICAL":                 "2/4/2025",
        "ACE HEATING COOLING & PLUMBING LLC DBA ACE HOME SERVICES":      "7/3/2024",
        "WILLARD EAST":                                                  "4/2/2025",
        "SNYDER AIR CONDITIONING PLUMBING & ELECTRIC":                   "10/29/2024",
        "A#1 AIR PLUMBING & ELECTRICAL":                                 "4/14/2025",
        "WILLARD WEST":                                                  "3/21/2025",
        "BASSETT SERVICES HEATING & COOLING LLC":                        "8/20/2025",
        "CJS HEATING AND AIR":                                           "8/18/2025",
        "CHAMPIONS GREENVILLE":                                          "8/11/2025",
        "RIGHT NOW HEATING COOLING PLUMBING LLC":                        "7/8/2024",
        "BLAZE TRIANGLE LLC":                                            "3/26/2025",
        "ACE AIR CONDITIONING AND PLUMBING":                             "5/15/2025",
        "PIONEER COMFORT NASHVILLE":                                     "12/23/2025",
        "LEGACY HEATING AND AIR INC":                                    "8/19/2025",
        "AXSOM AIR OF LA INC":                                           "8/23/2018",
        "DIRECT HEATING AND COOLING LLC":                                "11/15/2024",
        "COTE'S MECHANICAL LLC":                                         "5/19/2025",
        "TIN MAN HEATING COOLING PLUMBING & ELECTRICAL":                 "9/25/2025",
        "KORTE DOES IT ALL INC":                                         "8/25/2025",
        "CHAMPIONS FLAT ROCK":                                           "8/11/2025",
        "EVANS AIR CONDITIONING PLUMBING & ELECTRICAL":                  "4/28/2025",
        "GEES HEATING COOLING PLUMBING AND ELECTRICAL LLC":              "12/18/2025",
        "HENCO PLUMBING SERVICES":                                       "1/21/2025",
        "PRIDE MECHANICAL":                                              "8/28/2024",
        "RENFROW HEATING COOLING PLUMBING AND ELECTRICAL":               "10/22/2025",
        "SUNSET HEATING COOLING AND PLUMBING LLC":                       "1/18/2025",
        "TRINITY AIR HEATING COOLING PLUMBING & ELECTRICAL":             "12/18/2025",
    }.items()
}

def _lookup_enrolled_date(dealer_name: str, did: str = "") -> str:
    """Return enrolled date: static lookup → AC activation date fallback."""
    norm = _norm_dealer(dealer_name)
    # Exact match
    if norm in _APEX_ENROLLED_DATES:
        return _APEX_ENROLLED_DATES[norm]
    # Partial match (name contains a key or key contains the name)
    for key, val in _APEX_ENROLLED_DATES.items():
        if key in norm or norm in key:
            return val
    # Fallback: AC activation date via dealer_id
    if did:
        for aid, ddid in _account_to_dealer.items():
            if str(ddid) == str(did):
                return (_account_to_activation_date.get(aid, "") or "")[:10]
    return ""

# ── Daily Sales Dump processing helpers ───────────────────────────────────────

_DUMP_ENCODING_TRIES = [("utf-16", "\t"), ("utf-8", "\t"), ("utf-8", ","), ("latin-1", "\t"), ("latin-1", ",")]

def _read_dump(content: bytes):
    """Try several encodings/separators; return the first DataFrame that looks like a dump."""
    if _pd is None:
        raise HTTPException(500, "pandas not installed on server")
    for enc, sep in _DUMP_ENCODING_TRIES:
        try:
            df = _pd.read_csv(_io.BytesIO(content), encoding=enc, sep=sep)
            # Must have at least Dealer Id and inserted_time
            cols_lower = {c.strip().lower() for c in df.columns}
            if "dealer id" in cols_lower and "inserted_time" in cols_lower:
                # Rename columns to strip whitespace
                df.columns = [c.strip() for c in df.columns]
                return df
        except Exception:
            continue
    raise HTTPException(400, "Could not parse file as a Daily Sales Dump (expected tab-separated UTF-16 with Dealer Id + inserted_time columns)")

def _partner_filter_list(partner_filter: str) -> list:
    """Parse comma-separated partner filter into list of lowercase strings. Empty = match all."""
    return [p.strip().lower() for p in partner_filter.split(",") if p.strip()] if partner_filter else []

def _build_apex_dealer_id_set(partner_filter: str = "") -> set:
    """Return str dealer_ids for strategic partner contractors. Empty filter = all partners."""
    pf_list = _partner_filter_list(partner_filter)
    dealer_ids = set()
    for aid, sp_val in _account_to_strategic_partners.items():
        if not sp_val:
            continue
        if _account_to_type.get(aid, "").strip().lower() != "contractor":
            continue
        if pf_list and not any(pf in sp_val.lower() for pf in pf_list):
            continue
        did = _account_to_dealer.get(aid, "")
        if did:
            dealer_ids.add(str(did))
    return dealer_ids

async def _fetch_apex_dealer_ids_from_ac(partner_filter: str = "") -> set:
    """
    Fallback: query AC directly for strategic-partner dealer IDs when the
    in-memory index is empty or hasn't been built yet.
    Scans accountCustomFieldData for CF132 (Strategic Partners), optionally
    filtering to a specific partner value, then looks up CF18 (Dealer ID).
    Empty partner_filter = any non-empty strategic partner value.
    """
    pf_list = _partner_filter_list(partner_filter)
    matching_acct_ids: set = set()

    # Pass 1 — find accounts with matching Strategic Partners (CF132)
    offset = 0
    while True:
        resp = await ac_get("accountCustomFieldData", {"limit": 100, "offset": offset})
        items = resp.get("accountCustomFieldData", [])
        if not items:
            break
        for item in items:
            if str(item.get("customFieldId", "")) != "132":
                continue
            val = (item.get("fieldValue") or "").strip()
            if not val:
                continue
            if pf_list and not any(pf in val.lower() for pf in pf_list):
                continue
            aid = str(item.get("accountId", ""))
            if aid:
                matching_acct_ids.add(aid)
        offset += len(items)
        total  = int(resp.get("meta", {}).get("total", 0))
        if offset >= total:
            break

    if not matching_acct_ids:
        return set()

    # Filter to Contractor accounts only (CF76) using in-memory index or AC query
    contractor_ids: set = {aid for aid in matching_acct_ids
                           if _account_to_type.get(aid, "").strip().lower() == "contractor"}
    # If index is cold, fall back to querying CF76
    if not contractor_ids:
        offset = 0
        while True:
            resp = await ac_get("accountCustomFieldData", {"limit": 100, "offset": offset})
            items = resp.get("accountCustomFieldData", [])
            if not items:
                break
            for item in items:
                if str(item.get("customFieldId", "")) != "76":
                    continue
                aid = str(item.get("accountId", ""))
                if aid not in matching_acct_ids:
                    continue
                if (item.get("fieldValue") or "").strip().lower() == "contractor":
                    contractor_ids.add(aid)
            offset += len(items)
            total = int(resp.get("meta", {}).get("total", 0))
            if offset >= total:
                break
    matching_acct_ids = contractor_ids if contractor_ids else matching_acct_ids

    # Pass 2 — get Dealer IDs (CF18) for those accounts
    dealer_ids: set = set()
    # First try from already-loaded index
    for aid in matching_acct_ids:
        did = _account_to_dealer.get(aid, "")
        if did:
            dealer_ids.add(str(did))

    # If index missing entries, query CF18 directly
    if len(dealer_ids) < len(matching_acct_ids) // 2:
        offset = 0
        while True:
            resp = await ac_get("accountCustomFieldData", {"limit": 100, "offset": offset})
            items = resp.get("accountCustomFieldData", [])
            if not items:
                break
            for item in items:
                if str(item.get("customFieldId", "")) != "18":
                    continue
                aid = str(item.get("accountId", ""))
                if aid not in matching_acct_ids:
                    continue
                did = str(item.get("fieldValue") or "").strip()
                if did and did not in ("0", ""):
                    dealer_ids.add(did)
            offset += len(items)
            total  = int(resp.get("meta", {}).get("total", 0))
            if offset >= total:
                break

    return dealer_ids

def _parse_dollar(s) -> float:
    if s is None or (isinstance(s, float) and _pd.isna(s)):
        return 0.0
    return float(str(s).replace("$", "").replace(",", "").strip() or 0)

def _dump_to_production(df: "_pd.DataFrame", apex_ids: set) -> dict:
    """
    Filter a Daily Sales Dump DataFrame to APEX dealers and aggregate by month.
    Returns: {month_label: [production_row, ...]}
    """
    import re

    # Normalise Dealer Id to string
    df = df.copy()
    df["_did_str"] = df["Dealer Id"].apply(lambda x: str(int(x)) if _pd.notna(x) and str(x).replace(".0","").isdigit() else "")
    apex_rows = df[df["_did_str"].isin(apex_ids)].copy()

    # Build did→name, did→account_id, did→strategic_partner from AC index
    did_to_ac_name: dict = {}
    did_to_account_id: dict = {}
    did_to_sp: dict = {}
    for aid, ddid in _account_to_dealer.items():
        sdid = str(ddid)
        if sdid in apex_ids and sdid not in did_to_ac_name:
            did_to_ac_name[sdid] = _account_to_name.get(aid, "")
            did_to_account_id[sdid] = aid
            did_to_sp[sdid] = _account_to_strategic_partners.get(aid, "")

    # Parse dates (even if apex_rows is empty we still want zero rows per month)
    if not apex_rows.empty:
        apex_rows["_date"] = _pd.to_datetime(apex_rows["inserted_time"], errors="coerce", dayfirst=False)
        apex_rows["_month_label"] = apex_rows["_date"].dt.strftime("%B %Y")
        month_labels = sorted(apex_rows["_month_label"].dropna().unique().tolist(),
                              key=lambda lbl: (_pd.to_datetime("1 " + lbl, format="%d %B %Y", errors="coerce") or _pd.Timestamp.max))
    else:
        month_labels = []

    # Parse NIA dollar column (funded amount)
    nia_col = next((c for c in apex_rows.columns if c.strip().lower() == "nia"), None) if not apex_rows.empty else None
    if nia_col:
        apex_rows["_nia"] = apex_rows[nia_col].apply(_parse_dollar)
    else:
        if not apex_rows.empty:
            apex_rows["_nia"] = 0.0

    MONTH_ORDER = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]

    def _zero_row(did, name):
        return {"dealer": name or f"Dealer {did}", "dealer_id": did,
                "account_id": did_to_account_id.get(did, ""),
                "strategic_partner": did_to_sp.get(did, ""),
                "apps": 0, "approved": 0, "pending": 0, "rpas": 0, "nia": 0, "revenue": 0.0}

    results = {}
    for month_label in month_labels:
        mdf = apex_rows[apex_rows["_month_label"] == month_label]
        # Primary apps only for counting (avoid re-submissions inflating)
        primary = mdf[mdf["Primary App"] == 1] if "Primary App" in mdf.columns else mdf
        prod_rows = []
        active_dids = set()
        for dealer_name, grp in primary.groupby("Contractor Name"):
            funded = grp[grp["App Sub Status"].str.upper().str.strip() == "FUNDED"]
            # Approved = Pre-Approved OR Funded (funded implies approved even via Further Review path)
            approved = grp[
                (grp["Response Description"].str.strip() == "Pre-Approved") |
                (grp["App Sub Status"].str.upper().str.strip() == "FUNDED")
            ]
            pending = grp[grp["Processing Status Description"].str.strip().isin(
                ["Needs Review", "Decision Pending", "Partially Approved", "Fully Approved", "Preapproved"]
            )]
            nia_count = grp[grp.get("Shrinkage", _pd.Series(dtype=str)).str.strip().str.startswith("5.", na=False)]
            did = str(grp["_did_str"].iloc[0])
            active_dids.add(did)
            prod_rows.append({
                "dealer":            dealer_name,
                "dealer_id":         did,
                "account_id":        did_to_account_id.get(did, ""),
                "strategic_partner": did_to_sp.get(did, ""),
                "apps":              int(len(grp)),
                "approved":          int(len(approved)),
                "pending":           int(len(pending)),
                "rpas":              int(len(funded)),
                "nia":               int(len(nia_count)),
                "revenue":           round(float(funded["_nia"].sum()), 2),
            })
        # Add enrolled dealers with zero activity this month
        for did in sorted(apex_ids - active_dids):
            name = did_to_ac_name.get(did, "")
            if name:  # only add if we know the dealer name
                prod_rows.append(_zero_row(did, name))
        # Sort: active (revenue desc, name) then zeros (name)
        prod_rows.sort(key=lambda r: (r["apps"] == 0, -r["revenue"], r["dealer"].lower()))
        results[month_label] = prod_rows

    # Sort months chronologically
    def _month_sort(label):
        parts = label.split()
        m = MONTH_ORDER.index(parts[0]) if parts[0] in MONTH_ORDER else 99
        y = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 9999
        return (y, m)

    return dict(sorted(results.items(), key=lambda kv: _month_sort(kv[0])))


def _dump_to_rollup(df: "_pd.DataFrame", apex_ids: set) -> list:
    """
    Aggregate the *entire* dump (all months) per dealer into a single rollup row.
    Returns a list of rollup rows compatible with the rollup upload format.
    Enrolled date is pulled from the in-memory AC index when available.
    """
    df = df.copy()
    df["_did_str"] = df["Dealer Id"].apply(
        lambda x: str(int(x)) if _pd.notna(x) and str(x).replace(".0", "").isdigit() else ""
    )
    apex_rows = df[df["_did_str"].isin(apex_ids)].copy()
    if apex_rows.empty:
        return []

    # Build did→account_id and did→strategic_partner lookup
    did_to_account_id: dict = {}
    did_to_sp: dict = {}
    for aid, ddid in _account_to_dealer.items():
        sdid = str(ddid)
        if sdid in apex_ids and sdid not in did_to_account_id:
            did_to_account_id[sdid] = aid
            did_to_sp[sdid] = _account_to_strategic_partners.get(aid, "")

    # Primary apps only
    primary = apex_rows[apex_rows["Primary App"] == 1] if "Primary App" in apex_rows.columns else apex_rows

    # NIA dollar column
    nia_col = next((c for c in primary.columns if c.strip().lower() == "nia"), None)
    if nia_col:
        primary = primary.copy()
        primary["_nia"] = primary[nia_col].apply(_parse_dollar)
    else:
        primary = primary.copy()
        primary["_nia"] = 0.0

    primary["_date"] = _pd.to_datetime(primary["inserted_time"], errors="coerce")

    rollup_rows = []
    for dealer_name, grp in primary.groupby("Contractor Name"):
        funded   = grp[grp["App Sub Status"].str.upper().str.strip() == "FUNDED"]
        # Approved = Pre-Approved OR Funded (funded implies approved even via Further Review path)
        approved = grp[
            (grp["Response Description"].str.strip() == "Pre-Approved") |
            (grp["App Sub Status"].str.upper().str.strip() == "FUNDED")
        ]

        total_apps    = int(len(grp))
        total_rpas    = int(len(funded))
        total_revenue = round(float(funded["_nia"].sum()), 2)
        # Store as 0–100 so pctBar() can display directly
        par = round(len(approved) / total_apps * 100, 1) if total_apps > 0 else 0.0
        tur = round(total_rpas / len(approved) * 100, 1) if len(approved) > 0 else 0.0

        last_dt = grp["_date"].max()
        last_app_date = last_dt.strftime("%Y-%m-%d") if _pd.notna(last_dt) else ""

        did = str(grp["_did_str"].iloc[0])
        enrolled_date = _lookup_enrolled_date(dealer_name, did)
        if enrolled_date:
            try:
                # Normalise to YYYY-MM-DD if stored as M/D/YYYY
                from datetime import datetime as _dt
                for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
                    try:
                        enrolled_date = _dt.strptime(enrolled_date, fmt).strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        pass
            except Exception:
                pass

        rollup_rows.append({
            "dealer":            dealer_name,
            "dealer_id":         did,
            "account_id":        did_to_account_id.get(did, ""),
            "strategic_partner": did_to_sp.get(did, ""),
            "enrolled_date":     enrolled_date,
            "last_app_date":     last_app_date,
            "ttm_apps":          total_apps,
            "ttm_rpas":          total_rpas,
            "ttm_revenue":       total_revenue,
            "pre_approval_rate": par,
            "take_up_rate":      tur,
        })

    # Add enrolled dealers with zero activity in the entire dump period
    active_dids = {r["dealer_id"] for r in rollup_rows}
    for did in sorted(apex_ids - active_dids):
        for aid, ddid in _account_to_dealer.items():
            if str(ddid) == did:
                name = _account_to_name.get(aid, "")
                if name:
                    enrolled_date = _lookup_enrolled_date(name, did)
                    if enrolled_date:
                        try:
                            from datetime import datetime as _dt
                            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
                                try:
                                    enrolled_date = _dt.strptime(enrolled_date, fmt).strftime("%Y-%m-%d")
                                    break
                                except ValueError:
                                    pass
                        except Exception:
                            pass
                    rollup_rows.append({
                        "dealer":            name,
                        "dealer_id":         did,
                        "account_id":        did_to_account_id.get(did, ""),
                        "strategic_partner": did_to_sp.get(did, ""),
                        "enrolled_date":     enrolled_date,
                        "last_app_date":     "",
                        "ttm_apps":          0,
                        "ttm_rpas":          0,
                        "ttm_revenue":       0.0,
                        "pre_approval_rate": 0.0,
                        "take_up_rate":      0.0,
                    })
                break

    # Sort: active dealers (revenue desc, name) then zeros (name)
    rollup_rows.sort(key=lambda r: (r["ttm_apps"] == 0, -r["ttm_revenue"], r["dealer"].lower()))
    return rollup_rows


_MONTH_TO_Q = {
    "January": 1, "February": 1, "March": 1,
    "April": 2,   "May": 2,      "June": 2,
    "July": 3,    "August": 3,   "September": 3,
    "October": 4, "November": 4, "December": 4,
}

def _aggregate_to_quarters(monthly: dict) -> dict:
    """
    Roll up monthly production rows into quarterly periods.
    Returns {quarter_label: [aggregated_prod_rows]} e.g. {"Q1 2026": [...]}
    """
    buckets: dict = {}  # q_label -> {dealer_id -> row}
    for month_label, prod_rows in monthly.items():
        parts = month_label.split()
        if len(parts) != 2:
            continue
        month_name, year = parts[0], parts[1]
        q_num = _MONTH_TO_Q.get(month_name)
        if not q_num:
            continue
        q_label = f"Q{q_num} {year}"
        bucket = buckets.setdefault(q_label, {})
        for row in prod_rows:
            did = row["dealer_id"]
            if did not in bucket:
                bucket[did] = {
                    "dealer":    row["dealer"],
                    "dealer_id": did,
                    "apps":     0, "approved": 0, "pending": 0,
                    "rpas":     0, "nia":      0, "revenue": 0.0,
                }
            qr = bucket[did]
            qr["apps"]     += row.get("apps", 0)
            qr["approved"] += row.get("approved", 0)
            qr["pending"]  += row.get("pending", 0)
            qr["rpas"]     += row.get("rpas", 0)
            qr["nia"]      += row.get("nia", 0)
            qr["revenue"]  = round(qr["revenue"] + row.get("revenue", 0.0), 2)

    result = {}
    for q_label, bucket in sorted(buckets.items()):
        rows = sorted(bucket.values(), key=lambda r: (-r["revenue"], r["dealer"]))
        result[q_label] = rows
    return result


@app.post("/api/apex/upload/daily-dump")
async def apex_upload_daily_dump(
    file: UploadFile = File(...),
    partner: str = "",
    dealer_ids: str = Form(default=""),   # comma-separated dealer IDs from client roster
    admin=Depends(_require_admin),
):
    """
    Accept a Daily Sales Dump (TSV, UTF-16) and automatically build production
    data for each calendar month found in the file, filtered to APEX dealers.
    Creates / updates a period for each month detected.
    """
    content = await file.read()
    df = _read_dump(content)

    # Priority 1 — dealer IDs passed by the client (already loaded from AC roster in the UI)
    source = "client_roster"
    if dealer_ids.strip():
        apex_ids = {d.strip() for d in dealer_ids.split(",") if d.strip()}
    else:
        # Priority 2 — in-memory index
        apex_ids = _build_apex_dealer_id_set(partner)
        source = "index"

    if not apex_ids:
        # Priority 3 — direct AC API fallback (slow but reliable)
        print(f"[apex-dump] falling back to AC API query "
              f"(strategic_partners index size: {len(_account_to_strategic_partners)})")
        apex_ids = await _fetch_apex_dealer_ids_from_ac(partner)
        source = "ac_api"

    if not apex_ids:
        partner_label = partner or "All Strategic Partners"
        raise HTTPException(400,
            f"No strategic partner dealers found (partner filter: {partner_label!r}). "
            f"Strategic Partners index size: {len(_account_to_strategic_partners)} accounts. "
            f"Make sure the dealer index has been built and accounts have the Strategic Partners field set."
        )

    monthly = _dump_to_production(df, apex_ids)
    if not monthly:
        raise HTTPException(400, "No matching APEX dealer rows found in this file.")

    # Build rollup (all months aggregated) once for the whole dump
    rollup_rows = _dump_to_rollup(df, apex_ids)
    # Build quarterly aggregates
    quarterly = _aggregate_to_quarters(monthly)
    now_iso = datetime.now().isoformat()

    data = _load_apex_data()
    created = updated = 0
    periods_touched = []

    # Store monthly periods
    for month_label, prod_rows in monthly.items():
        if month_label not in data["periods"]:
            data["periods"][month_label] = {}
            created += 1
        else:
            updated += 1
        data["periods"][month_label]["production"] = prod_rows
        data["periods"][month_label]["production_uploaded_at"] = now_iso
        data["periods"][month_label]["production_source"] = "daily_dump"
        data["periods"][month_label]["period_type"] = "monthly"
        if rollup_rows:
            data["periods"][month_label]["rollup"] = rollup_rows
            data["periods"][month_label]["rollup_uploaded_at"] = now_iso
            data["periods"][month_label]["rollup_source"] = "daily_dump"
        periods_touched.append(month_label)

    # Store quarterly periods
    for q_label, q_prod_rows in quarterly.items():
        if q_label not in data["periods"]:
            data["periods"][q_label] = {}
            created += 1
        else:
            updated += 1
        data["periods"][q_label]["production"] = q_prod_rows
        data["periods"][q_label]["production_uploaded_at"] = now_iso
        data["periods"][q_label]["production_source"] = "daily_dump"
        data["periods"][q_label]["period_type"] = "quarterly"
        if rollup_rows:
            data["periods"][q_label]["rollup"] = rollup_rows
            data["periods"][q_label]["rollup_uploaded_at"] = now_iso
            data["periods"][q_label]["rollup_source"] = "daily_dump"
        periods_touched.append(q_label)

    _save_apex_data(data)

    summary = {ok: len(v) for ok, v in monthly.items()}
    print(f"[apex-dump] {admin} → processed {sum(summary.values())} dealer-months, "
          f"{len(quarterly)} quarters (dealer_ids source={source}, count={len(apex_ids)}, rollup_rows={len(rollup_rows)})")
    return {
        "ok": True,
        "periods_created":    created,
        "periods_updated":    updated,
        "periods":            periods_touched,
        "dealers_per_month":  summary,
        "quarters":           list(quarterly.keys()),
        "rollup_dealers":     len(rollup_rows),
        "apex_dealer_ids_used": len(apex_ids),
        "dealer_id_source":   source,
    }


_apex_partners_cache: list | None = None

@app.get("/api/apex/partners")
async def apex_list_partners(admin=Depends(_require_admin)):
    """Return all distinct Strategic Partners values found on Contractor accounts."""
    global _apex_partners_cache
    if _apex_partners_cache is not None:
        return {"partners": _apex_partners_cache}
    partners: set = set()
    for aid, sp_val in _account_to_strategic_partners.items():
        if not sp_val:
            continue
        if _account_to_type.get(aid, "").strip().lower() != "contractor":
            continue
        for p in sp_val.split(","):
            p = p.strip()
            if p:
                partners.add(p)
    _apex_partners_cache = sorted(partners, key=str.lower)
    return {"partners": _apex_partners_cache}


@app.get("/api/apex/dealers")
async def apex_get_dealers(partner: str = "", admin=Depends(_require_admin)):
    """Return Contractor accounts with a Strategic Partners value, optionally filtered to one partner."""
    pf_list = _partner_filter_list(partner)

    rows = []
    for aid, sp_val in _account_to_strategic_partners.items():
        if not sp_val:
            continue
        # Contractors only
        if _account_to_type.get(aid, "").strip().lower() != "contractor":
            continue
        if pf_list and not any(pf in sp_val.lower() for pf in pf_list):
            continue
        name = _account_to_name.get(aid, "")
        did  = str(_account_to_dealer.get(aid, ""))
        rows.append({
            "account_id":        aid,
            "name":              name,
            "dealer_id":         did,
            "region":            _account_to_region.get(aid, ""),
            "status":            _account_to_status.get(aid, ""),
            "activation_date":   _account_to_activation_date.get(aid, ""),
            "enrolled_date":     _lookup_enrolled_date(name, did),
            "last_app_date":     _account_to_last_app.get(aid, ""),
            "last_rpa_date":     _account_to_last_rpa.get(aid, ""),
            "strategic_partner": sp_val,
        })

    rows.sort(key=lambda r: r["name"].lower())
    return {"dealers": rows, "count": len(rows)}

_apex_data_cache: dict | None = None

def _load_apex_data() -> dict:
    global _apex_data_cache
    if _apex_data_cache is not None:
        return _apex_data_cache
    if os.path.exists(_APEX_FILE):
        try:
            with open(_APEX_FILE) as f:
                _apex_data_cache = json.load(f)
                return _apex_data_cache
        except Exception:
            pass
    _apex_data_cache = {"periods": {}}
    return _apex_data_cache

def _save_apex_data(data: dict):
    global _apex_data_cache
    _apex_data_cache = data          # update cache in-place so reads are instant
    dirpath = os.path.dirname(_APEX_FILE)
    if dirpath:
        os.makedirs(dirpath, exist_ok=True)
    with open(_APEX_FILE, "w") as f:
        json.dump(data, f, indent=2)

def _parse_upload(content: bytes, filename: str) -> "_pd.DataFrame":
    if _pd is None:
        raise HTTPException(500, "pandas not installed on server")
    fn = (filename or "").lower()
    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        return _pd.read_excel(_io.BytesIO(content))
    # CSV — try utf-8, fall back to latin-1
    try:
        return _pd.read_csv(_io.StringIO(content.decode("utf-8")))
    except UnicodeDecodeError:
        return _pd.read_csv(_io.StringIO(content.decode("latin-1")))

def _map_cols(df: "_pd.DataFrame", col_map: dict) -> dict:
    """Return {target: actual_col} for each target that could be matched."""
    lower_cols = {c.strip().lower(): c for c in df.columns}
    result = {}
    for target, options in col_map.items():
        for opt in options:
            if opt in lower_cols:
                result[target] = lower_cols[opt]
                break
    return result

def _df_to_rows(df: "_pd.DataFrame", mapped: dict) -> list:
    rows = []
    for _, row in df.iterrows():
        r = {}
        for target, col in mapped.items():
            val = row.get(col)
            try:
                if _pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            # Convert numpy types to plain Python
            if val is not None:
                try:
                    val = val.item()
                except AttributeError:
                    pass
            r[target] = val
        rows.append(r)
    return rows

@app.get("/apex-review")  # legacy redirect
async def apex_review_redirect():
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/strategic-partner-report", status_code=301)

@app.get("/strategic-partner-report")
async def apex_review_page(admin=Depends(_require_admin)):
    return FileResponse("static/reports/apex-review.html")

@app.get("/api/apex/periods")
async def apex_list_periods(admin=Depends(_require_admin)):
    data = _load_apex_data()
    periods_raw = data.get("periods", {})
    periods = [
        {
            "label": k,
            "type":  v.get("period_type", "quarterly" if k.startswith("Q") else "monthly"),
            "has_production": bool(v.get("production")),
            "has_rollup":     bool(v.get("rollup")),
        }
        for k, v in periods_raw.items()
    ]
    return {"periods": periods}

@app.get("/api/apex/data/{period:path}")
async def apex_get_period(period: str, admin=Depends(_require_admin)):
    data = _load_apex_data()
    period_data = data.get("periods", {}).get(period)
    if not period_data:
        raise HTTPException(404, "Period not found")
    return period_data

@app.post("/api/apex/upload/{period:path}/production")
async def apex_upload_production(
    period: str,
    file: UploadFile = File(...),
    admin=Depends(_require_admin),
):
    content = await file.read()
    df = _parse_upload(content, file.filename or "")
    col_map = {
        "dealer":    ["dealer", "dealer name", "name", "company", "contractor", "location"],
        "tenure":    ["tenure", "dealer tenure"],
        "apps":      ["apps", "applications", "app count", "total apps"],
        "approved":  ["approved", "approvals", "approved apps"],
        "pending":   ["pending", "pending apps"],
        "rpas":      ["rpas", "rpa", "rpa count", "total rpas"],
        "nia":       ["nia", "not interested", "not available"],
        "revenue":   ["revenue", "funded", "funded amount", "net invoice", "amount",
                      "funded revenue", "net funded", "total funded"],
    }
    mapped = _map_cols(df, col_map)
    if "dealer" not in mapped:
        raise HTTPException(400, f"Could not find dealer/name column. Columns found: {list(df.columns)}")
    rows = _df_to_rows(df, mapped)
    # strip empty rows
    rows = [r for r in rows if r.get("dealer")]

    data = _load_apex_data()
    data["periods"].setdefault(period, {})
    data["periods"][period]["production"] = rows
    data["periods"][period]["production_cols"] = mapped
    data["periods"][period]["production_uploaded_at"] = datetime.now().isoformat()
    _save_apex_data(data)
    return {"ok": True, "rows": len(rows), "columns_mapped": mapped}

@app.post("/api/apex/upload/{period:path}/rollup")
async def apex_upload_rollup(
    period: str,
    file: UploadFile = File(...),
    admin=Depends(_require_admin),
):
    content = await file.read()
    df = _parse_upload(content, file.filename or "")
    col_map = {
        "dealer":            ["dealer", "dealer name", "name", "company", "contractor", "location"],
        "enrolled_date":     ["enrolled date", "enrolled", "enrollment date", "activation date",
                              "partner activation date", "activated date"],
        "last_app_date":     ["last app date", "last app", "last application date", "last application"],
        "ttm_apps":          ["ttm apps", "ttm applications", "apps", "applications", "total apps"],
        "ttm_rpas":          ["ttm rpas", "ttm rpa", "rpas", "rpa", "total rpas"],
        "ttm_revenue":       ["ttm net invoice", "ttm revenue", "ttm funded", "net invoice",
                              "revenue", "funded", "total funded", "net funded"],
        "pre_approval_rate": ["pre-approval rate", "pre approval rate", "approval rate",
                              "ttm pre-approval rate", "ttm approval rate"],
        "take_up_rate":      ["take up rate", "takeup rate", "ttm take up rate",
                              "ttm takeup rate", "conversion rate"],
    }
    mapped = _map_cols(df, col_map)
    if "dealer" not in mapped:
        raise HTTPException(400, f"Could not find dealer/name column. Columns found: {list(df.columns)}")
    rows = _df_to_rows(df, mapped)
    rows = [r for r in rows if r.get("dealer")]

    data = _load_apex_data()
    data["periods"].setdefault(period, {})
    data["periods"][period]["rollup"] = rows
    data["periods"][period]["rollup_cols"] = mapped
    data["periods"][period]["rollup_uploaded_at"] = datetime.now().isoformat()
    _save_apex_data(data)
    return {"ok": True, "rows": len(rows), "columns_mapped": mapped}

@app.post("/api/apex/meta/{period:path}")
async def apex_save_meta(period: str, body: dict = Body(...), admin=Depends(_require_admin)):
    data = _load_apex_data()
    data["periods"].setdefault(period, {})
    for key in ("new_locations", "highlights", "discussion", "improvement"):
        if key in body:
            data["periods"][period][key] = body[key]
    data["periods"][period]["meta_updated_at"] = datetime.now().isoformat()
    _save_apex_data(data)
    return {"ok": True}

@app.delete("/api/apex/periods/{period:path}")
async def apex_delete_period(period: str, admin=Depends(_require_admin)):
    data = _load_apex_data()
    if period in data.get("periods", {}):
        del data["periods"][period]
        _save_apex_data(data)
    return {"ok": True}

@app.get("/api/apex/export/{period:path}/csv")
async def apex_export_csv(period: str, table: str = "production", admin=Depends(_require_admin)):
    data = _load_apex_data()
    period_data = data.get("periods", {}).get(period, {})
    rows = period_data.get(table, [])
    if not rows:
        raise HTTPException(404, "No data for this period/table")
    output = _io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    filename = f"apex_{table}_{period.replace(' ', '_')}.csv"
    return StreamingResponse(
        _io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# Admin Analytics Dashboards (admin-only)
# ══════════════════════════════════════════════════════════════════════════════

def _slp_get_field(slp: dict, field_id: str) -> str:
    """Extract a field value from a raw SLP record."""
    for f in slp.get("fields", []):
        if f.get("id") == field_id:
            return (f.get("value") or "").strip()
    return ""

def _parse_iso_date(s: str):
    """Parse an ISO8601 date string, returning a date object or None."""
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        return None

def _days_since(d) -> Optional[int]:
    """Return days since a date (date object or ISO string). None if unparseable."""
    if d is None:
        return None
    if isinstance(d, str):
        d = _parse_iso_date(d)
    if d is None:
        return None
    try:
        return (date.today() - d).days
    except Exception:
        return None


# ── Page route ────────────────────────────────────────────────────────────────

@app.get("/admin/dashboards")
async def admin_dashboards_page(admin=Depends(_require_admin)):
    return FileResponse("static/admin-dashboards.html")


# ── 1. Onboarding Pipeline ────────────────────────────────────────────────────

@app.get("/api/admin/dashboard/onboarding-pipeline")
async def dash_onboarding_pipeline(admin=Depends(_require_admin)):
    all_slps = list(await get_slp_cache())
    today = date.today()

    STATUS_ORDER = [
        "Not Started",
        "In Progress – Signed Contract Needed",
        "In Progress – Other",
        "Pending - Training Not Completed",
        "Pending - Waiting on Online Reviews",
        "Waiting_on_BDR_Approval",
        "On Indefinite Hold - Agreement/Documents Not Signed",
        "Account on Hold (Suspended)",
        "Contractor Activated",
        "Deactivated",
        "Deactivated for Dormancy",
        "Declined by Onboarding",
        "Not Active",
    ]

    buckets: dict = {s: {"count": 0, "days_list": [], "accounts": []} for s in STATUS_ORDER}
    buckets["(other)"] = {"count": 0, "days_list": [], "accounts": []}

    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        status = _slp_get_field(slp, "slp-status-detail") or "(other)"
        if status not in buckets:
            status = "(other)"

        activation_date = _slp_get_field(slp, "contractor-activated-date")
        created_raw = slp.get("cdate") or slp.get("created_at") or ""
        ref_date = _parse_iso_date(activation_date) or _parse_iso_date(created_raw)
        days = _days_since(ref_date)

        aid = get_account_id(slp)
        account_name = _account_to_name.get(str(aid), "") if aid else ""
        platform = _slp_get_field(slp, "platform")
        bdr = _slp_get_field(slp, "assigned-bdr")

        buckets[status]["count"] += 1
        if days is not None:
            buckets[status]["days_list"].append(days)
        buckets[status]["accounts"].append({
            "account_id":   str(aid) if aid else "",
            "account_name": account_name,
            "platform":     platform,
            "bdr":          bdr,
            "days_in_status": days,
            "activation_date": activation_date,
        })

    pipeline = []
    for status in STATUS_ORDER + ["(other)"]:
        b = buckets[status]
        if b["count"] == 0:
            continue
        dl = b["days_list"]
        pipeline.append({
            "status":       status,
            "count":        b["count"],
            "avg_days":     round(sum(dl) / len(dl)) if dl else None,
            "min_days":     min(dl) if dl else None,
            "max_days":     max(dl) if dl else None,
            "accounts":     sorted(b["accounts"], key=lambda x: (x["days_in_status"] or 0), reverse=True)[:50],
        })

    total = sum(b["count"] for b in buckets.values())
    activated = buckets.get("Contractor Activated", {}).get("count", 0)
    return {
        "total_slps":       total,
        "total_activated":  activated,
        "pipeline":         pipeline,
        "generated_at":     today.isoformat(),
    }


# ── 2. BDR Performance ────────────────────────────────────────────────────────

@app.get("/api/admin/dashboard/bdr-performance")
async def dash_bdr_performance(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    admin=Depends(_require_admin),
):
    all_slps = list(await get_slp_cache())
    today = date.today()

    # Default range: YTD
    from_d = _parse_iso_date(from_date) or date(today.year, 1, 1)
    to_d   = _parse_iso_date(to_date)   or today

    this_month_start = date(today.year, today.month, 1)
    this_q = ((today.month - 1) // 3) * 3 + 1
    this_q_start = date(today.year, this_q, 1)
    this_year_start = date(today.year, 1, 1)

    bdr_data: dict = {}

    def _get_bdr(bdr_raw: str) -> str:
        return bdr_raw.strip() if bdr_raw.strip() else "(Unassigned)"

    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        bdr = _get_bdr(_slp_get_field(slp, "assigned-bdr"))
        status = _slp_get_field(slp, "slp-status-detail")
        activation_date_str = _slp_get_field(slp, "contractor-activated-date")
        activation_date = _parse_iso_date(activation_date_str)

        if bdr not in bdr_data:
            bdr_data[bdr] = {
                "bdr": bdr,
                "activations_custom_range": 0,
                "activations_this_month":   0,
                "activations_this_quarter": 0,
                "activations_this_year":    0,
                "pipeline_count":           0,
                "waiting_approval":         0,
                "pending_statuses": [],
            }

        # Count activations in various windows
        if activation_date:
            if from_d <= activation_date <= to_d:
                bdr_data[bdr]["activations_custom_range"] += 1
            if activation_date >= this_month_start:
                bdr_data[bdr]["activations_this_month"] += 1
            if activation_date >= this_q_start:
                bdr_data[bdr]["activations_this_quarter"] += 1
            if activation_date >= this_year_start:
                bdr_data[bdr]["activations_this_year"] += 1

        # Pipeline (pending/in-progress)
        PIPELINE_STATUSES = {
            "Not Started", "In Progress – Signed Contract Needed", "In Progress – Other",
            "Pending - Training Not Completed", "Pending - Waiting on Online Reviews",
            "On Indefinite Hold - Agreement/Documents Not Signed",
        }
        if status in PIPELINE_STATUSES:
            bdr_data[bdr]["pipeline_count"] += 1

        if status == "Waiting_on_BDR_Approval":
            aid = get_account_id(slp)
            name = _account_to_name.get(str(aid), "") if aid else ""
            bdr_data[bdr]["waiting_approval"] += 1
            bdr_data[bdr]["pending_statuses"].append({
                "account_id":   str(aid) if aid else "",
                "account_name": name,
                "platform":     _slp_get_field(slp, "platform"),
            })

    results = sorted(bdr_data.values(), key=lambda x: x["activations_this_year"], reverse=True)
    return {
        "from_date":       from_d.isoformat(),
        "to_date":         to_d.isoformat(),
        "bdr_performance": results,
        "generated_at":    today.isoformat(),
    }


# ── 3. Dormancy Risk ──────────────────────────────────────────────────────────

@app.get("/api/admin/dashboard/dormancy-risk")
async def dash_dormancy_risk(
    threshold: int = Query(90, ge=1),
    admin=Depends(_require_admin),
):
    today = date.today()
    results = []

    # Use _account_to_last_app which is built from SLP data + CF140 fallback
    for aid, last_app_str in _account_to_last_app.items():
        if not last_app_str:
            continue
        d = _parse_iso_date(last_app_str)
        if d is None:
            continue
        days = (today - d).days
        if days < threshold:
            continue
        owner_uid = _account_to_owner.get(aid, "")
        results.append({
            "account_id":   aid,
            "account_name": _account_to_name.get(aid, f"Account {aid}"),
            "owner":        _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else "",
            "region":       _account_to_region.get(aid, ""),
            "platform":     _account_to_platform.get(aid, ""),
            "dealer_id":    _account_to_dealer.get(aid, ""),
            "last_app_date": last_app_str,
            "days_since_app": days,
        })

    results.sort(key=lambda x: x["days_since_app"], reverse=True)

    over60  = sum(1 for r in results if r["days_since_app"] > 60)
    over90  = sum(1 for r in results if r["days_since_app"] > 90)
    over180 = sum(1 for r in results if r["days_since_app"] > 180)

    return {
        "threshold_days": threshold,
        "total_at_risk":  len(results),
        "over_60_days":   over60,
        "over_90_days":   over90,
        "over_180_days":  over180,
        "accounts":       results,
        "generated_at":   today.isoformat(),
    }


# ── 4. AM Workload ────────────────────────────────────────────────────────────

@app.get("/api/admin/dashboard/am-workload")
async def dash_am_workload(admin=Depends(_require_admin)):
    today = date.today()
    STALE_DAYS = 90

    am_data: dict = {}

    for aid, name in _account_to_name.items():
        owner_uid = _account_to_owner.get(aid, "")
        owner_name = _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else "(Unassigned)"

        last_app_str = _account_to_last_app.get(aid, "")
        last_app_d   = _parse_iso_date(last_app_str)

        if last_app_d is None:
            activity_status = "never_active"
        elif (today - last_app_d).days > STALE_DAYS:
            activity_status = "stale"
        else:
            activity_status = "active"

        if owner_name not in am_data:
            am_data[owner_name] = {
                "am": owner_name,
                "total_accounts": 0,
                "active":         0,
                "stale":          0,
                "never_active":   0,
                "account_list":   [],
            }

        am_data[owner_name]["total_accounts"] += 1
        am_data[owner_name][activity_status]  += 1
        am_data[owner_name]["account_list"].append({
            "account_id":     aid,
            "account_name":   name,
            "activity_status": activity_status,
            "last_app_date":   last_app_str,
            "region":          _account_to_region.get(aid, ""),
        })

    results = sorted(am_data.values(), key=lambda x: x["total_accounts"], reverse=True)
    # Trim account lists to top 100 per AM to keep response size reasonable
    for r in results:
        r["account_list"] = sorted(r["account_list"], key=lambda a: a["last_app_date"] or "", reverse=True)[:100]

    return {
        "stale_threshold_days": STALE_DAYS,
        "total_accounts":        len(_account_to_name),
        "am_workload":           results,
        "generated_at":          today.isoformat(),
    }


# ── 5. Activation Trends ──────────────────────────────────────────────────────

@app.get("/api/admin/dashboard/activation-trends")
async def dash_activation_trends(admin=Depends(_require_admin)):
    all_slps = list(await get_slp_cache())
    today = date.today()

    # Build last 12 month labels
    months = []
    for i in range(11, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{m:02d}")

    # month_key → platform → count
    by_month: dict = {m: {} for m in months}
    months_set = set(months)

    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        if _slp_get_field(slp, "slp-status-detail") != "Contractor Activated":
            continue

        act_str = _slp_get_field(slp, "contractor-activated-date")
        if not act_str:
            continue

        d = _parse_iso_date(act_str)
        if d is None:
            continue

        month_key = f"{d.year}-{d.month:02d}"
        if month_key not in months_set:
            continue

        platform = _slp_get_field(slp, "platform") or "Unknown"
        by_month[month_key][platform] = by_month[month_key].get(platform, 0) + 1

    # All platforms seen
    all_platforms: set = set()
    for v in by_month.values():
        all_platforms |= set(v.keys())
    all_platforms = sorted(all_platforms)

    trend = []
    for m in months:
        row = {"month": m, "total": sum(by_month[m].values())}
        for p in all_platforms:
            row[p] = by_month[m].get(p, 0)
        trend.append(row)

    return {
        "months":       months,
        "platforms":    all_platforms,
        "trend":        trend,
        "generated_at": today.isoformat(),
    }


# ── 6. Channel / Program Mix ──────────────────────────────────────────────────

@app.get("/api/admin/dashboard/channel-mix")
async def dash_channel_mix(admin=Depends(_require_admin)):
    all_slps = list(await get_slp_cache())
    today = date.today()

    this_month = f"{today.year}-{today.month:02d}"
    prev_m = today.month - 1 or 12
    prev_y = today.year - (1 if today.month == 1 else 0)
    prev_month = f"{prev_y}-{prev_m:02d}"

    platform_data: dict = {}

    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        if _slp_get_field(slp, "slp-status-detail") != "Contractor Activated":
            continue

        platform = _slp_get_field(slp, "platform") or "Unknown"
        act_str  = _slp_get_field(slp, "contractor-activated-date")
        act_d    = _parse_iso_date(act_str)

        if platform not in platform_data:
            platform_data[platform] = {
                "platform":       platform,
                "total_activated": 0,
                "this_month":      0,
                "prev_month":      0,
            }

        platform_data[platform]["total_activated"] += 1
        if act_d:
            mk = f"{act_d.year}-{act_d.month:02d}"
            if mk == this_month:
                platform_data[platform]["this_month"] += 1
            elif mk == prev_month:
                platform_data[platform]["prev_month"] += 1

    results = sorted(platform_data.values(), key=lambda x: x["total_activated"], reverse=True)
    for r in results:
        prev = r["prev_month"]
        curr = r["this_month"]
        if prev > 0:
            r["mom_change_pct"] = round((curr - prev) / prev * 100, 1)
        else:
            r["mom_change_pct"] = None

    return {
        "this_month":   this_month,
        "prev_month":   prev_month,
        "channel_mix":  results,
        "generated_at": today.isoformat(),
    }


# ── 7. Dealer ID Integrity ────────────────────────────────────────────────────

@app.get("/api/admin/dashboard/dealer-id-integrity")
async def dash_dealer_id_integrity(admin=Depends(_require_admin)):
    all_slps = list(await get_slp_cache())
    today = date.today()

    # Build set of all SLP dealer IDs
    slp_dealer_ids: set = set()
    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)
        d = _slp_get_field(slp, "dealer-id")
        if d:
            slp_dealer_ids.add(d)

    missing_cf18: list = []
    mismatch: list = []
    dealer_id_count: dict = {}

    for aid, name in _account_to_name.items():
        cf18 = _account_to_dealer.get(aid, "").strip()
        owner_uid = _account_to_owner.get(aid, "")
        owner = _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else ""

        if not cf18:
            missing_cf18.append({
                "account_id":   aid,
                "account_name": name,
                "owner":        owner,
                "region":       _account_to_region.get(aid, ""),
            })
        else:
            # Track duplicates
            dealer_id_count[cf18] = dealer_id_count.get(cf18, [])
            dealer_id_count[cf18].append({"account_id": aid, "account_name": name, "owner": owner})

            # Check if CF18 matches any SLP dealer-id
            if cf18 not in slp_dealer_ids:
                mismatch.append({
                    "account_id":   aid,
                    "account_name": name,
                    "dealer_id":    cf18,
                    "owner":        owner,
                    "region":       _account_to_region.get(aid, ""),
                })

    duplicates = [
        {"dealer_id": did, "accounts": accts}
        for did, accts in dealer_id_count.items()
        if len(accts) > 1
    ]
    duplicates.sort(key=lambda x: len(x["accounts"]), reverse=True)

    return {
        "missing_dealer_id":        len(missing_cf18),
        "dealer_id_mismatch":       len(mismatch),
        "duplicate_dealer_ids":     len(duplicates),
        "missing_cf18_accounts":    sorted(missing_cf18, key=lambda x: x["account_name"])[:200],
        "mismatch_accounts":        sorted(mismatch, key=lambda x: x["account_name"])[:200],
        "duplicate_dealer_id_list": duplicates[:100],
        "generated_at":             today.isoformat(),
    }


# ── 8. Oracle ID Coverage ─────────────────────────────────────────────────────

@app.get("/api/admin/dashboard/oracle-coverage")
async def dash_oracle_coverage(admin=Depends(_require_admin)):
    all_slps = list(await get_slp_cache())
    today = date.today()

    missing_by_platform: dict = {}
    total_activated = 0
    total_missing   = 0

    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        if _slp_get_field(slp, "slp-status-detail") != "Contractor Activated":
            continue

        total_activated += 1
        oracle = _slp_get_field(slp, "oracle-producer-ids")
        if oracle:
            continue

        total_missing += 1
        platform = _slp_get_field(slp, "platform") or "Unknown"
        if platform not in missing_by_platform:
            missing_by_platform[platform] = {"platform": platform, "count": 0, "accounts": []}

        aid = get_account_id(slp)
        name = _account_to_name.get(str(aid), "") if aid else ""
        missing_by_platform[platform]["count"] += 1
        missing_by_platform[platform]["accounts"].append({
            "account_id":   str(aid) if aid else "",
            "account_name": name,
            "dealer_id":    _slp_get_field(slp, "dealer-id"),
            "bdr":          _slp_get_field(slp, "assigned-bdr"),
            "region":       _account_to_region.get(str(aid), "") if aid else "",
        })

    results = sorted(missing_by_platform.values(), key=lambda x: x["count"], reverse=True)
    pct_missing = round(total_missing / total_activated * 100, 1) if total_activated else 0

    return {
        "total_activated":      total_activated,
        "total_missing_oracle": total_missing,
        "pct_missing":          pct_missing,
        "by_platform":          results,
        "generated_at":         today.isoformat(),
    }


# ── 9. Contact Coverage ───────────────────────────────────────────────────────

@app.get("/api/admin/dashboard/contact-coverage")
async def dash_contact_coverage(admin=Depends(_require_admin)):
    today = date.today()

    # Fetch account contacts in parallel with contact details
    ac_contacts_raw, contacts_raw = await asyncio.gather(
        ac_get_all("accountContacts", "accountContacts", {"limit": 100}),
        ac_get_all("contacts", "contacts", {"limit": 100}),
    )

    # Map contact_id → email
    contact_email: dict = {}
    for c in contacts_raw:
        cid  = str(c.get("id", ""))
        em   = (c.get("email") or "").strip()
        contact_email[cid] = em

    # Map account_id → list of contact_ids
    acct_contacts: dict = defaultdict(list)
    for ac in ac_contacts_raw:
        aid = str(ac.get("account", "") or "")
        cid = str(ac.get("contact", "") or "")
        if aid and aid != "0" and cid and cid != "0":
            acct_contacts[aid].append(cid)

    no_contacts: list = []
    contacts_no_email: list = []

    for aid, name in _account_to_name.items():
        owner_uid = _account_to_owner.get(aid, "")
        owner = _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else ""
        region = _account_to_region.get(aid, "")
        cids = acct_contacts.get(aid, [])

        if not cids:
            no_contacts.append({
                "account_id":   aid,
                "account_name": name,
                "owner":        owner,
                "region":       region,
            })
        else:
            has_email = any(contact_email.get(cid, "") for cid in cids)
            if not has_email:
                contacts_no_email.append({
                    "account_id":     aid,
                    "account_name":   name,
                    "owner":          owner,
                    "region":         region,
                    "contact_count":  len(cids),
                })

    total_accounts = len(_account_to_name)
    return {
        "total_accounts":          total_accounts,
        "no_contacts_count":       len(no_contacts),
        "contacts_no_email_count": len(contacts_no_email),
        "no_contacts":             sorted(no_contacts, key=lambda x: x["account_name"])[:300],
        "contacts_no_email":       sorted(contacts_no_email, key=lambda x: x["account_name"])[:300],
        "generated_at":            today.isoformat(),
    }


# ── 10. New Account Velocity ──────────────────────────────────────────────────

@app.get("/api/admin/dashboard/new-account-velocity")
async def dash_new_account_velocity(
    days: int = Query(90, ge=1, le=365),
    admin=Depends(_require_admin),
):
    today = date.today()
    cutoff = today - timedelta(days=days)

    # Fetch all accounts with cdate
    all_accounts_raw = await ac_get_all("accounts", "accounts", {"limit": 100})

    # Build cdate map
    acct_cdate: dict = {}
    for a in all_accounts_raw:
        aid = str(a.get("id", ""))
        cdate_raw = a.get("cdate") or ""
        d = _parse_iso_date(cdate_raw)
        if d:
            acct_cdate[aid] = d

    # SLP map: account_id → earliest SLP creation date + earliest activation date
    all_slps = list(await get_slp_cache())
    acct_first_slp: dict = {}
    acct_first_activation: dict = {}

    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        aid = get_account_id(slp)
        if not aid:
            continue
        aid = str(aid)

        slp_cdate = _parse_iso_date(slp.get("cdate") or "")
        act_date  = _parse_iso_date(_slp_get_field(slp, "contractor-activated-date"))

        if slp_cdate:
            if aid not in acct_first_slp or slp_cdate < acct_first_slp[aid]:
                acct_first_slp[aid] = slp_cdate
        if act_date:
            if aid not in acct_first_activation or act_date < acct_first_activation[aid]:
                acct_first_activation[aid] = act_date

    new_accounts = []
    for aid, cdate in acct_cdate.items():
        if cdate < cutoff:
            continue
        first_slp  = acct_first_slp.get(aid)
        first_act  = acct_first_activation.get(aid)

        days_to_slp  = (first_slp  - cdate).days if first_slp  and first_slp  >= cdate else None
        days_to_act  = (first_act  - cdate).days if first_act  and first_act  >= cdate else None

        owner_uid = _account_to_owner.get(aid, "")
        new_accounts.append({
            "account_id":      aid,
            "account_name":    _account_to_name.get(aid, f"Account {aid}"),
            "created_date":    cdate.isoformat(),
            "days_ago":        (today - cdate).days,
            "owner":           _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else "",
            "region":          _account_to_region.get(aid, ""),
            "first_slp_date":  first_slp.isoformat() if first_slp else None,
            "first_act_date":  first_act.isoformat() if first_act else None,
            "days_to_first_slp": days_to_slp,
            "days_to_activation": days_to_act,
        })

    new_accounts.sort(key=lambda x: x["created_date"], reverse=True)

    days_to_slp_vals = [x["days_to_first_slp"] for x in new_accounts if x["days_to_first_slp"] is not None]
    days_to_act_vals = [x["days_to_activation"] for x in new_accounts if x["days_to_activation"] is not None]

    return {
        "window_days":          days,
        "new_accounts_count":   len(new_accounts),
        "with_slp_count":       sum(1 for x in new_accounts if x["first_slp_date"]),
        "activated_count":      sum(1 for x in new_accounts if x["first_act_date"]),
        "avg_days_to_slp":      round(sum(days_to_slp_vals)/len(days_to_slp_vals), 1) if days_to_slp_vals else None,
        "avg_days_to_act":      round(sum(days_to_act_vals)/len(days_to_act_vals), 1) if days_to_act_vals else None,
        "accounts":             new_accounts[:500],
        "generated_at":         today.isoformat(),
    }


# ── 11. Multi-Channel Opportunity ─────────────────────────────────────────────

@app.get("/api/admin/dashboard/multi-channel")
async def dash_multi_channel(
    base_platform: str = Query("360 Finance"),
    target_platform: str = Query("Microf"),
    admin=Depends(_require_admin),
):
    all_slps = list(await get_slp_cache())
    today = date.today()

    # Map account_id → set of activated platforms
    acct_platforms: dict = defaultdict(set)

    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        if _slp_get_field(slp, "slp-status-detail") != "Contractor Activated":
            continue

        aid = get_account_id(slp)
        if not aid:
            continue

        platform = _slp_get_field(slp, "platform")
        if platform:
            acct_platforms[str(aid)].add(platform)

    opportunities = []
    for aid, platforms in acct_platforms.items():
        # Has base_platform activated but NOT target_platform
        has_base   = any(base_platform.lower() in p.lower() for p in platforms)
        has_target = any(target_platform.lower() in p.lower() for p in platforms)
        if has_base and not has_target:
            owner_uid = _account_to_owner.get(aid, "")
            opportunities.append({
                "account_id":        aid,
                "account_name":      _account_to_name.get(aid, f"Account {aid}"),
                "owner":             _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else "",
                "region":            _account_to_region.get(aid, ""),
                "activated_platforms": sorted(platforms),
                "dealer_id":         _account_to_dealer.get(aid, ""),
            })

    opportunities.sort(key=lambda x: x["account_name"])

    # Build unique platform list from all activated SLPs
    all_platforms: set = set()
    seen_ids2: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids2:
            continue
        seen_ids2.add(sid)
        if _slp_get_field(slp, "slp-status-detail") == "Contractor Activated":
            p = _slp_get_field(slp, "platform")
            if p:
                all_platforms.add(p)

    return {
        "base_platform":    base_platform,
        "target_platform":  target_platform,
        "opportunity_count": len(opportunities),
        "opportunities":    opportunities[:500],
        "available_platforms": sorted(all_platforms),
        "generated_at":     today.isoformat(),
    }


# ── 12. Geographic Coverage ───────────────────────────────────────────────────

@app.get("/api/admin/dashboard/geographic")
async def dash_geographic(admin=Depends(_require_admin)):
    all_slps = list(await get_slp_cache())
    today = date.today()

    # Count activated dealers by state (using SLP doing-business-in-states field)
    state_counts: dict = {}

    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        if _slp_get_field(slp, "slp-status-detail") != "Contractor Activated":
            continue

        states_str = _slp_get_field(slp, "doing-business-in-states")
        if not states_str:
            # Fall back to account state_prov
            aid = get_account_id(slp)
            if aid:
                states_str = _account_to_state_prov.get(str(aid), "")

        if states_str:
            for st in states_str.replace(",", " ").split():
                st = st.strip().upper()
                if len(st) == 2 and st.isalpha():
                    state_counts[st] = state_counts.get(st, 0) + 1

    sorted_states = sorted(state_counts.items(), key=lambda x: x[1], reverse=True)

    return {
        "state_coverage":  [{"state": s, "count": c} for s, c in sorted_states],
        "total_states":    len(state_counts),
        "top_10_states":   [{"state": s, "count": c} for s, c in sorted_states[:10]],
        "bottom_10_states": [{"state": s, "count": c} for s, c in sorted_states[-10:]],
        "generated_at":    today.isoformat(),
    }


# ── 13. Training-to-Activation Funnel ────────────────────────────────────────

@app.get("/api/admin/dashboard/training-funnel")
async def dash_training_funnel(admin=Depends(_require_admin)):
    today = date.today()

    # Fetch training records
    training_records_raw = await ac_get_all(
        f"customObjects/records/{TRAINING_SCHEMA_ID}", "records", {}
    )

    # Build account → latest training date
    acct_training: dict = {}
    for r in training_records_raw:
        fmap = {f["id"]: (f.get("value") or "").strip() for f in r.get("fields", [])}
        train_date_str = fmap.get("date-of-training", "")
        train_date = _parse_iso_date(train_date_str)
        if not train_date:
            continue

        for aid in r.get("relationships", {}).get("account", []):
            aid = str(aid)
            if aid not in acct_training or train_date > acct_training[aid]["date"]:
                acct_training[aid] = {
                    "date":          train_date,
                    "date_str":      train_date.isoformat(),
                    "training_type": fmap.get("training-type", ""),
                    "trained_by":    fmap.get("trained-by", ""),
                }

    # Build account → first activation date from SLP cache
    all_slps = list(await get_slp_cache())
    acct_activation: dict = {}
    seen_ids: set = set()
    for slp in all_slps:
        sid = slp.get("id", "")
        if sid in seen_ids:
            continue
        seen_ids.add(sid)

        if _slp_get_field(slp, "slp-status-detail") != "Contractor Activated":
            continue

        aid = get_account_id(slp)
        if not aid:
            continue
        aid = str(aid)

        act_date = _parse_iso_date(_slp_get_field(slp, "contractor-activated-date"))
        if act_date:
            if aid not in acct_activation or act_date < acct_activation[aid]:
                acct_activation[aid] = act_date

    trained_and_activated  = []
    trained_not_activated  = []
    activated_not_trained  = []
    days_to_act_vals = []

    for aid, train_info in acct_training.items():
        owner_uid = _account_to_owner.get(aid, "")
        row = {
            "account_id":    aid,
            "account_name":  _account_to_name.get(aid, f"Account {aid}"),
            "owner":         _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else "",
            "region":        _account_to_region.get(aid, ""),
            "training_date": train_info["date_str"],
            "training_type": train_info["training_type"],
            "trained_by":    train_info["trained_by"],
        }
        act_d = acct_activation.get(aid)
        if act_d:
            days = (act_d - train_info["date"]).days
            row["activation_date"]     = act_d.isoformat()
            row["days_train_to_act"]   = days
            row["days_since_training"] = (today - train_info["date"]).days
            trained_and_activated.append(row)
            if days >= 0:
                days_to_act_vals.append(days)
        else:
            row["activation_date"]     = None
            row["days_train_to_act"]   = None
            row["days_since_training"] = (today - train_info["date"]).days
            trained_not_activated.append(row)

    # Activated but never trained
    for aid, act_d in acct_activation.items():
        if aid not in acct_training:
            owner_uid = _account_to_owner.get(aid, "")
            activated_not_trained.append({
                "account_id":       aid,
                "account_name":     _account_to_name.get(aid, f"Account {aid}"),
                "owner":            _user_id_to_name.get(owner_uid, owner_uid) if owner_uid else "",
                "region":           _account_to_region.get(aid, ""),
                "activation_date":  act_d.isoformat(),
            })

    trained_not_activated.sort(key=lambda x: x["days_since_training"], reverse=True)
    trained_and_activated.sort(key=lambda x: x["days_train_to_act"] or 0, reverse=True)

    return {
        "trained_and_activated_count": len(trained_and_activated),
        "trained_not_activated_count": len(trained_not_activated),
        "activated_not_trained_count": len(activated_not_trained),
        "avg_days_train_to_activation": round(sum(days_to_act_vals)/len(days_to_act_vals), 1) if days_to_act_vals else None,
        "trained_and_activated":  trained_and_activated[:200],
        "trained_not_activated":  trained_not_activated[:200],
        "activated_not_trained":  activated_not_trained[:200],
        "generated_at":           today.isoformat(),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
